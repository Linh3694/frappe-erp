# Copyright (c) 2026, Wellspring International School
# Export Excel thiết bị IT từ Frappe (sau cutover)

import io
import os
import tempfile

import frappe
from frappe import _
from frappe.utils import cint, now_datetime

from erp.api.erp_inventory.inventory_helpers import (
	DEVICE_TYPES,
	normalize_device_type,
	resolve_user_link,
	room_to_fe,
	user_to_fe,
)


STATUS_LABELS = {
	"Active": "Đang sử dụng",
	"Standby": "Sẵn sàng bàn giao",
	"Broken": "Hỏng",
	"PendingDocumentation": "Thiếu biên bản",
}

# Cột specs theo loại thiết bị: (header, spec_key)
SPEC_COLUMNS = {
	"laptop": [("Processor", "processor"), ("RAM", "ram"), ("Ổ cứng", "storage"), ("Màn hình", "display")],
	"monitor": [("Màn hình", "display")],
	"printer": [("Địa chỉ IP", "ip"), ("RAM", "ram"), ("Ổ cứng", "storage"), ("Màn hình", "display")],
	"projector": [],
	"phone": [("IMEI 1", "imei1"), ("IMEI 2", "imei2"), ("Số điện thoại", "phone_number")],
	"tool": [],
}

BASE_HEADERS = ["Tên thiết bị", "Loại thiết bị", "Hãng sản xuất", "Serial", "Năm sản xuất", "Trạng thái"]
TAIL_HEADERS = ["Người sử dụng", "Email", "Phòng"]


def _columns_for(dt):
	"""Bộ cột chuẩn dùng chung cho export + template import."""
	return BASE_HEADERS + [h for h, _ in SPEC_COLUMNS.get(dt, [])] + TAIL_HEADERS


def _device_specs_dict(doc):
	dt = doc.device_type
	table_map = {
		"laptop": "specs_laptop",
		"monitor": "specs_monitor",
		"printer": "specs_printer",
		"projector": "specs_projector",
		"phone": "specs_phone",
		"tool": "specs_tool",
	}
	rows = doc.get(table_map.get(dt)) or []
	if not rows:
		return {}
	row = rows[0]
	out = {}
	for fn in ("processor", "ram", "storage", "display", "ip", "imei1", "imei2", "phone_number"):
		if hasattr(row, fn):
			out[fn] = getattr(row, fn) or ""
	return out


@frappe.whitelist(allow_guest=False)
def export_devices_excel(device_type=None):
	"""Export Excel — parity exportController.exportDevices."""
	try:
		import openpyxl
		from openpyxl.styles import Alignment, Font, PatternFill
	except ImportError:
		frappe.throw(_("Thiếu openpyxl"))

	dt = normalize_device_type(device_type)
	devices = frappe.get_all(
		"ERP Inventory Device",
		filters={"device_type": dt},
		fields=["name"],
		order_by="name_display asc",
	)
	# Export tất cả thiết bị (kể cả chưa bàn giao)
	exported = [frappe.get_doc("ERP Inventory Device", d.name) for d in devices]

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = f"Danh sách {dt}"
	headers = ["STT"] + _columns_for(dt)
	ws.append(headers)
	header_fill = PatternFill(start_color="002855", end_color="002855", fill_type="solid")
	for cell in ws[1]:
		cell.font = Font(bold=True, color="FFFFFF")
		cell.fill = header_fill
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

	for idx, doc in enumerate(exported, start=1):
		# Ưu tiên người giữ hiện tại, fallback về assigned_users[0]
		holder_user_name = doc.current_holder_user
		if not holder_user_name and doc.assigned_users:
			holder_user_name = doc.assigned_users[0].user
		user = user_to_fe(holder_user_name) if holder_user_name else None
		room_code = frappe.db.get_value("ERP Administrative Room", doc.room, "physical_code") if doc.room else ""
		specs = _device_specs_dict(doc)
		row = [
			idx,
			doc.name_display,
			doc.device_subtype or doc.device_type,
			doc.manufacturer or "",
			doc.serial,
			doc.release_year or "",
			STATUS_LABELS.get(doc.status, doc.status),
		]
		for _, spec_key in SPEC_COLUMNS.get(dt, []):
			row.append(specs.get(spec_key, ""))
		row.extend([
			(user or {}).get("fullname", ""),
			(user or {}).get("email", ""),
			room_code or "",
		])
		ws.append(row)

	info_row = len(exported) + 3
	ws.cell(row=info_row, column=1, value=f"Tổng thiết bị đã export: {len(exported)}")

	tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
	wb.save(tmp.name)
	tmp.close()

	frappe.local.response.filename = f"thiet-bi-{dt}-{now_datetime().date()}.xlsx"
	frappe.local.response.filecontent = open(tmp.name, "rb").read()
	frappe.local.response.type = "download"
	os.unlink(tmp.name)
	return


@frappe.whitelist(allow_guest=False)
def download_import_template(device_type=None):
	"""Template import — parity getImportTemplate."""
	try:
		import openpyxl
	except ImportError:
		frappe.throw(_("Thiếu openpyxl"))

	dt = normalize_device_type(device_type)
	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = f"Nhập {dt}"
	cols = _columns_for(dt)
	ws.append(cols)

	# Dòng mẫu khớp đúng thứ tự cột
	sample_map = {
		"Tên thiết bị": "Tên thiết bị mẫu",
		"Loại thiết bị": "Laptop" if dt == "laptop" else "",
		"Hãng sản xuất": "Dell",
		"Serial": "SN-XXXXXX",
		"Năm sản xuất": 2024,
		"Trạng thái": "Standby",
		"Người sử dụng": "Nguyễn Văn A",
		"Email": "nguyen.van.a@wellspring.edu.vn",
		"Phòng": "P101",
	}
	ws.append([sample_map.get(c, "") for c in cols])

	guide = wb.create_sheet("Hướng dẫn")
	guide.append(["Cột", "Mô tả", "Bắt buộc", "Ghi chú"])
	guide.append(["Serial", "Số serial duy nhất", "Có", ""])
	guide.append(["Email", "Email người sử dụng — dùng để match tài khoản", "Không", "Để trống nếu chưa bàn giao"])
	guide.append(["Người sử dụng", "Tên hiển thị — chỉ để tham khảo", "Không", "KHÔNG dùng khi import, hệ thống match theo cột Email"])
	guide.append(["Phòng", "Mã phòng (physical_code)", "Không", "Map theo danh sách phòng trong hệ thống"])

	tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
	wb.save(tmp.name)
	tmp.close()
	frappe.local.response.filename = f"template-import-{dt}.xlsx"
	frappe.local.response.filecontent = open(tmp.name, "rb").read()
	frappe.local.response.type = "download"
	os.unlink(tmp.name)
