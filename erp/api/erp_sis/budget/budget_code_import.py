"""
Import Excel mã ngân sách — phòng ban dùng mã đơn vị (unit_code), nhiều giá trị cách nhau dấu phẩy.
"""

import os
import tempfile
from typing import Any, Dict, List, Optional, Tuple

import frappe
from frappe import _
from frappe.utils.xlsxutils import read_xlsx_file_from_attached_file

from erp.utils.api_response import error_response, success_response, validation_error_response

from .utils import CODE_DT, ORG_UNIT_DT, _is_finance

# Cột file mẫu (tiếng Việt)
IMPORT_COLUMNS = [
    "Mã ngân sách",
    "Khoản mục",
    "Mã cấp trên",
    "Đang hoạt động",
    "Phòng ban áp dụng",
]

_ROW_FIELD_KEYS = {
    "budget_code": ("Mã ngân sách", "budget_code", "Ma ngan sach"),
    "account_item": ("Khoản mục", "account_item", "Khoan muc"),
    "parent_budget_code": ("Mã cấp trên", "parent_budget_code", "Ma cap tren"),
    "is_active": ("Đang hoạt động", "is_active", "Trạng thái", "Trang thai"),
    "unit_codes": (
        "Phòng ban áp dụng",
        "Mã đơn vị",
        "unit_codes",
        "applicable_departments",
        "Phong ban ap dung",
    ),
}


def _cell_str(value: Any) -> str:
    """Chuẩn hoá ô Excel thành chuỗi — giữ số nguyên không thập phân."""
    if value is None:
        return ""
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    s = str(value).strip()
    return "" if s.lower() == "nan" else s


def _row_get(row: Dict[str, Any], field: str) -> str:
    for key in _ROW_FIELD_KEYS[field]:
        if key in row:
            val = _cell_str(row[key])
            if val:
                return val
    return ""


def _parse_bool_cell(value: Any, default: bool = True) -> bool:
    if value is None or _cell_str(value) == "":
        return default
    s = _cell_str(value).lower()
    if s in ("1", "true", "yes", "có", "co", "active", "hoạt động", "dang hoat dong"):
        return True
    if s in ("0", "false", "no", "không", "khong", "inactive", "ngưng", "ngung"):
        return False
    return default


def _import_excel_to_rows(file_content: bytes) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    data = read_xlsx_file_from_attached_file(fcontent=file_content)
    if not data:
        return rows
    header = [str(h).strip() if h is not None else "" for h in data[0]]
    for raw in data[1:]:
        if not any(raw):
            continue
        rows.append(
            {
                header[idx]: (raw[idx] if idx < len(raw) else None)
                for idx in range(len(header))
                if header[idx]
            }
        )
    return rows


def _resolve_code_doc_name(budget_code: str) -> Optional[str]:
    """Tra docname từ mã ngân sách (budget_code)."""
    if not budget_code:
        return None
    return frappe.db.get_value(CODE_DT, {"budget_code": budget_code}, "name")


def _resolve_unit_codes(raw: str) -> Tuple[List[str], List[str]]:
    """Parse mã đơn vị (unit_code) → danh sách docname phòng ban; trả thêm lỗi không tìm thấy."""
    if not raw:
        return [], []
    codes = [c.strip() for c in raw.split(",") if c.strip()]
    departments: List[str] = []
    missing: List[str] = []
    for code in codes:
        dept = frappe.db.get_value(ORG_UNIT_DT, {"unit_code": code}, "name")
        if dept:
            if dept not in departments:
                departments.append(dept)
        else:
            missing.append(code)
    return departments, missing


def _upsert_import_row(
    budget_code: str,
    account_item: str,
    parent_doc_name: Optional[str],
    is_active: bool,
    departments: List[str],
):
    """Tạo mới hoặc cập nhật 1 mã ngân sách từ dòng import."""
    existing_name = _resolve_code_doc_name(budget_code)
    doc = frappe.get_doc(CODE_DT, existing_name) if existing_name else frappe.new_doc(CODE_DT)

    doc.budget_code = budget_code
    doc.account_item = account_item or doc.account_item or ""
    doc.parent_budget_code = parent_doc_name
    doc.is_active = 1 if is_active else 0

    doc.set("applicable_departments", [])
    for dept in departments:
        doc.append("applicable_departments", {"department": dept})

    doc.save(ignore_permissions=True)


def _parse_import_row(row: Dict[str, Any], row_num: int) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
    """Parse 1 dòng Excel → payload hoặc thông báo lỗi."""
    budget_code = _row_get(row, "budget_code")
    if not budget_code:
        return None, f"Dòng {row_num}: thiếu Mã ngân sách"

    parent_code = _row_get(row, "parent_budget_code")
    unit_codes_raw = _row_get(row, "unit_codes")
    departments, missing_units = _resolve_unit_codes(unit_codes_raw)
    if missing_units:
        return None, (
            f"Dòng {row_num}: không tìm thấy mã đơn vị: {', '.join(missing_units)}"
        )

    is_active_raw = None
    for key in _ROW_FIELD_KEYS["is_active"]:
        if key in row and _cell_str(row[key]):
            is_active_raw = row[key]
            break

    return {
        "row_num": row_num,
        "budget_code": budget_code,
        "account_item": _row_get(row, "account_item"),
        "parent_code": parent_code,
        "is_active": _parse_bool_cell(is_active_raw, True),
        "departments": departments,
    }, None


@frappe.whitelist(allow_guest=False)
def download_budget_code_import_template():
    """Tải file mẫu Excel import mã ngân sách."""
    if not _is_finance():
        return error_response("Bạn không có quyền quản lý mã ngân sách")

    try:
        import openpyxl
    except ImportError:
        return error_response(_("Thiếu openpyxl để tạo file mẫu"))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mã ngân sách"
    ws.append(IMPORT_COLUMNS)

    # Dòng mẫu — mã dạng text để giữ số 0 đầu
    sample_rows = [
        ["02", "Nhóm chi phí hành chính tổng hợp", "", "Có", ""],
        ["0211", "CP Điện, nước, viễn thông", "02", "Có", ""],
        ["021101", "CP Điện", "0211", "Có", "TS"],
    ]
    for sample in sample_rows:
        ws.append(sample)

    # Ép cột mã sang text
    for row_idx in range(2, 2 + len(sample_rows)):
        for col_idx in (1, 3):  # Mã ngân sách, Mã cấp trên
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.number_format = "@"

    guide = wb.create_sheet("Hướng dẫn")
    guide.append(["Cột", "Mô tả", "Bắt buộc", "Ghi chú"])
    guide.append(["Mã ngân sách", "Mã duy nhất toàn hệ thống", "Có", "Nên định dạng Text trong Excel"])
    guide.append(["Khoản mục", "Tên khoản mục ngân sách", "Không", ""])
    guide.append(["Mã cấp trên", "Mã ngân sách của nhóm cha", "Không", "Để trống nếu là cấp 1"])
    guide.append(["Đang hoạt động", "Có / Không", "Không", "Mặc định: Có"])
    guide.append([
        "Phòng ban áp dụng",
        "Mã đơn vị (unit_code) từ Sơ đồ tổ chức",
        "Không",
        "Chỉ gán được ở mã cấp 4; nhiều mã cách nhau dấu phẩy (VD: TS, MKT)",
    ])

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    tmp.close()
    frappe.local.response.filename = "template-import-ma-ngan-sach.xlsx"
    frappe.local.response.filecontent = open(tmp.name, "rb").read()
    frappe.local.response.type = "download"
    os.unlink(tmp.name)


@frappe.whitelist(allow_guest=False)
def import_budget_codes_excel():
    """Import mã ngân sách từ file Excel (multipart field `file`)."""
    if not _is_finance():
        return error_response("Bạn không có quyền quản lý mã ngân sách")

    if not frappe.request or "file" not in frappe.request.files:
        return validation_error_response("Thiếu file Excel", {"file": ["required"]})

    file = frappe.request.files["file"]
    if not file:
        return validation_error_response("File rỗng", {"file": ["empty"]})

    try:
        rows = _import_excel_to_rows(file.stream.read())
    except Exception as ex:
        return validation_error_response(
            f"Không đọc được file Excel: {ex}", {"file": ["invalid"]}
        )

    if not rows:
        return validation_error_response("File không có dữ liệu", {"file": ["empty"]})

    parsed: List[Dict[str, Any]] = []
    errors: List[Dict[str, Any]] = []
    seen_codes: set[str] = set()

    for idx, row in enumerate(rows, start=2):
        payload, err = _parse_import_row(row, idx)
        if err:
            errors.append({"row": idx, "error": err})
            continue
        code_key = payload["budget_code"].lower()
        if code_key in seen_codes:
            errors.append({"row": idx, "error": f"Dòng {idx}: trùng mã '{payload['budget_code']}' trong file"})
            continue
        seen_codes.add(code_key)
        parsed.append(payload)

    success_count = 0

    # Import nhiều vòng để xử lý cha-con trong cùng file (tối đa 4 cấp)
    pending = list(parsed)
    for _pass in range(4):
        if not pending:
            break
        still_pending: List[Dict[str, Any]] = []
        for item in pending:
            parent_doc = None
            parent_code = item.get("parent_code") or ""
            if parent_code:
                parent_doc = _resolve_code_doc_name(parent_code)
                if not parent_doc:
                    still_pending.append(item)
                    continue
                if parent_doc == _resolve_code_doc_name(item["budget_code"]):
                    errors.append({
                        "row": item["row_num"],
                        "error": f"Dòng {item['row_num']}: mã không thể là cấp trên của chính nó",
                    })
                    continue

            try:
                _upsert_import_row(
                    budget_code=item["budget_code"],
                    account_item=item.get("account_item") or "",
                    parent_doc_name=parent_doc,
                    is_active=item.get("is_active", True),
                    departments=item.get("departments") or [],
                )
                success_count += 1
            except frappe.ValidationError as ex:
                errors.append({"row": item["row_num"], "error": f"Dòng {item['row_num']}: {ex}"})
            except Exception as ex:
                errors.append({"row": item["row_num"], "error": f"Dòng {item['row_num']}: {ex}"})
        pending = still_pending

    for item in pending:
        parent_code = item.get("parent_code") or ""
        errors.append({
            "row": item["row_num"],
            "error": f"Dòng {item['row_num']}: không tìm thấy mã cấp trên '{parent_code}'",
        })

    if success_count:
        frappe.db.commit()
    else:
        frappe.db.rollback()

    return success_response(
        data={
            "success_count": success_count,
            "error_count": len(errors),
            "total_count": len(rows),
            "errors": errors,
        },
        message=f"Đã import {success_count}/{len(parsed)} mã ngân sách",
    )
