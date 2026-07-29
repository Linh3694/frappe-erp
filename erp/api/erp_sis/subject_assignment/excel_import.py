# Copyright (c) 2026, Wellspring International School and contributors
# For license information, please see license.txt

"""
Nhập phân công giảng dạy từ Excel (SIS-164).

Luồng: parse -> resolve tên sang ID -> đối chiếu với dữ liệu hiện có -> ghi qua
batch_update_assignments. Cùng một endpoint phục vụ cả kiểm tra thử (dry_run=1) lẫn
ghi thật, để những gì người dùng thấy ở bước xem trước đúng bằng những gì sẽ chạy.

Quy tắc đối chiếu một dòng với dữ liệu hiện có, xét trong cùng (giáo viên, lớp, môn, năm học):
- Khớp đúng khoảng ngày            -> KHÔNG ĐỔI (bỏ qua, giữ tính idempotent)
- Đè lên đúng MỘT phân công đã có  -> CẬP NHẬT khoảng ngày của phân công đó
- Đè lên NHIỀU phân công           -> lỗi, không đoán bừa sửa cái nào
- Không đè cái nào                 -> TẠO MỚI (một đợt mới)

Nhờ vậy xuất file ra, sửa một ngày rồi nạp lại là sửa đúng đợt đó, còn thêm một dòng
không chồng ngày là thêm đợt mới.

⚠️ Chỉ THÊM và CẬP NHẬT. Không xoá phân công qua file — xoá kéo theo gỡ giáo viên khỏi
thời khoá biểu nên phải làm có chủ đích trên UI.
"""

from datetime import date, datetime
from io import BytesIO
from typing import Dict, List, Optional, Tuple

import frappe

from erp.utils.api_response import error_response, forbidden_response, single_item_response
from erp.utils.campus_utils import get_current_campus_from_context

from .assignment_validator import (
	fetch_existing_ranges,
	find_overlaps_in_batch,
	format_range,
	make_key,
	ranges_overlap,
	resolve_effective_range,
)
from .excel_template import (
	COL_CLASS,
	COL_END,
	COL_START,
	COL_SUBJECT,
	COL_TEACHER_CODE,
	COL_TEACHER_NAME,
	HEADERS,
	REQUIRED_COLUMNS,
	SHEET_DATA,
	normalize_header,
	normalize_lookup,
)
from .utils import build_teacher_code_lookup

# Giới hạn để một file hỏng (thừa hàng nghìn dòng rỗng có định dạng) không treo worker
MAX_ROWS = 5000


# ============================================================
# Parse
# ============================================================

def _coerce_date(value, label: str) -> Tuple[Optional[date], Optional[str]]:
	"""
	Đọc một ô ngày. Chấp nhận kiểu Ngày của Excel lẫn chuỗi người dùng gõ tay.

	Trả (date, None) hoặc (None, thông báo lỗi).
	"""
	if value is None or (isinstance(value, str) and not value.strip()):
		return None, None

	if isinstance(value, datetime):
		return value.date(), None
	if isinstance(value, date):
		return value, None

	text = str(value).strip()
	for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
		try:
			return datetime.strptime(text, fmt).date(), None
		except ValueError:
			continue

	return None, f"{label} '{text}' không đúng định dạng ngày (dd/mm/yyyy)"


def parse_workbook(stream: BytesIO) -> Tuple[List[Dict], List[str]]:
	"""
	Đọc sheet dữ liệu thành list dict, giữ nguyên SỐ DÒNG EXCEL THẬT để báo lỗi.

	Trả (rows, errors). errors ở đây là lỗi cấu trúc file, gặp là dừng luôn.
	"""
	from openpyxl import load_workbook

	errors: List[str] = []

	try:
		wb = load_workbook(stream, data_only=True, read_only=True)
	except Exception as e:
		return [], [f"Không đọc được file Excel: {str(e)}"]

	if SHEET_DATA in wb.sheetnames:
		ws = wb[SHEET_DATA]
	else:
		# Người dùng có thể đổi tên sheet; rơi về sheet đầu tiên thay vì từ chối thẳng.
		ws = wb[wb.sheetnames[0]]

	rows_iter = ws.iter_rows(values_only=True)
	try:
		header_row = next(rows_iter)
	except StopIteration:
		return [], ["File rỗng, không có dòng tiêu đề"]

	headers = [normalize_header(h) for h in header_row]
	position = {h: i for i, h in enumerate(headers) if h}

	missing = [h for h in HEADERS if h not in position]
	if missing:
		errors.append(
			"File thiếu cột bắt buộc: " + ", ".join(missing)
			+ ". Hãy tải lại file mẫu và giữ nguyên dòng tiêu đề."
		)
		return [], errors

	rows: List[Dict] = []
	for offset, raw in enumerate(rows_iter):
		excel_row = offset + 2  # +1 vì 0-based, +1 vì đã lấy mất dòng tiêu đề

		if excel_row - 1 > MAX_ROWS:
			errors.append(f"File vượt quá {MAX_ROWS} dòng dữ liệu, hãy tách nhỏ file.")
			break

		values = {h: (raw[i] if i < len(raw) else None) for h, i in position.items()}

		# Bỏ qua dòng trống hoàn toàn (Excel hay để lại hàng rỗng có định dạng)
		if all(
			v is None or (isinstance(v, str) and not v.strip())
			for v in values.values()
		):
			continue

		rows.append({"excel_row": excel_row, "values": values})

	wb.close()
	return rows, errors


# ============================================================
# Resolve
# ============================================================

def build_class_lookup(campus_id: str, school_year_id: str, education_stage_id: str) -> Dict:
	"""
	Bảng tra tên lớp -> class_id, trong đúng campus + năm học + cấp học.

	Trả kèm `other_year` và `other_stage` để phân biệt "không có lớp này" với
	"lớp có tồn tại nhưng thuộc năm học / cấp học khác" — hai lỗi này cần câu thông báo
	khác nhau, nếu gộp làm một người dùng sẽ đi tìm sai chỗ.
	"""
	rows = frappe.db.sql(
		"""
		SELECT
			c.name, c.title, c.short_title, c.school_year_id,
			eg.education_stage_id
		FROM `tabSIS Class` c
		LEFT JOIN `tabSIS Education Grade` eg ON eg.name = c.education_grade
		WHERE c.campus_id = %(campus_id)s
		""",
		{"campus_id": campus_id},
		as_dict=True,
	)

	exact: Dict[str, str] = {}
	other_year: Dict[str, str] = {}
	other_stage: Dict[str, str] = {}

	for row in rows:
		titles = {normalize_lookup(row.get("short_title")), normalize_lookup(row.get("title"))}
		titles.discard("")

		in_year = row.get("school_year_id") == school_year_id
		in_stage = row.get("education_stage_id") == education_stage_id

		for title in titles:
			if in_year and in_stage:
				exact[title] = row["name"]
			elif in_stage and not in_year:
				other_year.setdefault(title, row.get("school_year_id") or "")
			elif in_year and not in_stage:
				other_stage.setdefault(title, row.get("education_stage_id") or "")

	return {"exact": exact, "other_year": other_year, "other_stage": other_stage}


def build_subject_lookup(campus_id: str, education_stage_id: str) -> Dict:
	"""Bảng tra tên môn -> actual_subject_id trong campus + cấp học."""
	rows = frappe.get_all(
		"SIS Actual Subject",
		filters={"campus_id": campus_id},
		fields=["name", "title_vn", "education_stage_id"],
	)

	exact: Dict[str, str] = {}
	other_stage: Dict[str, str] = {}
	for row in rows:
		title = normalize_lookup(row.get("title_vn"))
		if not title:
			continue
		if row.get("education_stage_id") == education_stage_id:
			exact[title] = row["name"]
		else:
			other_stage.setdefault(title, row.get("education_stage_id") or "")

	return {"exact": exact, "other_stage": other_stage}


def resolve_rows(
	rows: List[Dict],
	campus_id: str,
	school_year_id: str,
	education_stage_id: str,
) -> Tuple[List[Dict], List[Dict]]:
	"""
	Đổi tên trong file thành ID hệ thống.

	Trả (resolved, errors). Mỗi lỗi là {"row": <số dòng excel>, "message": str}.
	"""
	teacher_lookup = build_teacher_code_lookup(campus_id)
	class_lookup = build_class_lookup(campus_id, school_year_id, education_stage_id)
	subject_lookup = build_subject_lookup(campus_id, education_stage_id)

	resolved: List[Dict] = []
	errors: List[Dict] = []

	for row in rows:
		excel_row = row["excel_row"]
		values = row["values"]
		row_errors: List[str] = []

		# --- Cột bắt buộc ---
		for column in REQUIRED_COLUMNS:
			raw = values.get(column)
			if raw is None or not str(raw).strip():
				row_errors.append(f"Thiếu '{column}'")

		if row_errors:
			errors.append({"row": excel_row, "message": "; ".join(row_errors)})
			continue

		# --- Giáo viên ---
		code = normalize_lookup(values.get(COL_TEACHER_CODE))
		teacher_id = teacher_lookup.get(code)
		if not teacher_id:
			row_errors.append(
				f"Mã GV '{str(values.get(COL_TEACHER_CODE)).strip()}' không tìm thấy "
				f"trong cơ sở này (hoặc trùng ở nhiều giáo viên)"
			)

		# --- Lớp ---
		class_title = normalize_lookup(values.get(COL_CLASS))
		class_id = class_lookup["exact"].get(class_title)
		if not class_id:
			if class_title in class_lookup["other_year"]:
				row_errors.append(
					f"Lớp '{str(values.get(COL_CLASS)).strip()}' thuộc năm học khác "
					f"({class_lookup['other_year'][class_title]}) — kiểm tra lại Năm học ở bước 1"
				)
			elif class_title in class_lookup["other_stage"]:
				row_errors.append(
					f"Lớp '{str(values.get(COL_CLASS)).strip()}' không thuộc cấp học đã chọn "
					f"— mỗi file chỉ dành cho một cấp học"
				)
			else:
				row_errors.append(f"Không tìm thấy lớp '{str(values.get(COL_CLASS)).strip()}'")

		# --- Môn ---
		subject_title = normalize_lookup(values.get(COL_SUBJECT))
		subject_id = subject_lookup["exact"].get(subject_title)
		if not subject_id:
			if subject_title in subject_lookup["other_stage"]:
				row_errors.append(
					f"Môn '{str(values.get(COL_SUBJECT)).strip()}' không thuộc cấp học đã chọn"
				)
			else:
				row_errors.append(f"Không tìm thấy môn '{str(values.get(COL_SUBJECT)).strip()}'")

		# --- Ngày ---
		start_date, start_error = _coerce_date(values.get(COL_START), COL_START)
		end_date, end_error = _coerce_date(values.get(COL_END), COL_END)
		if start_error:
			row_errors.append(start_error)
		if end_error:
			row_errors.append(end_error)
		if start_date and end_date and end_date < start_date:
			row_errors.append("Ngày kết thúc sớm hơn ngày bắt đầu")
		if end_date and not start_date:
			row_errors.append("Có ngày kết thúc thì phải có ngày bắt đầu")

		if row_errors:
			errors.append({"row": excel_row, "message": "; ".join(row_errors)})
			continue

		resolved.append({
			"index": excel_row,
			"excel_row": excel_row,
			"teacher_id": teacher_id,
			"teacher_label": str(values.get(COL_TEACHER_NAME) or values.get(COL_TEACHER_CODE) or ""),
			"class_id": class_id,
			"actual_subject_id": subject_id,
			"school_year_id": school_year_id,
			"campus_id": campus_id,
			# Không có ngày nào = cả năm, đúng quy ước của template
			"application_type": "full_year" if not start_date else "from_date",
			"start_date": start_date,
			"end_date": end_date,
		})

	return resolved, errors


# ============================================================
# Đối chiếu với dữ liệu hiện có
# ============================================================

def classify_rows(resolved: List[Dict]) -> Dict:
	"""
	Phân loại từng dòng thành create / update / unchanged / error.

	Xem docstring đầu module cho luật đối chiếu.
	"""
	errors: List[Dict] = []

	# Chồng lấn ngay giữa các dòng trong chính file
	in_file_conflicts = find_overlaps_in_batch(resolved)
	conflicted_rows = set()
	for conflict in in_file_conflicts:
		conflicted_rows.add(conflict["index"])
		errors.append({
			"row": conflict["index"],
			"message": (
				f"Khoảng {conflict['range']} chồng lên dòng {conflict['other_index']} "
				f"({conflict['other_range']}) — cùng giáo viên, lớp và môn"
			),
		})

	usable = [row for row in resolved if row["index"] not in conflicted_rows]

	buckets = fetch_existing_ranges(
		[
			make_key(
				row["teacher_id"],
				row["class_id"],
				row["actual_subject_id"],
				row["school_year_id"],
			)
			for row in usable
		],
		campus_id=usable[0]["campus_id"] if usable else None,
	)

	to_create: List[Dict] = []
	to_update: List[Dict] = []
	unchanged: List[Dict] = []

	for row in usable:
		key = make_key(
			row["teacher_id"],
			row["class_id"],
			row["actual_subject_id"],
			row["school_year_id"],
		)
		start, end = resolve_effective_range(
			row["start_date"],
			row["end_date"],
			row["application_type"],
			row["school_year_id"],
		)
		row["_start"], row["_end"] = start, end

		existing = buckets.get(key, [])
		exact = [e for e in existing if e["_start"] == start and e["_end"] == end]
		if exact:
			row["assignment_id"] = exact[0]["name"]
			unchanged.append(row)
			continue

		overlapping = [
			e for e in existing if ranges_overlap(start, end, e["_start"], e["_end"])
		]

		if len(overlapping) == 1:
			# Sửa ngày của đúng đợt đang đè -> cập nhật, không tạo thêm
			row["assignment_id"] = overlapping[0]["name"]
			to_update.append(row)
		elif len(overlapping) > 1:
			detail = ", ".join(
				format_range(e["_start"], e["_end"], row["school_year_id"])
				for e in overlapping
			)
			errors.append({
				"row": row["excel_row"],
				"message": (
					f"Khoảng {format_range(start, end, row['school_year_id'])} đè lên "
					f"{len(overlapping)} phân công đã có ({detail}) — không xác định được "
					f"cần sửa cái nào. Hãy sửa trực tiếp trên màn hình chi tiết giáo viên."
				),
			})
		else:
			to_create.append(row)

	return {
		"create": to_create,
		"update": to_update,
		"unchanged": unchanged,
		"errors": errors,
	}


# ============================================================
# Ghi
# ============================================================

def _payload(row: Dict, action: str) -> Dict:
	"""Một phần tử assignments cho batch_update_assignments."""
	payload = {
		"action": action,
		"class_id": row["class_id"],
		"actual_subject_id": row["actual_subject_id"],
		"school_year_id": row["school_year_id"],
		"application_type": row["application_type"],
		"start_date": row["start_date"].isoformat() if row.get("start_date") else None,
		"end_date": row["end_date"].isoformat() if row.get("end_date") else None,
	}
	if action == "update":
		payload["assignment_id"] = row["assignment_id"]
	return payload


def execute_import(classified: Dict, on_conflict: str = "skip") -> Dict:
	"""
	Ghi các dòng create/update, gom theo giáo viên.

	Mỗi giáo viên là một transaction riêng (batch_update_assignments tự begin/commit),
	nên một giáo viên hỏng không kéo đổ cả file.

	Xung đột giáo viên trên thời khoá biểu (một tiết đã đủ hai giáo viên):
	- on_conflict="overwrite": thử lại với replace_teacher_map khoá theo subject_id, đẩy
	  giáo viên mới vào slot 1. Khoá theo subject_id là cố ý — nó sống sót qua vòng
	  rollback/retry, còn khoá theo row_id thì không (xem timetable_sync_v2:280).
	- on_conflict="skip": bỏ các dòng thuộc môn bị xung đột rồi thử lại.
	"""
	from .batch_operations import _batch_update_assignments_internal

	by_teacher: Dict[str, List[Dict]] = {}
	for row in classified["create"]:
		by_teacher.setdefault(row["teacher_id"], []).append((row, "create"))
	for row in classified["update"]:
		by_teacher.setdefault(row["teacher_id"], []).append((row, "update"))

	stats = {"created": 0, "updated": 0, "skipped_conflict": 0, "failed": 0}
	logs: List[str] = []
	failures: List[Dict] = []

	for teacher_id, items in by_teacher.items():
		assignments = [_payload(row, action) for row, action in items]

		result = _batch_update_assignments_internal(teacher_id, assignments)

		if not result.get("success") and result.get("error_type") == "teacher_conflict":
			conflicts = result.get("conflicts", [])
			conflict_subjects = {
				c.get("conflict_key") for c in conflicts if c.get("conflict_key")
			}

			if on_conflict == "overwrite":
				replace_map = {subject: "teacher_1" for subject in conflict_subjects}
				result = _batch_update_assignments_internal(
					teacher_id, assignments, replace_teacher_map=replace_map
				)
				logs.append(
					f"Giáo viên {teacher_id}: ghi đè {len(conflict_subjects)} môn bị xung đột"
				)
			else:
				kept = [
					(row, action)
					for row, action in items
					if row["actual_subject_id"] not in conflict_subjects
				]
				dropped = len(items) - len(kept)
				stats["skipped_conflict"] += dropped
				logs.append(
					f"Giáo viên {teacher_id}: bỏ qua {dropped} dòng do xung đột thời khoá biểu"
				)
				for row, _action in items:
					if row["actual_subject_id"] in conflict_subjects:
						failures.append({
							"row": row["excel_row"],
							"message": "Bỏ qua do tiết học đã đủ giáo viên (xung đột thời khoá biểu)",
						})
				if not kept:
					continue
				assignments = [_payload(row, action) for row, action in kept]
				result = _batch_update_assignments_internal(teacher_id, assignments)

		if result.get("success"):
			teacher_stats = result.get("stats", {})
			stats["created"] += teacher_stats.get("created", 0)
			stats["updated"] += teacher_stats.get("updated", 0)
		else:
			stats["failed"] += len(assignments)
			message = result.get("message") or "Lỗi không xác định"
			logs.append(f"Giáo viên {teacher_id}: {message}")
			for row, _action in items:
				failures.append({"row": row["excel_row"], "message": message})

	return {"stats": stats, "logs": logs, "failures": failures}


# ============================================================
# Endpoint
# ============================================================

def _form_value(name: str, default=None):
	data = frappe.local.form_dict or {}
	value = data.get(name)
	if value is None and hasattr(frappe.request, "form"):
		value = frappe.request.form.get(name)
	return value if value not in (None, "") else default


@frappe.whitelist(allow_guest=False, methods=["POST"])
def import_subject_assignments():
	"""
	Nhập phân công giảng dạy từ file Excel.

	FormData:
		file                 (bắt buộc) .xlsx theo template
		campus_id            mặc định lấy theo context
		school_year_id       (bắt buộc)
		education_stage_id   (bắt buộc) — mỗi file một cấp học
		dry_run              "true" để chỉ kiểm tra, không ghi
		on_conflict          "skip" (mặc định) | "overwrite"
	"""
	try:
		campus_id = _form_value("campus_id") or get_current_campus_from_context()
		school_year_id = _form_value("school_year_id")
		education_stage_id = _form_value("education_stage_id")
		dry_run = str(_form_value("dry_run", "false")).lower() in ("1", "true", "yes")
		on_conflict = str(_form_value("on_conflict", "skip")).lower()

		if on_conflict not in ("skip", "overwrite"):
			return error_response(
				"on_conflict chỉ nhận 'skip' hoặc 'overwrite'", code="VALIDATION_ERROR"
			)

		if not all([campus_id, school_year_id, education_stage_id]):
			return error_response(
				"Thiếu tham số: cần campus_id, school_year_id và education_stage_id",
				code="VALIDATION_ERROR",
			)

		user_campus = get_current_campus_from_context()
		if user_campus and user_campus != campus_id:
			return forbidden_response("Access denied: Campus mismatch")

		files = frappe.request.files
		if not files or "file" not in files:
			return error_response("Chưa chọn file", code="VALIDATION_ERROR")

		stream = BytesIO(files["file"].read())

		rows, structure_errors = parse_workbook(stream)
		if structure_errors:
			return error_response(
				"File không đúng định dạng",
				errors={"errors": structure_errors},
				code="VALIDATION_ERROR",
			)

		resolved, resolve_errors = resolve_rows(
			rows, campus_id, school_year_id, education_stage_id
		)
		classified = classify_rows(resolved)

		all_errors = sorted(
			resolve_errors + classified["errors"], key=lambda e: e.get("row") or 0
		)

		summary = {
			"total_rows": len(rows),
			"valid_rows": len(classified["create"])
			+ len(classified["update"])
			+ len(classified["unchanged"]),
			"to_create": len(classified["create"]),
			"to_update": len(classified["update"]),
			"unchanged": len(classified["unchanged"]),
			"error_rows": len(all_errors),
			"errors": all_errors,
			"dry_run": dry_run,
			"on_conflict": on_conflict,
		}

		if dry_run:
			return single_item_response(summary, "Kiểm tra file hoàn tất")

		if not classified["create"] and not classified["update"]:
			return single_item_response(
				{**summary, "stats": {"created": 0, "updated": 0}},
				"Không có thay đổi nào để ghi",
			)

		outcome = execute_import(classified, on_conflict=on_conflict)

		return single_item_response(
			{
				**summary,
				"stats": outcome["stats"],
				"logs": outcome["logs"],
				# Lỗi phát sinh lúc ghi (xung đột thời khoá biểu, batch lỗi) gộp chung
				# vào danh sách lỗi để FE chỉ phải hiển thị một chỗ.
				"errors": sorted(
					all_errors + outcome["failures"], key=lambda e: e.get("row") or 0
				),
			},
			"Nhập phân công hoàn tất",
		)

	except Exception as e:
		frappe.log_error(f"Import subject assignments failed: {str(e)}")
		return error_response(f"Lỗi khi nhập file: {str(e)}")
