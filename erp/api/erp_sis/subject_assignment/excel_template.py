# Copyright (c) 2026, Wellspring International School and contributors
# For license information, please see license.txt

"""
Template Excel cho phân công giảng dạy (SIS-162).

Một dòng = một bản ghi SIS Subject Assignment:

    Họ tên giáo viên | Mã GV | Môn học | Lớp | Ngày bắt đầu | Ngày kết thúc

Quy ước ngày:
- Trống cả hai       -> cả năm học (application_type = full_year)
- Có bắt đầu, trống kết thúc -> từ ngày đó đến hết năm học
- Có cả hai          -> đúng khoảng đó

⚠️ Export và import dùng CHUNG file này. Đổi thứ tự/tên cột ở đây là đổi cho cả hai chiều;
đừng khai lại danh sách cột ở excel_export.py hay excel_import.py, nếu không file xuất ra
sẽ không nạp lại được bằng chính chức năng import.
"""

from typing import Dict, List, Optional

# ===== Cấu trúc file =====

SHEET_DATA = "Phân công"
SHEET_GUIDE = "Hướng dẫn"
SHEET_CATALOG = "Danh mục"

COL_TEACHER_NAME = "Họ tên giáo viên"
COL_TEACHER_CODE = "Mã GV"
COL_SUBJECT = "Môn học"
COL_CLASS = "Lớp"
COL_START = "Ngày bắt đầu"
COL_END = "Ngày kết thúc"

HEADERS = [
	COL_TEACHER_NAME,
	COL_TEACHER_CODE,
	COL_SUBJECT,
	COL_CLASS,
	COL_START,
	COL_END,
]

# Cột bắt buộc phải có giá trị thì dòng mới xử lý được
REQUIRED_COLUMNS = [COL_TEACHER_CODE, COL_SUBJECT, COL_CLASS]

DATE_FORMAT = "DD/MM/YYYY"
FONT_NAME = "Arial"

COLUMN_WIDTHS = {
	COL_TEACHER_NAME: 26,
	COL_TEACHER_CODE: 14,
	COL_SUBJECT: 36,
	COL_CLASS: 12,
	COL_START: 16,
	COL_END: 16,
}

GUIDE_RULES = [
	"1. MỖI DÒNG = MỘT phân công (một giáo viên dạy một môn cho một lớp trong một khoảng thời gian).",
	"   Giáo viên dạy nhiều môn / nhiều lớp thì lặp lại nhiều dòng.",
	"2. Bắt buộc: Mã GV, Môn học, Lớp. Thiếu một trong ba thì dòng đó bị loại.",
	"3. ĐỂ TRỐNG cả 'Ngày bắt đầu' và 'Ngày kết thúc' = phân công CẢ NĂM HỌC.",
	"4. Có 'Ngày bắt đầu', trống 'Ngày kết thúc' = áp dụng từ ngày đó đến hết năm học.",
	"5. Có cả hai = phân công theo đợt, áp dụng đúng trong khoảng đó.",
	"6. Cùng một giáo viên CÓ THỂ dạy cùng lớp + cùng môn ở nhiều đợt rời nhau — cứ tách",
	"   thành nhiều dòng. Các đợt KHÔNG được chồng ngày lên nhau.",
	"7. Định dạng ngày: dd/mm/yyyy. Nên nhập bằng kiểu Ngày của Excel, không nhập dạng chữ.",
	"8. Ngày dạy trong tuần mặc định là TẤT CẢ các ngày có tiết của môn đó. Muốn giới hạn",
	"   theo thứ thì sửa trên màn hình chi tiết giáo viên sau khi nhập.",
	"9. Tên Môn học và Lớp phải khớp danh mục ở sheet 'Danh mục' (đúng dấu tiếng Việt,",
	"   đúng dấu '/' và '&'). Không phân biệt hoa/thường.",
	"10. Giáo viên chưa có tài khoản trên hệ thống thì phải tạo trước khi nhập file.",
	"11. Mỗi file chỉ dành cho MỘT cấp học, đúng cấp đã chọn ở bước 1.",
]

GUIDE_WARNING = (
	"LƯU Ý: Nhập file chỉ THÊM và CẬP NHẬT phân công. Việc XOÁ phân công phải làm trực tiếp "
	"trên màn hình chi tiết giáo viên, vì xoá phân công kéo theo gỡ giáo viên khỏi thời khoá biểu."
)


def normalize_header(value) -> str:
	"""Chuẩn hoá tiêu đề cột khi đọc file: bỏ khoảng trắng thừa, gộp khoảng trắng liên tiếp."""
	return " ".join(str(value or "").split())


def normalize_lookup(value) -> str:
	"""Khoá tra cứu cho tên lớp / tên môn / mã GV: trim + gộp khoảng trắng + lowercase."""
	return " ".join(str(value or "").split()).lower()


def build_workbook(
	rows: List[Dict],
	meta: Dict,
	catalog: Optional[Dict[str, List[str]]] = None,
):
	"""
	Dựng workbook theo template.

	Args:
		rows: list dict theo đúng HEADERS (giá trị ngày là date hoặc None)
		meta: {"campus": str, "school_year": str, "education_stage": str, "exported_at": str}
		catalog: {"Lớp hợp lệ": [...], "Môn học hợp lệ": [...], "Giáo viên hợp lệ (Mã GV)": [...]}

	Returns:
		openpyxl.Workbook
	"""
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
	from openpyxl.utils import get_column_letter

	hdr_fill = PatternFill("solid", fgColor="1F3864")
	key_fill = PatternFill("solid", fgColor="E7E6E6")
	body_fill = PatternFill("solid", fgColor="FFF9E0")
	note_fill = PatternFill("solid", fgColor="FFFF00")
	thin = Side(style="thin", color="BFBFBF")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)

	wb = Workbook()

	# ----- Sheet dữ liệu -----
	ws = wb.active
	ws.title = SHEET_DATA

	for col, header in enumerate(HEADERS, start=1):
		cell = ws.cell(row=1, column=col, value=header)
		cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
		cell.fill = hdr_fill
		cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
		cell.border = border

	for row_idx, row in enumerate(rows, start=2):
		for col_idx, header in enumerate(HEADERS, start=1):
			cell = ws.cell(row=row_idx, column=col_idx, value=row.get(header))
			cell.font = Font(name=FONT_NAME, size=10)
			cell.border = border
			# Hai cột đầu là khoá nhận diện giáo viên -> tô khác vùng nhập liệu
			cell.fill = key_fill if col_idx <= 2 else body_fill
			if header in (COL_START, COL_END):
				cell.number_format = DATE_FORMAT
				cell.alignment = Alignment(horizontal="center", vertical="center")
			else:
				cell.alignment = Alignment(horizontal="left", vertical="center")

	ws.freeze_panes = "A2"
	ws.row_dimensions[1].height = 30
	for col, header in enumerate(HEADERS, start=1):
		ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTHS.get(header, 18)
	if rows:
		ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{len(rows) + 1}"

	# ----- Sheet hướng dẫn -----
	guide = wb.create_sheet(SHEET_GUIDE)
	guide["A1"] = "PHÂN CÔNG GIẢNG DẠY — FILE NHẬP LIỆU"
	guide["A1"].font = Font(name=FONT_NAME, size=14, bold=True, color="1F3864")

	line = 3
	guide.cell(row=line, column=1, value="Thông tin file (hệ thống điền — không sửa)").font = Font(
		name=FONT_NAME, size=11, bold=True
	)
	line += 1
	for label, key in (
		("Cơ sở", "campus"),
		("Năm học", "school_year"),
		("Cấp học", "education_stage"),
		("Ngày xuất file", "exported_at"),
	):
		guide.cell(row=line, column=1, value=label).font = Font(name=FONT_NAME, size=10, bold=True)
		guide.cell(row=line, column=2, value=meta.get(key) or "").font = Font(
			name=FONT_NAME, size=10, color="808080"
		)
		line += 1

	line += 1
	guide.cell(row=line, column=1, value="Chú giải màu").font = Font(name=FONT_NAME, size=11, bold=True)
	line += 1
	for fill, text in (
		(hdr_fill, "Dòng tiêu đề — KHÔNG sửa tên cột, KHÔNG đổi thứ tự, KHÔNG thêm cột lạ."),
		(key_fill, "Cột khoá — Mã GV là khoá nhận diện giáo viên. Họ tên chỉ để đối chiếu."),
		(body_fill, "Vùng cần điền."),
	):
		swatch = guide.cell(row=line, column=1, value="")
		swatch.fill = fill
		swatch.border = border
		guide.cell(row=line, column=2, value=text).font = Font(name=FONT_NAME, size=10)
		line += 1

	line += 1
	guide.cell(row=line, column=1, value="Quy tắc điền").font = Font(name=FONT_NAME, size=11, bold=True)
	line += 1
	for text in GUIDE_RULES:
		cell = guide.cell(row=line, column=1, value=text)
		cell.font = Font(name=FONT_NAME, size=10)
		cell.alignment = Alignment(vertical="center", wrap_text=True)
		guide.merge_cells(start_row=line, start_column=1, end_row=line, end_column=6)
		guide.row_dimensions[line].height = 17
		line += 1

	line += 1
	warn = guide.cell(row=line, column=1, value=GUIDE_WARNING)
	warn.font = Font(name=FONT_NAME, size=10, bold=True)
	warn.fill = note_fill
	warn.alignment = Alignment(vertical="center", wrap_text=True)
	guide.merge_cells(start_row=line, start_column=1, end_row=line + 1, end_column=6)
	guide.row_dimensions[line].height = 28

	guide.column_dimensions["A"].width = 30
	guide.column_dimensions["B"].width = 62
	for letter in ("C", "D", "E", "F"):
		guide.column_dimensions[letter].width = 12

	# ----- Sheet danh mục -----
	cat = wb.create_sheet(SHEET_CATALOG)
	col = 1
	max_len = 0
	for title, values in (catalog or {}).items():
		header = cat.cell(row=1, column=col, value=title)
		header.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
		header.fill = hdr_fill
		header.alignment = Alignment(horizontal="center", vertical="center")
		header.border = border
		for idx, value in enumerate(values, start=2):
			cell = cat.cell(row=idx, column=col, value=value)
			cell.font = Font(name=FONT_NAME, size=10)
			cell.border = border
		cat.column_dimensions[get_column_letter(col)].width = 36
		max_len = max(max_len, len(values))
		col += 2  # chừa một cột trống giữa các khối cho dễ đọc

	note = cat.cell(
		row=max_len + 3,
		column=1,
		value="Danh mục sinh theo Cơ sở / Năm học / Cấp học đã chọn. Không sửa tay.",
	)
	note.font = Font(name=FONT_NAME, size=9, italic=True, color="808080")

	return wb
