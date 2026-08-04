# -*- coding: utf-8 -*-
"""
Nhập Excel hàng loạt cho các danh mục Bus.

Chạy đồng bộ và trả kết quả ngay: dữ liệu danh mục chỉ ở mức vài trăm dòng nên
không cần job nền như erp/api/bulk_import.py.
"""

from io import BytesIO
from typing import Any, Callable, Dict, List, Optional, Tuple

import frappe
from openpyxl import load_workbook

from erp.utils.api_response import error_response, success_response, validation_error_response
from erp.utils.campus_utils import get_current_campus_from_context

from .bus_import_columns import (
    BUS_IMPORT_SPECS,
    BUS_WEEKDAYS,
    ImportSpec,
    friendly_unique_error,
    missing_headers,
    normalize_cell,
    parse_pickup_order,
    parse_row,
    weekly_trip_types,
)

# Số dòng lỗi / bỏ qua tối đa gửi về UI — tránh trả về danh sách quá dài
MAX_PREVIEW_ROWS = 200


class DuplicateHeaderError(ValueError):
    """File Excel có tiêu đề cột bị lặp."""

    def __init__(self, headers: List[str]):
        self.headers = headers
        super().__init__(", ".join(headers))


def _resolve_campus() -> str:
    """Campus của người dùng hiện tại; giữ fallback campus-1 như các API Bus khác."""
    return get_current_campus_from_context() or "campus-1"


def _campus_filter(campus_id: str) -> List[Optional[str]]:
    """Nhận cả bản ghi chưa gán campus.

    Danh mục Bus có bản ghi cũ để trống `campus_id` (xem `get_all_bus_drivers`), lọc chặt
    theo campus sẽ báo "không tìm thấy" dù bản ghi vẫn nằm đó.
    """
    return ["in", [campus_id, None, ""]]


def _read_rows(
    file_content: bytes, spec: ImportSpec
) -> Tuple[List[str], List[Tuple[int, Dict[str, Any]]]]:
    """Đọc sheet phù hợp trong file .xlsx → tiêu đề và dữ liệu kèm số dòng Excel."""
    workbook = load_workbook(BytesIO(file_content), data_only=True)
    worksheets = workbook.worksheets
    if not worksheets:
        workbook.close()
        return [], []

    selected_sheet = worksheets[0]
    for worksheet in worksheets:
        first_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        candidate_headers = [str(cell).strip() if cell is not None else "" for cell in first_row]
        if not missing_headers(spec, candidate_headers):
            selected_sheet = worksheet
            break

    data = [list(row) for row in selected_sheet.iter_rows(values_only=True)]
    workbook.close()
    if not data:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in data[0]]
    seen_headers = set()
    duplicate_headers = []
    for header in headers:
        if not header:
            continue
        if header in seen_headers and header not in duplicate_headers:
            duplicate_headers.append(header)
        seen_headers.add(header)
    if duplicate_headers:
        raise DuplicateHeaderError(duplicate_headers)

    rows: List[Tuple[int, Dict[str, Any]]] = []
    for row_num, raw in enumerate(data[1:], start=2):
        if not any(normalize_cell(cell) for cell in raw):
            continue
        rows.append(
            (
                row_num,
                {
                    headers[idx]: (raw[idx] if idx < len(raw) else None)
                    for idx in range(len(headers))
                    if headers[idx]
                },
            )
        )
    return headers, rows


def _existing_doc_name(spec: ImportSpec, values: Dict[str, str], campus_id: str) -> Optional[str]:
    """Bản ghi đã có theo khóa chống trùng của danh mục (trong cùng campus)."""
    filters: Dict[str, Any] = {"campus_id": campus_id}
    for field in spec.dedupe_fields:
        value = values.get(field)
        if not value:
            return None
        filters[field] = value
    return frappe.db.get_value(spec.doctype, filters, "name")


def _validation_message(ex: Exception) -> Optional[str]:
    """Câu thông báo của `frappe.throw` trong validate của doctype, đã bỏ thẻ HTML.

    Chỉ nhận ValidationError: lỗi hệ thống (SQL, thiếu quyền, lập trình) vẫn phải rơi
    xuống nhánh generic + log_error, không đưa nguyên văn ra cho người dùng.
    """
    if not isinstance(ex, frappe.exceptions.ValidationError):
        return None
    message = frappe.utils.strip_html(str(ex) or "").strip()
    return message or None


def _run_import(
    spec: ImportSpec,
    rows: List[Tuple[int, Dict[str, Any]]],
    campus_id: str,
    row_handler: Optional[Callable[[Dict[str, str], int, str], Tuple[Any, Optional[str], bool]]] = None,
) -> Dict[str, Any]:
    """
    Duyệt từng dòng: chuẩn hóa, bỏ qua dòng trùng, tạo bản ghi, gom kết quả.

    ``row_handler`` trả về một payload, hoặc danh sách payload khi một dòng Excel
    tương ứng nhiều bản ghi. Số đếm thành công luôn tính theo DÒNG.
    """
    errors: List[Dict[str, Any]] = []
    skipped: List[Dict[str, Any]] = []
    success_count = 0
    seen_keys = set()

    for row_num, row in rows:
        values, parse_error = parse_row(spec, row, row_num)
        if parse_error:
            errors.append({"row": row_num, "error": parse_error})
            continue

        dedupe_key = tuple(values.get(f, "") for f in spec.dedupe_fields)
        if dedupe_key in seen_keys:
            skipped.append({"row": row_num, "error": f"Dòng {row_num}: trùng với một dòng khác trong file"})
            continue
        seen_keys.add(dedupe_key)

        if row_handler:
            payload, handler_error, should_skip = row_handler(values, row_num, campus_id)
            if should_skip:
                skipped.append({"row": row_num, "error": handler_error or f"Dòng {row_num}: đã tồn tại"})
                continue
            if handler_error:
                errors.append({"row": row_num, "error": handler_error})
                continue
        else:
            if _existing_doc_name(spec, values, campus_id):
                skipped.append({"row": row_num, "error": f"Dòng {row_num}: đã tồn tại trong hệ thống"})
                continue
            payload = dict(values)
            payload["campus_id"] = campus_id

        # row_handler có thể trả nhiều payload cho một dòng (vd. học sinh trong tuyến:
        # một dòng Excel nở ra T2–T6 × lượt Đón/Trả). Cả nhóm nằm trong cùng savepoint
        # nên một dòng hoặc vào trọn vẹn, hoặc không vào gì.
        payloads = payload if isinstance(payload, list) else [payload]

        save_point = f"bus_import_row_{row_num}"
        frappe.db.savepoint(save_point=save_point)
        try:
            for item in payloads:
                doc = frappe.get_doc({"doctype": spec.doctype, **item})
                doc.insert()
            success_count += 1
        except Exception as ex:  # noqa: BLE001 — gom mọi lỗi tạo doc về mức dòng
            frappe.db.rollback(save_point=save_point)
            duplicate_message = friendly_unique_error(str(ex), spec, row_num)
            validation_message = _validation_message(ex)
            if duplicate_message:
                skipped.append({"row": row_num, "error": duplicate_message})
            elif validation_message:
                # Doctype tự chặn bằng frappe.throw — câu tiếng Việt đó nêu đúng lý do,
                # hữu ích hơn nhiều so với câu generic bên dưới
                errors.append({"row": row_num, "error": f"Dòng {row_num}: {validation_message}"})
            else:
                frappe.log_error(
                    title=f"Lỗi import Excel {spec.doctype}, dòng {row_num}",
                    message=frappe.get_traceback(),
                )
                errors.append(
                    {
                        "row": row_num,
                        "error": (
                            f"Dòng {row_num}: không tạo được bản ghi, "
                            "vui lòng kiểm tra lại dữ liệu dòng này"
                        ),
                    }
                )

    frappe.db.commit()

    return {
        "total_rows": len(rows),
        "success_count": success_count,
        "error_count": len(errors),
        "skipped_count": len(skipped),
        "errors_preview": errors[:MAX_PREVIEW_ROWS],
        "skipped_preview": skipped[:MAX_PREVIEW_ROWS],
        "errors_truncated": len(errors) > MAX_PREVIEW_ROWS,
        "skipped_truncated": len(skipped) > MAX_PREVIEW_ROWS,
    }


def _summary_message(spec: ImportSpec, result: Dict[str, Any]) -> str:
    parts = [f"Đã nhập {result['success_count']} {spec.entity_label}"]
    if result["skipped_count"]:
        parts.append(f"bỏ qua {result['skipped_count']} dòng đã có")
    if result["error_count"]:
        parts.append(f"{result['error_count']} dòng lỗi")
    return ", ".join(parts)


def _import_by_key(spec_key: str, row_handler=None):
    """Khung xử lý chung cho mọi endpoint import Bus."""
    spec = BUS_IMPORT_SPECS[spec_key]

    if not frappe.request or "file" not in frappe.request.files:
        return validation_error_response("Thiếu file Excel", {"file": ["required"]})

    uploaded = frappe.request.files["file"]
    if not uploaded:
        return validation_error_response("File rỗng", {"file": ["empty"]})

    try:
        headers, rows = _read_rows(uploaded.stream.read(), spec)
    except DuplicateHeaderError as ex:
        return validation_error_response(
            f"File có cột bị lặp: {', '.join(ex.headers)}",
            {"file": ["duplicate_columns"]},
        )
    except Exception:  # noqa: BLE001
        # Chi tiết lỗi thư viện đọc file chỉ để tra cứu, không đưa ra cho người dùng
        frappe.log_error(
            title=f"Lỗi đọc file Excel import {spec.doctype}",
            message=frappe.get_traceback(),
        )
        return validation_error_response(
            "File Excel không đúng định dạng hoặc bị hỏng", {"file": ["invalid"]}
        )

    if not rows:
        return validation_error_response("File không có dữ liệu", {"file": ["empty"]})

    lacking = missing_headers(spec, headers)
    if lacking:
        return validation_error_response(
            f"File thiếu cột bắt buộc: {', '.join(lacking)}", {"file": ["missing_columns"]}
        )

    try:
        result = _run_import(spec, rows, _resolve_campus(), row_handler=row_handler)
    except Exception:  # noqa: BLE001
        frappe.db.rollback()
        frappe.log_error(
            title=f"Lỗi import Excel {spec.doctype}",
            message=frappe.get_traceback(),
        )
        return error_response("Nhập Excel thất bại, vui lòng thử lại hoặc liên hệ quản trị")

    return success_response(data=result, message=_summary_message(spec, result))


@frappe.whitelist()
def import_bus_drivers():
    """Nhập danh sách tài xế từ Excel."""
    return _import_by_key("driver")


@frappe.whitelist()
def import_bus_monitors():
    """Nhập danh sách giám sát từ Excel."""
    return _import_by_key("monitor")


@frappe.whitelist()
def import_bus_transportations():
    """Nhập danh sách phương tiện từ Excel."""
    return _import_by_key("transportation")


@frappe.whitelist()
def import_bus_pickup_points():
    """Nhập danh sách điểm đón từ Excel."""
    return _import_by_key("pickup_point")


def _lookup_student_for_bus(student_code: str, campus_id: str) -> Optional[Dict[str, Any]]:
    """Hồ sơ học sinh gốc + lớp + năm học, tra theo mã học sinh trong campus."""
    found = frappe.db.sql(
        """
        SELECT
            s.name AS student_id,
            s.student_name AS full_name,
            s.student_code,
            cs.name AS class_student_id,
            cs.class_id,
            cs.school_year_id
        FROM `tabCRM Student` s
        INNER JOIN `tabSIS Class Student` cs ON s.name = cs.student_id
        WHERE s.student_code = %s
            AND s.campus_id = %s
            AND cs.class_type = 'regular'
        ORDER BY cs.creation DESC, cs.name DESC
        LIMIT 1
        """,
        (student_code, campus_id),
        as_dict=True,
    )
    return found[0] if found else None


def _student_row_handler(
    values: Dict[str, str], row_num: int, campus_id: str
) -> Tuple[Optional[Dict[str, Any]], Optional[str], bool]:
    """Dựng payload SIS Bus Student từ mã học sinh; tên tuyến (nếu có) tra sang route_id."""
    student_code = values.get("student_code", "")

    if frappe.db.exists("SIS Bus Student", {"student_code": student_code, "campus_id": campus_id}):
        return None, f"Dòng {row_num}: học sinh {student_code} đã có trong danh sách xe buýt", True

    student = _lookup_student_for_bus(student_code, campus_id)
    if not student:
        return (
            None,
            f"Dòng {row_num}: không tìm thấy học sinh có mã '{student_code}' đã được xếp lớp trong campus",
            False,
        )

    payload: Dict[str, Any] = {
        "full_name": student["full_name"],
        "student_code": student["student_code"],
        "class_id": student["class_id"],
        "status": values.get("status") or "Active",
        "campus_id": campus_id,
        "school_year_id": student["school_year_id"],
    }

    route_name = values.get("route_name", "")
    if route_name:
        # Ưu tiên tuyến của năm học học sinh đang theo — tuyến các năm hay trùng tên
        route_id = frappe.db.get_value(
            "SIS Bus Route",
            {
                "route_name": route_name,
                "campus_id": _campus_filter(campus_id),
                "school_year_id": student["school_year_id"],
            },
            "name",
        ) or frappe.db.get_value(
            "SIS Bus Route",
            {"route_name": route_name, "campus_id": _campus_filter(campus_id)},
            "name",
        )
        if not route_id:
            return None, f"Dòng {row_num}: không tìm thấy tuyến đường '{route_name}'", False
        payload["route_id"] = route_id

    return payload, None, False


@frappe.whitelist()
def import_bus_students():
    """Nhập danh sách học sinh đi xe buýt từ Excel (đối chiếu mã học sinh với CRM Student)."""
    return _import_by_key("student", row_handler=_student_row_handler)


def _resolve_school_year(campus_id: str) -> Optional[str]:
    """Năm học đang bật của campus; không có thì lấy năm học đang bật bất kỳ."""
    for filters in ({"campus_id": campus_id, "is_enable": 1}, {"is_enable": 1}):
        rows = frappe.get_all(
            "SIS School Year", filters=filters, fields=["name"], order_by="creation desc", limit=1
        )
        if rows:
            return rows[0].name
    return None


def _route_row_handler(
    values: Dict[str, str], row_num: int, campus_id: str
) -> Tuple[Any, Optional[str], bool]:
    """Dựng payload SIS Bus Route: biển số → xe, CCCD → nhân sự."""
    school_year_id = _resolve_school_year(campus_id)
    if not school_year_id:
        return None, f"Dòng {row_num}: chưa có năm học nào đang hoạt động", False

    route_name = values.get("route_name", "")
    # Chỉ coi là trùng khi cùng NĂM HỌC — tuyến năm trước thường trùng tên với năm nay
    if frappe.db.exists(
        "SIS Bus Route",
        {
            "route_name": route_name,
            "campus_id": _campus_filter(campus_id),
            "school_year_id": school_year_id,
        },
    ):
        return (
            None,
            f"Dòng {row_num}: tuyến '{route_name}' đã có trong năm học này",
            True,
        )

    license_plate = values.get("license_plate", "")
    vehicle_id = frappe.db.get_value(
        "SIS Bus Transportation",
        {"license_plate": license_plate, "campus_id": _campus_filter(campus_id)},
        "name",
    )
    if not vehicle_id:
        return None, f"Dòng {row_num}: không tìm thấy xe có biển số '{license_plate}'", False

    driver_citizen_id = values.get("driver_citizen_id", "")
    driver_id = frappe.db.get_value(
        "SIS Bus Driver",
        {"citizen_id": driver_citizen_id, "campus_id": _campus_filter(campus_id)},
        "name",
    )
    if not driver_id:
        return None, f"Dòng {row_num}: không tìm thấy tài xế có CCCD '{driver_citizen_id}'", False

    monitor_ids: List[str] = []
    for index, field in enumerate(("monitor1_citizen_id", "monitor2_citizen_id"), start=1):
        citizen_id = values.get(field, "")
        if not citizen_id:
            monitor_ids.append("")
            continue
        monitor_id = frappe.db.get_value(
            "SIS Bus Monitor",
            {"citizen_id": citizen_id, "campus_id": _campus_filter(campus_id)},
            "name",
        )
        if not monitor_id:
            return (
                None,
                f"Dòng {row_num}: không tìm thấy giám sát {index} có CCCD '{citizen_id}'",
                False,
            )
        monitor_ids.append(monitor_id)

    payload: Dict[str, Any] = {
        "route_name": route_name,
        "vehicle_code": values.get("vehicle_code", ""),
        "vehicle_id": vehicle_id,
        "driver_id": driver_id,
        "monitor1_id": monitor_ids[0],
        "monitor2_id": monitor_ids[1] or None,
        "status": values.get("status") or "Active",
        "campus_id": campus_id,
        "school_year_id": school_year_id,
    }
    return payload, None, False


def _route_student_row_handler(
    values: Dict[str, str], row_num: int, campus_id: str
) -> Tuple[Any, Optional[str], bool]:
    """
    Một dòng Excel → các bản ghi SIS Bus Route Student cho cả tuần học.

    Cấu hình chung cho cả tuần: nở ra T2–T6; lượt Đón/Trả suy từ việc cột
    'Điểm đón' / 'Điểm trả' có được điền hay không.
    """
    school_year_id = _resolve_school_year(campus_id)
    if not school_year_id:
        return None, f"Dòng {row_num}: chưa có năm học nào đang hoạt động", False

    route_name = values.get("route_name", "")
    # Phải bám năm học đang hoạt động, nếu không học sinh sẽ chui vào tuyến năm trước
    # (tuyến các năm hay trùng tên nhau)
    route_id = frappe.db.get_value(
        "SIS Bus Route",
        {
            "route_name": route_name,
            "campus_id": _campus_filter(campus_id),
            "school_year_id": school_year_id,
        },
        "name",
    )
    if not route_id:
        return (
            None,
            f"Dòng {row_num}: không tìm thấy tuyến đường '{route_name}' trong năm học này",
            False,
        )

    student_code = values.get("student_code", "")
    student = _lookup_student_for_bus(student_code, campus_id)
    if not student:
        return (
            None,
            f"Dòng {row_num}: không tìm thấy học sinh có mã '{student_code}' đã được xếp lớp trong campus",
            False,
        )

    if frappe.db.exists(
        "SIS Bus Route Student", {"route_id": route_id, "student_id": student["student_id"]}
    ):
        return (
            None,
            f"Dòng {row_num}: học sinh {student_code} đã có trong tuyến '{route_name}'",
            True,
        )

    trip_types, trip_error = weekly_trip_types(values, row_num)
    if trip_error:
        return None, trip_error, False

    pickup_order, order_error = parse_pickup_order(values.get("pickup_order", ""), row_num)
    if order_error:
        return None, order_error, False

    pickup_location = (values.get("pickup_location") or "").strip()
    drop_off_location = (values.get("drop_off_location") or "").strip()

    payloads: List[Dict[str, Any]] = []
    for weekday in BUS_WEEKDAYS:
        for trip_type in trip_types:
            payloads.append(
                {
                    "route_id": route_id,
                    "campus_id": campus_id,
                    "student_id": student["student_id"],
                    "class_student_id": student.get("class_student_id") or "",
                    "weekday": weekday,
                    "trip_type": trip_type,
                    "pickup_order": pickup_order,
                    # Lượt Đón đưa học sinh tới trường, lượt Trả xuất phát từ trường —
                    # giống mặc định của form thêm tuyến trên giao diện.
                    "pickup_location": pickup_location if trip_type == "Đón" else "Trường học",
                    "drop_off_location": drop_off_location if trip_type == "Trả" else "Trường học",
                }
            )
    return payloads, None, False


@frappe.whitelist()
def import_bus_routes():
    """Nhập danh sách tuyến đường từ Excel (chưa gồm học sinh)."""
    return _import_by_key("route", row_handler=_route_row_handler)


@frappe.whitelist()
def import_bus_route_students():
    """Nhập học sinh vào tuyến từ Excel — cấu hình chung cho cả tuần (T2–T6)."""
    return _import_by_key("route_student", row_handler=_route_student_row_handler)
