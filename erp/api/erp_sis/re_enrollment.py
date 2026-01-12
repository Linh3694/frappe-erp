"""
Admin Re-enrollment API
Handles re-enrollment management for admin/admission staff

API endpoints cho admin quản lý cấu hình và đơn tái ghi danh.
"""

import frappe
from frappe import _
from frappe.utils import nowdate, getdate, now
import json
import os
import requests
import uuid
from erp.utils.api_response import (
    validation_error_response, 
    list_response, 
    error_response, 
    success_response, 
    single_item_response,
    not_found_response
)

# URL pdf-service (có thể cấu hình trong site_config.json)
PDF_SERVICE_URL = frappe.conf.get("pdf_service_url", "http://172.16.20.113:5020")


def _check_admin_permission():
    """Kiểm tra quyền admin"""
    user_roles = frappe.get_roles(frappe.session.user)
    allowed_roles = ['System Manager', 'SIS Manager', 'Registrar', 'SIS BOD']
    
    if not any(role in user_roles for role in allowed_roles):
        return False
    return True


def _resolve_campus_id(campus_id):
    """
    Chuyển đổi campus_id từ format frontend (campus-1) sang format database (CAMPUS-00001)
    """
    if not campus_id:
        return None
    
    # Nếu đã đúng format CAMPUS-xxxxx thì return luôn
    if campus_id.startswith("CAMPUS-"):
        if frappe.db.exists("SIS Campus", campus_id):
            return campus_id
    
    # Nếu là format campus-1, campus-2, etc.
    if campus_id.startswith("campus-"):
        try:
            campus_index = int(campus_id.split("-")[1])
            mapped_campus = f"CAMPUS-{campus_index:05d}"
            if frappe.db.exists("SIS Campus", mapped_campus):
                return mapped_campus
        except (ValueError, IndexError):
            pass
    
    # Thử tìm theo name trực tiếp
    if frappe.db.exists("SIS Campus", campus_id):
        return campus_id
    
    # Thử tìm campus đầu tiên
    first_campus = frappe.db.get_value("SIS Campus", {}, "name", order_by="creation asc")
    if first_campus:
        return first_campus
    
    return None


def _auto_create_student_records(config_id, source_school_year_id, campus_id, logs=None, finance_year_id=None):
    """
    Tự động tạo re-enrollment records cho tất cả học sinh.
    Có thể lấy từ SIS Class Student hoặc SIS Finance Student.
    
    Args:
        config_id: ID của config tái ghi danh
        source_school_year_id: Năm học nguồn (năm học hiện tại của học sinh)
        campus_id: Campus ID
        logs: List để ghi log
        finance_year_id: (Optional) Nếu có, lấy học sinh từ SIS Finance Student thay vì SIS Class Student
    
    Điều kiện:
        - Nếu finance_year_id: lấy từ SIS Finance Student
        - Nếu không: lấy từ SIS Class Student với school_year = source_school_year_id
    """
    if logs is None:
        logs = []
    
    created_count = 0
    
    try:
        students = []
        
        # Nếu có finance_year_id thì lấy từ SIS Finance Student
        if finance_year_id:
            logs.append(f"Lấy học sinh từ Năm tài chính: {finance_year_id}")
            
            # Kiểm tra finance_year có tồn tại và thuộc đúng campus
            finance_year = frappe.db.get_value(
                "SIS Finance Year",
                finance_year_id,
                ["name", "campus_id"],
                as_dict=True
            )
            
            if not finance_year:
                logs.append(f"Không tìm thấy năm tài chính: {finance_year_id}")
                return 0
            
            if finance_year.campus_id != campus_id:
                logs.append(f"Năm tài chính không thuộc campus {campus_id}")
                return 0
            
            # Lấy học sinh từ SIS Finance Student
            students = frappe.db.sql("""
                SELECT DISTINCT 
                    fs.student_id,
                    s.student_name,
                    s.student_code,
                    fs.class_title,
                    fs.name as finance_student_id
                FROM `tabSIS Finance Student` fs
                INNER JOIN `tabCRM Student` s ON fs.student_id = s.name
                WHERE fs.finance_year_id = %(finance_year_id)s
            """, {
                "finance_year_id": finance_year_id
            }, as_dict=True)
            
            logs.append(f"Tìm thấy {len(students)} học sinh từ năm tài chính")
        else:
            # Lấy từ SIS Class Student như cũ
            logs.append(f"Lấy học sinh từ năm học nguồn: {source_school_year_id}, Campus: {campus_id}")
            
            # Lấy danh sách học sinh đã xếp lớp REGULAR trong năm học nguồn tại campus này
            # Chỉ lấy lớp regular, không lấy lớp mixed
            students = frappe.db.sql("""
                SELECT DISTINCT 
                    cs.student_id,
                    s.student_name,
                    s.student_code,
                    c.name as class_name,
                    c.title as class_title
                FROM `tabSIS Class Student` cs
                INNER JOIN `tabSIS Class` c ON cs.class_id = c.name
                INNER JOIN `tabCRM Student` s ON cs.student_id = s.name
                WHERE c.school_year_id = %(school_year_id)s
                  AND c.campus_id = %(campus_id)s
                  AND (c.class_type = 'regular' OR c.class_type IS NULL OR c.class_type = '')
            """, {
                "school_year_id": source_school_year_id,
                "campus_id": campus_id
            }, as_dict=True)
            
            logs.append(f"Tìm thấy {len(students)} học sinh đã xếp lớp")
        
        # Tạo records cho từng học sinh
        for student in students:
            try:
                # Kiểm tra đã có record chưa
                existing = frappe.db.exists("SIS Re-enrollment", {
                    "config_id": config_id,
                    "student_id": student.student_id
                })
                
                if existing:
                    continue
                
                # Tạo record mới (chưa có decision, chưa submit)
                re_doc = frappe.get_doc({
                    "doctype": "SIS Re-enrollment",
                    "config_id": config_id,
                    "student_id": student.student_id,
                    "student_name": student.student_name,
                    "student_code": student.student_code,
                    "current_class": student.class_title or student.class_name,
                    "campus_id": campus_id,
                    "status": "pending"
                    # decision và submitted_at để trống = chưa làm đơn
                })
                # Skip validation thời gian khi admin tạo records
                re_doc.flags.skip_config_validation = True
                re_doc.insert(ignore_permissions=True)
                created_count += 1
                
            except Exception as e:
                logs.append(f"Lỗi tạo record cho {student.student_code}: {str(e)}")
                continue
        
        frappe.db.commit()
        
    except Exception as e:
        logs.append(f"Lỗi auto-create records: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Auto Create Re-enrollment Records Error")
    
    return created_count


# ==================== CONFIG APIs ====================

@frappe.whitelist()
def get_configs():
    """
    Lấy danh sách tất cả cấu hình tái ghi danh.
    Có thể filter theo campus_id.
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        # Lấy filter từ query params
        campus_id = frappe.request.args.get('campus_id')
        
        filters = {}
        if campus_id:
            filters["campus_id"] = campus_id
        
        configs = frappe.get_all(
            "SIS Re-enrollment Config",
            filters=filters,
            fields=["*"],  # Lấy tất cả fields để tránh lỗi khi chưa migrate
            order_by="modified desc"
        )
        
        # Thêm thông tin bổ sung cho mỗi config
        for config in configs:
            # Tên năm học
            school_year = frappe.db.get_value(
                "SIS School Year",
                config.get("school_year_id"),
                ["title_vn", "title_en"],
                as_dict=True
            )
            config["school_year_name"] = school_year.title_vn if school_year else None
            
            # Tên campus (SIS Campus dùng title_vn, không phải title)
            campus_info = frappe.db.get_value(
                "SIS Campus", 
                config.get("campus_id"), 
                ["title_vn", "title_en"],
                as_dict=True
            )
            config["campus_name"] = campus_info.title_vn if campus_info else None
            
            # Đếm số đơn
            submission_count = frappe.db.count(
                "SIS Re-enrollment",
                {"config_id": config.name}
            )
            config["submission_count"] = submission_count
            
            # Số mức ưu đãi
            discount_count = frappe.db.count(
                "SIS Re-enrollment Discount",
                {"parent": config.name}
            )
            config["discount_count"] = discount_count
        
        logs.append(f"Tìm thấy {len(configs)} cấu hình")
        
        return list_response(configs)
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Admin Get Re-enrollment Configs Error")
        return error_response(
            message=f"Lỗi khi lấy danh sách cấu hình: {str(e)}",
            logs=logs
        )


@frappe.whitelist()
def get_config(config_id=None):
    """
    Lấy chi tiết một cấu hình tái ghi danh.
    Bao gồm bảng ưu đãi.
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        if not config_id:
            config_id = frappe.request.args.get('config_id')
        
        if not config_id:
            return validation_error_response(
                "Thiếu config_id",
                {"config_id": ["Config ID là bắt buộc"]}
            )
        
        # Kiểm tra config tồn tại
        if not frappe.db.exists("SIS Re-enrollment Config", config_id):
            return not_found_response("Không tìm thấy cấu hình tái ghi danh")
        
        config = frappe.get_doc("SIS Re-enrollment Config", config_id)
        
        # Lấy bảng ưu đãi
        discounts = frappe.get_all(
            "SIS Re-enrollment Discount",
            filters={"parent": config_id},
            fields=["name", "deadline", "description", "annual_discount", "semester_discount"],
            order_by="deadline asc"
        )
        
        # Lấy câu hỏi khảo sát từ config document
        questions = []
        for q in config.questions:
            # Parse options từ JSON
            options = []
            if q.options_json:
                try:
                    options = json.loads(q.options_json)
                except:
                    options = []
            
            questions.append({
                "name": q.name,
                "question_vn": q.question_vn,
                "question_en": q.question_en,
                "question_type": q.question_type,
                "is_required": q.is_required,
                "sort_order": q.sort_order,
                "options": options
            })
        
        # Tên năm học
        school_year = frappe.db.get_value(
            "SIS School Year",
            config.school_year_id,
            ["title_vn", "title_en"],
            as_dict=True
        )
        
        # Tên campus
        # SIS Campus dùng title_vn thay vì title
        campus_info = frappe.db.get_value("SIS Campus", config.campus_id, ["title_vn", "title_en"], as_dict=True)
        campus_name = campus_info.title_vn if campus_info else None
        
        logs.append(f"Lấy config: {config_id}")
        
        return single_item_response(
            data={
                "name": config.name,
                "title": config.title,
                "source_school_year_id": config.source_school_year_id,
                "school_year_id": config.school_year_id,
                "school_year_name_vn": school_year.title_vn if school_year else None,
                "school_year_name_en": school_year.title_en if school_year else None,
                "campus_id": config.campus_id,
                "campus_name": campus_name,
                "is_active": config.is_active,
                "start_date": str(config.start_date) if config.start_date else None,
                "end_date": str(config.end_date) if config.end_date else None,
                "service_document": config.service_document,
                "service_document_images": json.loads(config.service_document_images) if config.service_document_images else [],
                "agreement_text": config.agreement_text,
                "agreement_text_en": config.agreement_text_en,
                "discounts": discounts,
                "questions": questions,
                "created_by": config.created_by,
                "created_at": str(config.created_at) if config.created_at else None
            },
            message="Lấy cấu hình thành công"
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Admin Get Re-enrollment Config Error")
        return error_response(
            message=f"Lỗi khi lấy cấu hình: {str(e)}",
            logs=logs
        )


@frappe.whitelist()
def create_config():
    """
    Tạo cấu hình tái ghi danh mới.
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        # Lấy data từ request
        if frappe.request.is_json:
            data = frappe.request.json or {}
        else:
            data = frappe.form_dict
        
        logs.append(f"Tạo config mới: {json.dumps(data, default=str)}")
        
        # Validate required fields
        required_fields = ['title', 'source_school_year_id', 'school_year_id', 'campus_id', 'start_date', 'end_date']
        for field in required_fields:
            if field not in data or not data[field]:
                return validation_error_response(
                    f"Thiếu trường bắt buộc: {field}",
                    {field: [f"Trường {field} là bắt buộc"]}
                )
        
        # Resolve campus_id từ format frontend sang format database
        resolved_campus_id = _resolve_campus_id(data['campus_id'])
        if not resolved_campus_id:
            return validation_error_response(
                f"Không tìm thấy Campus: {data['campus_id']}",
                {"campus_id": ["Campus không tồn tại trong hệ thống"]}
            )
        data['campus_id'] = resolved_campus_id
        logs.append(f"Resolved campus_id: {resolved_campus_id}")
        
        # Xử lý service_document_images
        service_document_images = data.get('service_document_images')
        if service_document_images and isinstance(service_document_images, list):
            service_document_images = json.dumps(service_document_images)
        elif service_document_images and isinstance(service_document_images, str):
            # Đã là JSON string rồi
            pass
        else:
            service_document_images = None
        
        # Tạo document
        config_doc = frappe.get_doc({
            "doctype": "SIS Re-enrollment Config",
            "title": data['title'],
            "source_school_year_id": data['source_school_year_id'],
            "school_year_id": data['school_year_id'],
            "campus_id": data['campus_id'],
            "is_active": data.get('is_active', 0),
            "start_date": data['start_date'],
            "end_date": data['end_date'],
            "service_document": data.get('service_document'),
            "service_document_images": service_document_images,
            "agreement_text": data.get('agreement_text'),
            "agreement_text_en": data.get('agreement_text_en'),
            "finance_year_id": data.get('finance_year_id')  # Năm tài chính (optional)
        })
        
        # Thêm bảng ưu đãi nếu có
        discounts = data.get('discounts', [])
        if isinstance(discounts, str):
            discounts = json.loads(discounts)
        
        for discount in discounts:
            config_doc.append("discounts", {
                "deadline": discount.get('deadline'),
                "description": discount.get('description'),
                "annual_discount": discount.get('annual_discount', 0),
                "semester_discount": discount.get('semester_discount', 0)
            })
        
        # Thêm câu hỏi khảo sát nếu có
        questions_data = data.get('questions', [])
        if isinstance(questions_data, str):
            questions_data = json.loads(questions_data)
        
        # Thêm questions với options (lưu dưới dạng JSON)
        for idx, question in enumerate(questions_data):
            options_data = question.get('options', [])
            
            config_doc.append("questions", {
                "question_vn": question.get('question_vn'),
                "question_en": question.get('question_en'),
                "question_type": question.get('question_type', 'single_choice'),
                "is_required": question.get('is_required', 1),
                "sort_order": question.get('sort_order', idx),
                "options_json": json.dumps(options_data) if options_data else None
            })
        
        config_doc.insert()
        frappe.db.commit()
        
        logs.append(f"Đã tạo config: {config_doc.name}")
        
        # Auto-create re-enrollment records cho tất cả học sinh
        # Ưu tiên lấy từ Finance Year nếu có, không thì lấy từ SIS Class Student
        created_count = _auto_create_student_records(
            config_doc.name, 
            data['source_school_year_id'],  # Dùng năm học nguồn để lấy danh sách học sinh
            resolved_campus_id,
            logs,
            finance_year_id=data.get('finance_year_id')  # Nếu có thì lấy từ Finance Year
        )
        
        logs.append(f"Đã tạo {created_count} records cho học sinh")
        
        return success_response(
            data={
                "name": config_doc.name,
                "title": config_doc.title,
                "student_records_created": created_count
            },
            message=f"Tạo cấu hình tái ghi danh thành công. Đã tạo {created_count} đơn cho học sinh.",
            logs=logs
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Admin Create Re-enrollment Config Error")
        return error_response(
            message=f"Lỗi khi tạo cấu hình: {str(e)}",
            logs=logs
        )


@frappe.whitelist()
def update_config():
    """
    Cập nhật cấu hình tái ghi danh.
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        # Lấy data từ request
        if frappe.request.is_json:
            data = frappe.request.json or {}
        else:
            data = frappe.form_dict
        
        config_id = data.get('name') or data.get('config_id')
        if not config_id:
            return validation_error_response(
                "Thiếu config_id",
                {"config_id": ["Config ID là bắt buộc"]}
            )
        
        logs.append(f"Cập nhật config: {config_id}")
        
        # Lấy document
        if not frappe.db.exists("SIS Re-enrollment Config", config_id):
            return not_found_response("Không tìm thấy cấu hình")
        
        config_doc = frappe.get_doc("SIS Re-enrollment Config", config_id)
        
        # Resolve campus_id nếu có
        if 'campus_id' in data and data['campus_id']:
            resolved_campus_id = _resolve_campus_id(data['campus_id'])
            if resolved_campus_id:
                data['campus_id'] = resolved_campus_id
                logs.append(f"Resolved campus_id: {resolved_campus_id}")
        
        # Update các trường
        update_fields = ['title', 'source_school_year_id', 'school_year_id', 'campus_id', 'is_active', 
                        'start_date', 'end_date', 'service_document', 'agreement_text', 'agreement_text_en']
        
        for field in update_fields:
            if field in data:
                config_doc.set(field, data[field])
        
        # Xử lý riêng service_document_images vì cần convert từ list sang JSON string
        if 'service_document_images' in data:
            service_document_images = data['service_document_images']
            if isinstance(service_document_images, list):
                config_doc.service_document_images = json.dumps(service_document_images)
            elif isinstance(service_document_images, str):
                config_doc.service_document_images = service_document_images
            else:
                config_doc.service_document_images = None
        
        # Update bảng ưu đãi nếu có
        if 'discounts' in data:
            discounts = data['discounts']
            if isinstance(discounts, str):
                discounts = json.loads(discounts)
            
            # Xóa discounts cũ
            config_doc.discounts = []
            
            # Thêm discounts mới
            for discount in discounts:
                config_doc.append("discounts", {
                    "deadline": discount.get('deadline'),
                    "description": discount.get('description'),
                    "annual_discount": discount.get('annual_discount', 0),
                    "semester_discount": discount.get('semester_discount', 0)
                })
        
        # Update câu hỏi khảo sát nếu có
        if 'questions' in data:
            questions_data = data['questions']
            if isinstance(questions_data, str):
                questions_data = json.loads(questions_data)
            
            logs.append(f"Updating {len(questions_data)} questions")
            
            # Xóa questions cũ
            config_doc.questions = []
            
            # Thêm questions mới với options (lưu dưới dạng JSON)
            for idx, question in enumerate(questions_data):
                options_data = question.get('options', [])
                logs.append(f"Question {idx}: {len(options_data)} options")
                
                config_doc.append("questions", {
                    "question_vn": question.get('question_vn'),
                    "question_en": question.get('question_en'),
                    "question_type": question.get('question_type', 'single_choice'),
                    "is_required": question.get('is_required', 1),
                    "sort_order": question.get('sort_order', idx),
                    "options_json": json.dumps(options_data) if options_data else None
                })
        
        config_doc.save()
        frappe.db.commit()
        
        logs.append(f"Đã cập nhật config: {config_id}")
        
        return success_response(
            data={"name": config_doc.name},
            message="Cập nhật cấu hình thành công",
            logs=logs
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Admin Update Re-enrollment Config Error")
        return error_response(
            message=f"Lỗi khi cập nhật cấu hình: {str(e)}",
            logs=logs
        )


@frappe.whitelist()
def toggle_config_active():
    """
    Bật/tắt cấu hình tái ghi danh.
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        # Lấy data
        if frappe.request.is_json:
            data = frappe.request.json or {}
        else:
            data = frappe.form_dict
        
        config_id = data.get('config_id')
        is_active = data.get('is_active')
        
        if not config_id:
            return validation_error_response(
                "Thiếu config_id",
                {"config_id": ["Config ID là bắt buộc"]}
            )
        
        logs.append(f"Toggle config {config_id} -> is_active={is_active}")
        
        config_doc = frappe.get_doc("SIS Re-enrollment Config", config_id)
        config_doc.is_active = 1 if is_active else 0
        config_doc.save()
        
        frappe.db.commit()
        
        status_text = "đã mở" if is_active else "đã đóng"
        
        return success_response(
            data={"name": config_id, "is_active": config_doc.is_active},
            message=f"Đợt tái ghi danh {status_text}",
            logs=logs
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Admin Toggle Config Active Error")
        return error_response(
            message=f"Lỗi: {str(e)}",
            logs=logs
        )


@frappe.whitelist()
def delete_config():
    """
    Xóa cấu hình tái ghi danh.
    Chỉ xóa được nếu chưa có đơn nào.
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        # Lấy data
        if frappe.request.is_json:
            data = frappe.request.json or {}
        else:
            data = frappe.form_dict
        
        config_id = data.get('config_id')
        
        if not config_id:
            return validation_error_response(
                "Thiếu config_id",
                {"config_id": ["Config ID là bắt buộc"]}
            )
        
        # Kiểm tra có đơn nào không
        submission_count = frappe.db.count(
            "SIS Re-enrollment",
            {"config_id": config_id}
        )
        
        if submission_count > 0:
            return error_response(
                f"Không thể xóa vì đã có {submission_count} đơn tái ghi danh",
                logs=logs
            )
        
        frappe.delete_doc("SIS Re-enrollment Config", config_id)
        frappe.db.commit()
        
        logs.append(f"Đã xóa config: {config_id}")
        
        return success_response(
            message="Xóa cấu hình thành công",
            logs=logs
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Admin Delete Config Error")
        return error_response(
            message=f"Lỗi khi xóa: {str(e)}",
            logs=logs
        )


# ==================== SUBMISSION APIs ====================

@frappe.whitelist()
def get_submissions():
    """
    Lấy danh sách đơn tái ghi danh.
    Có thể filter theo config_id, status, decision.
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        # Lấy filters từ query params
        config_id = frappe.request.args.get('config_id')
        status = frappe.request.args.get('status')
        decision = frappe.request.args.get('decision')
        search = frappe.request.args.get('search')
        page = int(frappe.request.args.get('page', 1))
        page_size = int(frappe.request.args.get('page_size', 50))
        
        # Build query
        conditions = []
        values = {}
        
        if config_id:
            conditions.append("re.config_id = %(config_id)s")
            values["config_id"] = config_id
        
        # Xử lý status đặc biệt: not_submitted = decision IS NULL hoặc rỗng
        if status == 'not_submitted':
            conditions.append("(re.decision IS NULL OR re.decision = '')")
        elif status:
            conditions.append("re.status = %(status)s")
            values["status"] = status
        
        if decision:
            conditions.append("re.decision = %(decision)s")
            values["decision"] = decision
        
        # Search by student name or code
        if search:
            conditions.append("(re.student_name LIKE %(search)s OR re.student_code LIKE %(search)s)")
            values["search"] = f"%{search}%"
        
        where_clause = " AND ".join(conditions) if conditions else "1=1"
        
        # Count total
        total_query = f"""
            SELECT COUNT(*) as total
            FROM `tabSIS Re-enrollment` re
            WHERE {where_clause}
        """
        total = frappe.db.sql(total_query, values, as_dict=True)[0].total
        
        # Get submissions with pagination
        offset = (page - 1) * page_size
        query = f"""
            SELECT 
                re.name, re.config_id, re.student_id, re.student_name, re.student_code,
                re.guardian_id, re.guardian_name, g.phone_number as guardian_phone, g.email as guardian_email, 
                re.current_class, re.campus_id,
                re.decision, re.payment_type, re.not_re_enroll_reason,
                re.payment_status, re.selected_discount_id, re.selected_discount_name, re.selected_discount_percent,
                re.submitted_at, re.modified_by_admin, re.admin_modified_at
            FROM `tabSIS Re-enrollment` re
            LEFT JOIN `tabCRM Guardian` g ON re.guardian_id = g.name
            WHERE {where_clause}
            ORDER BY re.submitted_at DESC
            LIMIT {page_size} OFFSET {offset}
        """
        
        submissions = frappe.db.sql(query, values, as_dict=True)
        
        # Thêm display values và answers
        for sub in submissions:
            # Decision display
            decision_map = {
                "re_enroll": "Tái ghi danh",
                "considering": "Cân nhắc",
                "not_re_enroll": "Không tái ghi danh"
            }
            sub["decision_display"] = decision_map.get(sub.decision, "Chưa làm đơn")
            
            # Payment type display
            if sub.payment_type:
                sub["payment_display"] = "Đóng theo năm" if sub.payment_type == 'annual' else "Đóng theo kỳ"
            
            # Payment status display
            payment_status_map = {
                "unpaid": "Chưa đóng",
                "paid": "Đã đóng",
                "refunded": "Hoàn tiền"
            }
            sub["payment_status_display"] = payment_status_map.get(sub.payment_status, "-")
            
            # Lấy answers (câu trả lời khảo sát)
            answers = frappe.get_all(
                "SIS Re-enrollment Answer",
                filters={"parent": sub.name},
                fields=["question_id", "question_text_vn", "question_text_en", 
                        "selected_options", "selected_options_text_vn", "selected_options_text_en"]
            )
            sub["answers"] = answers
        
        logs.append(f"Tìm thấy {len(submissions)} / {total} đơn")
        
        return success_response(
            data={
                "items": submissions,
                "total": total,
                "page": page,
                "page_size": page_size,
                "total_pages": (total + page_size - 1) // page_size
            },
            message="Lấy danh sách đơn thành công",
            logs=logs
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Admin Get Submissions Error")
        return error_response(
            message=f"Lỗi: {str(e)}",
            logs=logs
        )


@frappe.whitelist()
def get_submission(submission_id=None):
    """
    Lấy chi tiết một đơn tái ghi danh.
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        if not submission_id:
            submission_id = frappe.request.args.get('submission_id')
        
        if not submission_id:
            return validation_error_response(
                "Thiếu submission_id",
                {"submission_id": ["Submission ID là bắt buộc"]}
            )
        
        if not frappe.db.exists("SIS Re-enrollment", submission_id):
            return not_found_response("Không tìm thấy đơn tái ghi danh")
        
        submission = frappe.get_doc("SIS Re-enrollment", submission_id)
        
        # Lấy thông tin config
        config_info = frappe.db.get_value(
            "SIS Re-enrollment Config",
            submission.config_id,
            ["title", "school_year_id"],
            as_dict=True
        )
        
        # Lấy thông tin guardian (phone, email) từ CRM Guardian
        guardian_info = frappe.db.get_value(
            "CRM Guardian",
            submission.guardian_id,
            ["phone_number", "email"],
            as_dict=True
        ) if submission.guardian_id else None
        
        return single_item_response(
            data={
                "name": submission.name,
                "config_id": submission.config_id,
                "config_title": config_info.title if config_info else None,
                "student_id": submission.student_id,
                "student_name": submission.student_name,
                "student_code": submission.student_code,
                "guardian_id": submission.guardian_id,
                "guardian_name": submission.guardian_name,
                "guardian_phone": guardian_info.phone_number if guardian_info else None,
                "guardian_email": guardian_info.email if guardian_info else None,
                "current_class": submission.current_class,
                "decision": submission.decision,
                "decision_display": "Tái ghi danh" if submission.decision == 're_enroll' else "Không tái ghi danh",
                "payment_type": submission.payment_type,
                "payment_display": "Đóng theo năm" if submission.payment_type == 'annual' else ("Đóng theo kỳ" if submission.payment_type == 'semester' else None),
                "selected_discount_deadline": str(submission.selected_discount_deadline) if submission.selected_discount_deadline else None,
                "not_re_enroll_reason": submission.not_re_enroll_reason,
                "agreement_accepted": submission.agreement_accepted,
                "status": submission.status,
                "submitted_at": str(submission.submitted_at) if submission.submitted_at else None,
                "modified_by_admin": submission.modified_by_admin,
                "admin_modified_at": str(submission.admin_modified_at) if submission.admin_modified_at else None,
                "admin_notes": submission.admin_notes
            },
            message="Lấy thông tin đơn thành công"
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Admin Get Submission Error")
        return error_response(
            message=f"Lỗi: {str(e)}",
            logs=logs
        )


@frappe.whitelist()
def update_submission():
    """
    Admin sửa đơn tái ghi danh cho phụ huynh.
    Ghi nhận admin đã sửa.
    
    POST body:
    {
        "submission_id": "SIS-REENROLL-00001",
        "decision": "re_enroll" | "considering" | "not_re_enroll",
        "payment_type": "annual" | "semester",
        "selected_discount_id": "...",
        "not_re_enroll_reason": "...",
        "payment_status": "unpaid" | "paid" | "refunded",
        "notes": [{"note": "...", "created_by_name": "..."}]
    }
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        # Lấy data
        if frappe.request.is_json:
            data = frappe.request.json or {}
        else:
            data = frappe.form_dict
        
        submission_id = data.get('submission_id') or data.get('name')
        if not submission_id:
            return validation_error_response(
                "Thiếu submission_id",
                {"submission_id": ["Submission ID là bắt buộc"]}
            )
        
        logs.append(f"Admin sửa đơn: {submission_id}")
        
        if not frappe.db.exists("SIS Re-enrollment", submission_id):
            return not_found_response("Không tìm thấy đơn tái ghi danh")
        
        submission = frappe.get_doc("SIS Re-enrollment", submission_id)
        
        # Lưu lại giá trị cũ để so sánh và tạo log
        old_values = {
            'decision': submission.decision,
            'payment_type': submission.payment_type,
            'selected_discount_id': submission.selected_discount_id,
            'not_re_enroll_reason': submission.not_re_enroll_reason,
            'payment_status': submission.payment_status
        }
        
        # Lưu lại answers cũ để so sánh
        old_answers = []
        for ans in submission.answers:
            old_answers.append({
                'question_id': ans.question_id,
                'selected_options_text_vn': ans.selected_options_text_vn
            })
        
        # Lưu lại các system_log cũ (để không bị mất khi clear notes)
        old_system_logs = []
        for note in submission.notes:
            if note.note_type == 'system_log':
                old_system_logs.append({
                    'note_type': note.note_type,
                    'note': note.note,
                    'created_by_user': note.created_by_user,
                    'created_by_name': note.created_by_name,
                    'created_at': note.created_at
                })
        
        # Các trường admin có thể sửa
        updatable_fields = ['decision', 'payment_type', 'selected_discount_id', 
                          'not_re_enroll_reason', 'payment_status']
        
        for field in updatable_fields:
            if field in data:
                submission.set(field, data[field])
        
        # Xử lý discount info nếu chọn
        if data.get('selected_discount_id') and data.get('decision') == 're_enroll':
            # Lấy config để tìm thông tin discount
            config = frappe.get_doc("SIS Re-enrollment Config", submission.config_id)
            payment_type = data.get('payment_type') or submission.payment_type
            for discount in config.discounts:
                if discount.name == data.get('selected_discount_id'):
                    submission.selected_discount_name = discount.description
                    submission.selected_discount_deadline = discount.deadline
                    # Lưu % giảm dựa trên payment_type
                    if payment_type == 'annual':
                        submission.selected_discount_percent = discount.annual_discount
                    else:
                        submission.selected_discount_percent = discount.semester_discount
                    break
        
        # Xử lý notes - clear và thêm mới
        notes_data = data.get('notes', [])
        
        # Clear existing notes
        submission.notes = []
        
        # Lấy full name của user hiện tại
        current_user = frappe.session.user
        current_user_name = frappe.db.get_value("User", current_user, "full_name") or current_user
        
        # Thêm lại các system_log cũ trước
        for old_log in old_system_logs:
            submission.append("notes", old_log)
        
        # Add manual notes từ frontend (chỉ lấy manual_note, bỏ qua system_log vì đã thêm ở trên)
        if notes_data:
            for note in notes_data:
                # Bỏ qua system_log từ frontend (đã có từ old_system_logs)
                if note.get('note_type') == 'system_log':
                    continue
                    
                # Convert ISO datetime to MySQL format nếu cần
                created_at = note.get('created_at')
                if created_at:
                    try:
                        # Parse ISO format và convert sang MySQL format
                        from datetime import datetime
                        if 'T' in str(created_at):
                            dt = datetime.fromisoformat(created_at.replace('Z', '+00:00'))
                            created_at = dt.strftime('%Y-%m-%d %H:%M:%S')
                    except:
                        created_at = now()
                else:
                    created_at = now()
                
                submission.append("notes", {
                    "note_type": "manual_note",
                    "note": note.get('note'),
                    "created_by_user": note.get('created_by_user') or current_user,
                    "created_by_name": note.get('created_by_name') or current_user_name,
                    "created_at": created_at
                })
        
        # Clear các field không cần thiết dựa trên decision
        decision = data.get('decision')
        if decision == 're_enroll':
            submission.not_re_enroll_reason = None
        elif decision in ['considering', 'not_re_enroll']:
            submission.payment_type = None
            submission.selected_discount_id = None
            submission.selected_discount_name = None
            submission.selected_discount_deadline = None
        
        # Xử lý answers (câu trả lời khảo sát)
        answers_data = data.get('answers', [])
        if answers_data:
            # Clear existing answers
            submission.answers = []
            
            # Add new answers
            for answer in answers_data:
                selected_options = answer.get('selected_options', [])
                if isinstance(selected_options, list):
                    selected_options_json = json.dumps(selected_options)
                else:
                    selected_options_json = selected_options
                
                submission.append("answers", {
                    "question_id": answer.get('question_id'),
                    "question_text_vn": answer.get('question_text_vn'),
                    "question_text_en": answer.get('question_text_en'),
                    "selected_options": selected_options_json,
                    "selected_options_text_vn": answer.get('selected_options_text_vn'),
                    "selected_options_text_en": answer.get('selected_options_text_en')
                })
        
        # Tạo log hệ thống về thay đổi của admin
        changes = []
        important_changes = []  # Các thay đổi quan trọng cần gửi notification (Trạng thái, Khảo sát)
        decision_map = {
            're_enroll': 'Tái ghi danh',
            'considering': 'Cân nhắc', 
            'not_re_enroll': 'Không tái ghi danh'
        }
        payment_type_map = {'annual': 'Đóng theo năm', 'semester': 'Đóng theo kỳ'}
        payment_status_map = {'unpaid': 'Chưa đóng', 'paid': 'Đã đóng', 'refunded': 'Hoàn tiền'}
        
        # So sánh decision (QUAN TRỌNG - gửi notification)
        new_decision = data.get('decision') or submission.decision
        if old_values['decision'] != new_decision:
            old_display = decision_map.get(old_values['decision'], old_values['decision'] or 'Chưa có')
            new_display = decision_map.get(new_decision, new_decision)
            changes.append(f"• Quyết định: {old_display} → {new_display}")
            important_changes.append('decision')
        
        # So sánh payment_type (QUAN TRỌNG - gửi notification)
        new_payment_type = data.get('payment_type') or submission.payment_type
        if old_values['payment_type'] != new_payment_type and new_decision == 're_enroll':
            old_display = payment_type_map.get(old_values['payment_type'], old_values['payment_type'] or 'Chưa có')
            new_display = payment_type_map.get(new_payment_type, new_payment_type)
            changes.append(f"• Phương thức: {old_display} → {new_display}")
            important_changes.append('payment_type')
        
        # So sánh payment_status (KHÔNG gửi notification - chỉ log)
        new_payment_status = data.get('payment_status') or submission.payment_status
        if old_values['payment_status'] != new_payment_status:
            old_display = payment_status_map.get(old_values['payment_status'], old_values['payment_status'] or 'Chưa có')
            new_display = payment_status_map.get(new_payment_status, new_payment_status)
            changes.append(f"• Thanh toán: {old_display} → {new_display}")
            # Không thêm vào important_changes
        
        # So sánh discount (QUAN TRỌNG - gửi notification)
        new_discount_id = data.get('selected_discount_id') or submission.selected_discount_id
        if old_values['selected_discount_id'] != new_discount_id:
            changes.append(f"• Ưu đãi: Đã cập nhật")
            important_changes.append('discount')
        
        # So sánh lý do (QUAN TRỌNG - gửi notification)
        new_reason = data.get('not_re_enroll_reason') or submission.not_re_enroll_reason
        if old_values['not_re_enroll_reason'] != new_reason and new_decision in ['considering', 'not_re_enroll']:
            changes.append(f"• Lý do: Đã cập nhật")
            important_changes.append('reason')
        
        # So sánh answers (QUAN TRỌNG - gửi notification)
        new_answers_data = data.get('answers', [])
        answers_changed = False
        
        if new_answers_data:
            # So sánh số lượng
            if len(new_answers_data) != len(old_answers):
                answers_changed = True
            else:
                # So sánh nội dung từng câu trả lời
                for new_ans in new_answers_data:
                    q_id = new_ans.get('question_id')
                    new_text = new_ans.get('selected_options_text_vn', '')
                    
                    # Tìm câu trả lời cũ tương ứng
                    old_ans = next((a for a in old_answers if a['question_id'] == q_id), None)
                    
                    if not old_ans:
                        # Câu hỏi mới
                        answers_changed = True
                        break
                    elif old_ans['selected_options_text_vn'] != new_text:
                        # Nội dung khác
                        answers_changed = True
                        break
        
        if answers_changed:
            changes.append(f"• Khảo sát: Đã cập nhật câu trả lời")
            important_changes.append('answers')
        
        # Nếu có thay đổi thì tạo log
        if changes:
            log_content = f"Admin {current_user_name} đã cập nhật đơn:\n" + "\n".join(changes)
            submission.append("notes", {
                "note_type": "system_log",
                "note": log_content,
                "created_by_user": current_user,
                "created_by_name": current_user_name,
                "created_at": now()
            })
        
        # Ghi nhận admin sửa
        submission.modified_by_admin = frappe.session.user
        submission.admin_modified_at = now()
        
        # Cập nhật submitted_at nếu chưa có
        if not submission.submitted_at:
            submission.submitted_at = now()
        
        submission.save()
        frappe.db.commit()
        
        logs.append(f"Đã cập nhật đơn: {submission_id}")
        
        # Chỉ gửi thông báo khi có thay đổi quan trọng (Trạng thái hoặc Khảo sát)
        # KHÔNG gửi khi chỉ thay đổi payment_status hoặc thêm note
        if important_changes:
            try:
                from erp.api.parent_portal.re_enrollment import _create_re_enrollment_announcement
                
                # Lấy thông tin năm học
                school_year = ""
                try:
                    config_doc = frappe.get_doc("SIS Re-enrollment Config", submission.config_id)
                    school_year_info = frappe.db.get_value(
                        "SIS School Year",
                        config_doc.school_year_id,
                        ["name_vn", "name_en"],
                        as_dict=True
                    )
                    if school_year_info:
                        school_year = school_year_info.name_vn or school_year_info.name_en or ""
                except:
                    pass
                
                # Lấy answers từ submission để gửi vào announcement
                answers_for_announcement = []
                for answer in submission.answers:
                    answers_for_announcement.append({
                        'question_text_vn': answer.question_text_vn,
                        'question_text_en': answer.question_text_en,
                        'selected_options_text_vn': answer.selected_options_text_vn,
                        'selected_options_text_en': answer.selected_options_text_en
                    })
                
                _create_re_enrollment_announcement(
                    student_id=submission.student_id,
                    student_name=submission.student_name,
                    student_code=submission.student_code,
                    submission_data={
                        'decision': submission.decision,
                        'payment_type': submission.payment_type,
                        'discount_name': submission.selected_discount_name,
                        'discount_percent': submission.selected_discount_percent,
                        'reason': submission.not_re_enroll_reason,
                        'school_year': school_year,
                        'submitted_at': str(now()),
                        'status': submission.status or 'pending',
                        'answers': answers_for_announcement  # Câu trả lời khảo sát
                    },
                    is_update=True
                )
                logs.append(f"Đã gửi thông báo cập nhật cho phụ huynh (thay đổi: {', '.join(important_changes)})")
            except Exception as notif_err:
                logs.append(f"Lỗi gửi thông báo: {str(notif_err)}")
                frappe.logger().error(f"Error sending admin update notification: {str(notif_err)}")
        else:
            logs.append("Không gửi thông báo (chỉ thay đổi thanh toán/ghi chú)")
        
        return success_response(
            data={"name": submission.name},
            message="Cập nhật đơn thành công",
            logs=logs
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Admin Update Submission Error")
        return error_response(
            message=f"Lỗi khi cập nhật đơn: {str(e)}",
            logs=logs
        )


@frappe.whitelist()
def bulk_update_status():
    """
    Cập nhật trạng thái hàng loạt.
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        # Lấy data
        if frappe.request.is_json:
            data = frappe.request.json or {}
        else:
            data = frappe.form_dict
        
        submission_ids = data.get('submission_ids', [])
        new_status = data.get('status')
        
        if isinstance(submission_ids, str):
            submission_ids = json.loads(submission_ids)
        
        if not submission_ids:
            return validation_error_response(
                "Thiếu danh sách đơn",
                {"submission_ids": ["Danh sách đơn là bắt buộc"]}
            )
        
        if not new_status or new_status not in ['pending', 'approved', 'rejected']:
            return validation_error_response(
                "Trạng thái không hợp lệ",
                {"status": ["Trạng thái phải là pending, approved hoặc rejected"]}
            )
        
        updated_count = 0
        for sub_id in submission_ids:
            try:
                submission = frappe.get_doc("SIS Re-enrollment", sub_id)
                submission.status = new_status
                submission.modified_by_admin = frappe.session.user
                submission.admin_modified_at = now()
                submission.save()
                updated_count += 1
            except Exception as e:
                logs.append(f"Lỗi cập nhật {sub_id}: {str(e)}")
        
        frappe.db.commit()
        
        logs.append(f"Đã cập nhật {updated_count}/{len(submission_ids)} đơn")
        
        return success_response(
            data={"updated_count": updated_count},
            message=f"Đã cập nhật {updated_count} đơn",
            logs=logs
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Admin Bulk Update Status Error")
        return error_response(
            message=f"Lỗi: {str(e)}",
            logs=logs
        )


# ==================== STATISTICS API ====================

@frappe.whitelist()
def get_statistics():
    """
    Lấy thống kê tái ghi danh.
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        config_id = frappe.request.args.get('config_id')
        
        if not config_id:
            return validation_error_response(
                "Thiếu config_id",
                {"config_id": ["Config ID là bắt buộc"]}
            )
        
        # Thống kê theo quyết định
        decision_stats = frappe.db.sql("""
            SELECT 
                decision,
                COUNT(*) as count
            FROM `tabSIS Re-enrollment`
            WHERE config_id = %s
            GROUP BY decision
        """, config_id, as_dict=True)
        
        # Thống kê theo trạng thái
        status_stats = frappe.db.sql("""
            SELECT 
                status,
                COUNT(*) as count
            FROM `tabSIS Re-enrollment`
            WHERE config_id = %s
            GROUP BY status
        """, config_id, as_dict=True)
        
        # Thống kê theo phương thức thanh toán (chỉ cho những người tái ghi danh)
        payment_stats = frappe.db.sql("""
            SELECT 
                payment_type,
                COUNT(*) as count
            FROM `tabSIS Re-enrollment`
            WHERE config_id = %s AND decision = 're_enroll'
            GROUP BY payment_type
        """, config_id, as_dict=True)
        
        # Tổng số
        total = frappe.db.count("SIS Re-enrollment", {"config_id": config_id})
        
        # Lấy thông tin config
        config = frappe.db.get_value(
            "SIS Re-enrollment Config",
            config_id,
            ["title", "campus_id"],
            as_dict=True
        )
        
        # Số học sinh trong campus (để so sánh)
        if config:
            # Đếm số học sinh active trong campus
            total_students = frappe.db.sql("""
                SELECT COUNT(DISTINCT cs.student_id) as count
                FROM `tabSIS Class Student` cs
                INNER JOIN `tabSIS Class` c ON cs.class_id = c.name
                WHERE c.campus_id = %s
            """, config.campus_id, as_dict=True)[0].count
        else:
            total_students = 0
        
        # Chuyển đổi thành dict dễ sử dụng
        decision_dict = {item.decision: item.count for item in decision_stats if item.decision}
        status_dict = {item.status: item.count for item in status_stats}
        payment_dict = {item.payment_type: item.count for item in payment_stats if item.payment_type}
        
        # Đếm số chưa làm đơn (decision = NULL hoặc rỗng)
        not_submitted_count = frappe.db.sql("""
            SELECT COUNT(*) as count
            FROM `tabSIS Re-enrollment`
            WHERE config_id = %s AND (decision IS NULL OR decision = '')
        """, config_id, as_dict=True)[0].count
        
        logs.append(f"Thống kê cho config {config_id}")
        
        return success_response(
            data={
                "total_submissions": total,
                "total_students_in_campus": total_students,
                "not_submitted": not_submitted_count,
                "by_decision": {
                    "re_enroll": decision_dict.get("re_enroll", 0),
                    "considering": decision_dict.get("considering", 0),
                    "not_re_enroll": decision_dict.get("not_re_enroll", 0)
                },
                "by_status": {
                    "pending": status_dict.get("pending", 0),
                    "approved": status_dict.get("approved", 0),
                    "rejected": status_dict.get("rejected", 0)
                },
                "by_payment_type": {
                    "annual": payment_dict.get("annual", 0),
                    "semester": payment_dict.get("semester", 0)
                }
            },
            message="Lấy thống kê thành công",
            logs=logs
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Admin Get Statistics Error")
        return error_response(
            message=f"Lỗi: {str(e)}",
            logs=logs
        )


@frappe.whitelist()
def export_submissions():
    """
    Export danh sách đơn ra CSV/Excel.
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        config_id = frappe.request.args.get('config_id')
        
        if not config_id:
            return validation_error_response(
                "Thiếu config_id",
                {"config_id": ["Config ID là bắt buộc"]}
            )
        
        submissions = frappe.db.sql("""
            SELECT 
                re.student_code as 'Mã học sinh',
                re.student_name as 'Tên học sinh',
                re.current_class as 'Lớp hiện tại',
                re.guardian_name as 'Phụ huynh',
                CASE re.decision 
                    WHEN 're_enroll' THEN 'Tái ghi danh'
                    WHEN 'not_re_enroll' THEN 'Không tái ghi danh'
                END as 'Quyết định',
                CASE re.payment_type
                    WHEN 'annual' THEN 'Đóng theo năm'
                    WHEN 'semester' THEN 'Đóng theo kỳ'
                    ELSE ''
                END as 'Phương thức thanh toán',
                re.not_re_enroll_reason as 'Lý do không tái ghi danh',
                CASE re.status
                    WHEN 'pending' THEN 'Chờ xử lý'
                    WHEN 'approved' THEN 'Đã duyệt'
                    WHEN 'rejected' THEN 'Từ chối'
                END as 'Trạng thái',
                DATE_FORMAT(re.submitted_at, '%%d/%%m/%%Y %%H:%%i') as 'Ngày nộp'
            FROM `tabSIS Re-enrollment` re
            WHERE re.config_id = %s
            ORDER BY re.student_name ASC
        """, config_id, as_dict=True)
        
        logs.append(f"Export {len(submissions)} đơn")
        
        return success_response(
            data={"items": submissions},
            message=f"Xuất {len(submissions)} đơn thành công",
            logs=logs
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Admin Export Submissions Error")
        return error_response(
            message=f"Lỗi: {str(e)}",
            logs=logs
        )


# ==================== UPLOAD PDF API ====================

@frappe.whitelist(allow_guest=False, methods=['POST'])
def upload_service_document():
    """
    Upload file PDF dịch vụ học sinh và convert sang ảnh.
    
    POST body (multipart/form-data):
    {
        "config_id": "SIS-REENROLL-CFG-00001" (optional - nếu muốn update config sau khi upload),
        "file": <PDF file>
    }
    
    Trả về:
    {
        "success": true,
        "data": {
            "file_url": "/files/...",
            "custom_name": "re-enrollment-xxxx",
            "images": ["url1", "url2", ...]
        }
    }
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        # Lấy config_id nếu có
        data = frappe.request.form
        config_id = data.get('config_id')
        
        logs.append(f"Upload PDF cho config: {config_id or 'new'}")
        
        # Kiểm tra có file không
        if not frappe.request.files:
            return validation_error_response(
                "Thiếu file PDF",
                {"file": ["Vui lòng chọn file PDF để upload"]}
            )
        
        # Lấy file từ request
        file_obj = None
        for file_key, f in frappe.request.files.items():
            if file_key == 'file':
                file_obj = f
                break
        
        if not file_obj:
            return validation_error_response(
                "Thiếu file PDF",
                {"file": ["Vui lòng chọn file PDF để upload"]}
            )
        
        # Kiểm tra file là PDF
        filename = file_obj.filename
        if not filename.lower().endswith('.pdf'):
            return validation_error_response(
                "Chỉ chấp nhận file PDF",
                {"file": ["Vui lòng upload file có định dạng PDF"]}
            )
        
        logs.append(f"File: {filename}")
        
        # Đọc nội dung file
        file_content = file_obj.stream.read()
        file_obj.stream.seek(0)  # Reset để có thể đọc lại
        
        # Tạo custom name unique
        custom_name = f"re-enrollment-{uuid.uuid4().hex[:8]}"
        
        logs.append(f"Custom name: {custom_name}")
        
        # 1. Lưu file vào Frappe
        file_doc = frappe.get_doc({
            "doctype": "File",
            "file_name": filename,
            "folder": "Home/Attachments",
            "content": file_content,
            "is_private": 0  # Public để dễ truy cập
        })
        file_doc.insert(ignore_permissions=True)
        frappe.db.commit()
        
        file_url = file_doc.file_url
        logs.append(f"Saved file: {file_url}")
        
        # 2. Gọi PDF Service để convert sang ảnh
        images = []
        try:
            # Upload file đến pdf-service
            pdf_service_url = frappe.conf.get("pdf_service_url", "http://172.16.20.113:5020")
            
            # Reset stream để upload
            file_obj.stream.seek(0)
            
            upload_response = requests.post(
                f"{pdf_service_url}/api/pdfs/upload-pdf",
                files={"pdfFile": (filename, file_obj.stream, "application/pdf")},
                data={"customName": custom_name, "uploader": frappe.session.user},
                timeout=120  # 2 phút cho convert
            )
            
            if upload_response.status_code == 200:
                upload_result = upload_response.json()
                logs.append(f"PDF Service upload: {json.dumps(upload_result)}")
                
                # Lấy danh sách ảnh
                images_response = requests.get(
                    f"{pdf_service_url}/api/pdfs/get-images/{custom_name}",
                    timeout=30
                )
                
                if images_response.status_code == 200:
                    images_result = images_response.json()
                    images = images_result.get("images", [])
                    logs.append(f"Got {len(images)} images from PDF")
                else:
                    logs.append(f"Failed to get images: {images_response.status_code}")
            else:
                logs.append(f"PDF Service error: {upload_response.status_code} - {upload_response.text}")
                
        except requests.exceptions.RequestException as e:
            # PDF service không chạy hoặc lỗi mạng
            logs.append(f"PDF Service unavailable: {str(e)}")
            # Không throw lỗi, vẫn lưu file nhưng không có ảnh
        
        # 3. Nếu có config_id, update config
        if config_id and frappe.db.exists("SIS Re-enrollment Config", config_id):
            config_doc = frappe.get_doc("SIS Re-enrollment Config", config_id)
            config_doc.service_document = file_url
            config_doc.service_document_images = json.dumps(images) if images else None
            config_doc.save()
            frappe.db.commit()
            logs.append(f"Updated config {config_id}")
        
        return success_response(
            data={
                "file_url": file_url,
                "custom_name": custom_name,
                "images": images,
                "image_count": len(images)
            },
            message=f"Upload thành công, đã convert thành {len(images)} ảnh" if images else "Upload thành công (chưa convert ảnh)",
            logs=logs
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Upload Service Document Error")
        return error_response(
            message=f"Lỗi khi upload file: {str(e)}",
            logs=logs
        )


@frappe.whitelist()
def get_service_document_images(config_id=None):
    """
    Lấy danh sách ảnh từ service document của một config.
    """
    logs = []
    
    try:
        if not config_id:
            config_id = frappe.request.args.get('config_id')
        
        if not config_id:
            return validation_error_response(
                "Thiếu config_id",
                {"config_id": ["Config ID là bắt buộc"]}
            )
        
        if not frappe.db.exists("SIS Re-enrollment Config", config_id):
            return not_found_response("Không tìm thấy cấu hình")
        
        config = frappe.get_doc("SIS Re-enrollment Config", config_id)
        
        images = []
        if config.service_document_images:
            try:
                images = json.loads(config.service_document_images)
            except:
                images = []
        
        return success_response(
            data={
                "service_document": config.service_document,
                "images": images,
                "image_count": len(images)
            },
            message="Lấy danh sách ảnh thành công",
            logs=logs
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        return error_response(
            message=f"Lỗi: {str(e)}",
            logs=logs
        )


# ==================== SYNC STUDENTS API ====================

@frappe.whitelist(allow_guest=False, methods=['POST'])
def sync_students():
    """
    Đồng bộ danh sách học sinh cho một config tái ghi danh.
    Tạo records cho học sinh mới chưa có trong danh sách.
    
    POST body:
    {
        "config_id": "SIS-REENROLL-CFG-00001"
    }
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        # Lấy data từ request
        if frappe.request.is_json:
            data = frappe.request.json or {}
        else:
            data = frappe.form_dict
        
        config_id = data.get('config_id')
        
        if not config_id:
            return validation_error_response(
                "Thiếu config_id",
                {"config_id": ["Config ID là bắt buộc"]}
            )
        
        # Lấy thông tin config
        if not frappe.db.exists("SIS Re-enrollment Config", config_id):
            return not_found_response("Không tìm thấy cấu hình")
        
        config = frappe.get_doc("SIS Re-enrollment Config", config_id)
        
        # Có thể override finance_year_id từ request
        finance_year_id = data.get('finance_year_id') or getattr(config, 'finance_year_id', None)
        
        logs.append(f"Sync students cho config: {config_id}")
        logs.append(f"Source school year: {config.source_school_year_id}, Campus: {config.campus_id}")
        if finance_year_id:
            logs.append(f"Finance Year: {finance_year_id}")
        
        # Gọi hàm auto-create với năm học nguồn hoặc năm tài chính
        created_count = _auto_create_student_records(
            config_id,
            config.source_school_year_id,  # Dùng năm học nguồn để lấy danh sách học sinh
            config.campus_id,
            logs,
            finance_year_id=finance_year_id  # Nếu có thì lấy từ Finance Year
        )
        
        return success_response(
            data={
                "created_count": created_count,
                "config_id": config_id
            },
            message=f"Đã đồng bộ thành công. Tạo thêm {created_count} đơn mới.",
            logs=logs
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Sync Students Error")
        return error_response(
            message=f"Lỗi khi đồng bộ: {str(e)}",
            logs=logs
        )


@frappe.whitelist(allow_guest=False, methods=['POST'])
def update_submission_decision():
    """
    Cập nhật quyết định cho một đơn tái ghi danh.
    Cho phép switch giữa các trạng thái: re_enroll, considering, not_re_enroll
    
    POST body:
    {
        "submission_id": "SIS-REENROLL-00001",
        "decision": "re_enroll" | "considering" | "not_re_enroll",
        "payment_type": "annual" | "semester" (required nếu decision = re_enroll),
        "selected_discount_id": "...", (optional - ID ưu đãi đã chọn nếu re_enroll),
        "reason": "..." (required nếu decision = considering hoặc not_re_enroll)
    }
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        # Lấy data từ request
        if frappe.request.is_json:
            data = frappe.request.json or {}
        else:
            data = frappe.form_dict
        
        submission_id = data.get('submission_id') or data.get('name')
        decision = data.get('decision')
        
        if not submission_id:
            return validation_error_response(
                "Thiếu submission_id",
                {"submission_id": ["Submission ID là bắt buộc"]}
            )
        
        if not decision or decision not in ['re_enroll', 'considering', 'not_re_enroll']:
            return validation_error_response(
                "Decision không hợp lệ",
                {"decision": ["Phải là re_enroll, considering hoặc not_re_enroll"]}
            )
        
        # Lấy submission
        if not frappe.db.exists("SIS Re-enrollment", submission_id):
            return not_found_response("Không tìm thấy đơn tái ghi danh")
        
        submission = frappe.get_doc("SIS Re-enrollment", submission_id)
        
        # Lưu decision cũ để tạo log
        old_decision = submission.decision
        
        logs.append(f"Update decision cho {submission_id}: {submission.decision} -> {decision}")
        
        # Validate theo loại decision
        if decision == 're_enroll':
            payment_type = data.get('payment_type')
            if not payment_type or payment_type not in ['annual', 'semester']:
                return validation_error_response(
                    "Thiếu phương thức thanh toán",
                    {"payment_type": ["Vui lòng chọn đóng theo năm hoặc theo kỳ"]}
                )
            
            submission.decision = decision
            submission.payment_type = payment_type
            submission.selected_discount_deadline = data.get('selected_discount_deadline')
            submission.not_re_enroll_reason = None  # Clear reason nếu đổi sang re_enroll
            
        else:  # considering hoặc not_re_enroll
            reason = data.get('reason') or data.get('not_re_enroll_reason')
            if not reason:
                return validation_error_response(
                    "Thiếu lý do",
                    {"reason": ["Vui lòng nhập lý do"]}
                )
            
            submission.decision = decision
            submission.not_re_enroll_reason = reason
            submission.payment_type = None  # Clear payment nếu không re_enroll
            submission.selected_discount_deadline = None
        
        # Ghi nhận thời gian submit nếu lần đầu có decision
        if not submission.submitted_at:
            submission.submitted_at = now()
        
        # Tạo log hệ thống về thay đổi decision
        decision_map = {
            're_enroll': 'Tái ghi danh',
            'considering': 'Cân nhắc', 
            'not_re_enroll': 'Không tái ghi danh'
        }
        current_user = frappe.session.user
        current_user_name = frappe.db.get_value("User", current_user, "full_name") or current_user
        
        old_display = decision_map.get(old_decision, old_decision or 'Chưa có')
        new_display = decision_map.get(decision, decision)
        
        log_content = f"Admin {current_user_name} đã cập nhật quyết định:\n• Quyết định: {old_display} → {new_display}"
        
        submission.append("notes", {
            "note_type": "system_log",
            "note": log_content,
            "created_by_user": current_user,
            "created_by_name": current_user_name,
            "created_at": now()
        })
        
        # Ghi nhận admin sửa
        submission.modified_by_admin = frappe.session.user
        submission.admin_modified_at = now()
        
        submission.save()
        frappe.db.commit()
        
        logs.append(f"Đã cập nhật decision thành công")
        
        # Gửi thông báo cho phụ huynh về việc cập nhật quyết định
        try:
            from erp.api.parent_portal.re_enrollment import _create_re_enrollment_announcement
            
            # Lấy thông tin năm học
            school_year = ""
            try:
                config_doc = frappe.get_doc("SIS Re-enrollment Config", submission.config_id)
                school_year_info = frappe.db.get_value(
                    "SIS School Year",
                    config_doc.school_year_id,
                    ["name_vn", "name_en"],
                    as_dict=True
                )
                if school_year_info:
                    school_year = school_year_info.name_vn or school_year_info.name_en or ""
            except:
                pass
            
            # Lấy answers từ submission để gửi vào announcement
            answers_for_announcement = []
            for answer in submission.answers:
                answers_for_announcement.append({
                    'question_text_vn': answer.question_text_vn,
                    'question_text_en': answer.question_text_en,
                    'selected_options_text_vn': answer.selected_options_text_vn,
                    'selected_options_text_en': answer.selected_options_text_en
                })
            
            _create_re_enrollment_announcement(
                student_id=submission.student_id,
                student_name=submission.student_name,
                student_code=submission.student_code,
                submission_data={
                    'decision': submission.decision,
                    'payment_type': submission.payment_type,
                    'discount_name': submission.selected_discount_name,
                    'discount_percent': submission.selected_discount_percent,
                    'reason': submission.not_re_enroll_reason,
                    'school_year': school_year,
                    'submitted_at': str(now()),
                    'status': submission.status or 'pending',
                    'answers': answers_for_announcement  # Câu trả lời khảo sát
                },
                is_update=True
            )
            logs.append("Đã gửi thông báo cập nhật quyết định cho phụ huynh")
        except Exception as notif_err:
            logs.append(f"Lỗi gửi thông báo: {str(notif_err)}")
            frappe.logger().error(f"Error sending decision update notification: {str(notif_err)}")
        
        return success_response(
            data={
                "name": submission.name,
                "decision": submission.decision,
                "payment_type": submission.payment_type,
                "reason": submission.not_re_enroll_reason
            },
            message="Cập nhật quyết định thành công",
            logs=logs
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Update Submission Decision Error")
        return error_response(
            message=f"Lỗi: {str(e)}",
            logs=logs
        )


# ==================== IMPORT PAYMENT STATUS API ====================

@frappe.whitelist(allow_guest=False, methods=['POST'])
def import_payment_status():
    """
    Import trạng thái thanh toán từ file Excel.
    
    File Excel gồm 2 cột:
    - Cột A: "Mã Học Sinh" (student_code, VD: WS02024251)
    - Cột B: "Tình Trạng" (Đã đóng tiền / Chưa đóng tiền / Hoàn tiền)
    
    POST body (multipart/form-data):
    {
        "config_id": "SIS-REENROLL-CFG-00001",
        "file": <Excel file>
    }
    
    Trả về:
    {
        "success": true,
        "data": {
            "success_count": 10,
            "error_count": 2,
            "total_count": 12,
            "errors": ["Dòng 3: Mã học sinh WS12345 không tồn tại", ...]
        }
    }
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        # Lấy config_id từ form data
        data = frappe.request.form
        config_id = data.get('config_id')
        
        if not config_id:
            return validation_error_response(
                "Thiếu config_id",
                {"config_id": ["Config ID là bắt buộc"]}
            )
        
        logs.append(f"Import payment status cho config: {config_id}")
        
        # Kiểm tra config tồn tại
        if not frappe.db.exists("SIS Re-enrollment Config", config_id):
            return not_found_response("Không tìm thấy cấu hình tái ghi danh")
        
        # Kiểm tra có file không
        if not frappe.request.files:
            return validation_error_response(
                "Thiếu file Excel",
                {"file": ["Vui lòng chọn file Excel để upload"]}
            )
        
        # Lấy file từ request
        file_obj = None
        for file_key, f in frappe.request.files.items():
            if file_key == 'file':
                file_obj = f
                break
        
        if not file_obj:
            return validation_error_response(
                "Thiếu file Excel",
                {"file": ["Vui lòng chọn file Excel để upload"]}
            )
        
        # Kiểm tra file là Excel
        filename = file_obj.filename
        if not (filename.lower().endswith('.xlsx') or filename.lower().endswith('.xls')):
            return validation_error_response(
                "Chỉ chấp nhận file Excel",
                {"file": ["Vui lòng upload file có định dạng .xlsx hoặc .xls"]}
            )
        
        logs.append(f"File: {filename}")
        
        # Đọc file Excel bằng openpyxl
        try:
            from openpyxl import load_workbook
            from io import BytesIO
            
            file_content = file_obj.stream.read()
            workbook = load_workbook(filename=BytesIO(file_content))
            sheet = workbook.active
            
        except Exception as e:
            logs.append(f"Lỗi đọc file Excel: {str(e)}")
            return error_response(
                message=f"Không thể đọc file Excel: {str(e)}",
                logs=logs
            )
        
        # Mapping trạng thái tiếng Việt sang giá trị database
        status_map = {
            'đã đóng tiền': 'paid',
            'đã đóng': 'paid',
            'paid': 'paid',
            'chưa đóng tiền': 'unpaid',
            'chưa đóng': 'unpaid',
            'unpaid': 'unpaid',
            'hoàn tiền': 'refunded',
            'refunded': 'refunded',
        }
        
        success_count = 0
        error_count = 0
        errors = []
        errors_preview = []  # Chi tiết lỗi với số dòng
        
        # Lấy user hiện tại để ghi log
        current_user = frappe.session.user
        current_user_name = frappe.db.get_value("User", current_user, "full_name") or current_user
        
        # Đọc từng dòng (bỏ qua header dòng 1)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
            # Bỏ qua dòng trống
            if not row or not row[0]:
                continue
            
            student_code = str(row[0]).strip() if row[0] else ''
            status_text = str(row[1]).strip().lower() if row[1] else ''
            
            # Validate dữ liệu
            if not student_code:
                error_msg = f"Dòng {row_idx}: Thiếu mã học sinh"
                errors.append(error_msg)
                errors_preview.append({
                    "row": row_idx,
                    "error": "Thiếu mã học sinh",
                    "data": {"Mã Học Sinh": student_code, "Tình Trạng": row[1] if len(row) > 1 else ''}
                })
                error_count += 1
                continue
            
            if not status_text:
                error_msg = f"Dòng {row_idx}: Thiếu tình trạng thanh toán cho {student_code}"
                errors.append(error_msg)
                errors_preview.append({
                    "row": row_idx,
                    "error": f"Thiếu tình trạng thanh toán",
                    "data": {"Mã Học Sinh": student_code, "Tình Trạng": ''}
                })
                error_count += 1
                continue
            
            # Map trạng thái
            payment_status = status_map.get(status_text)
            if not payment_status:
                error_msg = f"Dòng {row_idx}: Tình trạng '{row[1]}' không hợp lệ cho {student_code}. Chỉ chấp nhận: Đã đóng tiền, Chưa đóng tiền, Hoàn tiền"
                errors.append(error_msg)
                errors_preview.append({
                    "row": row_idx,
                    "error": f"Tình trạng '{row[1]}' không hợp lệ. Chỉ chấp nhận: Đã đóng tiền, Chưa đóng tiền, Hoàn tiền",
                    "data": {"Mã Học Sinh": student_code, "Tình Trạng": row[1] if len(row) > 1 else ''}
                })
                error_count += 1
                continue
            
            # Tìm submission theo student_code trong config hiện tại
            submission = frappe.db.get_value(
                "SIS Re-enrollment",
                {"config_id": config_id, "student_code": student_code},
                ["name", "payment_status"],
                as_dict=True
            )
            
            if not submission:
                error_msg = f"Dòng {row_idx}: Mã học sinh {student_code} không tồn tại trong đợt tái ghi danh này"
                errors.append(error_msg)
                errors_preview.append({
                    "row": row_idx,
                    "error": f"Mã học sinh {student_code} không tồn tại trong đợt tái ghi danh này",
                    "data": {"Mã Học Sinh": student_code, "Tình Trạng": row[1] if len(row) > 1 else ''}
                })
                error_count += 1
                continue
            
            # Cập nhật payment_status
            try:
                old_status = submission.payment_status
                
                # Chỉ cập nhật nếu khác trạng thái cũ
                if old_status != payment_status:
                    submission_doc = frappe.get_doc("SIS Re-enrollment", submission.name)
                    submission_doc.payment_status = payment_status
                    
                    # Ghi log thay đổi
                    payment_status_map = {'unpaid': 'Chưa đóng', 'paid': 'Đã đóng', 'refunded': 'Hoàn tiền'}
                    old_display = payment_status_map.get(old_status, old_status or 'Chưa có')
                    new_display = payment_status_map.get(payment_status, payment_status)
                    
                    log_content = f"Admin {current_user_name} đã cập nhật từ file Excel:\n• Thanh toán: {old_display} → {new_display}"
                    
                    submission_doc.append("notes", {
                        "note_type": "system_log",
                        "note": log_content,
                        "created_by_user": current_user,
                        "created_by_name": current_user_name,
                        "created_at": now()
                    })
                    
                    submission_doc.modified_by_admin = current_user
                    submission_doc.admin_modified_at = now()
                    submission_doc.save(ignore_permissions=True)
                
                success_count += 1
                
            except Exception as e:
                error_msg = f"Dòng {row_idx}: Lỗi cập nhật {student_code}: {str(e)}"
                errors.append(error_msg)
                errors_preview.append({
                    "row": row_idx,
                    "error": f"Lỗi cập nhật: {str(e)}",
                    "data": {"Mã Học Sinh": student_code, "Tình Trạng": row[1] if len(row) > 1 else ''}
                })
                error_count += 1
                continue
        
        frappe.db.commit()
        
        total_count = success_count + error_count
        logs.append(f"Kết quả: {success_count}/{total_count} thành công, {error_count} lỗi")
        
        return success_response(
            data={
                "success_count": success_count,
                "error_count": error_count,
                "total_count": total_count,
                "errors": errors,
                "errors_preview": errors_preview[:20]  # Chỉ trả về 20 lỗi đầu tiên
            },
            message=f"Import hoàn tất: {success_count}/{total_count} thành công" + (f", {error_count} lỗi" if error_count > 0 else ""),
            logs=logs
        )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Import Payment Status Error")
        return error_response(
            message=f"Lỗi khi import: {str(e)}",
            logs=logs
        )


# ==================== SEND REMINDER NOTIFICATION API ====================

@frappe.whitelist(allow_guest=False, methods=['POST'])
def send_reminder_notification():
    """
    Gửi push notification nhắc phụ huynh làm đơn tái ghi danh.
    
    POST body:
    {
        "submission_ids": ["SIS-REENROLL-00001", ...],  # Danh sách ID đơn cần nhắc
        "message": "Nội dung tin nhắn tùy chỉnh"       # Nội dung do user nhập
    }
    
    Trả về:
    {
        "success": true,
        "data": {
            "success_count": 10,
            "failed_count": 2,
            "total_students": 12
        }
    }
    """
    logs = []
    
    try:
        if not _check_admin_permission():
            return error_response("Bạn không có quyền truy cập", logs=logs)
        
        # Lấy data từ request
        if frappe.request.is_json:
            data = frappe.request.json or {}
        else:
            data = frappe.form_dict
        
        submission_ids = data.get('submission_ids', [])
        message = data.get('message', '')
        
        if isinstance(submission_ids, str):
            submission_ids = json.loads(submission_ids)
        
        if not submission_ids:
            return validation_error_response(
                "Thiếu danh sách đơn",
                {"submission_ids": ["Danh sách đơn là bắt buộc"]}
            )
        
        if not message or not message.strip():
            return validation_error_response(
                "Thiếu nội dung tin nhắn",
                {"message": ["Nội dung tin nhắn là bắt buộc"]}
            )
        
        # Giới hạn độ dài message (150 ký tự để hiển thị tốt trên mobile)
        message = message.strip()[:150]
        
        logs.append(f"Gửi nhắc nhở cho {len(submission_ids)} đơn")
        logs.append(f"Message: {message}")
        
        # Lấy danh sách student_id từ submission_ids
        student_ids = []
        for sub_id in submission_ids:
            student_id = frappe.db.get_value("SIS Re-enrollment", sub_id, "student_id")
            if student_id:
                student_ids.append(student_id)
        
        if not student_ids:
            return error_response("Không tìm thấy học sinh nào", logs=logs)
        
        logs.append(f"Tìm thấy {len(student_ids)} học sinh")
        
        # Gửi push notification
        try:
            from erp.utils.notification_handler import send_bulk_parent_notifications
            
            # Sử dụng notification_type = "reminder" (là giá trị hợp lệ trong ERP Notification)
            result = send_bulk_parent_notifications(
                recipient_type="reminder",
                recipients_data={
                    "student_ids": student_ids
                },
                title="Tái ghi danh",
                body=message,
                icon="/icon.png",
                data={
                    "type": "reminder",
                    "subtype": "re_enrollment",
                    "url": "/re-enrollment"  # URL trên parent-portal
                }
            )
            
            logs.append(f"Kết quả gửi: {result}")
            
            return success_response(
                data={
                    "success_count": result.get("success_count", 0),
                    "failed_count": result.get("failed_count", 0),
                    "total_students": len(student_ids),
                    "total_parents": result.get("total_parents", 0)
                },
                message=f"Đã gửi thông báo đến {result.get('success_count', 0)} phụ huynh",
                logs=logs
            )
            
        except Exception as notif_err:
            logs.append(f"Lỗi gửi notification: {str(notif_err)}")
            frappe.log_error(frappe.get_traceback(), "Send Re-enrollment Reminder Error")
            return error_response(
                message=f"Lỗi khi gửi thông báo: {str(notif_err)}",
                logs=logs
            )
        
    except Exception as e:
        logs.append(f"Lỗi: {str(e)}")
        frappe.log_error(frappe.get_traceback(), "Send Reminder Notification Error")
        return error_response(
            message=f"Lỗi: {str(e)}",
            logs=logs
        )

