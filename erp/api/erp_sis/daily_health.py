# Copyright (c) 2025, Wellspring International School and contributors
# For license information, please see license.txt

"""
Daily Health Visit API for SIS
Handles student visits to health clinic
"""

import frappe
from frappe import _
from frappe.utils import today, now, get_datetime, nowtime
import json
from erp.utils.api_response import (
    success_response,
    error_response,
    validation_error_response
)


def _check_teacher_permission():
    """Check if user has teacher permission"""
    user_roles = frappe.get_roles()
    allowed_roles = ["System Manager", "SIS Manager", "SIS Teacher", "SIS Administrative"]

    if not any(role in allowed_roles for role in user_roles):
        frappe.throw(_("Bạn không có quyền truy cập API này"), frappe.PermissionError)


def _check_medical_permission():
    """Check if user has medical staff permission (Y tế từ chối)"""
    user_roles = frappe.get_roles()
    allowed_roles = ["System Manager", "SIS Manager", "SIS Medical"]

    if not any(role in allowed_roles for role in user_roles):
        frappe.throw(_("Bạn không có quyền truy cập API này"), frappe.PermissionError)


def _get_request_data():
    """Get request data from various sources"""
    data = {}
    
    # 1. Lấy từ URL query params (GET request)
    if hasattr(frappe, 'request') and hasattr(frappe.request, 'args') and frappe.request.args:
        data.update(dict(frappe.request.args))
    
    # 2. Lấy từ form_dict (query params và form data)
    if frappe.local.form_dict:
        data.update(dict(frappe.local.form_dict))
    
    # 3. Merge với JSON body nếu có (POST request)
    if hasattr(frappe.request, 'is_json') and frappe.request.is_json:
        json_data = frappe.request.json or {}
        data.update(json_data)
    else:
        try:
            if hasattr(frappe.request, 'data') and frappe.request.data:
                raw = frappe.request.data
                body = raw.decode('utf-8') if isinstance(raw, (bytes, bytearray)) else raw
                if body:
                    parsed = json.loads(body)
                    if isinstance(parsed, dict):
                        data.update(parsed)
        except (json.JSONDecodeError, AttributeError, TypeError):
            pass
    
    return data


@frappe.whitelist(allow_guest=False)
def report_student_to_clinic():
    """
    Giáo viên báo cáo học sinh xuống Y tế
    Params:
        - student_id: ID học sinh (required)
        - class_id: ID lớp (required)
        - reason: Lý do xuống Y tế (required)
        - leave_class_time: Thời gian rời lớp (optional, default: now)
        - period: Tên tiết học (optional) - nếu có sẽ tự động cập nhật attendance thành excused
        - initial_status: Trạng thái ban đầu (optional, default: left_class)
                         - "left_class": Đã rời lớp (giáo viên báo)
                         - "at_clinic": Đang ở Y tế (Y tế tự tạo khi học sinh trực tiếp xuống)
    """
    try:
        _check_teacher_permission()
        
        data = _get_request_data()
        
        student_id = data.get("student_id")
        class_id = data.get("class_id")
        reason = data.get("reason")
        leave_class_time = data.get("leave_class_time") or nowtime()
        period = data.get("period")  # Tên tiết học (optional)
        report_date = data.get("date")  # Ngày báo cáo (optional, dùng để validate)
        initial_status = data.get("initial_status", "left_class")  # Trạng thái ban đầu
        
        # Validate initial_status
        valid_statuses = ["left_class", "at_clinic"]
        if initial_status not in valid_statuses:
            initial_status = "left_class"
        
        # Chặn báo Y tế cho ngày quá khứ
        if report_date and report_date != today():
            return validation_error_response(
                "Chỉ được báo Y tế cho ngày hôm nay",
                {"date": ["Không thể báo Y tế cho ngày đã qua"]}
            )
        
        # Validation
        errors = {}
        if not student_id:
            errors["student_id"] = ["student_id là bắt buộc"]
        if not class_id:
            errors["class_id"] = ["class_id là bắt buộc"]
        if not reason:
            errors["reason"] = ["reason là bắt buộc"]
        
        if errors:
            return validation_error_response("Dữ liệu không hợp lệ", errors)
        
        # Lấy thông tin student và class
        student = frappe.db.get_value("CRM Student", student_id, ["student_name", "student_code"], as_dict=True)
        class_info = frappe.db.get_value("SIS Class", class_id, ["title", "campus_id"], as_dict=True)
        
        if not student:
            return validation_error_response("Học sinh không tồn tại", {"student_id": ["Học sinh không tồn tại"]})
        if not class_info:
            return validation_error_response("Lớp không tồn tại", {"class_id": ["Lớp không tồn tại"]})
        
        # Lấy thông tin user
        reported_by_user = frappe.session.user
        reported_by_name = ""
        try:
            user = frappe.db.get_value("User", frappe.session.user, ["full_name"], as_dict=True)
            if user:
                reported_by_name = user.get("full_name") or frappe.session.user
        except:
            reported_by_name = frappe.session.user
        
        # Tạo bản ghi visit
        visit_data = {
            "doctype": "SIS Daily Health Visit",
            "student_id": student_id,
            "student_name": student.get("student_name"),
            "student_code": student.get("student_code"),
            "class_id": class_id,
            "class_name": class_info.get("title"),
            "visit_date": today(),
            "reason": reason,
            "leave_class_time": leave_class_time,
            "status": initial_status,
            "reported_by": reported_by_user,
            "reported_by_name": reported_by_name
        }
        
        # Nếu Y tế tự tạo (at_clinic), cập nhật thêm thông tin tiếp nhận
        if initial_status == "at_clinic":
            visit_data["arrived_at_clinic_time"] = nowtime()
            visit_data["received_by"] = reported_by_user
            visit_data["received_by_name"] = reported_by_name
        
        visit = frappe.get_doc(visit_data)
        visit.insert()
        
        # Nếu có period, tự động cập nhật attendance thành excused và lưu attendance_record_id để revert khi hủy/từ chối
        attendance_updated = False
        attendance_record_id = None
        if period:
            try:
                existing_attendance = frappe.db.exists(
                    "SIS Class Attendance",
                    {
                        "student_id": student_id,
                        "class_id": class_id,
                        "date": today(),
                        "period": period
                    }
                )
                
                if existing_attendance:
                    # Cập nhật attendance hiện có thành excused
                    frappe.db.set_value(
                        "SIS Class Attendance",
                        existing_attendance,
                        "status",
                        "excused"
                    )
                    attendance_updated = True
                    attendance_record_id = existing_attendance
                else:
                    # Tạo mới attendance với status excused
                    attendance = frappe.get_doc({
                        "doctype": "SIS Class Attendance",
                        "student_id": student_id,
                        "student_code": student.get("student_code"),
                        "student_name": student.get("student_name"),
                        "class_id": class_id,
                        "date": today(),
                        "period": period,
                        "status": "excused",
                        "remarks": f"Xuống Y tế: {reason}",
                        "campus_id": class_info.get("campus_id"),
                        "recorded_by": reported_by_user
                    })
                    attendance.insert()
                    attendance_updated = True
                    attendance_record_id = attendance.name
                
                # Lưu attendance_record_id vào visit để revert khi cancel/reject
                if attendance_record_id:
                    visit.db_set("attendance_record_id", attendance_record_id)
                    
                frappe.logger().info(f"[report_student_to_clinic] Updated attendance to excused for student {student_id}, period {period}")
            except Exception as att_err:
                # Không fail toàn bộ request nếu không thể cập nhật attendance
                frappe.logger().warning(f"[report_student_to_clinic] Could not update attendance: {str(att_err)}")
        
        frappe.db.commit()
        
        # Gửi push notification cho Mobile Medical + Homeroom + Vice-homeroom
        try:
            from erp.api.erp_sis.daily_health_notification import notify_health_visit_created
            notify_health_visit_created(visit_name=visit.name)
        except Exception as notif_err:
            frappe.logger().warning(f"[report_student_to_clinic] Không gửi được notification: {str(notif_err)}")
        
        return success_response(
            data={
                "name": visit.name,
                "status": visit.status,
                "attendance_updated": attendance_updated
            },
            message="Đã báo cáo học sinh xuống Y tế"
        )
    
    except frappe.ValidationError as e:
        return validation_error_response(str(e), {})
    except Exception as e:
        frappe.db.rollback()
        frappe.logger().error(f"Error reporting student to clinic: {str(e)}")
        import traceback
        frappe.logger().error(traceback.format_exc())
        return error_response(
            message=f"Lỗi khi báo cáo học sinh xuống Y tế: {str(e)}",
            code="REPORT_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def get_daily_health_visits():
    """
    Lấy danh sách học sinh đang ở/đã ở Y tế (cho trang DailyHealth)
    Params:
        - date: Ngày (optional, default: today)
        - campus: Campus filter (optional)
        - education_stage: Filter theo cấp học (optional)
        - status: Filter theo status (optional)
        - search: Tìm kiếm theo tên/mã học sinh/lớp (optional)
    """
    try:
        _check_teacher_permission()
        
        data = frappe.local.form_dict
        request_args = frappe.request.args
        
        visit_date = data.get("date") or request_args.get("date") or today()
        campus = data.get("campus") or request_args.get("campus")
        education_stage = data.get("education_stage") or request_args.get("education_stage")
        status_filter = data.get("status") or request_args.get("status")
        search = data.get("search") or request_args.get("search")
        
        # Build filters
        filters = {"visit_date": visit_date}
        if status_filter:
            filters["status"] = status_filter
        
        # Nếu có education_stage filter (ưu tiên như Sổ đầu bài), lấy class_id theo cấp học
        if education_stage:
            # Lấy education grades thuộc stage này
            grade_ids = frappe.get_all(
                "SIS Education Grade",
                filters={"education_stage_id": education_stage},
                pluck="name"
            )
            if grade_ids:
                stage_class_ids = frappe.get_all(
                    "SIS Class",
                    filters={"education_grade": ["in", grade_ids]},
                    pluck="name"
                )
                if stage_class_ids:
                    filters["class_id"] = ["in", stage_class_ids]
                else:
                    return success_response(
                        data={"data": [], "total": 0},
                        message="Lấy danh sách học sinh xuống Y tế thành công"
                    )
            else:
                return success_response(
                    data={"data": [], "total": 0},
                    message="Lấy danh sách học sinh xuống Y tế thành công"
                )
        # Nếu có campus filter (fallback), lấy danh sách class_id
        elif campus:
            campus_class_ids = frappe.get_all(
                "SIS Class",
                filters={"campus_id": campus},
                pluck="name"
            )
            if campus_class_ids:
                filters["class_id"] = ["in", campus_class_ids]
            else:
                return success_response(
                    data={"data": [], "total": 0},
                    message="Lấy danh sách học sinh xuống Y tế thành công"
                )
        
        # Query visits
        visits = frappe.get_all(
            "SIS Daily Health Visit",
            filters=filters,
            fields=[
                "name", "student_id", "student_name", "student_code",
                "class_id", "class_name", "visit_date", "reason",
                "leave_class_time", "arrive_clinic_time", "leave_clinic_time",
                "status", "reported_by", "reported_by_name",
                "received_by", "received_by_name", "creation",
                "checkout_notes", "transfer_hospital",
                "accompanying_teacher", "accompanying_health_staff"
            ],
            order_by="leave_class_time desc"
        )
        
        # Filter theo search term
        if search:
            search_lower = search.lower()
            visits = [v for v in visits if 
                     search_lower in (v.get("student_name") or "").lower() or
                     search_lower in (v.get("student_code") or "").lower() or
                     search_lower in (v.get("class_name") or "").lower()]
        
        # Lấy ảnh học sinh cho mỗi visit
        current_school_year = frappe.db.get_value(
            "SIS School Year",
            {"is_enable": 1},
            "name"
        )
        
        for visit in visits:
            student_photo = None
            try:
                # Ưu tiên lấy ảnh theo năm học hiện tại
                if current_school_year:
                    photos = frappe.get_all(
                        "SIS Photo",
                        filters={
                            "student_id": visit.get("student_id"),
                            "school_year_id": current_school_year
                        },
                        fields=["photo"],
                        order_by="creation desc",
                        limit=1
                    )
                    if photos and photos[0].get("photo"):
                        student_photo = photos[0]["photo"]
                
                # Fallback - lấy ảnh mới nhất nếu không có ảnh năm học
                if not student_photo:
                    photos = frappe.get_all(
                        "SIS Photo",
                        filters={"student_id": visit.get("student_id")},
                        fields=["photo"],
                        order_by="creation desc",
                        limit=1
                    )
                    if photos and photos[0].get("photo"):
                        student_photo = photos[0]["photo"]
            except Exception as photo_err:
                frappe.logger().warning(f"Error fetching photo for student {visit.get('student_id')}: {str(photo_err)}")
            
            visit["student_photo"] = student_photo
        
        # Piggyback: kiểm tra visit quá 15 phút chưa chuyển trạng thái (rate limited)
        try:
            from erp.api.erp_sis.daily_health_notification import piggyback_check_stale_visits
            piggyback_check_stale_visits()
        except Exception:
            pass
        
        return success_response(
            data={"data": visits, "total": len(visits)},
            message="Lấy danh sách học sinh xuống Y tế thành công"
        )
    
    except Exception as e:
        frappe.logger().error(f"Error getting daily health visits: {str(e)}")
        return error_response(
            message=f"Lỗi khi lấy danh sách xuống Y tế: {str(e)}",
            code="GET_VISITS_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def receive_student_at_clinic():
    """
    Nhân viên Y tế tiếp nhận học sinh
    Params:
        - visit_id: ID của visit (required)
    """
    try:
        _check_teacher_permission()
        
        data = _get_request_data()
        
        visit_id = data.get("visit_id")
        
        if not visit_id:
            return validation_error_response("visit_id là bắt buộc", {"visit_id": ["visit_id là bắt buộc"]})
        
        # Lấy visit
        visit = frappe.get_doc("SIS Daily Health Visit", visit_id)
        
        # Lấy thông tin user
        received_by_user = frappe.session.user
        received_by_name = ""
        try:
            user = frappe.db.get_value("User", frappe.session.user, ["full_name"], as_dict=True)
            if user:
                received_by_name = user.get("full_name") or frappe.session.user
        except:
            received_by_name = frappe.session.user
        
        # Cập nhật visit
        visit.status = "at_clinic"
        visit.arrive_clinic_time = nowtime()
        visit.received_by = received_by_user
        visit.received_by_name = received_by_name
        visit.save()
        frappe.db.commit()
        
        # Gửi push notification cho Homeroom + Vice-homeroom + Reporter
        try:
            from erp.api.erp_sis.daily_health_notification import notify_health_visit_received
            notify_health_visit_received(visit_name=visit.name)
        except Exception as notif_err:
            frappe.logger().warning(f"[receive_student_at_clinic] Không gửi được notification: {str(notif_err)}")
        
        return success_response(
            data={"name": visit.name, "status": visit.status},
            message="Đã tiếp nhận học sinh tại phòng Y tế"
        )
    
    except frappe.DoesNotExistError:
        return error_response(
            message="Bản ghi xuống Y tế không tồn tại",
            code="NOT_FOUND"
        )
    except Exception as e:
        frappe.db.rollback()
        frappe.logger().error(f"Error receiving student at clinic: {str(e)}")
        return error_response(
            message=f"Lỗi khi tiếp nhận học sinh: {str(e)}",
            code="RECEIVE_ERROR"
        )


def _revert_attendance_for_visit(visit):
    """
    Revert attendance về present khi cancel/reject visit.
    Ưu tiên dùng attendance_record_id. Nếu null (visit cũ) thì fallback: query theo
    student+class+date+excused+remarks chứa "Xuống Y tế" - chỉ revert khi tìm đúng 1 bản ghi.
    """
    att_id = visit.get("attendance_record_id")
    if not att_id:
        # Fallback cho visit cũ (chưa có attendance_record_id)
        candidates = frappe.get_all(
            "SIS Class Attendance",
            filters={
                "student_id": visit.student_id,
                "class_id": visit.class_id,
                "date": visit.visit_date,
                "status": "excused",
                "remarks": ["like", "%Xuống Y tế%"]
            },
            pluck="name"
        )
        if len(candidates) == 1:
            att_id = candidates[0]
            frappe.logger().info(f"[_revert_attendance_for_visit] Fallback: tìm thấy 1 attendance {att_id} cho visit {visit.name}")
        else:
            frappe.logger().warning(
                f"[_revert_attendance_for_visit] Visit {visit.name} không có attendance_record_id, "
                f"fallback tìm thấy {len(candidates)} bản ghi - bỏ qua revert"
            )
            return False
    try:
        frappe.db.set_value(
            "SIS Class Attendance",
            att_id,
            "status",
            "present"
        )
        frappe.logger().info(f"[_revert_attendance_for_visit] Đã revert attendance {att_id} về present")
        return True
    except Exception as e:
        frappe.logger().warning(f"[_revert_attendance_for_visit] Không thể revert: {str(e)}")
        return False


@frappe.whitelist(allow_guest=False)
def cancel_health_visit():
    """
    GV hủy đơn báo Y tế (học sinh quay lại lớp / trốn đi chơi / không xuống Y tế)
    Chỉ cho phép khi status = left_class (chưa Y tế tiếp nhận)
    Params:
        - visit_id: ID của visit (required)
        - reason: Lý do hủy (optional)
    """
    try:
        _check_teacher_permission()
        
        data = _get_request_data()
        visit_id = data.get("visit_id")
        reason = data.get("reason", "")
        
        if not visit_id:
            return validation_error_response("visit_id là bắt buộc", {"visit_id": ["visit_id là bắt buộc"]})
        
        visit = frappe.get_doc("SIS Daily Health Visit", visit_id)
        
        if visit.status != "left_class":
            return error_response(
                message="Chỉ có thể hủy đơn khi chưa được Y tế tiếp nhận",
                code="INVALID_STATUS"
            )
        
        visit.status = "cancelled"
        if reason:
            visit.checkout_notes = f"[GV hủy] {reason}"
        visit.save()
        
        _revert_attendance_for_visit(visit)
        frappe.db.commit()
        
        # Gửi push notification cho Mobile Medical (Hủy y tế)
        try:
            from erp.api.erp_sis.daily_health_notification import notify_health_visit_cancelled
            notify_health_visit_cancelled(visit_name=visit.name)
        except Exception as notif_err:
            frappe.logger().warning(f"[cancel_health_visit] Không gửi được notification: {str(notif_err)}")
        
        return success_response(
            data={"name": visit.name, "status": visit.status},
            message="Đã hủy đơn báo Y tế"
        )
    
    except frappe.DoesNotExistError:
        return error_response(message="Bản ghi xuống Y tế không tồn tại", code="NOT_FOUND")
    except Exception as e:
        frappe.db.rollback()
        frappe.logger().error(f"Error cancelling health visit: {str(e)}")
        return error_response(
            message=f"Lỗi khi hủy đơn: {str(e)}",
            code="CANCEL_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def reject_health_visit():
    """
    Y tế từ chối tiếp nhận học sinh.
    Chuyển status = rejected (Từ chối) để phân biệt với returned (Đã về lớp sau khi khám).
    Không revert attendance - học sinh cần thời gian di chuyển về lớp → GV điểm danh lại khi thấy.
    Params:
        - visit_id: ID của visit (required)
        - reject_reason: Lý do từ chối (optional)
    """
    try:
        _check_medical_permission()
        
        data = _get_request_data()
        visit_id = data.get("visit_id")
        reject_reason = data.get("reject_reason", "")
        
        if not visit_id:
            return validation_error_response("visit_id là bắt buộc", {"visit_id": ["visit_id là bắt buộc"]})
        
        visit = frappe.get_doc("SIS Daily Health Visit", visit_id)
        
        if visit.status != "left_class":
            return error_response(
                message="Chỉ có thể từ chối khi đơn chưa được tiếp nhận",
                code="INVALID_STATUS"
            )
        
        # Chuyển sang rejected (Từ chối) - phân biệt với returned (Đã về lớp sau khi khám)
        # Học sinh đang trên đường về lớp, GV sẽ điểm danh lại khi thấy học sinh
        visit.status = "rejected"
        visit.leave_clinic_time = nowtime()
        if reject_reason:
            visit.checkout_notes = f"[Y tế từ chối] {reject_reason}"
        visit.save()
        
        frappe.db.commit()
        
        # Gửi push notification cho Homeroom + Reporter (Y tế từ chối - Hủy y tế)
        try:
            from erp.api.erp_sis.daily_health_notification import notify_health_visit_rejected
            notify_health_visit_rejected(visit_name=visit.name)
        except Exception as notif_err:
            frappe.logger().warning(f"[reject_health_visit] Không gửi được notification: {str(notif_err)}")
        
        return success_response(
            data={"name": visit.name, "status": visit.status},
            message="Đã từ chối tiếp nhận - học sinh đang về lớp, GV điểm danh lại khi thấy học sinh"
        )
    
    except frappe.DoesNotExistError:
        return error_response(message="Bản ghi xuống Y tế không tồn tại", code="NOT_FOUND")
    except Exception as e:
        frappe.db.rollback()
        frappe.logger().error(f"Error rejecting health visit: {str(e)}")
        return error_response(
            message=f"Lỗi khi từ chối: {str(e)}",
            code="REJECT_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def update_visit_reason():
    """
    Cập nhật lý do báo cáo y tế
    Params:
        - visit_id: ID của visit (required)
        - reason: Lý do mới (required)
    """
    try:
        _check_teacher_permission()
        
        data = _get_request_data()
        
        visit_id = data.get("visit_id")
        reason = data.get("reason")
        
        if not visit_id:
            return validation_error_response("visit_id là bắt buộc", {"visit_id": ["visit_id là bắt buộc"]})
        
        if not reason or not str(reason).strip():
            return validation_error_response("reason là bắt buộc", {"reason": ["reason là bắt buộc"]})
        
        # Lấy và cập nhật visit
        visit = frappe.get_doc("SIS Daily Health Visit", visit_id)
        visit.reason = str(reason).strip()
        visit.save()
        frappe.db.commit()
        
        return success_response(
            data={"name": visit.name, "reason": visit.reason},
            message="Đã cập nhật lý do báo cáo y tế"
        )
    
    except frappe.DoesNotExistError:
        return error_response(
            message="Bản ghi xuống Y tế không tồn tại",
            code="NOT_FOUND"
        )
    except Exception as e:
        frappe.db.rollback()
        frappe.logger().error(f"Error updating visit reason: {str(e)}")
        return error_response(
            message=f"Lỗi khi cập nhật lý do: {str(e)}",
            code="UPDATE_REASON_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def start_examination():
    """
    Bắt đầu khám - chuyển status sang examining
    Params:
        - visit_id: ID của visit (required)
    """
    try:
        _check_teacher_permission()
        
        data = _get_request_data()
        
        visit_id = data.get("visit_id")
        
        if not visit_id:
            return validation_error_response("visit_id là bắt buộc", {"visit_id": ["visit_id là bắt buộc"]})
        
        # Lấy visit
        visit = frappe.get_doc("SIS Daily Health Visit", visit_id)
        
        # Chỉ cho phép start examination khi status là at_clinic
        if visit.status not in ["at_clinic", "examining"]:
            return error_response(
                message="Chỉ có thể bắt đầu khám khi học sinh đã được tiếp nhận",
                code="INVALID_STATUS"
            )
        
        # Cập nhật visit status sang examining
        visit.status = "examining"
        visit.save()
        frappe.db.commit()
        
        return success_response(
            data={"name": visit.name, "status": visit.status},
            message="Đã bắt đầu khám"
        )
    
    except frappe.DoesNotExistError:
        return error_response(
            message="Bản ghi xuống Y tế không tồn tại",
            code="NOT_FOUND"
        )
    except Exception as e:
        frappe.db.rollback()
        frappe.logger().error(f"Error starting examination: {str(e)}")
        return error_response(
            message=f"Lỗi khi bắt đầu khám: {str(e)}",
            code="START_EXAM_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def create_health_examination():
    """
    Tạo hồ sơ khám mới
    Params:
        - visit_id: ID của visit (required)
        - symptoms: Triệu chứng (required)
        - images: List of image URLs (optional)
        - disease_classification: Phân loại bệnh (optional)
        - examination_notes: Nhận định ban đầu (optional)
        - treatment_type: Loại điều trị (optional)
        - treatment_details: Chi tiết chăm sóc y tế (optional)
        - notes: Ghi chú (optional)
        - medical_staff: NVYT thăm khám - User ID (optional)
        - clinic_checkin_time: Thời gian vào y tế (optional)
        - clinic_checkout_time: Thời gian về (optional)
        - diagnosis: Chẩn đoán (optional) - deprecated
        - treatment: Xử lý/Dặn dò (optional) - deprecated
        - outcome: Kết quả (optional)
    """
    try:
        _check_teacher_permission()
        
        data = _get_request_data()
        
        visit_id = data.get("visit_id")
        symptoms = data.get("symptoms")
        diet_history = data.get("diet_history", "")
        images = data.get("images", [])
        disease_classification = data.get("disease_classification", "")
        examination_notes = data.get("examination_notes", "")
        treatment_type = data.get("treatment_type", "")
        treatment_details = data.get("treatment_details", "")
        notes = data.get("notes", "")
        
        # Trường mới: NVYT thăm khám, thời gian vào/về
        medical_staff = data.get("medical_staff", "")
        clinic_checkin_time = data.get("clinic_checkin_time", "")
        clinic_checkout_time = data.get("clinic_checkout_time", "")
        
        # Backward compatibility
        diagnosis = data.get("diagnosis", "")
        treatment = data.get("treatment", "")
        outcome = data.get("outcome")
        
        # Validation
        errors = {}
        if not visit_id:
            errors["visit_id"] = ["visit_id là bắt buộc"]
        if not symptoms:
            errors["symptoms"] = ["symptoms là bắt buộc"]
        
        if errors:
            return validation_error_response("Dữ liệu không hợp lệ", errors)
        
        # Lấy visit để lấy thông tin student
        visit = frappe.get_doc("SIS Daily Health Visit", visit_id)
        
        # Lấy thông tin user hiện tại (fallback nếu không chọn NVYT)
        examined_by_user = frappe.session.user
        examined_by_name = ""
        try:
            user = frappe.db.get_value("User", frappe.session.user, ["full_name"], as_dict=True)
            if user:
                examined_by_name = user.get("full_name") or frappe.session.user
        except:
            examined_by_name = frappe.session.user
        
        # Resolve tên NVYT thăm khám
        medical_staff_name = ""
        if medical_staff:
            try:
                ms_user = frappe.db.get_value("User", medical_staff, ["full_name"], as_dict=True)
                if ms_user:
                    medical_staff_name = ms_user.get("full_name") or medical_staff
            except:
                medical_staff_name = medical_staff
        
        # Tạo examination
        doc_dict = {
            "doctype": "SIS Health Examination",
            "student_id": visit.student_id,
            "student_name": visit.student_name,
            "student_code": visit.student_code,
            "examination_date": today(),
            "visit_id": visit_id,
            "symptoms": symptoms,
            "diet_history": diet_history,
            "disease_classification": disease_classification,
            "examination_notes": examination_notes,
            "treatment_type": treatment_type,
            "treatment_details": treatment_details,
            "notes": notes,
            "diagnosis": diagnosis,  # Deprecated
            "treatment": treatment,  # Deprecated
            "outcome": outcome,
            "examined_by": examined_by_user,
            "examined_by_name": examined_by_name,
        }
        if medical_staff:
            doc_dict["medical_staff"] = medical_staff
            doc_dict["medical_staff_name"] = medical_staff_name
        if clinic_checkin_time:
            doc_dict["clinic_checkin_time"] = clinic_checkin_time
        if clinic_checkout_time:
            doc_dict["clinic_checkout_time"] = clinic_checkout_time
        exam = frappe.get_doc(doc_dict)
        
        # Add images as child table if provided
        if images and isinstance(images, list):
            for img_data in images:
                if isinstance(img_data, dict) and img_data.get("image"):
                    exam.append("images", {
                        "image": img_data.get("image"),
                        "description": img_data.get("description", "")
                    })
        
        exam.insert()
        
        # Frappe/MariaDB tự điền CURRENT_TIME cho các trường Time không được set.
        # Cần NULL hóa các trường Time mà user không nhập.
        time_fields_to_clear = [
            "followup_checkin_time", "followup_checkout_time",
            "followup_clinic_checkin_time", "followup_clinic_checkout_time",
        ]
        if not clinic_checkin_time:
            time_fields_to_clear.append("clinic_checkin_time")
        if not clinic_checkout_time:
            time_fields_to_clear.append("clinic_checkout_time")
        
        if time_fields_to_clear:
            set_clause = ", ".join([f"`{f}` = NULL" for f in time_fields_to_clear])
            frappe.db.sql(
                f"UPDATE `tabSIS Health Examination` SET {set_clause} WHERE name = %s",
                exam.name
            )
        
        frappe.db.commit()
        
        return success_response(
            data={"name": exam.name},
            message="Tạo hồ sơ khám thành công"
        )
    
    except frappe.DoesNotExistError:
        return error_response(
            message="Bản ghi xuống Y tế không tồn tại",
            code="NOT_FOUND"
        )
    except Exception as e:
        frappe.db.rollback()
        frappe.logger().error(f"Error creating health examination: {str(e)}")
        import traceback
        frappe.logger().error(traceback.format_exc())
        return error_response(
            message=f"Lỗi khi tạo hồ sơ khám: {str(e)}",
            code="CREATE_EXAM_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def update_health_examination():
    """
    Cập nhật hồ sơ khám
    """
    try:
        _check_teacher_permission()
        
        data = _get_request_data()
        
        frappe.logger().info(f"[update_health_examination] Received data: {data}")
        
        exam_id = data.get("exam_id")
        symptoms = data.get("symptoms")
        diet_history = data.get("diet_history")
        images = data.get("images")
        disease_classification = data.get("disease_classification")
        examination_notes = data.get("examination_notes")
        treatment_type = data.get("treatment_type")
        treatment_details = data.get("treatment_details")
        notes = data.get("notes")
        hospital_diagnosis = data.get("hospital_diagnosis")
        hospital_treatment = data.get("hospital_treatment")
        
        # Trường mới: NVYT thăm khám, thời gian vào/về
        medical_staff = data.get("medical_staff")
        clinic_checkin_time = data.get("clinic_checkin_time")
        clinic_checkout_time = data.get("clinic_checkout_time")
        
        # Hospital fields mới
        hospital_insurance = data.get("hospital_insurance")
        hospital_school_coordination = data.get("hospital_school_coordination")
        hospital_medical_staff = data.get("hospital_medical_staff")
        hospital_direction = data.get("hospital_direction")
        hospital_advance_cost = data.get("hospital_advance_cost")
        hospital_payer = data.get("hospital_payer")
        hospital_payer_other = data.get("hospital_payer_other")
        hospital_transport = data.get("hospital_transport")
        hospital_transport_other = data.get("hospital_transport_other")
        hospital_health_monitoring = data.get("hospital_health_monitoring")
        hospital_notes = data.get("hospital_notes")
        
        # Followup fields
        followup_checkin_time = data.get("followup_checkin_time")
        followup_examination = data.get("followup_examination")
        followup_treatment_details = data.get("followup_treatment_details")
        followup_checkout_time = data.get("followup_checkout_time")
        followup_outcome = data.get("followup_outcome")
        followup_notes = data.get("followup_notes")
        followup_transfer_hospital = data.get("followup_transfer_hospital")
        followup_accompanying_teacher = data.get("followup_accompanying_teacher")
        followup_accompanying_health_staff = data.get("followup_accompanying_health_staff")
        
        # Followup fields mới
        followup_clinic_checkin_time = data.get("followup_clinic_checkin_time")
        followup_clinic_checkout_time = data.get("followup_clinic_checkout_time")
        followup_is_scheduled_recheck = data.get("followup_is_scheduled_recheck")
        followup_medical_suggestion = data.get("followup_medical_suggestion")
        followup_medical_staff = data.get("followup_medical_staff")
        
        # Backward compatibility
        diagnosis = data.get("diagnosis")
        treatment = data.get("treatment")
        outcome = data.get("outcome")
        
        if not exam_id:
            return validation_error_response("exam_id là bắt buộc", {"exam_id": ["exam_id là bắt buộc"]})
        
        # Lấy examination
        exam = frappe.get_doc("SIS Health Examination", exam_id)
        
        frappe.logger().info(f"[update_health_examination] Found exam: {exam.name}")
        
        # --- Helper để set field ---
        def _set(field, value):
            if value is not None:
                setattr(exam, field, value if value else None)
        
        # Cập nhật các trường cơ bản
        if symptoms is not None:
            exam.symptoms = symptoms
        _set("diet_history", diet_history)
        _set("disease_classification", disease_classification)
        _set("examination_notes", examination_notes)
        _set("treatment_type", treatment_type)
        _set("treatment_details", treatment_details)
        _set("notes", notes)
        _set("diagnosis", diagnosis)
        _set("treatment", treatment)
        if outcome is not None:
            exam.outcome = outcome
        _set("hospital_diagnosis", hospital_diagnosis)
        _set("hospital_treatment", hospital_treatment)
        
        # Trường mới: NVYT thăm khám, thời gian vào/về
        _set("medical_staff", medical_staff)
        _set("clinic_checkin_time", clinic_checkin_time)
        _set("clinic_checkout_time", clinic_checkout_time)
        if medical_staff is not None and medical_staff:
            try:
                ms_user = frappe.db.get_value("User", medical_staff, ["full_name"], as_dict=True)
                exam.medical_staff_name = ms_user.get("full_name") if ms_user else medical_staff
            except:
                exam.medical_staff_name = medical_staff
        elif medical_staff is not None:
            exam.medical_staff_name = None
        
        # Hospital fields mới
        _set("hospital_insurance", hospital_insurance)
        _set("hospital_school_coordination", hospital_school_coordination)
        _set("hospital_direction", hospital_direction)
        _set("hospital_payer", hospital_payer)
        _set("hospital_payer_other", hospital_payer_other)
        _set("hospital_transport", hospital_transport)
        _set("hospital_transport_other", hospital_transport_other)
        _set("hospital_health_monitoring", hospital_health_monitoring)
        _set("hospital_notes", hospital_notes)
        if hospital_advance_cost is not None:
            exam.hospital_advance_cost = hospital_advance_cost if hospital_advance_cost else 0
        if hospital_medical_staff is not None:
            exam.hospital_medical_staff = hospital_medical_staff if hospital_medical_staff else None
            if hospital_medical_staff:
                try:
                    hms_user = frappe.db.get_value("User", hospital_medical_staff, ["full_name"], as_dict=True)
                    exam.hospital_medical_staff_name = hms_user.get("full_name") if hms_user else hospital_medical_staff
                except:
                    exam.hospital_medical_staff_name = hospital_medical_staff
            else:
                exam.hospital_medical_staff_name = None
        
        # Followup fields
        _set("followup_checkin_time", followup_checkin_time)
        _set("followup_examination", followup_examination)
        _set("followup_treatment_details", followup_treatment_details)
        _set("followup_checkout_time", followup_checkout_time)
        if followup_outcome is not None:
            exam.followup_outcome = followup_outcome if followup_outcome else None
        _set("followup_notes", followup_notes)
        _set("followup_transfer_hospital", followup_transfer_hospital)
        _set("followup_accompanying_teacher", followup_accompanying_teacher)
        _set("followup_accompanying_health_staff", followup_accompanying_health_staff)
        
        # Followup fields mới
        _set("followup_clinic_checkin_time", followup_clinic_checkin_time)
        _set("followup_clinic_checkout_time", followup_clinic_checkout_time)
        if followup_is_scheduled_recheck is not None:
            exam.followup_is_scheduled_recheck = 1 if followup_is_scheduled_recheck else 0
        _set("followup_medical_suggestion", followup_medical_suggestion)
        if followup_medical_staff is not None:
            exam.followup_medical_staff = followup_medical_staff if followup_medical_staff else None
            if followup_medical_staff:
                try:
                    fms_user = frappe.db.get_value("User", followup_medical_staff, ["full_name"], as_dict=True)
                    exam.followup_medical_staff_name = fms_user.get("full_name") if fms_user else followup_medical_staff
                except:
                    exam.followup_medical_staff_name = followup_medical_staff
            else:
                exam.followup_medical_staff_name = None
        
        # Khi followup outcome thay đổi, cập nhật visit status tương ứng
        # Đồng thời set leave_clinic_time và sync clinic_checkout_time cho tất cả exam liên quan
        # (tránh trường hợp PH đón/chuyển viện không ghi nhận giờ checkout trên Parent Portal)
        if followup_outcome in ("transferred", "picked_up") and exam.visit_id:
            try:
                visit_doc = frappe.get_doc("SIS Daily Health Visit", exam.visit_id)
                visit_doc.status = followup_outcome
                if not visit_doc.leave_clinic_time:
                    visit_doc.leave_clinic_time = nowtime()
                if followup_outcome == "transferred":
                    if followup_transfer_hospital:
                        visit_doc.transfer_hospital = followup_transfer_hospital
                    if followup_accompanying_teacher:
                        visit_doc.accompanying_teacher = followup_accompanying_teacher
                    if followup_accompanying_health_staff:
                        visit_doc.accompanying_health_staff = followup_accompanying_health_staff
                visit_doc.save()
                leave_time = visit_doc.leave_clinic_time
                if leave_time:
                    exam.clinic_checkout_time = leave_time  # Để exam.save() lưu đúng cho thăm khám ban đầu
                related_exams = frappe.get_all("SIS Health Examination", filters={"visit_id": exam.visit_id}, fields=["name"])
                for re in related_exams:
                    frappe.db.set_value("SIS Health Examination", re.name, "outcome", followup_outcome, update_modified=False)
                    if leave_time:
                        frappe.db.set_value("SIS Health Examination", re.name, "clinic_checkout_time", leave_time, update_modified=False)
            except Exception as e:
                frappe.logger().error(f"Error updating visit status for followup: {str(e)}")

        # Khi outcome thăm khám ban đầu thay đổi (transferred/picked_up/return_class) qua form cập nhật,
        # cập nhật visit và sync clinic_checkout_time - tránh thiếu giờ checkout trên Parent Portal
        if outcome in ("transferred", "picked_up", "return_class") and exam.visit_id:
            try:
                visit_doc = frappe.get_doc("SIS Daily Health Visit", exam.visit_id)
                visit_map = {"return_class": "returned", "picked_up": "picked_up", "transferred": "transferred"}
                visit_doc.status = visit_map.get(outcome, outcome)
                if not visit_doc.leave_clinic_time:
                    visit_doc.leave_clinic_time = nowtime()
                visit_doc.save()
                leave_time = visit_doc.leave_clinic_time
                if leave_time:
                    exam.clinic_checkout_time = leave_time  # Để exam.save() lưu đúng
                related_exams = frappe.get_all("SIS Health Examination", filters={"visit_id": exam.visit_id}, fields=["name"])
                for re in related_exams:
                    if leave_time and re.name != exam.name:
                        frappe.db.set_value("SIS Health Examination", re.name, "clinic_checkout_time", leave_time, update_modified=False)
            except Exception as e:
                frappe.logger().error(f"Error updating visit for initial outcome: {str(e)}")
        
        # Update images if provided (cho phép empty array để clear tất cả)
        if images is not None and isinstance(images, list):
            frappe.logger().info(f"[update_health_examination] Updating images: {images}")
            # Clear existing images
            exam.images = []
            # Add new images
            for img_data in images:
                if isinstance(img_data, dict) and img_data.get("image"):
                    exam.append("images", {
                        "image": img_data.get("image"),
                        "description": img_data.get("description", "")
                    })
        
        exam.save()
        frappe.db.commit()
        
        frappe.logger().info(f"[update_health_examination] Successfully updated exam: {exam.name}")
        
        return success_response(
            data={"name": exam.name},
            message="Cập nhật hồ sơ khám thành công"
        )
    
    except frappe.DoesNotExistError:
        return error_response(
            message="Hồ sơ khám không tồn tại",
            code="NOT_FOUND"
        )
    except Exception as e:
        frappe.db.rollback()
        frappe.logger().error(f"Error updating health examination: {str(e)}")
        import traceback
        frappe.logger().error(traceback.format_exc())
        return error_response(
            message=f"Lỗi khi cập nhật hồ sơ khám: {str(e)}",
            code="UPDATE_EXAM_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def get_student_examination_history():
    """
    Lấy lịch sử khám của học sinh
    Params:
        - student_id: ID học sinh (required)
        - limit: Số lượng record (optional, default: 50)
    """
    try:
        _check_teacher_permission()
        
        data = frappe.local.form_dict
        request_args = frappe.request.args
        
        student_id = data.get("student_id") or request_args.get("student_id")
        limit = int(data.get("limit") or request_args.get("limit") or 50)
        
        if not student_id:
            return validation_error_response("student_id là bắt buộc", {"student_id": ["student_id là bắt buộc"]})
        
        # Lấy lịch sử khám
        exams = frappe.get_all(
            "SIS Health Examination",
            filters={"student_id": student_id},
            fields=[
                "name", "examination_date", "symptoms", "diet_history", "diagnosis",
                "treatment", "outcome", "examined_by", "examined_by_name", "creation", "modified",
                "visit_id", "disease_classification", "examination_notes",
                "treatment_type", "treatment_details", "notes",
                "medical_staff", "medical_staff_name",
                "clinic_checkin_time", "clinic_checkout_time",
                "hospital_diagnosis", "hospital_treatment",
                "hospital_insurance", "hospital_school_coordination",
                "hospital_medical_staff", "hospital_medical_staff_name",
                "hospital_direction", "hospital_advance_cost",
                "hospital_payer", "hospital_payer_other",
                "hospital_transport", "hospital_transport_other",
                "hospital_health_monitoring", "hospital_notes",
                "followup_checkin_time", "followup_examination",
                "followup_treatment_details", "followup_checkout_time",
                "followup_outcome", "followup_notes",
                "followup_transfer_hospital", "followup_accompanying_teacher",
                "followup_accompanying_health_staff",
                "followup_clinic_checkin_time", "followup_clinic_checkout_time",
                "followup_is_scheduled_recheck", "followup_medical_suggestion",
                "followup_medical_staff", "followup_medical_staff_name",
                "sent_to_parent", "sent_to_parent_at"
            ],
            order_by="examination_date desc, creation desc",
            limit=limit
        )
        
        # Lấy images + visit_reason cho mỗi examination
        for exam in exams:
            exam_images = frappe.get_all(
                "SIS Examination Image",
                filters={"parent": exam.get("name")},
                fields=["image", "description"],
                order_by="idx"
            )
            for img in exam_images:
                if img.get("image"):
                    if img["image"].startswith("http://") or img["image"].startswith("https://"):
                        pass  # Đã là URL đầy đủ
                    elif img["image"].startswith("/"):
                        img["image"] = frappe.utils.get_url(img["image"])
                    else:
                        img["image"] = frappe.utils.get_url("/files/" + img["image"])
            exam["images"] = exam_images

            # Lấy reason từ visit tương ứng (mỗi lượt xuống Y tế có lý do riêng từ GV)
            visit_id = exam.get("visit_id")
            if visit_id:
                visit_reason = frappe.db.get_value("SIS Daily Health Visit", visit_id, "reason")
                exam["visit_reason"] = visit_reason or ""
            else:
                exam["visit_reason"] = ""
        
        return success_response(
            data={"data": exams, "total": len(exams)},
            message="Lấy lịch sử khám thành công"
        )
    
    except Exception as e:
        frappe.logger().error(f"Error getting student examination history: {str(e)}")
        return error_response(
            message=f"Lỗi khi lấy lịch sử khám: {str(e)}",
            code="GET_HISTORY_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def get_daily_health_report_data():
    """
    Lấy dữ liệu báo cáo y tế hàng ngày (cho xuất Excel).
    Trả về danh sách visits kèm examinations chi tiết.
    
    Params:
        - date: Ngày cần lấy dữ liệu (optional, dùng khi không có date_from/date_to)
        - date_from: Từ ngày (optional, dùng với date_to cho khoảng thời gian)
        - date_to: Đến ngày (optional)
        - only_transferred: Chỉ lấy trường hợp chuyển viện (optional, 1/true/"true")
    
    Returns:
        - data: Danh sách visits với examinations, mỗi examination là một dòng
    """
    try:
        _check_teacher_permission()
        
        data = frappe.local.form_dict
        request_args = frappe.request.args
        
        date_from = data.get("date_from") or request_args.get("date_from")
        date_to = data.get("date_to") or request_args.get("date_to")
        only_transferred = data.get("only_transferred") or request_args.get("only_transferred")
        only_transferred = only_transferred in (1, True, "1", "true", "True")
        
        # Nếu có date_from và date_to -> filter theo khoảng; ngược lại dùng date (hoặc today)
        if date_from and date_to:
            visit_filters = {"visit_date": ["between", [date_from, date_to]]}
        else:
            report_date = data.get("date") or request_args.get("date") or today()
            visit_filters = {"visit_date": report_date}
        
        if only_transferred:
            visit_filters["status"] = "transferred"
        
        # Lấy tất cả visits trong ngày/khoảng ngày
        visits = frappe.get_all(
            "SIS Daily Health Visit",
            filters=visit_filters,
            fields=[
                "name", "student_id", "student_name", "student_code",
                "class_id", "class_name", "visit_date", "reason",
                "leave_class_time", "arrive_clinic_time", "leave_clinic_time",
                "status", "reported_by_name", "received_by_name", "creation",
                "checkout_notes", "transfer_hospital",
                "accompanying_teacher", "accompanying_health_staff"
            ],
            order_by="leave_class_time asc"
        )
        
        if not visits:
            return success_response(
                data={"data": [], "total": 0},
                message="Không có dữ liệu trong ngày"
            )
        
        # Lấy danh sách visit_id
        visit_ids = [v.name for v in visits]
        
        # Lấy tất cả examinations của các visits trong ngày
        examinations = frappe.get_all(
            "SIS Health Examination",
            filters={"visit_id": ["in", visit_ids]},
            fields=[
                "name", "visit_id", "student_id", "symptoms", "diet_history",
                "disease_classification", "examination_notes",
                "treatment_type", "treatment_details", "notes",
                "outcome", "examined_by_name", "creation",
                "hospital_diagnosis", "hospital_treatment",
                "followup_checkin_time", "followup_examination",
                "followup_treatment_details", "followup_checkout_time",
                "followup_outcome", "followup_notes",
                "followup_transfer_hospital", "followup_accompanying_teacher",
                "followup_accompanying_health_staff"
            ],
            order_by="creation asc"
        )
        
        # Tạo map visit_id -> list examinations
        exam_map = {}
        for exam in examinations:
            vid = exam.get("visit_id")
            if vid not in exam_map:
                exam_map[vid] = []
            exam_map[vid].append(exam)
        
        # Map treatment_type -> label
        treatment_type_labels = {
            "first_aid": "Sơ cứu",
            "medication": "Cho thuốc",
            "rest": "Nghỉ ngơi",
            "other": "Khác"
        }
        
        # Map outcome -> label
        outcome_labels = {
            "return_class": "Về lớp",
            "picked_up": "Phụ huynh đón",
            "transferred": "Chuyển viện"
        }
        
        # Map status -> label
        status_labels = {
            "left_class": "Rời khỏi lớp, chờ tiếp nhận",
            "at_clinic": "Đã tiếp nhận",
            "examining": "Đang khám",
            "returned": "Đã về lớp",
            "picked_up": "Phụ huynh đón",
            "transferred": "Chuyển viện"
        }
        
        # Tạo danh sách kết quả - mỗi examination là một dòng
        result = []
        for visit in visits:
            visit_exams = exam_map.get(visit.name, [])
            
            base_data = {
                "visit_id": visit.name,
                "visit_date": str(visit.visit_date) if visit.visit_date else "",
                "student_name": visit.student_name,
                "student_code": visit.student_code,
                "class_name": visit.class_name,
                "reason": visit.reason,
                "leave_class_time": str(visit.leave_class_time)[:5] if visit.leave_class_time else "",
                "arrive_clinic_time": str(visit.arrive_clinic_time)[:5] if visit.arrive_clinic_time else "",
                "leave_clinic_time": str(visit.leave_clinic_time)[:5] if visit.leave_clinic_time else "",
                "status": status_labels.get(visit.status, visit.status),
                "reported_by_name": visit.reported_by_name or "",
                "received_by_name": visit.received_by_name or "",
                "checkout_notes": visit.checkout_notes or "",
                "transfer_hospital": visit.transfer_hospital or "",
                "accompanying_teacher": visit.accompanying_teacher or "",
                "accompanying_health_staff": visit.accompanying_health_staff or "",
            }
            
            if visit_exams:
                # Có examinations - tạo một dòng cho mỗi examination
                for exam in visit_exams:
                    row = {**base_data}
                    row["symptoms"] = exam.get("symptoms") or ""
                    row["diet_history"] = exam.get("diet_history") or ""
                    row["disease_classification"] = exam.get("disease_classification") or ""
                    row["treatment_type"] = treatment_type_labels.get(exam.get("treatment_type"), exam.get("treatment_type") or "")
                    row["treatment_details"] = exam.get("treatment_details") or ""
                    row["examination_notes"] = exam.get("examination_notes") or ""
                    row["notes"] = exam.get("notes") or ""
                    row["outcome"] = outcome_labels.get(exam.get("outcome"), exam.get("outcome") or "")
                    row["examined_by_name"] = exam.get("examined_by_name") or ""
                    row["hospital_diagnosis"] = exam.get("hospital_diagnosis") or ""
                    row["hospital_treatment"] = exam.get("hospital_treatment") or ""
                    row["followup_checkin_time"] = str(exam.get("followup_checkin_time"))[:5] if exam.get("followup_checkin_time") else ""
                    row["followup_examination"] = exam.get("followup_examination") or ""
                    row["followup_treatment_details"] = exam.get("followup_treatment_details") or ""
                    row["followup_checkout_time"] = str(exam.get("followup_checkout_time"))[:5] if exam.get("followup_checkout_time") else ""
                    row["followup_outcome"] = outcome_labels.get(exam.get("followup_outcome"), exam.get("followup_outcome") or "")
                    row["followup_notes"] = exam.get("followup_notes") or ""
                    row["followup_transfer_hospital"] = exam.get("followup_transfer_hospital") or ""
                    row["followup_accompanying_teacher"] = exam.get("followup_accompanying_teacher") or ""
                    row["followup_accompanying_health_staff"] = exam.get("followup_accompanying_health_staff") or ""
                    result.append(row)
            else:
                # Không có examination - vẫn tạo một dòng với thông tin visit
                row = {**base_data}
                row["symptoms"] = ""
                row["diet_history"] = ""
                row["disease_classification"] = ""
                row["treatment_type"] = ""
                row["treatment_details"] = ""
                row["examination_notes"] = ""
                row["notes"] = ""
                row["outcome"] = ""
                row["examined_by_name"] = ""
                row["hospital_diagnosis"] = ""
                row["hospital_treatment"] = ""
                row["followup_checkin_time"] = ""
                row["followup_examination"] = ""
                row["followup_treatment_details"] = ""
                row["followup_checkout_time"] = ""
                row["followup_outcome"] = ""
                row["followup_notes"] = ""
                row["followup_transfer_hospital"] = ""
                row["followup_accompanying_teacher"] = ""
                row["followup_accompanying_health_staff"] = ""
                result.append(row)
        
        return success_response(
            data={"data": result, "total": len(result)},
            message="Lấy dữ liệu báo cáo thành công"
        )
    
    except Exception as e:
        frappe.logger().error(f"Error getting daily health report data: {str(e)}")
        import traceback
        frappe.logger().error(traceback.format_exc())
        return error_response(
            message=f"Lỗi khi lấy dữ liệu báo cáo: {str(e)}",
            code="GET_REPORT_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def get_students_at_clinic():
    """
    Lấy danh sách học sinh đang ở Y tế cho một lớp (cho LessonLog điểm danh)
    Params:
        - class_id: ID lớp (required)
        - date: Ngày (optional, default: today)
        - current_time: Thời gian hiện tại (optional, để check xem học sinh đã về chưa)
    """
    try:
        _check_teacher_permission()
        
        data = frappe.local.form_dict
        request_args = frappe.request.args
        
        class_id = data.get("class_id") or request_args.get("class_id")
        visit_date = data.get("date") or request_args.get("date") or today()
        current_time = data.get("current_time") or request_args.get("current_time")
        
        if not class_id:
            return validation_error_response("class_id là bắt buộc", {"class_id": ["class_id là bắt buộc"]})
        
        # Lấy danh sách học sinh của lớp đang ở Y tế
        # Status: left_class hoặc at_clinic
        # Nếu đã returned/picked_up/transferred và có leave_clinic_time, check xem đã quay về chưa
        visits = frappe.get_all(
            "SIS Daily Health Visit",
            filters={
                "class_id": class_id,
                "visit_date": visit_date
            },
            fields=[
                "name", "student_id", "status", 
                "leave_class_time", "leave_clinic_time"
            ]
        )
        
        # Filter học sinh đang ở Y tế
        students_at_clinic = {}
        for visit in visits:
            # Nếu status là left_class hoặc at_clinic -> đang ở Y tế
            if visit.status in ["left_class", "at_clinic"]:
                students_at_clinic[visit.student_id] = {
                    "visit_id": visit.name,
                    "status": visit.status,
                    "leave_class_time": visit.leave_class_time
                }
            # Nếu đã về nhưng chưa check thời gian
            elif visit.status in ["returned", "picked_up", "transferred"]:
                # Nếu có current_time và leave_clinic_time, check xem đã về chưa
                if current_time and visit.leave_clinic_time:
                    if visit.leave_clinic_time > current_time:
                        # Chưa về, vẫn còn ở Y tế
                        students_at_clinic[visit.student_id] = {
                            "visit_id": visit.name,
                            "status": visit.status,
                            "leave_class_time": visit.leave_class_time
                        }
        
        return success_response(
            data={"students": students_at_clinic},
            message="Lấy danh sách học sinh ở Y tế thành công"
        )
    
    except Exception as e:
        frappe.logger().error(f"Error getting students at clinic: {str(e)}")
        return error_response(
            message=f"Lỗi khi lấy danh sách học sinh ở Y tế: {str(e)}",
            code="GET_STUDENTS_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def complete_health_visit():
    """
    Hoàn thành lượt xuống Y tế (học sinh rời phòng Y tế)
    Params:
        - visit_id: ID của visit (required)
        - outcome: Kết quả (required): returned/picked_up/transferred
        - leave_clinic_time: Thời gian rời Y tế (optional, default: now)
        - checkout_notes: Ghi chú/Dặn dò khi checkout (optional)
        - transfer_hospital: Bệnh viện chuyển tới (optional, khi chuyển viện)
        - accompanying_teacher: Thầy/ cô đi cùng (optional, khi chuyển viện)
        - accompanying_health_staff: NVYT đi cùng (optional, khi chuyển viện)
    """
    try:
        _check_teacher_permission()
        
        data = _get_request_data()
        
        visit_id = data.get("visit_id")
        outcome = data.get("outcome")
        leave_clinic_time = data.get("leave_clinic_time") or nowtime()
        checkout_notes = data.get("checkout_notes", "")
        transfer_hospital = data.get("transfer_hospital", "")
        accompanying_teacher = data.get("accompanying_teacher", "")
        accompanying_health_staff = data.get("accompanying_health_staff", "")
        
        # Validation
        errors = {}
        if not visit_id:
            errors["visit_id"] = ["visit_id là bắt buộc"]
        if not outcome:
            errors["outcome"] = ["outcome là bắt buộc"]
        elif outcome not in ["returned", "picked_up", "transferred"]:
            errors["outcome"] = ["outcome phải là returned, picked_up, hoặc transferred"]
        # checkout_notes: không bắt buộc (trường đã chốt)
        
        if errors:
            return validation_error_response("Dữ liệu không hợp lệ", errors)
        
        # Lấy visit
        visit = frappe.get_doc("SIS Daily Health Visit", visit_id)
        
        # Cập nhật visit
        visit.status = outcome
        visit.leave_clinic_time = leave_clinic_time
        visit.checkout_notes = checkout_notes
        if outcome == "transferred":
            visit.transfer_hospital = transfer_hospital
            visit.accompanying_teacher = accompanying_teacher
            visit.accompanying_health_staff = accompanying_health_staff
        visit.save()
        
        # Cập nhật outcome và clinic_checkout_time cho TẤT CẢ examinations liên quan đến visit này
        # Thời gian ra về (clinic_checkout_time) = thời gian checkout từ visit.leave_clinic_time
        exam_outcome_map = {
            "returned": "return_class",
            "picked_up": "picked_up",
            "transferred": "transferred"
        }
        exam_outcome = exam_outcome_map.get(outcome)
        related_exams = frappe.get_all(
            "SIS Health Examination",
            filters={"visit_id": visit_id},
            fields=["name"]
        )
        for exam in related_exams:
            if exam_outcome:
                frappe.db.set_value(
                    "SIS Health Examination",
                    exam.name,
                    "outcome",
                    exam_outcome,
                    update_modified=False
                )
            # Đồng bộ thời gian ra về từ checkout vào exam - dùng cho Thời gian lưu trú (Parent Portal, GVCN)
            frappe.db.set_value(
                "SIS Health Examination",
                exam.name,
                "clinic_checkout_time",
                leave_clinic_time,
                update_modified=False
            )
        
        frappe.db.commit()
        
        # Gửi push notification cho Homeroom + Vice-homeroom + Mobile Medical
        try:
            from erp.api.erp_sis.daily_health_notification import notify_health_visit_completed
            notify_health_visit_completed(visit_name=visit.name)
        except Exception as notif_err:
            frappe.logger().warning(f"[complete_health_visit] Không gửi được notification: {str(notif_err)}")
        
        return success_response(
            data={"name": visit.name, "status": visit.status},
            message="Đã hoàn thành lượt xuống Y tế"
        )
    
    except frappe.DoesNotExistError:
        return error_response(
            message="Bản ghi xuống Y tế không tồn tại",
            code="NOT_FOUND"
        )
    except Exception as e:
        frappe.db.rollback()
        frappe.logger().error(f"Error completing health visit: {str(e)}")
        return error_response(
            message=f"Lỗi khi hoàn thành lượt xuống Y tế: {str(e)}",
            code="COMPLETE_ERROR"
        )


def _get_period_times(class_id, period_name, visit_date):
    """
    Lấy thời gian bắt đầu và kết thúc của tiết học từ Schedule Group.
    Returns: (start_time, end_time) hoặc (None, None) nếu không tìm thấy
    """
    try:
        # Lấy education_stage_id từ class
        class_doc = frappe.db.get_value("SIS Class", class_id, ["education_stage_id"], as_dict=True)
        if not class_doc or not class_doc.get("education_stage_id"):
            frappe.logger().warning(f"[get_period_times] Cannot find education_stage_id for class {class_id}")
            return None, None
        
        education_stage_id = class_doc.get("education_stage_id")
        
        # Tìm Schedule Group đang active cho education_stage và date
        schedule_group = frappe.db.sql("""
            SELECT name FROM `tabSIS Schedule Group`
            WHERE education_stage_id = %s
              AND is_active = 1
              AND start_date <= %s
              AND end_date >= %s
            ORDER BY start_date DESC
            LIMIT 1
        """, (education_stage_id, visit_date, visit_date), as_dict=True)
        
        if not schedule_group:
            frappe.logger().warning(f"[get_period_times] Cannot find active schedule group for education_stage {education_stage_id} on {visit_date}")
            return None, None
        
        schedule_group_id = schedule_group[0].get("name")
        
        # Tìm Period với period_name khớp
        period = frappe.db.get_value(
            "SIS Timetable Column",
            {
                "schedule_id": schedule_group_id,
                "period_name": period_name
            },
            ["start_time", "end_time"],
            as_dict=True
        )
        
        if not period:
            frappe.logger().warning(f"[get_period_times] Cannot find period '{period_name}' in schedule group {schedule_group_id}")
            return None, None
        
        return period.get("start_time"), period.get("end_time")
    
    except Exception as e:
        frappe.logger().error(f"[get_period_times] Error: {str(e)}")
        return None, None


def _time_to_seconds(time_val):
    """
    Chuyển đổi time value (có thể là timedelta hoặc string) sang seconds
    """
    if time_val is None:
        return None
    
    # Nếu là timedelta (từ database)
    if hasattr(time_val, 'total_seconds'):
        return time_val.total_seconds()
    
    # Nếu là string (HH:MM hoặc HH:MM:SS)
    if isinstance(time_val, str):
        parts = time_val.split(':')
        if len(parts) >= 2:
            hours = int(parts[0])
            minutes = int(parts[1])
            seconds = int(parts[2]) if len(parts) > 2 else 0
            return hours * 3600 + minutes * 60 + seconds
    
    return None


@frappe.whitelist(allow_guest=False)
def get_health_status_for_period():
    """
    Lấy trạng thái Y tế của học sinh theo tiết học (cho LessonLog).
    Trả về tất cả visit liên quan đến tiết học, kể cả đã checkout.
    
    Params:
        - class_id: ID lớp (required)
        - date: Ngày (required)
        - period: Tên tiết học, VD: "Tiết 1" (required)
    
    Returns:
        students: Record<student_id, {visit_id, status, leave_class_time, leave_clinic_time}>
        
    Logic:
        - Nếu có thời gian tiết học:
          1. Học sinh rời lớp TRONG tiết → luôn hiển thị (kể cả đã checkout: returned, picked_up, transferred)
          2. Học sinh rời lớp TRƯỚC tiết nhưng chưa về hoặc về sau tiết bắt đầu → hiển thị
        - Nếu không có thời gian tiết: hiển thị tất cả visit trong ngày
    """
    try:
        _check_teacher_permission()
        
        data = frappe.local.form_dict
        request_args = frappe.request.args
        
        class_id = data.get("class_id") or request_args.get("class_id")
        visit_date = data.get("date") or request_args.get("date")
        period_name = data.get("period") or request_args.get("period")
        
        # Validation
        errors = {}
        if not class_id:
            errors["class_id"] = ["class_id là bắt buộc"]
        if not visit_date:
            errors["date"] = ["date là bắt buộc"]
        if not period_name:
            errors["period"] = ["period là bắt buộc"]
        
        if errors:
            return validation_error_response("Dữ liệu không hợp lệ", errors)
        
        # Lấy thời gian tiết học
        period_start, period_end = _get_period_times(class_id, period_name, visit_date)
        period_start_sec = _time_to_seconds(period_start)
        period_end_sec = _time_to_seconds(period_end)
        
        has_period_times = period_start_sec is not None and period_end_sec is not None
        
        if not has_period_times:
            frappe.logger().warning(f"[get_health_status_for_period] Cannot get period times for {period_name}, using fallback logic")
        
        # Query tất cả visits trong ngày của lớp
        visits = frappe.get_all(
            "SIS Daily Health Visit",
            filters={
                "class_id": class_id,
                "visit_date": visit_date
            },
            fields=[
                "name", "student_id", "status",
                "leave_class_time", "leave_clinic_time"
            ],
            order_by="leave_class_time desc"
        )
        
        # Filter học sinh có visit liên quan đến tiết học
        # Logic mới: Nếu học sinh rời lớp trong tiết đó, luôn hiển thị trạng thái (kể cả đã checkout)
        students_at_clinic = {}
        
        for visit in visits:
            # Bỏ qua visit đã hủy hoặc từ chối - không hiển thị trong LessonLog
            if visit.status in ("cancelled", "rejected"):
                continue
            
            student_id = visit.student_id
            
            # Nếu student đã có trong result (ưu tiên visit mới nhất), skip
            if student_id in students_at_clinic:
                continue
            
            leave_class_sec = _time_to_seconds(visit.leave_class_time)
            leave_clinic_sec = _time_to_seconds(visit.leave_clinic_time)
            
            should_include = False
            
            if has_period_times and leave_class_sec is not None:
                # Logic có thời gian tiết học:
                # Học sinh rời lớp trong khoảng thời gian tiết → luôn hiển thị (kể cả đã checkout)
                # Hoặc: rời lớp trước tiết nhưng vẫn ở Y tế trong tiết
                left_during_period = period_start_sec <= leave_class_sec < period_end_sec
                
                # Rời lớp trước tiết nhưng chưa về hoặc về trong/sau tiết
                left_before_but_still_at_clinic = (
                    leave_class_sec < period_start_sec and
                    (leave_clinic_sec is None or leave_clinic_sec > period_start_sec)
                )
                
                if left_during_period or left_before_but_still_at_clinic:
                    should_include = True
            else:
                # Fallback: không có thông tin tiết học, hiển thị tất cả visit trong ngày
                # Để giáo viên biết học sinh đã/đang ở Y tế
                should_include = True
            
            if should_include:
                students_at_clinic[student_id] = {
                    "visit_id": visit.name,
                    "status": visit.status,
                    "leave_class_time": str(visit.leave_class_time) if visit.leave_class_time else None,
                    "leave_clinic_time": str(visit.leave_clinic_time) if visit.leave_clinic_time else None
                }
        
        return success_response(
            data={"students": students_at_clinic},
            message="Lấy trạng thái Y tế theo tiết học thành công"
        )
    
    except Exception as e:
        frappe.logger().error(f"Error getting health status for period: {str(e)}")
        import traceback
        frappe.logger().error(traceback.format_exc())
        return error_response(
            message=f"Lỗi khi lấy trạng thái Y tế theo tiết học: {str(e)}",
            code="GET_HEALTH_STATUS_ERROR"
        )


# ==================== DISEASE CLASSIFICATION CRUD ====================

@frappe.whitelist()
def get_disease_classifications(campus: str = None):
    """
    Lấy danh sách phân loại bệnh theo campus
    """
    try:
        filters = {"enabled": 1}
        if campus:
            filters["campus"] = campus
        
        classifications = frappe.get_all(
            "SIS Disease Classification",
            filters=filters,
            fields=["name", "code", "title", "campus", "enabled", "creation", "modified"],
            order_by="code asc"
        )
        
        return success_response(
            data={"data": classifications, "total": len(classifications)},
            message="Lấy danh sách phân loại bệnh thành công"
        )
    
    except Exception as e:
        frappe.logger().error(f"Error getting disease classifications: {str(e)}")
        return error_response(
            message=f"Lỗi khi lấy danh sách phân loại bệnh: {str(e)}",
            code="GET_DISEASE_CLASSIFICATIONS_ERROR"
        )


@frappe.whitelist()
def create_disease_classification(code: str = None, title: str = None, campus: str = None):
    """
    Tạo mới phân loại bệnh
    """
    try:
        # Lấy dữ liệu từ form_dict hoặc request JSON
        import json
        if not code and frappe.request and frappe.request.data:
            try:
                data = json.loads(frappe.request.data)
                code = data.get('code')
                title = data.get('title')
                campus = data.get('campus')
            except json.JSONDecodeError:
                pass
        
        code = code or frappe.form_dict.get('code')
        title = title or frappe.form_dict.get('title')
        campus = campus or frappe.form_dict.get('campus')
        
        if not code or not title or not campus:
            return error_response(
                message="Mã bệnh, tên phân loại và trường học là bắt buộc",
                code="MISSING_REQUIRED_FIELDS"
            )
        
        # Kiểm tra trùng mã trong cùng campus
        existing = frappe.get_all(
            "SIS Disease Classification",
            filters={"code": code, "campus": campus},
            limit=1
        )
        if existing:
            return error_response(
                message="Mã bệnh này đã tồn tại",
                code="DUPLICATE_CODE"
            )
        
        doc = frappe.get_doc({
            "doctype": "SIS Disease Classification",
            "code": code,
            "title": title,
            "campus": campus,
            "enabled": 1
        })
        doc.insert(ignore_permissions=True)
        frappe.db.commit()
        
        return success_response(
            data={"name": doc.name, "code": doc.code, "title": doc.title},
            message="Tạo phân loại bệnh thành công"
        )
    
    except Exception as e:
        frappe.logger().error(f"Error creating disease classification: {str(e)}")
        return error_response(
            message=f"Lỗi khi tạo phân loại bệnh: {str(e)}",
            code="CREATE_DISEASE_CLASSIFICATION_ERROR"
        )


@frappe.whitelist()
def update_disease_classification(name: str = None, code: str = None, title: str = None, enabled: int = None):
    """
    Cập nhật phân loại bệnh
    """
    try:
        # Lấy dữ liệu từ request JSON
        import json
        data = {}
        if frappe.request and frappe.request.data:
            try:
                data = json.loads(frappe.request.data)
            except json.JSONDecodeError:
                pass
        
        name = name or data.get('name') or frappe.form_dict.get('name')
        code = code if code is not None else data.get('code') or frappe.form_dict.get('code')
        title = title if title is not None else data.get('title') or frappe.form_dict.get('title')
        enabled_val = data.get('enabled') or frappe.form_dict.get('enabled')
        if enabled is None and enabled_val is not None:
            enabled = int(enabled_val) if enabled_val not in [None, ''] else None
        
        if not name:
            return error_response(
                message="ID phân loại bệnh là bắt buộc",
                code="MISSING_REQUIRED_FIELDS"
            )
        
        doc = frappe.get_doc("SIS Disease Classification", name)
        
        if code is not None:
            # Kiểm tra trùng mã trong cùng campus (trừ chính nó)
            existing = frappe.get_all(
                "SIS Disease Classification",
                filters={"code": code, "campus": doc.campus, "name": ["!=", name]},
                limit=1
            )
            if existing:
                return error_response(
                    message="Mã bệnh này đã tồn tại",
                    code="DUPLICATE_CODE"
                )
            doc.code = code
        
        if title is not None:
            doc.title = title
        
        if enabled is not None:
            doc.enabled = enabled
        
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        
        return success_response(
            data={"name": doc.name, "code": doc.code, "title": doc.title},
            message="Cập nhật phân loại bệnh thành công"
        )
    
    except frappe.DoesNotExistError:
        return error_response(
            message="Không tìm thấy phân loại bệnh",
            code="CLASSIFICATION_NOT_FOUND"
        )
    except Exception as e:
        frappe.logger().error(f"Error updating disease classification: {str(e)}")
        return error_response(
            message=f"Lỗi khi cập nhật phân loại bệnh: {str(e)}",
            code="UPDATE_DISEASE_CLASSIFICATION_ERROR"
        )


@frappe.whitelist()
def delete_disease_classification(name: str = None):
    """
    Xóa phân loại bệnh
    """
    try:
        import json
        data = {}
        if frappe.request and frappe.request.data:
            try:
                data = json.loads(frappe.request.data)
            except json.JSONDecodeError:
                pass
        
        name = name or data.get('name') or frappe.form_dict.get('name')
        
        if not name:
            return error_response(
                message="ID phân loại bệnh là bắt buộc",
                code="MISSING_REQUIRED_FIELDS"
            )
        
        frappe.delete_doc("SIS Disease Classification", name, ignore_permissions=True)
        frappe.db.commit()
        
        return success_response(
            data={"name": name},
            message="Xóa phân loại bệnh thành công"
        )
    
    except frappe.DoesNotExistError:
        return error_response(
            message="Không tìm thấy phân loại bệnh",
            code="CLASSIFICATION_NOT_FOUND"
        )
    except Exception as e:
        frappe.logger().error(f"Error deleting disease classification: {str(e)}")
        return error_response(
            message=f"Lỗi khi xóa phân loại bệnh: {str(e)}",
            code="DELETE_DISEASE_CLASSIFICATION_ERROR"
        )


# ==================== MEDICINE CRUD ====================

@frappe.whitelist()
def get_medicines(campus: str = None):
    """
    Lấy danh sách thuốc theo campus
    """
    try:
        filters = {"enabled": 1}
        if campus:
            filters["campus"] = campus
        
        medicines = frappe.get_all(
            "SIS Medicine",
            filters=filters,
            fields=["name", "title", "unit", "description", "campus", "enabled", "creation", "modified"],
            order_by="title asc"
        )
        
        return success_response(
            data={"data": medicines, "total": len(medicines)},
            message="Lấy danh sách thuốc thành công"
        )
    
    except Exception as e:
        frappe.logger().error(f"Error getting medicines: {str(e)}")
        return error_response(
            message=f"Lỗi khi lấy danh sách thuốc: {str(e)}",
            code="GET_MEDICINES_ERROR"
        )


@frappe.whitelist()
def create_medicine(title: str = None, campus: str = None, unit: str = None, description: str = None):
    """
    Tạo mới thuốc
    """
    try:
        # Lấy dữ liệu từ request JSON
        import json
        data = {}
        if frappe.request and frappe.request.data:
            try:
                data = json.loads(frappe.request.data)
            except json.JSONDecodeError:
                pass
        
        title = title or data.get('title') or frappe.form_dict.get('title')
        campus = campus or data.get('campus') or frappe.form_dict.get('campus')
        unit = unit or data.get('unit') or frappe.form_dict.get('unit')
        description = description or data.get('description') or frappe.form_dict.get('description')
        
        if not title or not campus:
            return error_response(
                message="Tên thuốc và trường học là bắt buộc",
                code="MISSING_REQUIRED_FIELDS"
            )
        
        # Kiểm tra trùng tên trong cùng campus
        existing = frappe.get_all(
            "SIS Medicine",
            filters={"title": title, "campus": campus},
            limit=1
        )
        if existing:
            return error_response(
                message="Thuốc này đã tồn tại",
                code="DUPLICATE_MEDICINE"
            )
        
        doc = frappe.get_doc({
            "doctype": "SIS Medicine",
            "title": title,
            "unit": unit,
            "description": description,
            "campus": campus,
            "enabled": 1
        })
        doc.insert(ignore_permissions=True)
        frappe.db.commit()
        
        return success_response(
            data={"name": doc.name, "title": doc.title},
            message="Tạo thuốc thành công"
        )
    
    except Exception as e:
        frappe.logger().error(f"Error creating medicine: {str(e)}")
        return error_response(
            message=f"Lỗi khi tạo thuốc: {str(e)}",
            code="CREATE_MEDICINE_ERROR"
        )


@frappe.whitelist()
def update_medicine(name: str = None, title: str = None, unit: str = None, description: str = None, enabled: int = None):
    """
    Cập nhật thuốc
    """
    try:
        # Lấy dữ liệu từ request JSON
        import json
        data = {}
        if frappe.request and frappe.request.data:
            try:
                data = json.loads(frappe.request.data)
            except json.JSONDecodeError:
                pass
        
        name = name or data.get('name') or frappe.form_dict.get('name')
        title = title if title is not None else data.get('title') or frappe.form_dict.get('title')
        unit = unit if unit is not None else data.get('unit') or frappe.form_dict.get('unit')
        description = description if description is not None else data.get('description') or frappe.form_dict.get('description')
        enabled_val = data.get('enabled') or frappe.form_dict.get('enabled')
        if enabled is None and enabled_val is not None:
            enabled = int(enabled_val) if enabled_val not in [None, ''] else None
        
        if not name:
            return error_response(
                message="ID thuốc là bắt buộc",
                code="MISSING_REQUIRED_FIELDS"
            )
        
        doc = frappe.get_doc("SIS Medicine", name)
        
        if title is not None:
            # Kiểm tra trùng tên trong cùng campus (trừ chính nó)
            existing = frappe.get_all(
                "SIS Medicine",
                filters={"title": title, "campus": doc.campus, "name": ["!=", name]},
                limit=1
            )
            if existing:
                return error_response(
                    message="Thuốc này đã tồn tại",
                    code="DUPLICATE_MEDICINE"
                )
            doc.title = title
        
        if unit is not None:
            doc.unit = unit
        
        if description is not None:
            doc.description = description
        
        if enabled is not None:
            doc.enabled = enabled
        
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        
        return success_response(
            data={"name": doc.name, "title": doc.title},
            message="Cập nhật thuốc thành công"
        )
    
    except frappe.DoesNotExistError:
        return error_response(
            message="Không tìm thấy thuốc",
            code="MEDICINE_NOT_FOUND"
        )
    except Exception as e:
        frappe.logger().error(f"Error updating medicine: {str(e)}")
        return error_response(
            message=f"Lỗi khi cập nhật thuốc: {str(e)}",
            code="UPDATE_MEDICINE_ERROR"
        )


@frappe.whitelist()
def delete_medicine(name: str = None):
    """
    Xóa thuốc
    """
    try:
        import json
        data = {}
        if frappe.request and frappe.request.data:
            try:
                data = json.loads(frappe.request.data)
            except json.JSONDecodeError:
                pass
        
        name = name or data.get('name') or frappe.form_dict.get('name')
        
        if not name:
            return error_response(
                message="ID thuốc là bắt buộc",
                code="MISSING_REQUIRED_FIELDS"
            )
        
        frappe.delete_doc("SIS Medicine", name, ignore_permissions=True)
        frappe.db.commit()
        
        return success_response(
            data={"name": name},
            message="Xóa thuốc thành công"
        )
    
    except frappe.DoesNotExistError:
        return error_response(
            message="Không tìm thấy thuốc",
            code="MEDICINE_NOT_FOUND"
        )
    except Exception as e:
        frappe.logger().error(f"Error deleting medicine: {str(e)}")
        return error_response(
            message=f"Lỗi khi xóa thuốc: {str(e)}",
            code="DELETE_MEDICINE_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def import_disease_classifications_excel(campus=None):
    """
    Import phân loại bệnh từ file Excel.
    Cột A: Mã bệnh, Cột B: Tên phân loại
    """
    try:
        import openpyxl
        import io

        # Đọc campus từ nhiều nguồn: tham số hàm, form_dict, query string
        campus = (
            campus
            or frappe.form_dict.get('campus')
            or (frappe.request.args.get('campus') if frappe.request else None)
        )
        if not campus:
            return error_response(
                message="Thiếu thông tin trường học",
                code="MISSING_CAMPUS"
            )

        # Kiểm tra campus tồn tại
        if not frappe.db.exists("SIS Campus", campus):
            return error_response(
                message=f"Không tìm thấy trường học: {campus}",
                code="CAMPUS_NOT_FOUND"
            )

        # Lấy file từ request
        file_obj = frappe.request.files.get('file')
        if not file_obj:
            return error_response(
                message="Không tìm thấy file",
                code="MISSING_FILE"
            )

        file_data = file_obj.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        ws = wb.active

        success_count = 0
        error_list = []
        total_count = 0

        # Tải trước danh sách mã đã có để kiểm tra trùng nhanh (so sánh uppercase)
        existing_codes = set(
            r.code.strip().upper()
            for r in frappe.get_all("SIS Disease Classification", filters={"campus": campus}, fields=["code"])
            if r.code
        )
        # Theo dõi mã trong file Excel để phát hiện trùng nội bộ
        seen_in_file = set()

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Bỏ qua dòng rỗng hoàn toàn
            if not any(row):
                continue

            code = str(row[0]).strip() if row[0] is not None else ''
            title = str(row[1]).strip() if row[1] is not None else ''
            total_count += 1

            # Validate dữ liệu bắt buộc
            if not code:
                error_list.append(f"Dòng {row_idx}: Mã bệnh không được để trống")
                continue
            if not title:
                error_list.append(f"Dòng {row_idx}: Tên phân loại không được để trống")
                continue

            code_upper = code.upper()

            # Kiểm tra trùng trong cùng file Excel
            if code_upper in seen_in_file:
                error_list.append(f"Dòng {row_idx}: Mã bệnh '{code}' bị trùng trong file Excel")
                continue

            # Kiểm tra trùng mã với dữ liệu đã có trong hệ thống
            if code_upper in existing_codes:
                error_list.append(f"Dòng {row_idx}: Mã bệnh '{code}' đã tồn tại trong hệ thống")
                continue

            try:
                doc = frappe.get_doc({
                    "doctype": "SIS Disease Classification",
                    "code": code,
                    "title": title,
                    "campus": campus,
                    "enabled": 1
                })
                doc.insert(ignore_permissions=True)
                success_count += 1
                seen_in_file.add(code_upper)
                existing_codes.add(code_upper)
            except Exception as e:
                error_list.append(f"Dòng {row_idx}: {str(e)}")

        frappe.db.commit()

        message = f"Import phân loại bệnh hoàn tất"
        return success_response(
            data={
                "success_count": success_count,
                "total_count": total_count,
                "errors": error_list,
                "message": message
            },
            message=message
        )

    except Exception as e:
        frappe.logger().error(f"Error importing disease classifications: {str(e)}")
        return error_response(
            message=f"Lỗi khi import phân loại bệnh: {str(e)}",
            code="IMPORT_DISEASE_CLASSIFICATION_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def import_medicines_excel(campus=None):
    """
    Import thuốc từ file Excel.
    Cột A: Tên thuốc, Cột B: Đơn vị (tùy chọn), Cột C: Mô tả (tùy chọn)
    """
    try:
        import openpyxl
        import io

        # Đọc campus từ nhiều nguồn: tham số hàm, form_dict, query string
        campus = (
            campus
            or frappe.form_dict.get('campus')
            or (frappe.request.args.get('campus') if frappe.request else None)
        )
        if not campus:
            return error_response(
                message="Thiếu thông tin trường học",
                code="MISSING_CAMPUS"
            )

        # Kiểm tra campus tồn tại
        if not frappe.db.exists("SIS Campus", campus):
            return error_response(
                message=f"Không tìm thấy trường học: {campus}",
                code="CAMPUS_NOT_FOUND"
            )

        # Lấy file từ request
        file_obj = frappe.request.files.get('file')
        if not file_obj:
            return error_response(
                message="Không tìm thấy file",
                code="MISSING_FILE"
            )

        file_data = file_obj.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        ws = wb.active

        success_count = 0
        error_list = []
        total_count = 0

        # Tải trước danh sách tên thuốc đã có để kiểm tra trùng nhanh (so sánh lowercase)
        existing_titles = set(
            r.title.strip().lower()
            for r in frappe.get_all("SIS Medicine", filters={"campus": campus}, fields=["title"])
            if r.title
        )
        # Theo dõi tên trong file Excel để phát hiện trùng nội bộ
        seen_in_file = set()

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Bỏ qua dòng rỗng hoàn toàn
            if not any(row):
                continue

            title = str(row[0]).strip() if row[0] is not None else ''
            unit = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            description = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            total_count += 1

            # Validate dữ liệu bắt buộc
            if not title:
                error_list.append(f"Dòng {row_idx}: Tên thuốc không được để trống")
                continue

            title_lower = title.lower()

            # Kiểm tra trùng trong cùng file Excel
            if title_lower in seen_in_file:
                error_list.append(f"Dòng {row_idx}: Tên thuốc '{title}' bị trùng trong file Excel")
                continue

            # Kiểm tra trùng với dữ liệu đã có trong hệ thống
            if title_lower in existing_titles:
                error_list.append(f"Dòng {row_idx}: Tên thuốc '{title}' đã tồn tại trong hệ thống")
                continue

            try:
                doc = frappe.get_doc({
                    "doctype": "SIS Medicine",
                    "title": title,
                    "unit": unit or None,
                    "description": description or None,
                    "campus": campus,
                    "enabled": 1
                })
                doc.insert(ignore_permissions=True)
                success_count += 1
                # Thêm vào tập đã xử lý để tránh trùng các dòng sau trong file
                seen_in_file.add(title_lower)
                existing_titles.add(title_lower)
            except Exception as e:
                error_list.append(f"Dòng {row_idx}: {str(e)}")

        frappe.db.commit()

        message = f"Import thuốc hoàn tất"
        return success_response(
            data={
                "success_count": success_count,
                "total_count": total_count,
                "errors": error_list,
                "message": message
            },
            message=message
        )

    except Exception as e:
        frappe.logger().error(f"Error importing medicines: {str(e)}")
        return error_response(
            message=f"Lỗi khi import thuốc: {str(e)}",
            code="IMPORT_MEDICINE_ERROR"
        )


# ========================================================================================
# TEACHER HEALTH APIs - Cho giáo viên xem hồ sơ thăm khám và gửi đến phụ huynh
# ========================================================================================

@frappe.whitelist(allow_guest=False)
def get_class_health_examinations():
    """
    Lấy danh sách hồ sơ thăm khám của học sinh trong lớp theo ngày.
    Dành cho giáo viên chủ nhiệm xem.
    
    Params:
        - class_id: ID lớp (required)
        - date: Ngày cần lấy dữ liệu (optional, default: today)
    
    Returns:
        - data: Danh sách học sinh có hồ sơ thăm khám, grouped theo student
    """
    try:
        _check_teacher_permission()
        
        data = _get_request_data()
        class_id = data.get("class_id")
        date = data.get("date") or today()
        
        if not class_id:
            return validation_error_response("Thiếu class_id", {"class_id": ["class_id là bắt buộc"]})
        
        # Lấy danh sách student_id trong lớp
        class_students = frappe.get_all(
            "SIS Class Student",
            filters={"class_id": class_id},
            fields=["student_id"]
        )
        student_ids = [cs.student_id for cs in class_students]
        
        if not student_ids:
            return success_response(data={"data": []}, message="Không có học sinh trong lớp")
        
        # Lấy tất cả Daily Health Visits của các học sinh trong ngày
        visits = frappe.get_all(
            "SIS Daily Health Visit",
            filters={
                "student_id": ["in", student_ids],
                "visit_date": date
            },
            fields=[
                "name", "student_id", "student_name", "student_code",
                "visit_date", "reason", "leave_class_time", "arrive_clinic_time",
                "leave_clinic_time", "status", "reported_by_name", "received_by_name",
                "checkout_notes", "transfer_hospital", "accompanying_teacher", "accompanying_health_staff"
            ],
            order_by="creation desc"
        )
        
        # Lấy tất cả Health Examinations của các học sinh trong ngày
        examinations = frappe.get_all(
            "SIS Health Examination",
            filters={
                "student_id": ["in", student_ids],
                "examination_date": date
            },
            fields=[
                "name", "student_id", "student_name", "student_code",
                "examination_date", "visit_id", "symptoms", "disease_classification",
                "examination_notes", "treatment_type", "treatment_details", "notes",
                "outcome", "examined_by", "examined_by_name",
                "medical_staff", "medical_staff_name",
                "clinic_checkin_time", "clinic_checkout_time",
                "sent_to_parent", "sent_to_parent_at",
                "creation", "modified",
                "hospital_diagnosis", "hospital_treatment",
                "hospital_insurance", "hospital_school_coordination",
                "hospital_medical_staff", "hospital_medical_staff_name",
                "hospital_direction", "hospital_advance_cost",
                "hospital_payer", "hospital_payer_other",
                "hospital_transport", "hospital_transport_other",
                "hospital_health_monitoring", "hospital_notes",
                "followup_checkin_time", "followup_examination",
                "followup_treatment_details", "followup_checkout_time",
                "followup_outcome", "followup_notes",
                "followup_transfer_hospital", "followup_accompanying_teacher",
                "followup_accompanying_health_staff",
                "followup_clinic_checkin_time", "followup_clinic_checkout_time",
                "followup_is_scheduled_recheck", "followup_medical_suggestion",
                "followup_medical_staff", "followup_medical_staff_name"
            ],
            order_by="creation desc"
        )
        
        # Lấy năm học hiện tại để query ảnh học sinh
        current_school_year = frappe.db.get_value(
            "SIS School Year",
            {"is_enable": 1},
            "name"
        )
        
        # Lấy ảnh học sinh từ SIS Photo (batch query)
        student_photos = {}
        if student_ids:
            # Ưu tiên lấy ảnh theo năm học hiện tại
            if current_school_year:
                photos = frappe.get_all(
                    "SIS Photo",
                    filters={
                        "student_id": ["in", student_ids],
                        "school_year_id": current_school_year
                    },
                    fields=["student_id", "photo"],
                    order_by="creation desc"
                )
                for p in photos:
                    if p.student_id and p.photo and p.student_id not in student_photos:
                        student_photos[p.student_id] = p.photo
            
            # Fallback - lấy ảnh mới nhất cho các student chưa có ảnh
            missing_ids = [sid for sid in student_ids if sid not in student_photos]
            if missing_ids:
                photos = frappe.get_all(
                    "SIS Photo",
                    filters={"student_id": ["in", missing_ids]},
                    fields=["student_id", "photo"],
                    order_by="creation desc"
                )
                for p in photos:
                    if p.student_id and p.photo and p.student_id not in student_photos:
                        student_photos[p.student_id] = p.photo
        
        # Group theo student
        student_data = {}
        
        # Thêm visits
        for visit in visits:
            sid = visit.student_id
            if sid not in student_data:
                student_data[sid] = {
                    "student_id": sid,
                    "student_name": visit.student_name,
                    "student_code": visit.student_code,
                    "student_photo": student_photos.get(sid, ""),
                    "visits": [],
                    "examinations": []
                }
            student_data[sid]["visits"].append(visit)
        
        # Thêm examinations
        for exam in examinations:
            sid = exam.student_id
            if sid not in student_data:
                student_data[sid] = {
                    "student_id": sid,
                    "student_name": exam.student_name,
                    "student_code": exam.student_code,
                    "student_photo": student_photos.get(sid, ""),
                    "visits": [],
                    "examinations": []
                }
            
            # Lấy images cho exam và chuyển sang URL đầy đủ
            images = frappe.get_all(
                "SIS Examination Image",
                filters={"parent": exam.name},
                fields=["image", "description"]
            )
            for img in images:
                if img.get("image"):
                    if img["image"].startswith("http://") or img["image"].startswith("https://"):
                        pass  # Đã là URL đầy đủ
                    elif img["image"].startswith("/"):
                        img["image"] = frappe.utils.get_url(img["image"])
                    else:
                        img["image"] = frappe.utils.get_url("/files/" + img["image"])
            exam["images"] = images

            # Lấy reason và leave_clinic_time từ visit tương ứng
            visit_id = exam.get("visit_id")
            if visit_id:
                visit_data = frappe.db.get_value(
                    "SIS Daily Health Visit",
                    visit_id,
                    ["reason", "leave_clinic_time"],
                    as_dict=True
                )
                if visit_data:
                    exam["visit_reason"] = visit_data.get("reason") or ""
                    # Thời gian ra về: ưu tiên exam.clinic_checkout_time, fallback visit.leave_clinic_time (dữ liệu cũ)
                    exam["clinic_checkout_time"] = exam.get("clinic_checkout_time") or visit_data.get("leave_clinic_time") or ""
                else:
                    exam["visit_reason"] = ""
            else:
                exam["visit_reason"] = ""
            
            student_data[sid]["examinations"].append(exam)
        
        # Convert dict to list và sắp xếp theo tên
        result = sorted(student_data.values(), key=lambda x: x.get("student_name", ""))
        
        return success_response(
            data={"data": result},
            message=f"Lấy danh sách thăm khám thành công ({len(result)} học sinh)"
        )
        
    except frappe.PermissionError as e:
        return error_response(message=str(e), code="PERMISSION_DENIED")
    except Exception as e:
        frappe.logger().error(f"Error getting class health examinations: {str(e)}")
        import traceback
        frappe.logger().error(traceback.format_exc())
        return error_response(
            message=f"Lỗi khi lấy danh sách thăm khám: {str(e)}",
            code="GET_CLASS_HEALTH_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def send_exam_to_parent():
    """
    Gửi hồ sơ thăm khám đến phụ huynh qua notification.
    
    Params:
        - exam_ids: Danh sách ID hồ sơ thăm khám cần gửi (required)
    
    Returns:
        - success: True/False
        - message: Thông báo kết quả
        - sent_count: Số lượng đã gửi thành công
    """
    try:
        _check_teacher_permission()
        
        data = _get_request_data()
        exam_ids = data.get("exam_ids") or []
        
        if not exam_ids:
            return validation_error_response("Thiếu exam_ids", {"exam_ids": ["exam_ids là bắt buộc"]})
        
        if isinstance(exam_ids, str):
            exam_ids = [exam_ids]
        
        # Lấy thông tin các exam
        exams = frappe.get_all(
            "SIS Health Examination",
            filters={"name": ["in", exam_ids]},
            fields=["name", "student_id", "student_name", "student_code", "disease_classification", "sent_to_parent"]
        )
        
        if not exams:
            return validation_error_response("Không tìm thấy hồ sơ thăm khám", {"exam_ids": ["Không tìm thấy hồ sơ"]})
        
        # Lọc chỉ những exam chưa gửi
        exams_to_send = [e for e in exams if not e.sent_to_parent]
        
        if not exams_to_send:
            return success_response(
                data={"sent_count": 0},
                message="Tất cả hồ sơ đã được gửi trước đó"
            )
        
        # Import notification handler
        from erp.utils.notification_handler import send_bulk_parent_notifications
        
        # Group exams theo student để gửi notification gộp
        student_exams = {}
        for exam in exams_to_send:
            sid = exam.student_id
            if sid not in student_exams:
                student_exams[sid] = {
                    "student_name": exam.student_name,
                    "student_code": exam.student_code,
                    "exams": []
                }
            student_exams[sid]["exams"].append(exam)
        
        sent_count = 0
        notification_results = []
        
        for student_id, info in student_exams.items():
            student_name = info["student_name"]
            exam_names = [e.name for e in info["exams"]]
            
            # Chuẩn bị notification
            title = {
                "vi": f"Học sinh {student_name} có cập nhật mới về sức khỏe",
                "en": f"Student {student_name} has health update"
            }
            
            body = {
                "vi": f"Phòng Y tế đã ghi nhận thông tin thăm khám cho {student_name}. Nhấn để xem chi tiết.",
                "en": f"Health clinic has recorded examination information for {student_name}. Tap to view details."
            }
            
            try:
                result = send_bulk_parent_notifications(
                    recipient_type="health_examination",
                    recipients_data={"student_ids": [student_id]},
                    title=title,
                    body=body,
                    icon="/health-icon.png",
                    data={
                        "type": "health_examination",
                        "student_id": student_id,
                        "student_name": student_name,
                        "exam_ids": exam_names
                    }
                )
                
                notification_results.append({
                    "student_id": student_id,
                    "success": result.get("success", False),
                    "message": result.get("message", "")
                })
                
                if result.get("success"):
                    # Cập nhật sent_to_parent cho các exams
                    for exam_name in exam_names:
                        frappe.db.set_value(
                            "SIS Health Examination",
                            exam_name,
                            {
                                "sent_to_parent": 1,
                                "sent_to_parent_at": now()
                            },
                            update_modified=False
                        )
                        sent_count += 1
                        
            except Exception as notif_err:
                frappe.logger().error(f"Error sending notification for student {student_id}: {str(notif_err)}")
                notification_results.append({
                    "student_id": student_id,
                    "success": False,
                    "message": str(notif_err)
                })
        
        frappe.db.commit()
        
        return success_response(
            data={
                "sent_count": sent_count,
                "total_requested": len(exams_to_send),
                "results": notification_results
            },
            message=f"Đã gửi {sent_count}/{len(exams_to_send)} hồ sơ thăm khám đến phụ huynh"
        )
        
    except frappe.PermissionError as e:
        return error_response(message=str(e), code="PERMISSION_DENIED")
    except Exception as e:
        frappe.db.rollback()
        frappe.logger().error(f"Error sending exam to parent: {str(e)}")
        import traceback
        frappe.logger().error(traceback.format_exc())
        return error_response(
            message=f"Lỗi khi gửi hồ sơ đến phụ huynh: {str(e)}",
            code="SEND_EXAM_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def recall_exam_from_parent():
    """
    Thu hồi hồ sơ thăm khám đã gửi đến phụ huynh.
    Xóa notifications và cập nhật trạng thái sent_to_parent = 0.
    
    Params:
        - exam_ids: Danh sách ID hồ sơ thăm khám cần thu hồi (required)
    
    Returns:
        - success: True/False
        - message: Thông báo kết quả
        - recalled_count: Số lượng đã thu hồi thành công
    """
    try:
        _check_teacher_permission()
        
        data = _get_request_data()
        exam_ids = data.get("exam_ids") or []
        
        if not exam_ids:
            return validation_error_response("Thiếu exam_ids", {"exam_ids": ["exam_ids là bắt buộc"]})
        
        if isinstance(exam_ids, str):
            exam_ids = [exam_ids]
        
        # Lấy thông tin các exam đã gửi
        exams = frappe.get_all(
            "SIS Health Examination",
            filters={
                "name": ["in", exam_ids],
                "sent_to_parent": 1
            },
            fields=["name", "student_id"]
        )
        
        if not exams:
            return success_response(
                data={"recalled_count": 0},
                message="Không có hồ sơ nào cần thu hồi"
            )
        
        recalled_count = 0
        
        for exam in exams:
            exam_id = exam.name
            
            # Tìm và xóa notifications liên quan
            # Pattern: Tìm notifications có type health_examination và chứa exam_id trong data
            notifications = frappe.db.sql("""
                SELECT name FROM `tabERP Notification`
                WHERE notification_type = 'health_examination'
                AND (
                    data LIKE %(pattern1)s
                    OR data LIKE %(pattern2)s
                )
            """, {
                "pattern1": f'%"{exam_id}"%',
                "pattern2": f'%\'{exam_id}\'%'
            }, as_dict=True)
            
            for notif in notifications:
                try:
                    frappe.delete_doc("ERP Notification", notif.name, force=True, ignore_permissions=True)
                except Exception as del_err:
                    frappe.logger().warning(f"Could not delete notification {notif.name}: {str(del_err)}")
            
            # Cập nhật trạng thái exam
            frappe.db.set_value(
                "SIS Health Examination",
                exam_id,
                {
                    "sent_to_parent": 0,
                    "sent_to_parent_at": None
                },
                update_modified=False
            )
            recalled_count += 1
        
        frappe.db.commit()
        
        return success_response(
            data={"recalled_count": recalled_count},
            message=f"Đã thu hồi {recalled_count} hồ sơ thăm khám"
        )
        
    except frappe.PermissionError as e:
        return error_response(message=str(e), code="PERMISSION_DENIED")
    except Exception as e:
        frappe.db.rollback()
        frappe.logger().error(f"Error recalling exam from parent: {str(e)}")
        import traceback
        frappe.logger().error(traceback.format_exc())
        return error_response(
            message=f"Lỗi khi thu hồi hồ sơ: {str(e)}",
            code="RECALL_EXAM_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def get_parent_health_records():
    """
    API dành cho Parent Portal: Lấy danh sách hồ sơ thăm khám đã được gửi đến phụ huynh.
    
    Params:
        - student_id: ID học sinh (required)
    
    Returns:
        - data: Danh sách hồ sơ thăm khám grouped theo ngày
    """
    try:
        data = _get_request_data()
        student_id = data.get("student_id") or frappe.form_dict.get("student_id")
        
        if not student_id:
            return validation_error_response("Thiếu student_id", {"student_id": ["student_id là bắt buộc"]})
        
        # Lấy tất cả examinations đã gửi đến parent
        examinations = frappe.get_all(
            "SIS Health Examination",
            filters={
                "student_id": student_id,
                "sent_to_parent": 1
            },
            fields=[
                "name", "student_id", "student_name", "student_code",
                "examination_date", "visit_id", "symptoms", "diet_history",
                "disease_classification", "examination_notes",
                "treatment_type", "treatment_details", "notes",
                "outcome", "examined_by", "examined_by_name",
                "medical_staff", "medical_staff_name",
                "clinic_checkin_time", "clinic_checkout_time",
                "sent_to_parent_at",
                "hospital_diagnosis", "hospital_treatment",
                "hospital_insurance", "hospital_school_coordination",
                "hospital_medical_staff", "hospital_medical_staff_name",
                "hospital_direction", "hospital_advance_cost",
                "hospital_payer", "hospital_payer_other",
                "hospital_transport", "hospital_transport_other",
                "hospital_health_monitoring", "hospital_notes",
                "followup_checkin_time", "followup_examination",
                "followup_treatment_details", "followup_checkout_time",
                "followup_outcome", "followup_notes",
                "followup_transfer_hospital", "followup_accompanying_teacher",
                "followup_accompanying_health_staff",
                "followup_clinic_checkin_time", "followup_clinic_checkout_time",
                "followup_is_scheduled_recheck", "followup_medical_suggestion",
                "followup_medical_staff", "followup_medical_staff_name",
                "creation", "modified"
            ],
            order_by="examination_date desc, creation desc"
        )
        
        # Lấy images + checkout_notes từ visit cho từng exam, chuyển image sang URL đầy đủ
        for exam in examinations:
            images = frappe.get_all(
                "SIS Examination Image",
                filters={"parent": exam.name},
                fields=["image", "description"]
            )
            for img in images:
                if img.get("image"):
                    if img["image"].startswith("http://") or img["image"].startswith("https://"):
                        pass
                    elif img["image"].startswith("/"):
                        img["image"] = frappe.utils.get_url(img["image"])
                    else:
                        img["image"] = frappe.utils.get_url("/files/" + img["image"])
            exam["images"] = images
            
            if exam.get("visit_id"):
                visit_data = frappe.db.get_value(
                    "SIS Daily Health Visit",
                    exam["visit_id"],
                    ["checkout_notes", "reason", "transfer_hospital", "leave_clinic_time"],
                    as_dict=True
                )
                if visit_data:
                    exam["checkout_notes"] = visit_data.get("checkout_notes")
                    exam["visit_reason"] = visit_data.get("reason") or ""
                    # Bệnh viện chuyển tới từ visit (khi checkout chuyển viện) - dùng cho màn PH và GVCN
                    exam["transfer_hospital"] = visit_data.get("transfer_hospital") or ""
                    # Thời gian ra về: ưu tiên exam.clinic_checkout_time (nếu hợp lệ), fallback visit.leave_clinic_time
                    # Lưu ý: Frappe lưu Time rỗng thành "00:00:00" - coi là rỗng để fallback sang visit (fix Thời gian lưu trú không hiển thị)
                    raw_checkout = exam.get("clinic_checkout_time")
                    raw_str = (str(raw_checkout).strip() if raw_checkout else "")[:8]
                    if not raw_str or raw_str in ("00:00:00", "00:00"):
                        exam["clinic_checkout_time"] = visit_data.get("leave_clinic_time") or ""
                else:
                    exam["visit_reason"] = ""
                    exam["transfer_hospital"] = ""
        
        # Group theo ngày
        grouped = {}
        for exam in examinations:
            date_key = str(exam.examination_date)
            if date_key not in grouped:
                grouped[date_key] = []
            grouped[date_key].append(exam)
        
        # Convert to list sorted by date desc
        result = [
            {"date": date, "examinations": exams}
            for date, exams in sorted(grouped.items(), reverse=True)
        ]
        
        return success_response(
            data={"data": result, "total": len(examinations)},
            message=f"Lấy {len(examinations)} hồ sơ thăm khám thành công"
        )
        
    except Exception as e:
        frappe.logger().error(f"Error getting parent health records: {str(e)}")
        import traceback
        frappe.logger().error(traceback.format_exc())
        return error_response(
            message=f"Lỗi khi lấy hồ sơ thăm khám: {str(e)}",
            code="GET_PARENT_HEALTH_ERROR"
        )


# ========================================================================================
# FIRST AID CONFIG APIs
# ========================================================================================

@frappe.whitelist()
def get_first_aid_items(campus: str = None):
    """
    Lấy danh sách vật tư sơ cứu
    """
    try:
        filters = {"enabled": 1}
        
        # Lấy campus từ nhiều nguồn
        campus = campus or frappe.form_dict.get('campus')
        if campus:
            filters["campus"] = campus
        
        items = frappe.get_all(
            "SIS First Aid",
            filters=filters,
            fields=["name", "title", "unit", "description", "campus", "enabled", "creation", "modified"],
            order_by="title asc"
        )
        
        return success_response(
            data={"data": items, "total": len(items)},
            message="Lấy danh sách sơ cứu thành công"
        )
    
    except Exception as e:
        frappe.logger().error(f"Error getting first aid items: {str(e)}")
        return error_response(
            message=f"Lỗi khi lấy danh sách sơ cứu: {str(e)}",
            code="GET_FIRST_AID_ERROR"
        )


@frappe.whitelist()
def create_first_aid(title: str = None, campus: str = None, unit: str = None, description: str = None):
    """
    Tạo mới vật tư sơ cứu
    """
    try:
        # Lấy dữ liệu từ request JSON
        import json
        data = {}
        if frappe.request and frappe.request.data:
            try:
                data = json.loads(frappe.request.data)
            except json.JSONDecodeError:
                pass
        
        title = title or data.get('title') or frappe.form_dict.get('title')
        campus = campus or data.get('campus') or frappe.form_dict.get('campus')
        unit = unit or data.get('unit') or frappe.form_dict.get('unit')
        description = description or data.get('description') or frappe.form_dict.get('description')
        
        if not title or not campus:
            return error_response(
                message="Tên sơ cứu và trường học là bắt buộc",
                code="MISSING_REQUIRED_FIELDS"
            )
        
        # Kiểm tra trùng tên trong cùng campus
        existing = frappe.get_all(
            "SIS First Aid",
            filters={"title": title, "campus": campus},
            limit=1
        )
        if existing:
            return error_response(
                message="Vật tư sơ cứu này đã tồn tại",
                code="DUPLICATE_FIRST_AID"
            )
        
        doc = frappe.get_doc({
            "doctype": "SIS First Aid",
            "title": title,
            "unit": unit,
            "description": description,
            "campus": campus,
            "enabled": 1
        })
        doc.insert(ignore_permissions=True)
        frappe.db.commit()
        
        return success_response(
            data={"name": doc.name, "title": doc.title},
            message="Tạo vật tư sơ cứu thành công"
        )
    
    except Exception as e:
        frappe.logger().error(f"Error creating first aid: {str(e)}")
        return error_response(
            message=f"Lỗi khi tạo vật tư sơ cứu: {str(e)}",
            code="CREATE_FIRST_AID_ERROR"
        )


@frappe.whitelist()
def update_first_aid(name: str = None, title: str = None, unit: str = None, description: str = None, enabled: int = None):
    """
    Cập nhật vật tư sơ cứu
    """
    try:
        # Lấy dữ liệu từ request JSON
        import json
        data = {}
        if frappe.request and frappe.request.data:
            try:
                data = json.loads(frappe.request.data)
            except json.JSONDecodeError:
                pass
        
        name = name or data.get('name') or frappe.form_dict.get('name')
        title = title if title is not None else data.get('title')
        unit = unit if unit is not None else data.get('unit')
        description = description if description is not None else data.get('description')
        enabled = enabled if enabled is not None else data.get('enabled')
        
        if not name:
            return error_response(
                message="Thiếu ID vật tư sơ cứu",
                code="MISSING_NAME"
            )
        
        # Lấy document
        doc = frappe.get_doc("SIS First Aid", name)
        
        # Kiểm tra trùng tên nếu đổi title
        if title and title != doc.title:
            existing = frappe.get_all(
                "SIS First Aid",
                filters={
                    "title": title,
                    "campus": doc.campus,
                    "name": ["!=", name]
                },
                limit=1
            )
            if existing:
                return error_response(
                    message="Tên vật tư sơ cứu này đã tồn tại",
                    code="DUPLICATE_FIRST_AID"
                )
            doc.title = title
        
        if unit is not None:
            doc.unit = unit
        if description is not None:
            doc.description = description
        if enabled is not None:
            doc.enabled = enabled
        
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        
        return success_response(
            data={"name": doc.name, "title": doc.title},
            message="Cập nhật vật tư sơ cứu thành công"
        )
    
    except frappe.DoesNotExistError:
        return error_response(
            message="Vật tư sơ cứu không tồn tại",
            code="NOT_FOUND"
        )
    except Exception as e:
        frappe.logger().error(f"Error updating first aid: {str(e)}")
        return error_response(
            message=f"Lỗi khi cập nhật vật tư sơ cứu: {str(e)}",
            code="UPDATE_FIRST_AID_ERROR"
        )


@frappe.whitelist()
def delete_first_aid(name: str = None):
    """
    Xóa vật tư sơ cứu
    """
    try:
        # Lấy dữ liệu từ request JSON
        import json
        data = {}
        if frappe.request and frappe.request.data:
            try:
                data = json.loads(frappe.request.data)
            except json.JSONDecodeError:
                pass
        
        name = name or data.get('name') or frappe.form_dict.get('name')
        
        if not name:
            return error_response(
                message="Thiếu ID vật tư sơ cứu",
                code="MISSING_NAME"
            )
        
        # Kiểm tra tồn tại
        if not frappe.db.exists("SIS First Aid", name):
            return error_response(
                message="Vật tư sơ cứu không tồn tại",
                code="NOT_FOUND"
            )
        
        frappe.delete_doc("SIS First Aid", name, ignore_permissions=True)
        frappe.db.commit()
        
        return success_response(
            data={"name": name},
            message="Xóa vật tư sơ cứu thành công"
        )
    
    except Exception as e:
        frappe.logger().error(f"Error deleting first aid: {str(e)}")
        return error_response(
            message=f"Lỗi khi xóa vật tư sơ cứu: {str(e)}",
            code="DELETE_FIRST_AID_ERROR"
        )


@frappe.whitelist(allow_guest=False)
def import_first_aid_excel(campus=None):
    """
    Import vật tư sơ cứu từ file Excel.
    Cột A: Tên sơ cứu, Cột B: Đơn vị (tùy chọn), Cột C: Mô tả (tùy chọn)
    """
    try:
        import openpyxl
        import io

        # Đọc campus từ nhiều nguồn: tham số hàm, form_dict, query string
        campus = (
            campus
            or frappe.form_dict.get('campus')
            or (frappe.request.args.get('campus') if frappe.request else None)
        )
        if not campus:
            return error_response(
                message="Thiếu thông tin trường học",
                code="MISSING_CAMPUS"
            )

        # Kiểm tra campus tồn tại
        if not frappe.db.exists("SIS Campus", campus):
            return error_response(
                message=f"Không tìm thấy trường học: {campus}",
                code="CAMPUS_NOT_FOUND"
            )

        # Lấy file từ request
        file_obj = frappe.request.files.get('file')
        if not file_obj:
            return error_response(
                message="Không tìm thấy file",
                code="MISSING_FILE"
            )

        file_data = file_obj.read()
        wb = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True, data_only=True)
        ws = wb.active

        success_count = 0
        error_list = []
        total_count = 0

        # Tải trước danh sách tên sơ cứu đã có để kiểm tra trùng nhanh (so sánh lowercase)
        existing_titles = set(
            r.title.strip().lower()
            for r in frappe.get_all("SIS First Aid", filters={"campus": campus}, fields=["title"])
            if r.title
        )
        # Theo dõi tên trong file Excel để phát hiện trùng nội bộ
        seen_in_file = set()

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            # Bỏ qua dòng rỗng hoàn toàn
            if not any(row):
                continue

            title = str(row[0]).strip() if row[0] is not None else ''
            unit = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            description = str(row[2]).strip() if len(row) > 2 and row[2] is not None else ''
            total_count += 1

            # Validate dữ liệu bắt buộc
            if not title:
                error_list.append(f"Dòng {row_idx}: Tên sơ cứu không được để trống")
                continue

            title_lower = title.lower()

            # Kiểm tra trùng trong cùng file Excel
            if title_lower in seen_in_file:
                error_list.append(f"Dòng {row_idx}: Tên sơ cứu '{title}' bị trùng trong file Excel")
                continue

            # Kiểm tra trùng với dữ liệu đã có trong hệ thống
            if title_lower in existing_titles:
                error_list.append(f"Dòng {row_idx}: Tên sơ cứu '{title}' đã tồn tại trong hệ thống")
                continue

            try:
                doc = frappe.get_doc({
                    "doctype": "SIS First Aid",
                    "title": title,
                    "unit": unit or None,
                    "description": description or None,
                    "campus": campus,
                    "enabled": 1
                })
                doc.insert(ignore_permissions=True)
                success_count += 1
                # Thêm vào tập đã xử lý để tránh trùng các dòng sau trong file
                seen_in_file.add(title_lower)
                existing_titles.add(title_lower)
            except Exception as e:
                error_list.append(f"Dòng {row_idx}: {str(e)}")

        frappe.db.commit()

        message = f"Import vật tư sơ cứu hoàn tất"
        return success_response(
            data={
                "success_count": success_count,
                "total_count": total_count,
                "errors": error_list,
                "message": message
            },
            message=message
        )

    except Exception as e:
        frappe.logger().error(f"Error importing first aid items: {str(e)}")
        return error_response(
            message=f"Lỗi khi import vật tư sơ cứu: {str(e)}",
            code="IMPORT_FIRST_AID_ERROR"
        )
