"""Build grid preview kiểu Excel và export XLSX/ZIP cho TKB draft."""

from __future__ import annotations

import io
import json
import zipfile
from typing import Dict, List, Optional, Tuple

import frappe

DAY_ORDER = {"mon": 0, "tue": 1, "wed": 2, "thu": 3, "fri": 4, "sat": 5, "sun": 6}

DAY_LABEL_VN = {
	"mon": "Thứ 2",
	"tue": "Thứ 3",
	"wed": "Thứ 4",
	"thu": "Thứ 5",
	"fri": "Thứ 6",
	"sat": "Thứ 7",
	"sun": "Chủ nhật",
}


def draft_has_variant_index() -> bool:
	try:
		return bool(frappe.db.sql("SHOW COLUMNS FROM `tabSIS_TKB_Gen_Result` LIKE 'variant_index'"))
	except Exception:
		return False


def _parse_teacher_ids(raw) -> List[str]:
	if not raw:
		return []
	try:
		parsed = json.loads(raw) if isinstance(raw, str) else raw
		return [str(t) for t in parsed if t]
	except (json.JSONDecodeError, TypeError):
		return []


def _build_teacher_name_map(all_ids: List[str]) -> Dict[str, str]:
	if not all_ids:
		return {}
	rows = frappe.db.sql(
		"""
		SELECT t.name,
		       COALESCE(NULLIF(u.full_name, ''), u.first_name, t.user_id, t.name) AS full_name
		FROM `tabSIS Teacher` t
		LEFT JOIN `tabUser` u ON u.name = t.user_id
		WHERE t.name IN %(ids)s
		""",
		{"ids": all_ids},
		as_dict=True,
	)
	return {r["name"]: (r.get("full_name") or r["name"]) for r in rows}


def _fetch_raw_rows(session_id: str, variant_index: int) -> List[dict]:
	variant_clause = "AND r.variant_index = %(variant_index)s" if draft_has_variant_index() else ""
	return frappe.db.sql(
		f"""
		SELECT
			r.class_id, r.day_of_week, r.timetable_column_id,
			r.timetable_subject_id, r.teacher_ids, r.room_id, r.period_priority,
			ts.title_vn AS subject_title,
			tc.period_name, tc.start_time, tc.end_time,
			c.title AS class_title,
			COALESCE(NULLIF(room.physical_code, ''), room.title_vn, '') AS room_title
		FROM `tabSIS_TKB_Gen_Result` r
		LEFT JOIN `tabSIS Timetable Subject` ts ON ts.name = r.timetable_subject_id
		LEFT JOIN `tabSIS Timetable Column` tc ON tc.name = r.timetable_column_id
		LEFT JOIN `tabSIS Class` c ON c.name = r.class_id
		LEFT JOIN `tabERP Administrative Room` room ON room.name = r.room_id
		WHERE r.session_id = %(session_id)s {variant_clause}
		ORDER BY r.day_of_week, r.period_priority, c.title
		""",
		{"session_id": session_id, "variant_index": variant_index},
		as_dict=True,
	)


def _format_time(value) -> str:
	if not value:
		return ""
	text = str(value)
	return text[:5] if len(text) >= 5 else text


def build_excel_grid(session_id: str, variant_index: int = 0) -> dict:
	"""Dựng cấu trúc grid Thứ | Tiết | [lớp...] cho preview UI."""
	from .data_collector import TimetableDataCollector

	collector = TimetableDataCollector(session_id)
	inp = collector.collect()
	raw_rows = _fetch_raw_rows(session_id, variant_index)

	# Danh sách lớp — ưu tiên từ session scope, fallback từ draft
	class_map: Dict[str, str] = {c.name: c.title for c in inp.classes}
	for r in raw_rows:
		cid = r["class_id"]
		if cid not in class_map:
			class_map[cid] = r.get("class_title") or cid
	classes = [{"name": k, "title": v} for k, v in sorted(class_map.items(), key=lambda x: x[1])]

	if not classes:
		return {"classes": [], "rows": [], "variant_index": variant_index}

	# Batch resolve tên GV
	all_teacher_ids: set[str] = set()
	for r in raw_rows:
		all_teacher_ids.update(_parse_teacher_ids(r.get("teacher_ids")))
	teacher_name_map = _build_teacher_name_map(list(all_teacher_ids))

	# Gom dữ liệu ô theo (day, period_priority)
	slot_map: Dict[Tuple[str, int], dict] = {}
	for r in raw_rows:
		day = r.get("day_of_week") or "mon"
		priority = int(r.get("period_priority") or 0)
		key = (day, priority)
		if key not in slot_map:
			slot_map[key] = {"cells": {}}
		teacher_ids = _parse_teacher_ids(r.get("teacher_ids"))
		slot_map[key]["cells"][r["class_id"]] = {
			"subject_title": r.get("subject_title") or "",
			"teacher_names": [teacher_name_map.get(tid, tid) for tid in teacher_ids],
			"room_title": r.get("room_title") or "",
		}

	# Khung đầy đủ ngày × tiết (cho phép click ô trống)
	study_periods = sorted(
		[p for p in inp.periods if p.period_type == "study"],
		key=lambda x: x.period_priority,
	) or sorted(inp.periods, key=lambda x: x.period_priority)

	rows = []
	for day in inp.working_days:
		for period in study_periods:
			key = (day, period.period_priority)
			cells = slot_map.get(key, {}).get("cells", {})
			rows.append({
				"day_of_week": day,
				"day_label": DAY_LABEL_VN.get(day, day),
				"period_name": period.period_name or f"Tiết {period.period_priority}",
				"period_priority": period.period_priority,
				"timetable_column_id": period.name,
				"start_time": _format_time(period.start_time),
				"end_time": _format_time(period.end_time),
				"cells": cells,
			})

	return {
		"classes": classes,
		"rows": rows,
		"variant_index": variant_index,
	}


def _cell_text(cell: Optional[dict]) -> str:
	if not cell:
		return ""
	subject = (cell.get("subject_title") or "").strip()
	teachers = cell.get("teacher_names") or []
	teacher_line = " · ".join(t for t in teachers if t)
	room = (cell.get("room_title") or "").strip()
	lines = [subject]
	if teacher_line:
		lines.append(teacher_line)
	if room:
		lines.append(room)
	return "\n".join(lines)


def build_xlsx_bytes(session_id: str, variant_index: int, session_title: str = "") -> bytes:
	"""Tạo file Excel column-based cho 1 biến thể."""
	try:
		import openpyxl
		from openpyxl.styles import Alignment, Border, Font, Side
	except ImportError:
		frappe.throw("Thiếu openpyxl để export TKB")

	grid = build_excel_grid(session_id, variant_index)
	wb = openpyxl.Workbook()
	ws = wb.active
	title = (session_title or session_id or "TKB").strip()
	ws.title = f"Biến thể {variant_index + 1}"[:31]

	classes = grid["classes"]
	headers = ["Thứ", "Tiết"] + [c["title"] for c in classes]
	ws.append(headers)

	thin = Side(style="thin", color="CCCCCC")
	border = Border(left=thin, right=thin, top=thin, bottom=thin)
	header_font = Font(bold=True)
	center = Alignment(horizontal="center", vertical="center", wrap_text=True)
	left_wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

	for col_idx, _ in enumerate(headers, start=1):
		cell = ws.cell(row=1, column=col_idx)
		cell.font = header_font
		cell.alignment = center
		cell.border = border

	rows = grid["rows"]
	excel_row = 2
	day_merge_start: Optional[int] = None
	prev_day: Optional[str] = None

	for row_data in rows:
		day = row_data["day_of_week"]
		if day != prev_day:
			if day_merge_start is not None and excel_row - 1 > day_merge_start:
				ws.merge_cells(start_row=day_merge_start, start_column=1, end_row=excel_row - 1, end_column=1)
			day_merge_start = excel_row
			prev_day = day

		period_label = row_data["period_name"]
		if row_data.get("start_time") and row_data.get("end_time"):
			period_label = f"{period_label}\n{row_data['start_time']} – {row_data['end_time']}"

		# Chỉ ghi nhãn Thứ ở hàng đầu mỗi nhóm ngày (sẽ merge sau)
		ws.cell(
			row=excel_row,
			column=1,
			value=row_data["day_label"] if day_merge_start == excel_row else "",
		)
		ws.cell(row=excel_row, column=2, value=period_label)

		for col_offset, cls in enumerate(classes, start=3):
			cell_data = (row_data.get("cells") or {}).get(cls["name"])
			ws.cell(row=excel_row, column=col_offset, value=_cell_text(cell_data))

		for col_idx in range(1, len(headers) + 1):
			cell = ws.cell(row=excel_row, column=col_idx)
			cell.border = border
			cell.alignment = left_wrap if col_idx > 2 else center

		excel_row += 1

	if day_merge_start is not None and excel_row - 1 >= day_merge_start:
		ws.merge_cells(start_row=day_merge_start, start_column=1, end_row=excel_row - 1, end_column=1)

	ws.column_dimensions["A"].width = 10
	ws.column_dimensions["B"].width = 14
	for col_idx in range(3, len(headers) + 1):
		col_letter = openpyxl.utils.get_column_letter(col_idx)
		ws.column_dimensions[col_letter].width = 18

	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()


def build_zip_bytes(session_id: str, variant_indices: List[int]) -> Tuple[bytes, str]:
	"""Đóng gói nhiều biến thể thành ZIP — mỗi biến thể 1 file xlsx."""
	session_title = frappe.db.get_value("SIS Timetable Generation Session", session_id, "title") or session_id
	safe_title = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in session_title)[:40]

	buf = io.BytesIO()
	with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
		for idx in variant_indices:
			xlsx_data = build_xlsx_bytes(session_id, idx, session_title)
			filename = f"TKB-{safe_title}-bien-the-{idx + 1}.xlsx"
			zf.writestr(filename, xlsx_data)

	zip_name = f"TKB-{safe_title}-variants.zip"
	return buf.getvalue(), zip_name


def send_zip_download(session_id: str, variant_indices: List[int]) -> None:
	"""Trả ZIP qua frappe.local.response (download)."""
	content, filename = build_zip_bytes(session_id, variant_indices)
	frappe.local.response.filename = filename
	frappe.local.response.filecontent = content
	frappe.local.response.type = "download"
	frappe.local.response["Content-Type"] = "application/zip"
