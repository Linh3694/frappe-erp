# Copyright (c) 2026, Wellspring International School and contributors
# For license information, please see license.txt

"""
Sinh file mẫu import TKB theo ĐÚNG khung giờ đang áp dụng.

Vì sao cần: trước đây nút "Tải file mẫu" mở thẳng một file tĩnh
(public/Template/import-timetables.xlsx) với cột Tiết ghi cứng 'Tiết 1 + 2'… và danh
sách lớp ghi cứng của một năm học cũ. Mỗi lần trường đổi khung giờ (tạo SIS Schedule
mới) là file mẫu lệch tên tiết, người dùng điền vào rồi import — validator chấp nhận
nhầm cột của schedule đã nghỉ hưu và TKB hiển thị lệch tiết.

File mẫu giờ sinh từ chính dữ liệu import sẽ được đối chiếu: cột "Tiết" lấy nguyên văn
period_name của các study column thuộc schedule active, cột lớp lấy đúng lớp của campus
+ năm học + cấp học. Nhờ vậy file tải về luôn qua được bước kiểm tra nghiêm ngặt ở
import_validator._validate_period_references.
"""

from io import BytesIO
from typing import Dict, List, Optional

import frappe
from frappe.utils import now_datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from erp.utils.api_response import error_response

# Thứ 2 → Thứ 6, khớp lưới TKB mặc định (không kèm cuối tuần)
WEEKDAY_LABELS = ["Thứ 2", "Thứ 3", "Thứ 4", "Thứ 5", "Thứ 6"]

HEADER_FILL = PatternFill("solid", fgColor="F26B21")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LABEL_FONT = Font(bold=True, size=11)
THIN = Side(style="thin", color="D0D5DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _get_active_schedule(education_stage_id: str, campus_id: str, on_date) -> Optional[str]:
    """Khung giờ áp dụng cho ngày chỉ định — cùng điều kiện với validator/executor."""
    if not (education_stage_id and campus_id and on_date):
        return None
    return frappe.db.get_value(
        "SIS Schedule",
        {
            "education_stage_id": education_stage_id,
            "campus_id": campus_id,
            "is_active": 1,
            "start_date": ["<=", on_date],
            "end_date": [">=", on_date],
        },
        "name",
    )


def _get_period_names(education_stage_id: str, campus_id: str, start_date) -> List[str]:
    """
    Tên các tiết HỌC (bỏ giờ nghỉ) theo đúng thứ tự hiển thị.

    Ưu tiên schedule active, không có thì dùng cột legacy — giữ đúng thứ tự ưu tiên của
    import_validator để file mẫu không bao giờ mâu thuẫn với bước kiểm tra.
    """
    schedule_id = _get_active_schedule(education_stage_id, campus_id, start_date)
    if schedule_id:
        columns = frappe.get_all(
            "SIS Timetable Column",
            fields=["period_name", "period_priority"],
            filters={"schedule_id": schedule_id, "period_type": "study"},
            order_by="period_priority asc",
        )
    else:
        columns = frappe.get_all(
            "SIS Timetable Column",
            fields=["period_name", "period_priority"],
            filters={
                "education_stage_id": education_stage_id,
                "campus_id": campus_id,
                "period_type": "study",
                "schedule_id": ["is", "not set"],
            },
            order_by="period_priority asc",
        )

    return [c.period_name for c in columns if c.period_name]


def _get_class_titles(campus_id: str, school_year_id: str, education_stage_id: str) -> List[str]:
    """Lớp của cấp học, lấy short_title vì đó là thứ executor dùng để tra SIS Class."""
    grades = frappe.get_all(
        "SIS Education Grade",
        fields=["name"],
        filters={"education_stage_id": education_stage_id, "campus_id": campus_id},
    )
    if not grades:
        return []

    classes = frappe.get_all(
        "SIS Class",
        fields=["short_title", "title"],
        filters={
            "campus_id": campus_id,
            "school_year_id": school_year_id,
            "education_grade": ["in", [g.name for g in grades]],
        },
        order_by="short_title asc",
    )
    return [c.short_title or c.title for c in classes if (c.short_title or c.title)]


def _build_workbook(period_names: List[str], class_titles: List[str], meta: Dict) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "TKB"

    headers = ["Thứ", "Tiết"] + class_titles
    ws.append(headers)
    for idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    for day in WEEKDAY_LABELS:
        for period in period_names:
            ws.append([day, period] + [""] * len(class_titles))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.border = BORDER
            if cell.column <= 2:
                cell.font = LABEL_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 16
    for idx in range(3, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = 26
    ws.freeze_panes = "C2"

    # Sheet hướng dẫn: người dùng phải biết tên tiết là bắt buộc khớp nguyên văn
    guide = wb.create_sheet("Hướng dẫn")
    guide.append(["File mẫu sinh tự động — KHÔNG sửa cột 'Thứ' và 'Tiết'"])
    guide.append([])
    guide.append(["Campus", meta.get("campus", "")])
    guide.append(["Năm học", meta.get("school_year", "")])
    guide.append(["Cấp học", meta.get("education_stage", "")])
    guide.append(["Khung giờ áp dụng", meta.get("schedule", "")])
    guide.append(["Áp dụng từ ngày", meta.get("start_date", "")])
    guide.append(["Tạo lúc", meta.get("generated_at", "")])
    guide.append([])
    guide.append(["Lưu ý:"])
    guide.append(["1. Tên tiết phải khớp NGUYÊN VĂN với khung giờ đang áp dụng."])
    guide.append(["   Đổi khung giờ thì phải tải lại file mẫu, dùng file cũ sẽ bị từ chối."])
    guide.append(["2. Mỗi ô điền tên môn học đúng như trong hệ thống. Ô trống = không có tiết."])
    guide.append(["3. Không thêm/bớt/đổi tên cột lớp."])
    guide["A1"].font = Font(bold=True, size=12)
    guide.column_dimensions["A"].width = 30
    guide.column_dimensions["B"].width = 44

    return wb


@frappe.whitelist(allow_guest=False)
def download_import_template():
    """
    Tải file mẫu import TKB khớp khung giờ đang áp dụng.

    Query params: campus_id, school_year_id, education_stage_id, start_date (YYYY-MM-DD).
    start_date quyết định lấy schedule nào — cùng ngày mà người dùng sẽ nhập khi import.
    """
    try:
        args = frappe.local.form_dict
        campus_id = args.get("campus_id")
        school_year_id = args.get("school_year_id")
        education_stage_id = args.get("education_stage_id")
        start_date = args.get("start_date") or frappe.utils.nowdate()

        missing = [
            label
            for label, value in (
                ("campus_id", campus_id),
                ("school_year_id", school_year_id),
                ("education_stage_id", education_stage_id),
            )
            if not value
        ]
        if missing:
            return error_response(f"Thiếu tham số: {', '.join(missing)}")

        period_names = _get_period_names(education_stage_id, campus_id, start_date)
        if not period_names:
            return error_response(
                "Cấp học này chưa có khung giờ (SIS Schedule) nào áp dụng cho ngày "
                f"{start_date}. Cấu hình khung giờ trước khi tải file mẫu."
            )

        class_titles = _get_class_titles(campus_id, school_year_id, education_stage_id)
        if not class_titles:
            return error_response(
                "Không tìm thấy lớp nào của cấp học trong năm học đã chọn."
            )

        schedule_id = _get_active_schedule(education_stage_id, campus_id, start_date)
        meta = {
            "campus": frappe.db.get_value("SIS Campus", campus_id, "title_vn") or campus_id,
            "school_year": frappe.db.get_value("SIS School Year", school_year_id, "title_vn")
            or school_year_id,
            "education_stage": frappe.db.get_value(
                "SIS Education Stage", education_stage_id, "title_vn"
            )
            or education_stage_id,
            "schedule": schedule_id or "(chưa cấu hình — dùng danh sách tiết của cấp học)",
            "start_date": str(start_date),
            "generated_at": now_datetime().strftime("%d/%m/%Y %H:%M"),
        }

        workbook = _build_workbook(period_names, class_titles, meta)
        stream = BytesIO()
        workbook.save(stream)

        stamp = now_datetime().strftime("%Y%m%d-%H%M")
        frappe.response["filename"] = f"mau-import-tkb-{stamp}.xlsx"
        frappe.response["filecontent"] = stream.getvalue()
        frappe.response["type"] = "binary"

    except Exception as e:
        frappe.log_error(f"Download timetable import template failed: {str(e)}")
        return error_response(f"Lỗi khi tạo file mẫu: {str(e)}")
