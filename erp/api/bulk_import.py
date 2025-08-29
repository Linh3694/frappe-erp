import frappe
from frappe import _
from frappe.utils import now_datetime, get_datetime
import json
import os
from erp.utils.api_response import (
    success_response, error_response, paginated_response,
    single_item_response, validation_error_response,
    not_found_response, forbidden_response
)
from erp.utils.campus_utils import get_current_campus_from_context
import traceback


@frappe.whitelist(allow_guest=False, methods=['POST'])
def start_bulk_import():
    """
    Start a bulk import job for Excel file processing

    Expected parameters:
    - doctype: Target DocType (e.g., "CRM Student", "SIS Subject")
    - file_url: URL of uploaded Excel file
    - options: Optional import options (update_if_exists, dry_run, etc.)
    """
    try:
        # Get parameters from request
        data = _get_request_data()

        # Validate required parameters
        doctype = data.get("doctype")
        file_url = data.get("file_url")

        if not doctype:
            return validation_error_response(
                message="doctype parameter is required",
                errors={"doctype": ["Required field"]}
            )

        if not file_url:
            return validation_error_response(
                message="file_url parameter is required",
                errors={"file_url": ["Required field"]}
            )

        # Validate DocType exists
        if not frappe.db.exists("DocType", doctype):
            return validation_error_response(
                message=f"DocType '{doctype}' does not exist",
                errors={"doctype": ["Invalid DocType"]}
            )

        # Get current user's campus
        campus_id = get_current_campus_from_context()
        if not campus_id:
            return forbidden_response(
                message="Unable to determine user's campus. Please contact administrator.",
                code="NO_CAMPUS_ACCESS"
            )

        # Create bulk import job
        options = data.get("options", {})
        options_json = json.dumps(options) if options else None

        job = frappe.get_doc({
            "doctype": "SIS Bulk Import Job",
            "doctype_target": doctype,
            "file_url": file_url,
            "options_json": options_json,
            "status": "Queued",
            "campus_id": campus_id,
            "created_by": frappe.session.user
        })

        job.insert(ignore_permissions=True)
        frappe.db.commit()

        # Enqueue the bulk import worker
        frappe.enqueue(
            "erp.api.bulk_import.process_bulk_import",
            job_id=job.name,
            queue="long",  # Use long queue for bulk processing
            timeout=3600,  # 1 hour timeout
            now=False  # Allow queuing
        )

        frappe.logger().info(f"Bulk import job {job.name} created and queued for {doctype}")

        return single_item_response(
            data={"job_id": job.name},
            message="Bulk import job has been created and queued for processing"
        )

    except Exception as e:
        frappe.log_error(f"Error starting bulk import: {str(e)}")
        return error_response(
            message="Failed to start bulk import job",
            code="START_BULK_IMPORT_ERROR"
        )


@frappe.whitelist(allow_guest=False, methods=['GET'])
def get_bulk_import_status():
    """
    Get status of a bulk import job

    Expected parameters:
    - job_id: Job ID to check
    """
    try:
        # Get job_id from request
        job_id = frappe.form_dict.get("job_id") or frappe.local.form_dict.get("job_id")

        if not job_id:
            return validation_error_response(
                message="job_id parameter is required",
                errors={"job_id": ["Required field"]}
            )

        # Get job document
        try:
            job = frappe.get_doc("SIS Bulk Import Job", job_id)
        except frappe.DoesNotExistError:
            return not_found_response(
                message="Bulk import job not found",
                code="JOB_NOT_FOUND"
            )

        # Check if user has access to this job
        if job.created_by != frappe.session.user:
            return forbidden_response(
                message="You don't have permission to view this job",
                code="ACCESS_DENIED"
            )

        # Calculate progress percentage
        progress_percentage = job.get_progress_percentage()

        # Build response
        response_data = {
            "job_id": job.name,
            "status": job.status,
            "doctype_target": job.doctype_target,
            "total_rows": job.total_rows or 0,
            "processed_rows": job.processed_rows or 0,
            "success_count": job.success_count or 0,
            "error_count": job.error_count or 0,
            "progress_percentage": progress_percentage,
            "started_at": job.started_at,
            "finished_at": job.finished_at,
            "message": job.message,
            "error_file_url": job.error_file_url
        }

        return single_item_response(
            data=response_data,
            message="Bulk import status retrieved successfully"
        )

    except Exception as e:
        frappe.log_error(f"Error getting bulk import status: {str(e)}")
        return error_response(
            message="Failed to get bulk import status",
            code="GET_STATUS_ERROR"
        )


@frappe.whitelist(allow_guest=False, methods=['GET', 'POST'])
def download_template():
    """
    Download Excel template for a specific DocType

    Expected parameters:
    - doctype: Target DocType to generate template for
    """
    try:
        # Get doctype from request - try multiple sources
        doctype = (
            frappe.form_dict.get("doctype") or
            frappe.local.form_dict.get("doctype") or
            frappe.request.args.get("doctype") if hasattr(frappe.request, 'args') and frappe.request.args else None
        )

        # Log doctype resolution for debugging
        if not doctype:
            frappe.logger().info(f"Download template - doctype resolution failed. form_dict keys: {list(frappe.form_dict.keys())}")

        # TEMPORARY: For debugging, use hardcoded doctype if not found
        # Remove temporary fix - now requires proper doctype parameter

        if not doctype:
            return validation_error_response(
                message="doctype parameter is required",
                errors={"doctype": ["Required field"]}
            )

        # Validate DocType exists
        if not frappe.db.exists("DocType", doctype):
            return validation_error_response(
                message=f"DocType '{doctype}' does not exist",
                errors={"doctype": ["Invalid DocType"]}
            )

        # Generate template
        template_data = _generate_excel_template(doctype)

        if not template_data:
            return error_response(
                message="Failed to generate template for this DocType",
                code="TEMPLATE_GENERATION_ERROR"
            )

        # Try to generate Excel file if pandas is available
        excel_file_url = _generate_excel_file_from_template(template_data)

        if excel_file_url:
            # Return Excel file URL if generation successful
            return single_item_response(
                data={
                    "template_data": template_data,
                    "excel_file_url": excel_file_url,
                    "download_url": excel_file_url
                },
                message=f"Excel template generated successfully for {doctype}"
            )
        else:
            # Fallback to CSV data only
            return single_item_response(
                data=template_data,
                message=f"Template generated successfully for {doctype}"
            )

    except Exception as e:
        frappe.log_error(f"Error downloading template: {str(e)}")
        return error_response(
            message="Failed to download template",
            code="DOWNLOAD_TEMPLATE_ERROR"
        )


def _get_request_data():
    """Helper to get request data from various sources"""
    # Try JSON body first
    if frappe.request.data:
        try:
            return json.loads(frappe.request.data.decode('utf-8'))
        except (json.JSONDecodeError, UnicodeDecodeError):
            pass

    # Fallback to form_dict
    return frappe.local.form_dict or {}


def _generate_excel_file_from_template(template_data):
    """
    Generate Excel file from template data and upload to Frappe
    Returns file URL if successful, None otherwise
    """
    try:
        import pandas as pd
        from frappe.utils.file_manager import save_file
        import uuid

        # Create DataFrame for template with proper formatting
        headers = [field['fieldname'] for field in template_data['fields']]  # Use fieldname for data import
        display_headers = [field['label'] + (' *' if field.get('reqd') else '') for field in template_data['fields']]

        # Create sample data row
        sample_data = [template_data['sample_data'].get(field['fieldname'], '') for field in template_data['fields']]

        # Create instruction row
        instruction_data = ['← Fill your data starting from this row (delete this instruction row)' if i == 0 else '' for i in range(len(headers))]

        # Create multiple empty rows for data entry
        empty_data = [[''] * len(headers) for _ in range(5)]  # 5 empty rows for data entry

        # Combine all rows
        all_data = [
            ['TEMPLATE HEADER - DO NOT MODIFY'],
            headers,  # Actual field names for import
            ['DISPLAY HEADER (for reference)'],
            display_headers,  # Display headers with * for required
            ['SAMPLE DATA - DELETE AFTER COPYING'],
            sample_data,
            ['INSTRUCTION - DELETE THIS ROW'],
            instruction_data,
            ['YOUR DATA STARTS HERE'],
        ] + empty_data

        # Create DataFrame
        df = pd.DataFrame(all_data, columns=headers)

        # Create Excel file in memory with formatting
        from io import BytesIO
        output = BytesIO()

        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            # Get workbook and worksheet for formatting
            df.to_excel(writer, sheet_name='Import Data', index=False, header=False)
            workbook = writer.book
            worksheet = writer.sheets['Import Data']

            # Add formatting
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            # Define styles
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')
            instruction_fill = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
            required_font = Font(color='FF0000')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Format header rows
            for row_idx in range(1, 9):  # Format first 8 rows
                for col_idx in range(1, len(headers) + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)

                    if row_idx == 1:  # Template header
                        cell.font = Font(bold=True, color='FF0000', size=12)
                        cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                    elif row_idx == 2:  # Field names (technical)
                        cell.font = Font(bold=True, color='006600')
                        cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
                    elif row_idx == 4:  # Display headers
                        cell.font = Font(bold=True)
                        cell.fill = header_fill
                        cell.font = header_font
                    elif row_idx in [6, 8]:  # Sample data and instructions
                        cell.fill = instruction_fill
                        cell.font = Font(italic=True)

                    # Add border to all cells
                    cell.border = border
                    # Center align
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # Set column widths
            for col_num, column_title in enumerate(headers, 1):
                column_letter = chr(64 + col_num)  # A, B, C, ...
                if len(column_title) < 15:
                    worksheet.column_dimensions[column_letter].width = max(15, len(column_title) + 2)
                else:
                    worksheet.column_dimensions[column_letter].width = 20

            # Add data validation for select fields
            from openpyxl.worksheet.datavalidation import DataValidation

            for field_idx, field in enumerate(template_data['fields']):
                if field.get('fieldtype') == 'Select' and field.get('options'):
                    options = field['options'].split('\n')
                    if options:
                        dv = DataValidation(
                            type="list",
                            formula1=f'"{",".join(options)}"',
                            allow_blank=True
                        )
                        dv.error = 'Please select from the dropdown list'
                        dv.errorTitle = 'Invalid Selection'

                        # Apply to data entry rows (starting from row 10)
                        start_cell = worksheet.cell(row=10, column=field_idx + 1)
                        end_cell = worksheet.cell(row=15, column=field_idx + 1)  # 5 data rows
                        dv.add(f'{start_cell.coordinate}:{end_cell.coordinate}')
                        worksheet.add_data_validation(dv)

            # Create Instructions sheet
            instructions_data = []
            for field in template_data['fields']:
                instructions_data.append({
                    'Field Name': field['fieldname'],
                    'Display Label': field['label'],
                    'Required': 'Yes' if field.get('reqd') else 'No',
                    'Field Type': field.get('fieldtype', 'Data'),
                    'Options': field.get('options', ''),
                    'Description': f"{'Required field' if field.get('reqd') else 'Optional field'}"
                })

            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='Field Instructions', index=False)

            # Format instructions sheet
            instructions_sheet = writer.sheets['Field Instructions']
            for col_num, column_title in enumerate(instructions_df.columns, 1):
                column_letter = chr(64 + col_num)
                instructions_sheet.column_dimensions[column_letter].width = max(15, len(column_title) + 2)

            # Create Notes sheet with comprehensive instructions
            notes_data = [
                ['BULK IMPORT TEMPLATE'],
                [f'DocType: {template_data["doctype"]}'],
                [f'Generated: {frappe.utils.now()}'],
                [''],
                ['=== HOW TO USE THIS TEMPLATE ==='],
                ['1. DO NOT modify the first 8 rows (headers and instructions)'],
                ['2. Start entering your data from row 10 onwards'],
                ['3. Delete the instruction rows (rows 5, 7, 9) after reading'],
                ['4. For dropdown fields, use the validation dropdowns'],
                ['5. Date format: YYYY-MM-DD (e.g., 2024-12-25)'],
                ['6. Leave Campus ID empty to use current campus'],
                ['7. Required fields are marked with * in display headers'],
                ['8. Save the file after filling data'],
                ['9. Upload this file back to the system for import'],
                [''],
                ['=== REQUIRED FIELDS ==='],
            ] + [[f"- {field['label']}"] for field in template_data['fields'] if field.get('reqd')] + [
                [''],
                ['=== OPTIONAL FIELDS ==='],
            ] + [[f"- {field['label']}"] for field in template_data['fields'] if not field.get('reqd')] + [
                [''],
                ['=== FIELD TYPES ==='],
            ] + [[f"- {field['label']}: {field.get('fieldtype', 'Data')}"] for field in template_data['fields']] + [
                [''],
                ['=== SAMPLE DATA ==='],
            ] + [[f"- {field['fieldname']}: {template_data['sample_data'].get(field['fieldname'], '')}"] for field in template_data['fields']]

            notes_df = pd.DataFrame(notes_data, columns=['Instructions'])
            notes_df.to_excel(writer, sheet_name='How to Use', index=False, header=False)

            # Format notes sheet
            notes_sheet = writer.sheets['How to Use']
            notes_sheet.column_dimensions['A'].width = 60

            # Add color coding
            from openpyxl.styles import PatternFill
            for row_idx in range(1, len(notes_data) + 1):
                cell = notes_sheet.cell(row=row_idx, column=1)
                if 'REQUIRED' in str(cell.value):
                    cell.fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
                    cell.font = Font(bold=True, color='FF0000')
                elif 'HOW TO USE' in str(cell.value):
                    cell.fill = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
                    cell.font = Font(bold=True, color='006600')
                elif str(cell.value).startswith('==='):
                    cell.font = Font(bold=True)

        # Upload file to Frappe
        file_content = output.getvalue()
        filename = f"bulk_import_template_{template_data['doctype'].lower().replace(' ', '_')}_{uuid.uuid4().hex[:8]}.xlsx"

        file_doc = save_file(
            filename,
            file_content,
            "SIS Bulk Import Job",
            "",  # No parent doc
            is_private=0
        )

        return file_doc.file_url if file_doc else None

    except ImportError:
        frappe.logger().info("pandas/openpyxl not available for Excel generation")
        return None
    except Exception as e:
        frappe.logger().error(f"Error generating Excel file: {str(e)}")
        return None


def _generate_excel_template(doctype):
    """
    Generate Excel template based on DocType metadata

    Returns template structure with headers and sample data
    """
    try:
        # Get DocType metadata
        meta = frappe.get_meta(doctype)

        # Get fields for template (exclude system fields)
        template_fields = []
        sample_data = {}

        for field in meta.fields:
            # Skip system and non-user fields
            if field.fieldtype in ["Column Break", "Section Break", "Tab Break", "HTML", "Button"]:
                continue

            if field.hidden or field.read_only or field.fieldname in ["name", "owner", "creation", "modified", "modified_by", "docstatus"]:
                continue

            # Add field to template
            field_info = {
                "fieldname": field.fieldname,
                "label": field.label or field.fieldname,
                "fieldtype": field.fieldtype,
                "reqd": field.reqd or False,
                "options": field.options if field.fieldtype in ["Select", "Link"] else None
            }

            template_fields.append(field_info)

            # Add sample data
            if field.fieldtype == "Data":
                sample_data[field.fieldname] = f"Sample {field.label}"
            elif field.fieldtype == "Select" and field.options:
                options = field.options.split("\n")
                sample_data[field.fieldname] = options[0] if options else ""
            elif field.fieldtype == "Link":
                sample_data[field.fieldname] = f"Sample {field.options}"

        return {
            "doctype": doctype,
            "fields": template_fields,
            "sample_data": sample_data,
            "notes": [
                f"Template for {doctype} bulk import",
                "Required fields are marked with *",
                "Date format: YYYY-MM-DD",
                "Leave campus_id empty to use current user's campus"
            ]
        }

    except Exception as e:
        frappe.logger().error(f"Error generating template for {doctype}: {str(e)}")
        return None


@frappe.whitelist(allow_guest=False)
def process_bulk_import(job_id):
    """
    Worker function to process bulk import job
    This function runs in background queue
    """
    try:
        frappe.logger().info(f"Starting bulk import processing for job {job_id}")

        # Get job
        job = frappe.get_doc("SIS Bulk Import Job", job_id)
        job.status = "Running"
        job.started_at = now_datetime()
        job.save(ignore_permissions=True)
        frappe.db.commit()

        # Process the Excel file
        result = _process_excel_file(job)

        # Update job with results
        if result["success"]:
            job.mark_completed(
                message=result["message"],
                error_file_url=result.get("error_file_url")
            )
        else:
            job.mark_failed(
                message=result["message"],
                error_file_url=result.get("error_file_url")
            )

        frappe.logger().info(f"Bulk import job {job_id} completed with status: {job.status}")

    except Exception as e:
        error_msg = f"Bulk import processing failed: {str(e)}"
        frappe.logger().error(f"{error_msg}\n{traceback.format_exc()}")

        try:
            job = frappe.get_doc("SIS Bulk Import Job", job_id)
            job.mark_failed(message=error_msg)
        except Exception:
            frappe.logger().error(f"Failed to mark job {job_id} as failed")


def _process_excel_file(job):
    """
    Process Excel file for bulk import

    Returns:
        dict: {
            "success": bool,
            "message": str,
            "error_file_url": str (optional)
        }
    """
    try:
        import pandas as pd
        from frappe.utils.file_manager import save_file

        frappe.logger().info(f"Processing Excel file for job {job.name}")

        # Download file from URL
        file_path = job.file_url
        if not file_path.startswith("/"):
            file_path = frappe.get_site_path("public", file_path)

        # Read Excel file
        try:
            df = pd.read_excel(file_path)
        except Exception as e:
            return {
                "success": False,
                "message": f"Failed to read Excel file: {str(e)}"
            }

        # Validate that this is a template-generated file
        if len(df) < 10:
            return {
                "success": False,
                "message": "Excel file appears to be too short. Please use the downloaded template and fill data starting from row 10."
            }

        # Check if this looks like our template format
        first_row = df.iloc[0].fillna('').astype(str).tolist()
        expected_header = "TEMPLATE HEADER - DO NOT MODIFY"

        if len(first_row) == 0 or not any(expected_header in str(cell) for cell in first_row):
            frappe.logger().warning(f"Template validation failed. First row: {first_row[:3]}")
            return {
                "success": False,
                "message": "This doesn't appear to be a template-generated file. Please download the template first, then fill your data."
            }

        # Skip template header rows (first 9 rows contain template info, data starts from row 10)
        df = df.iloc[9:]  # Skip first 9 rows, keep from row 10 onwards

        # Remove completely empty rows
        df = df.dropna(how='all')

        total_rows = len(df)
        if total_rows == 0:
            return {
                "success": False,
                "message": "No data found in Excel file. Please fill data starting from row 10 in the template."
            }

        job.total_rows = total_rows
        job.save(ignore_permissions=True)
        frappe.db.commit()

        frappe.logger().info(f"Excel file has {total_rows} rows")

        # Process data in batches
        batch_size = 100
        success_count = 0
        error_count = 0
        errors = []

        options = job.get_options_dict()
        update_if_exists = options.get("update_if_exists", False)
        dry_run = options.get("dry_run", False)

        # Process data in batches, accounting for skipped rows
        for i in range(0, total_rows, batch_size):
            batch_df = df.iloc[i:i+batch_size]
            # Adjust row numbers for original Excel file (add back the 9 skipped rows + 1 for 1-indexing)
            original_start_index = i + 9 + 1  # +9 for skipped rows, +1 for Excel 1-indexing
            batch_result = _process_batch(job, batch_df, original_start_index, update_if_exists, dry_run)

            success_count += batch_result["success_count"]
            error_count += batch_result["error_count"]
            errors.extend(batch_result["errors"])

            # Update progress
            processed_rows = min(i + batch_size, total_rows)
            job.update_progress(
                processed_rows=processed_rows,
                success_count=success_count,
                error_count=error_count
            )

        # Generate error report if there are errors
        error_file_url = None
        if errors:
            error_file_url = _generate_error_report(job, errors)

        message = f"Import completed: {success_count} success, {error_count} errors"
        if dry_run:
            message = f"Dry run completed: {success_count} would be imported, {error_count} errors"

        return {
            "success": True,
            "message": message,
            "error_file_url": error_file_url
        }

    except Exception as e:
        error_msg = f"Error processing Excel file: {str(e)}"
        frappe.logger().error(f"{error_msg}\n{traceback.format_exc()}")
        return {
            "success": False,
            "message": error_msg
        }


def _process_batch(job, batch_df, start_index, update_if_exists, dry_run):
    """Process a batch of records"""
    success_count = 0
    error_count = 0
    errors = []

    for idx, row in batch_df.iterrows():
        row_num = start_index + idx + 2  # +2 because Excel is 1-indexed and has header

        try:
            # Convert row to dict and clean data
            row_data = {}
            for col in batch_df.columns:
                value = row[col]
                if pd.isna(value):
                    value = None
                row_data[col] = value

            # Process single record
            result = _process_single_record(job, row_data, row_num, update_if_exists, dry_run)

            if result["success"]:
                success_count += 1
            else:
                error_count += 1
                errors.append({
                    "row": row_num,
                    "data": row_data,
                    "error": result["error"]
                })

        except Exception as e:
            error_count += 1
            errors.append({
                "row": row_num,
                "data": str(row.to_dict())[:500],  # Limit data size
                "error": str(e)
            })

    return {
        "success_count": success_count,
        "error_count": error_count,
        "errors": errors
    }


def _process_single_record(job, row_data, row_num, update_if_exists, dry_run):
    """Process a single record"""
    try:
        doctype = job.doctype_target

        # Handle campus_id - use from file or default to user's campus
        campus_id = row_data.get("campus_id")
        if not campus_id:
            campus_id = job.campus_id

        # Build document data
        doc_data = {
            "doctype": doctype,
            "campus_id": campus_id
        }

        # Map Excel columns to DocType fields
        meta = frappe.get_meta(doctype)
        for field in meta.fields:
            if field.fieldname in row_data and field.fieldname not in ["name", "owner", "creation", "modified"]:
                doc_data[field.fieldname] = row_data[field.fieldname]

        # Check if record exists for update
        existing_doc = None
        if update_if_exists:
            # Try to find existing record based on unique fields
            existing_doc = _find_existing_record(doctype, doc_data)

        if dry_run:
            # Just validate without saving
            return {"success": True}

        if existing_doc:
            # Update existing record
            for key, value in doc_data.items():
                if key != "doctype" and value is not None:
                    existing_doc.set(key, value)
            existing_doc.save(ignore_permissions=True)
        else:
            # Create new record
            doc = frappe.get_doc(doc_data)
            doc.insert(ignore_permissions=True)

        frappe.db.commit()
        return {"success": True}

    except Exception as e:
        return {
            "success": False,
            "error": str(e)
        }


def _find_existing_record(doctype, doc_data):
    """Find existing record for update"""
    # This is a simple implementation - in production, you'd want more sophisticated
    # unique key detection based on DocType configuration

    if doctype == "CRM Student" and doc_data.get("student_code"):
        return frappe.db.exists(doctype, {"student_code": doc_data["student_code"]})

    if doctype == "SIS Subject" and doc_data.get("title"):
        return frappe.db.exists(doctype, {"title": doc_data["title"]})

    return None


def _generate_error_report(job, errors):
    """Generate Excel error report"""
    try:
        import pandas as pd
        from frappe.utils.file_manager import save_file

        # Create error DataFrame
        error_data = []
        for error in errors:
            row_data = {"__row_number": error["row"], "__error": error["error"]}
            if isinstance(error.get("data"), dict):
                row_data.update(error["data"])
            else:
                row_data["__original_data"] = str(error.get("data", ""))[:1000]
            error_data.append(row_data)

        df_errors = pd.DataFrame(error_data)

        # Save to temporary file
        temp_file_path = f"/tmp/bulk_import_errors_{job.name}.xlsx"
        df_errors.to_excel(temp_file_path, index=False)

        # Upload file to Frappe
        with open(temp_file_path, "rb") as f:
            file_doc = save_file(
                f"bulk_import_errors_{job.name}.xlsx",
                f.read(),
                "SIS Bulk Import Job",
                job.name,
                is_private=1
            )

        # Clean up temp file
        os.unlink(temp_file_path)

        return file_doc.file_url

    except Exception as e:
        frappe.logger().error(f"Failed to generate error report: {str(e)}")
        return None
