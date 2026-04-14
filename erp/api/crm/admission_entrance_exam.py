"""
API Khảo sát đầu vào — CRUD kỳ khảo sát, học sinh, điểm môn.
"""

import frappe

from erp.utils.api_response import (
    success_response,
    error_response,
    single_item_response,
    list_response,
    validation_error_response,
    not_found_response,
)
from erp.api.crm.utils import check_crm_permission, get_request_data


def _enrich_modified_by_name(items, modified_by_field="modified_by"):
    """Bổ sung modified_by_name (full_name từ User) cho mỗi item"""
    for item in items:
        user_id = item.get(modified_by_field)
        if user_id:
            full_name = frappe.db.get_value("User", user_id, "full_name")
            item["modified_by_name"] = full_name or user_id
        else:
            item["modified_by_name"] = None
    return items


VALID_ENTRANCE_STATUSES = frozenset(
    {"new", "schedule_notified", "not_attending", "exam_taken", "completed"}
)
VALID_EXAM_RESULTS = frozenset({"", "pass", "conditional_pass", "retake", "fail"})

# Nhãn hiển thị Excel nhập liệu — đồng bộ FE (entranceExamLabels)
_ENTRANCE_STATUS_LABEL_VN = {
    "new": "Đã đăng ký",
    "schedule_notified": "Thông báo lịch thi",
    "not_attending": "Không thi",
    "exam_taken": "Đã thi",
    "completed": "Hoàn thành",
}
_ENTRANCE_RESULT_LABEL_VN = {
    "": "—",
    "pass": "Đạt",
    "conditional_pass": "Đạt có điều kiện",
    "retake": "Thi lại",
    "fail": "Không đạt",
}


def _export_status_cell(status):
    return _ENTRANCE_STATUS_LABEL_VN.get(status or "new", status or "Đã đăng ký")


def _export_exam_result_cell(exam_result):
    er = exam_result or ""
    return _ENTRANCE_RESULT_LABEL_VN.get(er, er or "—")


def _export_ksdv_cell(ksdv_fee_paid):
    return "Đã đóng" if ksdv_fee_paid else "Chưa đóng"


def _get_lead_primary_phone(parent_lead_id):
    """SĐT phụ huynh: ưu tiên CRM Lead Phone có is_primary, không thì dòng đầu."""
    if not parent_lead_id:
        return None
    rows = frappe.get_all(
        "CRM Lead Phone",
        filters={"parent": parent_lead_id, "parenttype": "CRM Lead"},
        fields=["phone_number", "is_primary"],
        order_by="is_primary desc, idx asc",
        limit_page_length=1,
    )
    if rows:
        return rows[0].phone_number
    return None


def _enrich_lead_flat_fields(data, crm_lead_id):
    """Gắn thông tin từ CRM Lead: mã HS, phụ huynh chính, SĐT, PIC (hiển thị full_name)."""
    lead = frappe.db.get_value(
        "CRM Lead",
        crm_lead_id,
        [
            "crm_code",
            "student_name",
            "student_dob",
            "target_grade",
            "student_code",
            "guardian_name",
            "pic",
        ],
        as_dict=True,
    )
    if lead:
        data["crm_code"] = lead.get("crm_code") or crm_lead_id
        data["student_name"] = lead.get("student_name") or "-"
        data["student_dob"] = lead.get("student_dob")
        data["target_grade"] = lead.get("target_grade") or ""
        data["student_code"] = lead.get("student_code") or ""
        data["guardian_name"] = lead.get("guardian_name") or ""
        pic = lead.get("pic")
        data["pic_full_name"] = (
            frappe.db.get_value("User", pic, "full_name") if pic else None
        ) or (pic or "")
        data["primary_phone"] = _get_lead_primary_phone(crm_lead_id) or ""
    else:
        data["crm_code"] = crm_lead_id
        data["student_name"] = "-"
        data["student_dob"] = None
        data["target_grade"] = ""
        data["student_code"] = ""
        data["guardian_name"] = ""
        data["pic_full_name"] = ""
        data["primary_phone"] = ""
    return data


def _resolve_crm_lead_name(ref):
    """
    Chuẩn hóa tham chiếu CRM Lead: docname (name) hoặc mã crm_code.
    API tab CRM và thêm HS có thể gửi một trong hai — bản ghi Link luôn lưu name.
    """
    if not ref:
        return None
    ref = (ref or "").strip()
    if not ref:
        return None
    if frappe.db.exists("CRM Lead", ref):
        return ref
    lead_name = frappe.db.get_value("CRM Lead", {"crm_code": ref}, "name")
    return lead_name


# ========== KỲ KHẢO SÁT ==========


@frappe.whitelist()
def get_entrance_exams():
    """Danh sách kỳ khảo sát; filter school_year_id nếu có"""
    check_crm_permission()
    school_year_id = frappe.request.args.get("school_year_id")
    filters = {}
    if school_year_id and school_year_id != "all":
        filters["school_year_id"] = school_year_id
    items = frappe.get_all(
        "CRM Admission Entrance Exam",
        filters=filters,
        fields=[
            "name",
            "exam_name",
            "school_year_id",
            "exam_date",
            "exam_time",
            "student_count",
            "is_active",
            "modified",
            "modified_by",
        ],
        order_by="modified desc",
    )
    _enrich_modified_by_name(items)
    return list_response(items)


@frappe.whitelist()
def get_entrance_exam(exam_id=None):
    """Chi tiết một kỳ khảo sát"""
    check_crm_permission()
    exam_id = exam_id or frappe.request.args.get("exam_id")
    if not exam_id:
        return validation_error_response("Thiếu exam_id", {"exam_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Entrance Exam", exam_id):
        return not_found_response("Không tìm thấy kỳ khảo sát")
    doc = frappe.get_doc("CRM Admission Entrance Exam", exam_id)
    data = doc.as_dict()
    if doc.modified_by:
        data["modified_by_name"] = (
            frappe.db.get_value("User", doc.modified_by, "full_name") or doc.modified_by
        )
    return single_item_response(data, "Thành công")


@frappe.whitelist(methods=["POST"])
def create_entrance_exam():
    """Tạo kỳ khảo sát mới"""
    check_crm_permission()
    data = get_request_data()
    if not data.get("exam_name"):
        return validation_error_response("Thiếu exam_name", {"exam_name": ["Bắt buộc"]})
    try:
        doc = frappe.new_doc("CRM Admission Entrance Exam")
        doc.exam_name = (data.get("exam_name") or "").strip()
        doc.exam_date = data.get("exam_date") or None
        doc.exam_time = (data.get("exam_time") or "").strip() or None
        doc.student_count = data.get("student_count", 0) or 0
        doc.is_active = 1 if data.get("is_active", True) else 0
        doc.school_year_id = data.get("school_year_id") or None
        doc.insert(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Tạo kỳ khảo sát thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi tạo kỳ khảo sát: {str(e)}")


@frappe.whitelist(methods=["POST"])
def update_entrance_exam():
    """Cập nhật kỳ khảo sát"""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Entrance Exam", name):
        return not_found_response("Không tìm thấy kỳ khảo sát")
    try:
        doc = frappe.get_doc("CRM Admission Entrance Exam", name)
        if "exam_name" in data:
            doc.exam_name = data["exam_name"].strip()
        if "exam_date" in data:
            doc.exam_date = data["exam_date"] or None
        if "exam_time" in data:
            doc.exam_time = (data.get("exam_time") or "").strip() or None
        if "student_count" in data:
            doc.student_count = data["student_count"] or 0
        if "is_active" in data:
            doc.is_active = 1 if data["is_active"] else 0
        if "school_year_id" in data:
            doc.school_year_id = data["school_year_id"] or None
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Cập nhật thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi cập nhật: {str(e)}")


@frappe.whitelist(methods=["POST"])
def toggle_entrance_exam_active():
    """Bật/tắt kỳ khảo sát"""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    is_active = data.get("is_active", True)
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Entrance Exam", name):
        return not_found_response("Không tìm thấy kỳ khảo sát")
    try:
        doc = frappe.get_doc("CRM Admission Entrance Exam", name)
        doc.is_active = 1 if is_active else 0
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Cập nhật trạng thái thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi: {str(e)}")


@frappe.whitelist(methods=["POST"])
def delete_entrance_exam():
    """Xóa kỳ khảo sát và toàn bộ học sinh thuộc kỳ"""
    check_crm_permission()
    name = get_request_data().get("name")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Entrance Exam", name):
        return not_found_response("Không tìm thấy kỳ khảo sát")
    try:
        stu_names = frappe.get_all(
            "CRM Admission Entrance Exam Student",
            filters={"entrance_exam_id": name},
            pluck="name",
        )
        for stu in stu_names:
            frappe.delete_doc("CRM Admission Entrance Exam Student", stu, ignore_permissions=True)
        frappe.delete_doc("CRM Admission Entrance Exam", name, ignore_permissions=True)
        frappe.db.commit()
        return success_response(message="Xóa kỳ khảo sát thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi xóa: {str(e)}")


def _get_entrance_exam_student_summary(exam_id):
    """Tổng số học sinh + phân theo trạng thái (tùy dùng báo cáo)"""
    filters = {"entrance_exam_id": exam_id}
    total = frappe.db.count("CRM Admission Entrance Exam Student", filters=filters)
    out = {"total": total}
    for s in VALID_ENTRANCE_STATUSES:
        out[f"status_{s}"] = frappe.db.count(
            "CRM Admission Entrance Exam Student",
            filters={**filters, "status": s},
        )
    return out


def _enrich_entrance_student_row(item):
    """Gắn thông tin lead (mã CRM, HS, PH, SĐT, PIC) + modified_by_name."""
    _enrich_lead_flat_fields(item, item["crm_lead_id"])
    if item.get("modified_by"):
        item["modified_by_name"] = (
            frappe.db.get_value("User", item["modified_by"], "full_name") or item["modified_by"]
        )
    else:
        item["modified_by_name"] = None
    return item


def _serialize_entrance_exam_student_detail(record_id):
    """Chi tiết một bản ghi học sinh trong kỳ (kèm điểm môn) — dùng API detail và tab CRM."""
    if not record_id or not frappe.db.exists("CRM Admission Entrance Exam Student", record_id):
        return None
    doc = frappe.get_doc("CRM Admission Entrance Exam Student", record_id, ignore_permissions=True)
    exam = frappe.db.get_value(
        "CRM Admission Entrance Exam",
        doc.entrance_exam_id,
        ["exam_name", "exam_date", "exam_time", "school_year_id"],
        as_dict=True,
    )
    data = doc.as_dict()
    if doc.modified_by:
        data["modified_by_name"] = (
            frappe.db.get_value("User", doc.modified_by, "full_name") or doc.modified_by
        )
    else:
        data["modified_by_name"] = None
    _enrich_lead_flat_fields(data, doc.crm_lead_id)
    data["exam_name"] = (exam or {}).get("exam_name")
    data["exam_date"] = (exam or {}).get("exam_date")
    data["exam_time"] = (exam or {}).get("exam_time")
    data["school_year_id"] = (exam or {}).get("school_year_id")
    data["scores"] = [
        {"subject": row.subject, "score": row.score} for row in (doc.scores or [])
    ]
    return data


@frappe.whitelist()
def get_entrance_exam_students():
    """Danh sách học sinh trong một kỳ khảo sát"""
    check_crm_permission()
    exam_id = frappe.request.args.get("exam_id")
    if not exam_id:
        return validation_error_response("Thiếu exam_id", {"exam_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Entrance Exam", exam_id):
        return not_found_response("Không tìm thấy kỳ khảo sát")

    search = frappe.request.args.get("search")
    status_filter = frappe.request.args.get("status")

    filters = {"entrance_exam_id": exam_id}
    if status_filter and status_filter in VALID_ENTRANCE_STATUSES:
        filters["status"] = status_filter

    or_filters = None
    if search and search.strip():
        lead_rows = frappe.db.sql(
            """
            SELECT name FROM `tabCRM Lead`
            WHERE name LIKE %(s)s OR crm_code LIKE %(s)s OR student_name LIKE %(s)s
            """,
            {"s": f"%{search.strip()}%"},
            as_dict=True,
        )
        lead_names = [r["name"] for r in lead_rows]
        if not lead_names:
            return list_response(
                [],
                "Thành công",
                meta={"summary": _get_entrance_exam_student_summary(exam_id)},
            )
        or_filters = {"crm_lead_id": ["in", lead_names]}

    items = frappe.get_all(
        "CRM Admission Entrance Exam Student",
        filters=filters,
        or_filters=or_filters,
        fields=[
            "name",
            "entrance_exam_id",
            "crm_lead_id",
            "status",
            "exam_result",
            "ksdv_fee_paid",
            "result_link",
            "modified",
            "modified_by",
        ],
        order_by="modified desc",
    )
    for item in items:
        _enrich_entrance_student_row(item)

    return list_response(
        items,
        "Thành công",
        meta={"summary": _get_entrance_exam_student_summary(exam_id)},
    )


@frappe.whitelist(methods=["POST"])
def add_entrance_exam_student():
    """Thêm học sinh (CRM Lead) vào kỳ khảo sát"""
    check_crm_permission()
    data = get_request_data()
    exam_id = data.get("exam_id") or data.get("entrance_exam_id")
    crm_lead_raw = data.get("crm_lead_id")
    if not exam_id or not crm_lead_raw:
        return validation_error_response(
            "Thiếu exam_id hoặc crm_lead_id",
            {"exam_id": ["Bắt buộc"], "crm_lead_id": ["Bắt buộc"]},
        )
    if not frappe.db.exists("CRM Admission Entrance Exam", exam_id):
        return not_found_response("Không tìm thấy kỳ khảo sát")
    crm_lead_id = _resolve_crm_lead_name(crm_lead_raw)
    if not crm_lead_id:
        return not_found_response("Không tìm thấy CRM Lead")

    existing = frappe.db.exists(
        "CRM Admission Entrance Exam Student",
        {"entrance_exam_id": exam_id, "crm_lead_id": crm_lead_id},
    )
    if existing:
        return validation_error_response("Học sinh đã có trong kỳ khảo sát", {"crm_lead_id": ["Đã tồn tại"]})

    status = data.get("status") or "new"
    if status not in VALID_ENTRANCE_STATUSES:
        status = "new"

    try:
        doc = frappe.new_doc("CRM Admission Entrance Exam Student")
        doc.entrance_exam_id = exam_id
        doc.crm_lead_id = crm_lead_id
        doc.status = status
        er = data.get("exam_result")
        if er in VALID_EXAM_RESULTS:
            doc.exam_result = er or None
        doc.insert(ignore_permissions=True)
        frappe.db.commit()
        row = doc.as_dict()
        _enrich_entrance_student_row(row)
        return single_item_response(row, "Thêm học sinh thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi thêm học sinh: {str(e)}")


@frappe.whitelist(methods=["POST"])
def update_entrance_exam_student_status():
    """Cập nhật trạng thái học sinh trong kỳ"""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    new_status = data.get("status")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if new_status not in VALID_ENTRANCE_STATUSES:
        return validation_error_response("Trạng thái không hợp lệ", {"status": ["Không hợp lệ"]})
    if not frappe.db.exists("CRM Admission Entrance Exam Student", name):
        return not_found_response("Không tìm thấy bản ghi")
    try:
        doc = frappe.get_doc("CRM Admission Entrance Exam Student", name)
        doc.status = new_status
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        row = doc.as_dict()
        _enrich_entrance_student_row(row)
        return single_item_response(row, "Cập nhật trạng thái thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi: {str(e)}")


@frappe.whitelist(methods=["POST"])
def update_entrance_exam_student_result():
    """Cập nhật kết quả thi"""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    exam_result = data.get("exam_result")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if exam_result not in VALID_EXAM_RESULTS:
        return validation_error_response("Kết quả không hợp lệ", {"exam_result": ["Không hợp lệ"]})
    if not frappe.db.exists("CRM Admission Entrance Exam Student", name):
        return not_found_response("Không tìm thấy bản ghi")
    try:
        doc = frappe.get_doc("CRM Admission Entrance Exam Student", name)
        doc.exam_result = exam_result or None
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        row = doc.as_dict()
        _enrich_entrance_student_row(row)
        return single_item_response(row, "Cập nhật kết quả thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi: {str(e)}")


@frappe.whitelist(methods=["POST"])
def update_entrance_exam_student_meta():
    """
    Cập nhật trạng thái, kết quả, link kết quả — không chỉnh bảng điểm môn.
    Body: name (bắt buộc), status, exam_result, result_link (tùy chọn, gửi đủ 3 từ client).
    """
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Entrance Exam Student", name):
        return not_found_response("Không tìm thấy bản ghi")
    try:
        doc = frappe.get_doc("CRM Admission Entrance Exam Student", name)
        if "status" in data:
            new_status = data.get("status")
            if new_status not in VALID_ENTRANCE_STATUSES:
                return validation_error_response("Trạng thái không hợp lệ", {"status": ["Không hợp lệ"]})
            doc.status = new_status
        if "exam_result" in data:
            exam_result = data.get("exam_result")
            if exam_result not in VALID_EXAM_RESULTS:
                return validation_error_response("Kết quả không hợp lệ", {"exam_result": ["Không hợp lệ"]})
            doc.exam_result = exam_result or None
        if "result_link" in data:
            doc.result_link = (data.get("result_link") or "").strip()
        if "ksdv_fee_paid" in data:
            # Phí KSĐV đã đóng (checkbox)
            doc.ksdv_fee_paid = 1 if data.get("ksdv_fee_paid") else 0
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        row = doc.as_dict()
        _enrich_entrance_student_row(row)
        return single_item_response(row, "Cập nhật thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi: {str(e)}")


@frappe.whitelist(methods=["POST"])
def delete_entrance_exam_student():
    """Xóa học sinh khỏi kỳ khảo sát"""
    check_crm_permission()
    name = get_request_data().get("name")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Entrance Exam Student", name):
        return not_found_response("Không tìm thấy bản ghi")
    try:
        frappe.delete_doc("CRM Admission Entrance Exam Student", name, ignore_permissions=True)
        frappe.db.commit()
        return success_response(message="Xóa học sinh thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi xóa: {str(e)}")


@frappe.whitelist()
def get_entrance_exam_student_detail():
    """Chi tiết một học sinh trong kỳ (kèm điểm môn) — dùng trang chi tiết HS"""
    check_crm_permission()
    record_id = frappe.request.args.get("record_id") or frappe.request.args.get("name")
    if not record_id:
        return validation_error_response("Thiếu record_id", {"record_id": ["Bắt buộc"]})
    data = _serialize_entrance_exam_student_detail(record_id)
    if data is None:
        return not_found_response("Không tìm thấy bản ghi")
    return single_item_response(data, "Thành công")


@frappe.whitelist()
def get_entrance_exam_student_details_for_lead():
    """Danh sách chi tiết khảo sát đầu vào của một CRM Lead (tab Thông tin chung)."""
    check_crm_permission()
    crm_lead_raw = frappe.request.args.get("crm_lead_id")
    if not crm_lead_raw:
        return validation_error_response("Thiếu crm_lead_id", {"crm_lead_id": ["Bắt buộc"]})
    crm_lead_id = _resolve_crm_lead_name(crm_lead_raw)
    if not crm_lead_id:
        return not_found_response("Không tìm thấy CRM Lead")

    # get_all mặc định áp quyền đọc DocType; insert HS dùng ignore_permissions — có thể lệch → cần ignore ở đây
    row_names = frappe.get_all(
        "CRM Admission Entrance Exam Student",
        filters={"crm_lead_id": crm_lead_id},
        pluck="name",
        order_by="modified desc",
        ignore_permissions=True,
    )
    items = []
    for name in row_names:
        serialized = _serialize_entrance_exam_student_detail(name)
        if serialized:
            items.append(serialized)
    return list_response(items, "Thành công")


@frappe.whitelist(methods=["POST"])
def update_entrance_exam_scores():
    """Cập nhật bảng điểm theo môn (thay thế toàn bộ dòng con)"""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    scores = data.get("scores")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if scores is None:
        return validation_error_response("Thiếu scores", {"scores": ["Bắt buộc"]})
    if not isinstance(scores, list):
        return validation_error_response("scores phải là mảng", {"scores": ["Sai định dạng"]})
    if not frappe.db.exists("CRM Admission Entrance Exam Student", name):
        return not_found_response("Không tìm thấy bản ghi")
    try:
        doc = frappe.get_doc("CRM Admission Entrance Exam Student", name)
        if "result_link" in data:
            doc.result_link = (data.get("result_link") or "").strip()
        doc.scores = []
        for row in scores:
            subj = (row.get("subject") or "").strip()
            if not subj:
                continue
            doc.append(
                "scores",
                {
                    "subject": subj,
                    "score": row.get("score"),
                },
            )
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        out = doc.as_dict()
        out["scores"] = [{"subject": r.subject, "score": r.score} for r in doc.scores]
        _enrich_entrance_student_row(out)
        return single_item_response(out, "Lưu điểm thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi lưu điểm: {str(e)}")


def _entrance_exam_excel_header_and_data_start(rows):
    """File từ FE: 2 dòng tiêu đề (tiếng Việt + mã cột crm_code, status, …)."""
    if not rows:
        return None, 1
    r0 = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
    r1 = (
        [str(c).strip().lower() if c is not None else "" for c in rows[1]]
        if len(rows) > 1
        else []
    )

    def key_hits(row):
        keys = (
            "name",
            "crm_lead_id",
            "crm_code",
            "student_name",
            "status",
            "exam_result",
            "result_link",
            "ksdv_fee_paid",
        )
        return sum(1 for k in keys if k in row)

    if r1 and key_hits(r1) >= key_hits(r0) and key_hits(r1) >= 1:
        return r1, 2
    if key_hits(r0) >= 1:
        return r0, 1
    if r1:
        return r1, 2
    return r0, 1


def _norm_entrance_status_import(val):
    """Chuẩn hoá trạng thái kỳ (mã API hoặc nhãn tiếng Việt)."""
    if val is None or str(val).strip() == "":
        return None
    s = str(val).strip().lower()
    vn = {
        "đã đăng ký": "new",
        "thông báo lịch thi": "schedule_notified",
        "không thi": "not_attending",
        "đã thi": "exam_taken",
        "hoàn thành": "completed",
    }
    if s in vn:
        return vn[s]
    if s in VALID_ENTRANCE_STATUSES:
        return s
    return None


def _norm_entrance_exam_result_import(val):
    """Chuẩn hoá kết quả thi (mã API hoặc nhãn tiếng Việt như FE)."""
    if val is None:
        return ""
    raw = str(val).strip()
    if raw in ("", "-", "—", "none"):
        return ""
    s = raw.lower()
    vn = {
        "đạt": "pass",
        "đạt có điều kiện": "conditional_pass",
        "thi lại": "retake",
        "không đạt": "fail",
        "chưa có": "",
    }
    if s in vn:
        return vn[s]
    if s in VALID_EXAM_RESULTS:
        return s
    return None


def _norm_ksdv_fee_import(val):
    """Đã đóng / Chưa đóng (như FE) hoặc 0/1."""
    if val is None or str(val).strip() == "":
        return None
    s = str(val).strip().lower()
    if s in ("1", "true", "yes", "có", "y", "x", "đã đóng", "da dong"):
        return 1
    if s in ("0", "false", "no", "không", "chưa đóng", "chua dong"):
        return 0
    try:
        return 1 if int(float(s)) != 0 else 0
    except Exception:
        return None


@frappe.whitelist()
def export_entrance_exam_students_template():
    """
    Xuất template nhập liệu: Mã CRM, họ tên; trạng thái / kết quả / phí KSĐV hiển thị nhãn tiếng Việt (đồng bộ FE).
    """
    check_crm_permission()
    exam_id = frappe.request.args.get("exam_id")
    if not exam_id:
        return validation_error_response("Thiếu exam_id", {"exam_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Entrance Exam", exam_id):
        return not_found_response("Không tìm thấy kỳ khảo sát")

    items = frappe.get_all(
        "CRM Admission Entrance Exam Student",
        filters={"entrance_exam_id": exam_id},
        fields=[
            "name",
            "crm_lead_id",
            "status",
            "exam_result",
            "result_link",
            "ksdv_fee_paid",
        ],
        order_by="modified desc",
    )
    for item in items:
        lead = frappe.db.get_value(
            "CRM Lead",
            item["crm_lead_id"],
            ["crm_code", "student_name"],
            as_dict=True,
        )
        if lead:
            item["crm_code"] = lead.get("crm_code") or item["crm_lead_id"]
            item["student_name"] = lead.get("student_name") or ""
        else:
            item["crm_code"] = item["crm_lead_id"]
            item["student_name"] = ""

    headers = [
        "crm_code",
        "student_name",
        "status",
        "exam_result",
        "result_link",
        "ksdv_fee_paid",
    ]
    header_labels = [
        "Mã CRM",
        "Họ tên học sinh",
        "Trạng thái",
        "Kết quả thi",
        "Link kết quả",
        "Phí KSĐV",
    ]

    rows = []
    for r in items:
        st = r.get("status") or "new"
        er = r.get("exam_result") or ""
        rows.append(
            {
                "crm_code": r.get("crm_code", ""),
                "student_name": r.get("student_name", ""),
                "status": _export_status_cell(st),
                "exam_result": _export_exam_result_cell(er),
                "result_link": r.get("result_link") or "",
                "ksdv_fee_paid": _export_ksdv_cell(r.get("ksdv_fee_paid")),
            }
        )

    return success_response(
        message="OK",
        data={"headers": headers, "header_labels": header_labels, "rows": rows},
    )


@frappe.whitelist(methods=["POST"])
def import_entrance_exam_students_meta():
    """
    Nhập liệu hàng loạt: khớp dòng theo Mã CRM; ô Trạng thái / Kết quả / Phí KSĐV dùng nhãn tiếng Việt (hoặc mã API).
    """
    check_crm_permission()
    import io

    import openpyxl

    exam_id = frappe.form_dict.get("exam_id") or frappe.form_dict.get("entrance_exam_id")
    if not exam_id:
        return validation_error_response("Thiếu exam_id", {"exam_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Entrance Exam", exam_id):
        return not_found_response("Không tìm thấy kỳ khảo sát")

    file = frappe.request.files.get("file")
    if not file:
        return validation_error_response("Thiếu file", {"file": ["Bắt buộc"]})

    try:
        content = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()
    except Exception as e:
        return error_response(f"Lỗi đọc file Excel: {str(e)}")

    if not rows or len(rows) < 2:
        return error_response("File Excel trống hoặc không đủ dòng tiêu đề")

    headers, data_start_idx = _entrance_exam_excel_header_and_data_start(rows)
    if not headers:
        return error_response("Không đọc được dòng tiêu đề")

    def col(*names):
        for n in names:
            for i, h in enumerate(headers):
                if h == n:
                    return i
        return None

    idx_name = col("name")
    idx_lead = col("crm_lead_id", "crm_lead")
    idx_code = col("crm_code")
    idx_status = col("status")
    idx_result = col("exam_result")
    idx_link = col("result_link")
    idx_ksdv = col("ksdv_fee_paid")

    if idx_code is None:
        return error_response(
            "Không tìm thấy cột crm_code (Mã CRM). Tải lại template từ nút Nhập liệu."
        )

    success_count = 0
    error_count = 0
    errors = []

    for row_idx, row in enumerate(rows[data_start_idx:], start=data_start_idx + 1):
        if not row:
            continue
        name_val = row[idx_name] if idx_name is not None and idx_name < len(row) else None
        lead_val = row[idx_lead] if idx_lead is not None and idx_lead < len(row) else None
        code_val = row[idx_code] if idx_code is not None and idx_code < len(row) else None
        if not code_val and not name_val and not lead_val:
            continue
        if (
            not str(code_val or "").strip()
            and not str(name_val or "").strip()
            and not str(lead_val or "").strip()
        ):
            continue

        rec_name = None
        # Ưu tiên Mã CRM (template mới)
        if code_val and str(code_val).strip():
            lid = frappe.db.get_value("CRM Lead", {"crm_code": str(code_val).strip()}, "name")
            if lid:
                rec_name = frappe.db.get_value(
                    "CRM Admission Entrance Exam Student",
                    {"entrance_exam_id": exam_id, "crm_lead_id": lid},
                    "name",
                )

        if not rec_name and name_val and str(name_val).strip():
            cand = str(name_val).strip()
            if frappe.db.exists("CRM Admission Entrance Exam Student", cand):
                eid = frappe.db.get_value(
                    "CRM Admission Entrance Exam Student", cand, "entrance_exam_id"
                )
                if eid == exam_id:
                    rec_name = cand

        if not rec_name and lead_val:
            raw = str(lead_val).strip()
            lid = frappe.db.get_value("CRM Lead", raw, "name")
            if not lid:
                lid = frappe.db.get_value("CRM Lead", {"crm_code": raw}, "name")
            if lid:
                rec_name = frappe.db.get_value(
                    "CRM Admission Entrance Exam Student",
                    {"entrance_exam_id": exam_id, "crm_lead_id": lid},
                    "name",
                )

        if not rec_name:
            error_count += 1
            errors.append(
                f"Dòng {row_idx}: Không tìm thấy học sinh trong kỳ (kiểm tra Mã CRM)"
            )
            continue

        try:
            doc = frappe.get_doc("CRM Admission Entrance Exam Student", rec_name)
            if idx_status is not None and idx_status < len(row):
                cell_st = row[idx_status]
                if cell_st is not None and str(cell_st).strip() != "":
                    st = _norm_entrance_status_import(cell_st)
                    if st is None:
                        error_count += 1
                        errors.append(f"Dòng {row_idx}: Trạng thái không hợp lệ")
                        continue
                    doc.status = st
            if idx_result is not None and idx_result < len(row):
                cell_er = row[idx_result]
                if cell_er is None or str(cell_er).strip() == "":
                    doc.exam_result = None
                else:
                    er = _norm_entrance_exam_result_import(cell_er)
                    if er is None:
                        error_count += 1
                        errors.append(f"Dòng {row_idx}: Kết quả không hợp lệ")
                        continue
                    doc.exam_result = er or None
            if idx_link is not None and idx_link < len(row):
                v = row[idx_link]
                doc.result_link = (str(v).strip() if v is not None else "") or ""
            if idx_ksdv is not None and idx_ksdv < len(row):
                kv = _norm_ksdv_fee_import(row[idx_ksdv])
                if kv is not None:
                    doc.ksdv_fee_paid = kv
            doc.save(ignore_permissions=True)
            success_count += 1
        except Exception as ex:
            error_count += 1
            errors.append(f"Dòng {row_idx}: Lỗi cập nhật — {str(ex)[:80]}")

    frappe.db.commit()
    return success_response(
        message=f"Nhập liệu: {success_count} thành công, {error_count} lỗi",
        data={"success_count": success_count, "error_count": error_count, "errors": errors[:50]},
    )


@frappe.whitelist(methods=["POST"])
def add_entrance_exam_students_excel():
    """Import Excel: cột CRM Lead / crm_code — thêm học sinh vào kỳ"""
    check_crm_permission()
    import io

    import openpyxl

    exam_id = frappe.form_dict.get("exam_id") or frappe.form_dict.get("entrance_exam_id")
    if not exam_id:
        return validation_error_response("Thiếu exam_id", {"exam_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Entrance Exam", exam_id):
        return not_found_response("Không tìm thấy kỳ khảo sát")

    file = frappe.request.files.get("file")
    if not file:
        return validation_error_response("Thiếu file", {"file": ["Bắt buộc"]})

    try:
        content = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()
    except Exception as e:
        return error_response(f"Lỗi đọc file Excel: {str(e)}")

    if not rows:
        return error_response("File Excel trống")

    headers = [str(c).strip().lower() if c else "" for c in rows[0]]
    crm_col_idx = None
    for i, h in enumerate(headers):
        if h in ("crm_lead", "crm_lead_id", "crm_code", "crm id"):
            crm_col_idx = i
            break
    if crm_col_idx is None:
        return error_response(
            "Không tìm thấy cột CRM Lead trong file. Cần có cột: CRM Lead, CRM Lead ID hoặc CRM Code"
        )

    success_count = 0
    error_count = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        val = row[crm_col_idx] if crm_col_idx < len(row) else None
        if not val or not str(val).strip():
            continue
        lead_id = str(val).strip()

        lead = frappe.db.get_value("CRM Lead", lead_id, "name")
        if not lead:
            lead = frappe.db.get_value("CRM Lead", {"crm_code": lead_id}, "name")
        if not lead:
            error_count += 1
            errors.append(f"Dòng {row_idx}: Không tìm thấy CRM Lead '{lead_id}'")
            continue

        existing = frappe.db.exists(
            "CRM Admission Entrance Exam Student",
            {"entrance_exam_id": exam_id, "crm_lead_id": lead},
        )
        if existing:
            error_count += 1
            errors.append(f"Dòng {row_idx}: Học sinh đã có trong kỳ")
            continue

        try:
            doc = frappe.new_doc("CRM Admission Entrance Exam Student")
            doc.entrance_exam_id = exam_id
            doc.crm_lead_id = lead
            doc.status = "new"
            doc.insert(ignore_permissions=True)
            success_count += 1
        except Exception:
            error_count += 1
            errors.append(f"Dòng {row_idx}: Lỗi khi thêm")

    frappe.db.commit()
    return success_response(
        message=f"Import: {success_count} thành công, {error_count} lỗi",
        data={"success_count": success_count, "error_count": error_count, "errors": errors[:50]},
    )


def _format_mail_merge_date(value):
    """Định dạng ngày cho mail merge (theo cấu hình user Frappe)."""
    if not value:
        return ""
    try:
        from frappe.utils import format_date

        return format_date(value) or ""
    except Exception:
        return str(value)


@frappe.whitelist(methods=["POST"])
def send_entrance_exam_notification():
    """
    Gửi email thông báo kỳ khảo sát đầu vào cho phụ huynh (mail merge {{tags}}).
    notification_type: exam_schedule | exam_result
    """
    check_crm_permission()
    data = get_request_data()

    exam_id = data.get("exam_id")
    notification_type = (data.get("notification_type") or "").strip()
    subject_template = (data.get("subject") or "").strip()
    content_template = (data.get("content") or "").strip()
    student_record_ids = data.get("student_record_ids") or []

    if not exam_id:
        return validation_error_response("Thiếu exam_id", {"exam_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Entrance Exam", exam_id):
        return not_found_response("Không tìm thấy kỳ khảo sát")
    if notification_type not in ("exam_schedule", "exam_result"):
        return validation_error_response(
            "Loại thông báo không hợp lệ",
            {"notification_type": ['Chọn "exam_schedule" hoặc "exam_result"']},
        )
    if not subject_template or not content_template:
        return validation_error_response(
            "Thiếu tiêu đề hoặc nội dung",
            {"subject": ["Bắt buộc"], "content": ["Bắt buộc"]},
        )
    if not isinstance(student_record_ids, list) or len(student_record_ids) == 0:
        return validation_error_response(
            "Chọn ít nhất một học sinh",
            {"student_record_ids": ["Bắt buộc"]},
        )

    from frappe.utils import md_to_html

    from erp.utils.email_service import send_email_via_service

    exam = frappe.get_doc("CRM Admission Entrance Exam", exam_id)
    sender = frappe.conf.get("admission_sender_email") or "tuyensinh@wellspring.edu.vn"

    sent = 0
    errors = []

    for record_id in student_record_ids:
        if not record_id or not frappe.db.exists(
            "CRM Admission Entrance Exam Student", record_id
        ):
            errors.append({"record": record_id, "error": "Không tìm thấy bản ghi học sinh"})
            continue

        rec = frappe.get_doc("CRM Admission Entrance Exam Student", record_id)
        if rec.entrance_exam_id != exam_id:
            errors.append({"record": record_id, "error": "Học sinh không thuộc kỳ này"})
            continue

        lead = frappe.db.get_value(
            "CRM Lead",
            rec.crm_lead_id,
            [
                "guardian_email",
                "crm_code",
                "student_name",
                "student_dob",
                "guardian_name",
            ],
            as_dict=True,
        )
        if not lead:
            errors.append({"record": record_id, "error": "Không tìm thấy CRM Lead"})
            continue

        guardian_email = (lead.get("guardian_email") or "").strip()
        if not guardian_email:
            errors.append({"record": record_id, "error": "Không có email phụ huynh"})
            continue

        context = {
            "student_name": lead.get("student_name") or "",
            "crm_code": lead.get("crm_code") or rec.crm_lead_id,
            "student_dob": _format_mail_merge_date(lead.get("student_dob")),
            "guardian_name": lead.get("guardian_name") or "",
            "exam_date": _format_mail_merge_date(exam.exam_date),
            "exam_time": exam.exam_time or "",
        }
        if notification_type == "exam_result":
            context["status"] = _ENTRANCE_STATUS_LABEL_VN.get(
                rec.status or "new", rec.status or ""
            )
            er = rec.exam_result or ""
            context["exam_result"] = _ENTRANCE_RESULT_LABEL_VN.get(er, er or "—")

        final_subject = subject_template
        final_content = content_template
        for key, val in context.items():
            final_subject = final_subject.replace(f"{{{{{key}}}}}", str(val))
            final_content = final_content.replace(f"{{{{{key}}}}}", str(val))

        html_body = str(md_to_html(final_content) or final_content)

        result = send_email_via_service(
            to_list=[guardian_email],
            subject=final_subject,
            body=html_body,
            from_email=sender,
        )
        if result.get("success"):
            sent += 1
        else:
            errors.append(
                {
                    "record": record_id,
                    "error": result.get("message", "Gửi email thất bại"),
                }
            )

    return success_response(
        {"sent": sent, "errors": errors},
        message=f"Đã gửi {sent} email",
    )


# --- Mẫu email gửi thông báo kỳ KSĐV (DocType CRM Admission Entrance Exam Email Template) ---

_ENTRANCE_EXAM_EMAIL_TEMPLATE = "CRM Admission Entrance Exam Email Template"


@frappe.whitelist()
def get_entrance_exam_email_templates():
    """Danh sách mẫu email (lưu trong DB, dùng chung toàn trường)."""
    check_crm_permission()
    rows = frappe.get_all(
        _ENTRANCE_EXAM_EMAIL_TEMPLATE,
        fields=["name", "template_name", "notification_type", "subject", "body", "modified"],
        order_by="modified desc",
    )
    return list_response(rows)


@frappe.whitelist(methods=["POST"])
def create_entrance_exam_email_template():
    """Tạo mẫu email mới."""
    check_crm_permission()
    data = get_request_data()
    template_name = (data.get("template_name") or "").strip()
    notification_type = (data.get("notification_type") or "").strip()
    subject = (data.get("subject") or "").strip()
    body = data.get("body")
    if body is not None:
        body = str(body)
    else:
        body = ""

    if not template_name:
        return validation_error_response("Thiếu tên mẫu", {"template_name": ["Bắt buộc"]})
    if notification_type not in ("exam_schedule", "exam_result"):
        return validation_error_response(
            "Loại thông báo không hợp lệ",
            {"notification_type": ['Chọn "exam_schedule" hoặc "exam_result"']},
        )
    if not subject:
        return validation_error_response("Thiếu tiêu đề", {"subject": ["Bắt buộc"]})
    if not body.strip():
        return validation_error_response("Thiếu nội dung", {"body": ["Bắt buộc"]})

    try:
        doc = frappe.new_doc(_ENTRANCE_EXAM_EMAIL_TEMPLATE)
        doc.template_name = template_name
        doc.notification_type = notification_type
        doc.subject = subject
        doc.body = body
        doc.insert(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Đã tạo mẫu")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi tạo mẫu: {str(e)}")


@frappe.whitelist(methods=["POST"])
def update_entrance_exam_email_template():
    """Cập nhật mẫu email (theo `name` = ID bản ghi)."""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    if not name or not frappe.db.exists(_ENTRANCE_EXAM_EMAIL_TEMPLATE, name):
        return not_found_response("Không tìm thấy mẫu")

    template_name = (data.get("template_name") or "").strip()
    notification_type = (data.get("notification_type") or "").strip()
    subject = (data.get("subject") or "").strip()
    body = data.get("body")
    if body is not None:
        body = str(body)
    else:
        body = ""

    if not template_name:
        return validation_error_response("Thiếu tên mẫu", {"template_name": ["Bắt buộc"]})
    if notification_type not in ("exam_schedule", "exam_result"):
        return validation_error_response(
            "Loại thông báo không hợp lệ",
            {"notification_type": ['Chọn "exam_schedule" hoặc "exam_result"']},
        )
    if not subject:
        return validation_error_response("Thiếu tiêu đề", {"subject": ["Bắt buộc"]})
    if not body.strip():
        return validation_error_response("Thiếu nội dung", {"body": ["Bắt buộc"]})

    try:
        doc = frappe.get_doc(_ENTRANCE_EXAM_EMAIL_TEMPLATE, name)
        doc.template_name = template_name
        doc.notification_type = notification_type
        doc.subject = subject
        doc.body = body
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Đã cập nhật mẫu")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi cập nhật: {str(e)}")


@frappe.whitelist(methods=["POST"])
def delete_entrance_exam_email_template():
    """Xóa mẫu email."""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    if not name or not frappe.db.exists(_ENTRANCE_EXAM_EMAIL_TEMPLATE, name):
        return not_found_response("Không tìm thấy mẫu")
    try:
        frappe.delete_doc(_ENTRANCE_EXAM_EMAIL_TEMPLATE, name, ignore_permissions=True)
        frappe.db.commit()
        return success_response(message="Đã xóa mẫu")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi xóa: {str(e)}")
