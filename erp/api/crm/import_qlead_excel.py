"""
Import ho so tu file Excel mau (sheet QLead) thang vao buoc QLead.

Chay bang bench, KHONG co endpoint whitelist va KHONG co man hinh upload —
day la script van hanh mot lan, khong phai tinh nang cho nguoi dung cuoi.

    # xem truoc, khong ghi gi
    bench --site <site> execute erp.api.crm.import_qlead_excel.run \
        --kwargs "{'path': '/duong/dan/CRM_Mau_Import_QLead_Enrolled_v2.xlsx'}"

    # ghi that
    bench --site <site> execute erp.api.crm.import_qlead_excel.run \
        --kwargs "{'path': '...', 'dry_run': 0}"

Khac `bulk_import_leads` (chi doc 12 truong): script nay doc TOAN BO cot cua file mau —
PIC, nguon, ma hoc sinh, ly do tu choi, uu dai %, ca hai phu huynh, khoa ngan han,
va nhat ky cham soc (tao CRM Lead Note).

Chong trung theo cap (SDT da chuan hoa + ten hoc sinh) nen:
  - anh chi em dung chung SDT van vao duoc,
  - chay lai script lan hai khong tao ban ghi trung.
"""

import re
import unicodedata
from datetime import date, datetime, timedelta
from datetime import time as dtime

import frappe
from frappe.utils import flt

from erp.api.crm.utils import (
    STEP_STATUSES,
    generate_crm_code,
    normalize_phone_number,
    resolve_status_input,
    validate_phone_number,
)
from erp.utils.family_relationship import (
    find_guardian_by_phone,
    guardian_phone_matches,
)
from erp.utils.relationship_types import normalize as normalize_relationship

STEP = "QLead"
SHEET = "QLead"
HEADER_ROW = 2
FIRST_DATA_ROW = 3
EXAMPLE_PREFIX = "[VÍ DỤ]"

# Tieu de cot tren file mau -> key noi bo. Dong bo voi sheet «Từ điển cột».
COLUMNS = {
    # bat buoc
    "Họ tên học sinh *": "student_name",
    "PH1: SĐT chính *": "g1_phone_1",
    "Trạng thái *": "status",
    "Lớp dự tuyển *": "target_grade",
    # phan loai / phu trach
    "PIC Sales (email User)": "pic_sales",
    "Campus (mã SIS Campus)": "campus_id",
    "Phương thức nhận data": "data_source",
    "Nguồn 1 (nguồn chính)": "source_1",
    "Nguồn 2 (nguồn phụ)": "sub_source_1",
    "Nguồn 3 (ghi chú nguồn)": "source_note_1",
    "Người giới thiệu": "referrer",
    "Mã CBGV người giới thiệu": "staff_code",
    "SĐT người giới thiệu": "referrer_phone",
    # hoc sinh
    "Giới tính (Nam/Nữ)": "student_gender",
    "Ngày sinh (dd/mm/yyyy)": "student_dob",
    "Số định danh HS (CCCD/CMND)": "student_personal_id_number",
    "Mã học sinh": "student_code",
    "Lớp đang học": "current_grade",
    "Trường đang học": "current_school",
    "Hệ học": "study_program",
    "Năm học dự tuyển": "target_academic_year",
    "Học kỳ dự tuyển": "target_semester",
    # Cung tro vao student_note (Small Text). KHONG dung
    # student_special_characteristics: field do la Data/varchar(140), du lieu that
    # co o dai toi 3.264 ky tu nen se bi cat cut.
    "Lưu ý đặc biệt về HS/PH": "student_note",
    "Ghi chú học sinh": "student_note",        # tieu de cu, giu de file cu van chay
    # tu choi
    "Lý do từ chối (nhóm)": "reject_reason",
    "Mô tả chi tiết từ chối": "reject_detail",
    # uu dai
    "Ưu đãi học phí (%)": "tuition_fee_pct",
    "Ưu đãi phí dịch vụ (%)": "service_fee_pct",
    "Ưu đãi phí phát triển trường (%)": "dev_fee_pct",
    "Ưu đãi KSĐV (%)": "ksdv_pct",
    # khoa ngan han / su kien
    "Khóa ngắn hạn 1 - Tên": "course_1_name",
    "Khóa ngắn hạn 1 - Trạng thái": "course_1_status",
    "Khóa ngắn hạn 2 - Tên": "course_2_name",
    "Khóa ngắn hạn 2 - Trạng thái": "course_2_status",
    "Sự kiện đã tham gia": "event_1",
    # nhat ky cham soc
    "Ghi chú / Log chăm sóc": "care_log",
}

# Khoi phu huynh — sinh ra tu cung mot bo hau to nen PH1/PH2 luon doi xung.
_GUARDIAN_SUFFIX = {
    "Họ tên phụ huynh": "name",
    "Quan hệ với HS": "relationship",
    "SĐT chính": "phone_1",
    "SĐT phụ": "phone_2",
    "Email": "email",
    "Số CCCD/Hộ chiếu": "id_number",
    "Ngày sinh (dd/mm/yyyy)": "dob",
    "Nghề nghiệp": "occupation",
    "Chức vụ": "position",
    "Nơi công tác": "workplace",
    "Quốc tịch": "nationality",
    "Địa chỉ liên hệ": "address",
    "Ghi chú": "note",
}
for _slot in (1, 2):
    for _label, _key in _GUARDIAN_SUFFIX.items():
        _hdr = f"PH{_slot}: {_label}"
        if _slot == 1 and _key == "phone_1":
            _hdr += " *"          # cot bat buoc, tieu de co dau sao
        COLUMNS.setdefault(_hdr, f"g{_slot}_{_key}")

# PH1 -> truong phang tren CRM Lead; PH2 -> document CRM Guardian.
_G1_TO_LEAD = {
    "g1_name": "guardian_name",
    "g1_relationship": "relationship",
    "g1_email": "guardian_email",
    "g1_id_number": "guardian_id_number",
    "g1_dob": "guardian_dob",
    "g1_occupation": "guardian_occupation",
    "g1_position": "guardian_position",
    "g1_workplace": "guardian_workplace",
    "g1_nationality": "guardian_nationality",
    "g1_address": "guardian_address",
    "g1_note": "guardian_note",
}

# campus_id khong nam day: phai qua _Resolver.campus() de doi ten/short title -> docname.
_LEAD_SIMPLE_FIELDS = (
    "student_name", "student_gender", "student_personal_id_number", "student_code",
    "current_grade", "target_grade", "current_school", "study_program",
    "target_semester", "student_note", "data_source", "staff_code",
)
_PCT_FIELDS = ("tuition_fee_pct", "service_fee_pct", "dev_fee_pct", "ksdv_pct")

_VALID_STATUSES = STEP_STATUSES[STEP]


# ---------------------------------------------------------------- doc / chuan hoa


def _txt(v):
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    return str(v).strip()


def _pct(v):
    """
    Cot «(%)» chi nhan so. Chuoi mo ta uu dai ("giam 50% hoc phi ky I") -> None.

    Khong dung thang frappe.utils.flt: flt tra ve 0.0 cho chuoi khong parse duoc,
    nhu vay o mo ta se bien thanh uu dai 0% — sai lech im lang.
    """
    s = _txt(v).replace("%", "").replace(",", ".").strip()
    if not s or not re.fullmatch(r"\d+(\.\d+)?", s):
        return None
    return flt(s)


def _date(v):
    """dd/mm/yyyy | dd-mm-yyyy | YYYY-MM-DD | datetime -> YYYY-MM-DD."""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = _txt(v)
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:10], fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _phone(v):
    """-> +84xxxxxxxxx, hoac '' neu khong hop le."""
    s = _txt(v)
    if not s or not validate_phone_number(s):
        return ""
    return normalize_phone_number(s)


def _emails(v):
    s = _txt(v)
    if not s:
        return []
    out = []
    for p in re.split(r"[\s,;/]+", s):
        p = p.strip()
        # Chi nhan ky tu ASCII: Frappe validate email chat hon, du lieu that co o
        # 'nguyenhuonggiãng7@gmail.com' — regex cu cho qua roi ca DONG bi rollback.
        if p and re.fullmatch(r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}", p) \
                and p not in out:
            out.append(p)
    return out


def _slug(s):
    s = unicodedata.normalize("NFD", s or "")
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("đ", "d").replace("Đ", "D").lower()
    s = re.sub(r"[^a-z0-9]+", "-", s).strip("-")
    return s[:40] or "guardian"


# ---------------------------------------------------------------- tra cuu (co cache)


class _Resolver:
    """Gom cache tra cuu de khong query lai cho tung dong."""

    def __init__(self):
        self._user = {}
        self._source = {}
        self._source_note = {}
        self._year = {}
        self._campus = {}

    def user(self, raw):
        s = _txt(raw)
        if not s:
            return None
        if s in self._user:
            return self._user[s]
        found = s if frappe.db.exists("User", s) else None
        if not found:
            row = frappe.db.sql(
                "select name from `tabUser` where lower(ifnull(email,'')) = lower(%s) limit 1",
                (s,),
            )
            found = row[0][0] if row else None
        self._user[s] = found
        return found

    def _by_name_or_field(self, doctype, field, raw, cache):
        s = _txt(raw)
        if not s:
            return None
        if s in cache:
            return cache[s]
        found = s if frappe.db.exists(doctype, s) else frappe.db.get_value(
            doctype, {field: s}, "name"
        )
        cache[s] = found
        return found

    def source(self, raw):
        return self._by_name_or_field("CRM Source", "source_name", raw, self._source)

    def source_note(self, raw):
        return self._by_name_or_field("CRM Source Note", "note_name", raw, self._source_note)

    def campus(self, raw):
        """
        CRM Lead.campus_id la Link -> SIS Campus, luu DOCNAME (CAMPUS-00001).
        Nguoi nhap thuong ghi Short Title ("WSHN 01") hoac ten day du, nen nhan ca:
        docname | short_title | title_vn | title_en (bo qua hoa/thuong va khoang trang).
        """
        s = _txt(raw)
        if not s:
            return None
        if s in self._campus:
            return self._campus[s]
        found = s if frappe.db.exists("SIS Campus", s) else None
        if not found:
            key = " ".join(s.split()).lower()
            for row in frappe.get_all(
                "SIS Campus", fields=["name", "short_title", "title_vn", "title_en"],
                limit_page_length=0,
            ):
                if any(" ".join(str(row.get(f) or "").split()).lower() == key
                       for f in ("short_title", "title_vn", "title_en")):
                    found = row["name"]
                    break
        self._campus[s] = found
        return found

    def school_year(self, raw):
        s = _txt(raw)
        if not s:
            return None
        if s in self._year:
            return self._year[s]
        found = s if frappe.db.exists("SIS School Year", s) else None
        if not found:
            for f in ("title_vn", "title_en"):
                if not frappe.db.has_column("SIS School Year", f):
                    continue
                found = frappe.db.get_value("SIS School Year", {f: s}, "name")
                if found:
                    break
        self._year[s] = found
        return found


def _get_or_create_guardian(row, dry_run):
    """
    PH2 -> document CRM Guardian. Dinh danh theo SDT da chuan hoa (khop rang buoc
    duy nhat ma add_lead_guardian dang dung), nho vay hai anh em cung bo chi tao mot
    guardian. Thieu ten hoac SDT thi bo qua ca khoi, khong bao loi dong.
    """
    name = _txt(row.get("g2_name"))
    phone = _phone(row.get("g2_phone_1"))
    if not name or not phone:
        return None, None

    # Tra CA field phang LAN bang con `CRM Guardian Phone`: phu huynh da khai so thu
    # hai o lan nhap truoc thi khop bang con, neu chi tra field phang se tao Guardian
    # THU HAI cho cung mot nguoi (xem erp/utils/family_relationship.py).
    existing = find_guardian_by_phone(phone)
    if existing:
        return existing, "reused"
    if dry_run:
        return "(sẽ tạo mới)", "created"

    gid = f"{_slug(name)}-{frappe.generate_hash(length=6)}"
    while frappe.db.exists("CRM Guardian", {"guardian_id": gid}):
        gid = f"{_slug(name)}-{frappe.generate_hash(length=6)}"

    doc = frappe.get_doc({
        "doctype": "CRM Guardian",
        "guardian_id": gid,
        "guardian_name": name,
        # `phone` da la +84... (_phone goi normalize_phone_number). Cho nay tung viet
        # `e164` — bien khong ton tai trong scope nay, NameError lam ca dong bi rollback
        # o savepoint cua run() moi khi PH2 la nguoi moi.
        "phone_number": phone,
        "email": (_emails(row.get("g2_email")) or [""])[0],
        "id_number": _txt(row.get("g2_id_number")),
        "occupation": _txt(row.get("g2_occupation")),
        "position": _txt(row.get("g2_position")),
        "workplace": _txt(row.get("g2_workplace")),
        "address": _txt(row.get("g2_address")),
        "note": _txt(row.get("g2_note")),
        "dob": _date(row.get("g2_dob")),
    })
    doc.flags.ignore_validate = True
    doc.flags.ignore_mandatory = True
    doc.insert(ignore_permissions=True)

    # Bang con `phone_numbers` la cai giao dien CRM doc, va tu nay ca dedup cung tra
    # (find_guardian_by_phone). Truoc day chi so THU HAI duoc them, so chinh khong co
    # dong nao -> giao dien hien "Chua co du lieu" va so phu bi coi la khong chinh.
    # _add_child_contact tu bo trung va tu dat dung mot primary.
    changed = _add_child_contact(doc, "phone_numbers", "phone_number", phone)
    extra = _phone(row.get("g2_phone_2"))
    if extra and extra != phone:
        changed = _add_child_contact(doc, "phone_numbers", "phone_number", extra) or changed
    if changed:
        doc.save(ignore_permissions=True)
    return doc.name, "created"


# ---------------------------------------------------------------- nhat ky cham soc

# Moi muc bat dau bang «dd/m» dau dong: "9/6:", "09/06 :", "16/10-", co the kem nam.
_ENTRY_RX = re.compile(r"^\s*(\d{1,2})\s*[/.\-]\s*(\d{1,2})(?:\s*[/.\-]\s*(\d{2,4}))?\s*[:：\-]?\s*")

# Suy phuong thuc lien he tu noi dung (Select cua CRM Lead Note).
_METHOD_HINTS = (
    ("Zalo", ("zalo", "ntin", "nhan tin", "nhắn tin")),
    ("Email", ("email", "mail", "gui thu", "gửi thư")),
    ("Gap truc tiep", ("school tour", "schooltour", "open day", "openday", "tham quan",
                       "den truong", "đến trường", "gap truc tiep", "gặp trực tiếp",
                       "hoi thao", "hội thảo", "den tham", "đến thăm")),
)


def _guess_method(text):
    low = unicodedata.normalize("NFC", (text or "")).lower()
    for method, keys in _METHOD_HINTS:
        if any(k in low for k in keys):
            return method
    return "Goi dien"


def _parse_care_entries(body):
    """
    Tach nhat ky thanh cac muc. Tra ve list theo DUNG thu tu trong file
    (moi nhat truoc): [{'d','m','y','text','tag'}], d/m = None neu muc khong co ngay.
    Dong khong mo dau bang ngay duoc noi vao muc phia tren.
    """
    entries, cur = [], None
    for line in (body or "").splitlines():
        if not line.strip():
            continue
        m = _ENTRY_RX.match(line)
        if m:
            if cur:
                entries.append(cur)
            yr = m.group(3)
            if yr and len(yr) == 2:
                yr = "20" + yr
            cur = {
                "d": int(m.group(1)), "m": int(m.group(2)),
                "y": int(yr) if yr else None,
                "tag": f"{int(m.group(1)):02d}/{int(m.group(2)):02d}",
                "text": line[m.end():].strip(),
            }
        elif cur:
            cur["text"] = (cur["text"] + "\n" + line.strip()).strip()
        else:
            cur = {"d": None, "m": None, "y": None, "tag": "", "text": line.strip()}
    if cur:
        entries.append(cur)
    return entries


def _assign_years(entries, as_of, year_floor):
    """
    Log ghi moi-nhat-truoc nhung KHONG ghi nam. Di tu tren (moi) xuong duoi (cu),
    chan tran bang `as_of`; moi khi ngay lai "tien len" so voi muc truoc thi lui 1 nam.
    Muc nao ghi ro nam thi dung lam moc neo.

    Tra ve (list date|None cung do dai entries, suspicious: bool).
    `suspicious` = co muc bi day xuong duoi `year_floor` -> thu tu trong file kha nang
    bi PIC ghi lon hoac sai ngay; date bi kep lai o floor va can review tay.
    """
    out, prev, suspicious = [], as_of, False
    for e in entries:
        if not e["d"]:
            out.append(None)
            continue
        try:
            if e["y"]:
                cand = date(e["y"], e["m"], e["d"])
                # nam nhap tay cung co the sai (gap '2028' trong du lieu that):
                # ngoai khoang hop le thi bo, quay ve suy tu vi tri.
                if not (year_floor <= cand <= as_of):
                    suspicious = True
                    cand = date(prev.year, e["m"], e["d"])
                    if cand > prev:
                        cand = date(prev.year - 1, e["m"], e["d"])
            else:
                cand = date(prev.year, e["m"], e["d"])
                if cand > prev:
                    cand = date(prev.year - 1, e["m"], e["d"])
        except ValueError:          # 32/13, 30/02…
            out.append(None)
            continue
        if cand > as_of:
            suspicious = True
            cand = as_of
        if cand < year_floor:
            suspicious = True
            cand = year_floor
        out.append(cand)
        prev = cand
    return out, suspicious


def _add_care_notes(lead_doc, content, assignee, as_of, year_floor):
    """
    Tao mot CRM Lead Note cho TUNG muc trong nhat ky, lui `creation` ve ngay suy ra
    de tab ghi chu hien dung thu tu thoi gian (get_notes sort theo creation desc).

    Frappe luon ghi creation = now() luc insert, nen phai UPDATE lai sau do
    (update_modified=False de khong dung vao `modified`).

    Tra ve (so note da tao, suspicious).
    """
    body = _txt(content)
    if not body:
        return 0, False

    entries = _parse_care_entries(body)
    if not entries:
        return 0, False
    dates, suspicious = _assign_years(entries, as_of, year_floor)

    # muc khong co ngay -> gan cung ngay muc co ngay gan nhat (uu tien phia duoi)
    fallback = next((d for d in dates if d), as_of)

    made = 0
    for i, (e, d) in enumerate(zip(entries, dates)):
        text = e["text"].strip()
        if not text:
            continue
        tag = e["tag"] or "không rõ ngày"
        title = f"{tag} — {text.splitlines()[0]}"[:135]
        note = frappe.get_doc({
            "doctype": "CRM Lead Note",
            "lead": lead_doc.name,
            "campus_id": lead_doc.campus_id,
            "category": "Lich su",
            "title": title,
            "communication_method": _guess_method(text),
            "content": text.replace("\n", "<br>"),
            "assignee": assignee or frappe.session.user,
            "is_completed": 1,
        })
        note.insert(ignore_permissions=True)

        # 12:00 tru i giay: cung mot ngay thi muc moi hon van dung truoc
        stamp = datetime.combine(d or fallback, dtime(12, 0, 0)) - timedelta(seconds=i)
        frappe.db.set_value("CRM Lead Note", note.name, "creation", stamp,
                            update_modified=False)
        made += 1
    return made, suspicious


# ---------------------------------------------------------------- doc file


def _resolve_path(path):
    """
    Chap nhan ca hai kieu duong dan:
      - duong dan he thong: /home/frappe/.../file.xlsx hoac /tmp/file.xlsx
      - duong dan file cua Frappe (cai hien tren giao dien): /private/files/... , /files/...
        -> quy ve sites/<site>/private|public/files/...
    """
    import os

    p = (path or "").strip()
    if not p:
        frappe.throw("Thieu tham so path")
    if os.path.isabs(p) and os.path.isfile(p):
        return p

    tried = [p]
    for prefix, folder in (("/private/files/", "private"), ("private/files/", "private"),
                           ("/files/", "public"), ("files/", "public")):
        if p.startswith(prefix):
            cand = frappe.get_site_path(folder, "files", p.split("files/", 1)[1])
            if os.path.isfile(cand):
                return cand
            tried.append(cand)
            break
    else:
        # ten file tran -> thu ca hai thu muc
        for folder in ("private", "public"):
            cand = frappe.get_site_path(folder, "files", p)
            if os.path.isfile(cand):
                return cand
            tried.append(cand)

    frappe.throw("Khong tim thay file. Da thu:\n  " + "\n  ".join(tried))


def _read_rows(path, limit=0):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    if SHEET not in wb.sheetnames:
        frappe.throw(f"File khong co sheet '{SHEET}'")
    ws = wb[SHEET]

    header, unknown = {}, []
    for j, cell in enumerate(next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW)), start=1):
        h = _txt(cell.value)
        if not h:
            continue
        key = COLUMNS.get(h)
        if key:
            header[j] = key
        else:
            unknown.append(h)

    missing = [h for h, k in COLUMNS.items() if k in
               ("student_name", "g1_phone_1", "status", "target_grade")
               and k not in header.values()]
    if missing:
        frappe.throw("File thieu cot bat buoc: " + ", ".join(missing))

    rows = []
    for r_idx, r in enumerate(ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True),
                              start=FIRST_DATA_ROW):
        data = {}
        for j, key in header.items():
            if j - 1 < len(r):
                data[key] = r[j - 1]
        # Mot dong duoc coi la ho so khi co ten HS HOAC SDT. Khong dung "co bat ky o nao
        # co gia tri": nguoi nhap thuong keo cong thuc / hang so (Campus, Quan he) xuong
        # qua vung du lieu that, nhung dong do khong phai ho so.
        if not _txt(data.get("student_name")) and not _txt(data.get("g1_phone_1")):
            continue
        if _txt(data.get("student_name")).startswith(EXAMPLE_PREFIX):
            continue
        data["_row"] = r_idx
        rows.append(data)
        if limit and len(rows) >= limit:
            break
    wb.close()
    return rows, unknown


# ---------------------------------------------------------------- validate 1 dong


def _validate(row, rs):
    """-> (list loi chan cung, dict gia tri da resolve)."""
    errs, out = [], {}

    if not _txt(row.get("student_name")):
        errs.append("thieu Ho ten hoc sinh")

    phone = _phone(row.get("g1_phone_1"))
    if not phone:
        errs.append(f"SDT PH1 khong hop le: {_txt(row.get('g1_phone_1'))!r}")
    out["phone"] = phone

    raw_status = _txt(row.get("status"))
    status = resolve_status_input(raw_status, _VALID_STATUSES) if raw_status else ""
    if not status:
        errs.append("thieu Trang thai")
    elif status not in _VALID_STATUSES:
        errs.append(f"Trang thai khong hop le cho buoc {STEP}: {raw_status!r}")
    out["status"] = status

    ds = _txt(row.get("data_source"))
    if ds and ds not in ("Online", "Offline", "Doi tac", "AI Chatbot"):
        errs.append(f"Phuong thuc nhan data ngoai Select: {ds!r}")

    pic_raw = _txt(row.get("pic_sales"))
    pic = rs.user(pic_raw) if pic_raw else None
    if pic_raw and not pic:
        errs.append(f"PIC khong co User tuong ung: {pic_raw!r}")
    out["pic"] = pic

    src_raw = _txt(row.get("source_1"))
    src = rs.source(src_raw) if src_raw else None
    if src_raw and not src:
        errs.append(f"Nguon 1 khong co trong CRM Source: {src_raw!r}")
    out["source"] = src

    yr_raw = _txt(row.get("target_academic_year"))
    yr = rs.school_year(yr_raw) if yr_raw else None
    if yr_raw and not yr:
        errs.append(f"Nam hoc khong khop SIS School Year: {yr_raw!r}")
    out["year"] = yr

    cid_raw = _txt(row.get("campus_id"))
    campus = rs.campus(cid_raw) if cid_raw else None
    if cid_raw and not campus:
        errs.append(f"Campus khong khop SIS Campus (docname/short_title/ten): {cid_raw!r}")
    out["campus"] = campus

    if status == "Tu choi" and not _txt(row.get("reject_reason")):
        out["warn_no_reject_reason"] = True

    return errs, out


def _existing_lead(phone, student_name):
    """Trung khi CUNG SDT va CUNG ten hoc sinh — anh chi em dung chung so van vao duoc."""
    if not phone:
        return None
    rows = frappe.db.sql(
        """
        SELECT cl.name FROM `tabCRM Lead` cl
        INNER JOIN `tabCRM Lead Phone` clp ON clp.parent = cl.name
        WHERE clp.phone_number = %s
          AND LOWER(TRIM(IFNULL(cl.student_name,''))) = %s
        LIMIT 1
        """,
        (phone, " ".join((student_name or "").split()).lower()),
    )
    return rows[0][0] if rows else None


# ---------------------------------------------------------------- tao 1 ho so


def _build_lead(row, resolved, rs):
    doc = frappe.new_doc("CRM Lead")
    doc.step = STEP
    doc.status = resolved["status"]
    doc.crm_code = generate_crm_code()

    for f in _LEAD_SIMPLE_FIELDS:
        v = _txt(row.get(f))
        if v:
            doc.set(f, v)

    dob = _date(row.get("student_dob"))
    if dob:
        doc.student_dob = dob
    if resolved["year"]:
        doc.target_academic_year = resolved["year"]
    if resolved["pic"]:
        doc.pic_sales = resolved["pic"]
    if resolved["campus"]:
        doc.campus_id = resolved["campus"]

    for f in _PCT_FIELDS:
        v = _pct(row.get(f))
        if v is not None:
            doc.set(f, v)

    if doc.status == "Tu choi":
        doc.reject_reason = _txt(row.get("reject_reason"))
        doc.reject_detail = _txt(row.get("reject_detail"))

    # PH1 -> truong phang
    for key, field in _G1_TO_LEAD.items():
        v = _txt(row.get(key))
        if not v:
            continue
        doc.set(field, _date(v) if field == "guardian_dob" else v)

    # SDT: PH1 chinh (primary) + PH1 phu + SDT phu cua PH2 khong trung
    seen = []
    for raw, primary in ((row.get("g1_phone_1"), 1), (row.get("g1_phone_2"), 0)):
        p = _phone(raw)
        if p and p not in seen:
            seen.append(p)
            doc.append("phone_numbers", {"phone_number": p, "is_primary": primary})

    for addr in _emails(row.get("g1_email")):
        doc.append("emails", {"email_address": addr, "is_primary": 1 if not doc.emails else 0})

    if resolved["source"]:
        doc.append("source", {
            "source": resolved["source"],
            "sub_source": _txt(row.get("sub_source_1")),
            "source_note": rs.source_note(row.get("source_note_1")) or "",
        })

    for n in (1, 2):
        cname = _txt(row.get(f"course_{n}_name"))
        if cname:
            doc.append("courses", {
                "course_name": cname,
                "status": _txt(row.get(f"course_{n}_status")),
            })

    for ev in [e.strip() for e in _txt(row.get("event_1")).split(";") if e.strip()]:
        doc.append("events", {"event_name": ev})

    return doc


# ---------------------------------------------------------------- entry point


def run(path, dry_run=1, limit=0, commit_every=100, as_of=None, year_floor=None):
    """
    path         duong dan file .xlsx (sheet «QLead»)
    dry_run      1 = chi kiem tra va bao cao, KHONG ghi DB (mac dinh)
    limit        chi xu ly N dong dau, 0 = tat ca
    commit_every commit sau moi N ban ghi tao thanh cong
    as_of        'YYYY-MM-DD' — tran tren khi suy nam cho nhat ky (mac dinh: hom nay)
    year_floor   'YYYY-MM-DD' — san duoi; muc bi day xuong duoi moc nay se bi kep lai
                 va ho so duoc danh dau can review (mac dinh: 2024-07-01)

    Khi ghi that, danh sach ban ghi da tao duoc luu ra <path>.import-log.json
    de con duong lui — xem `undo()`.
    """
    from frappe.utils import getdate, nowdate

    dry_run = int(dry_run)
    limit = int(limit)
    commit_every = int(commit_every) or 100
    as_of_d = getdate(as_of) if as_of else getdate(nowdate())
    floor_d = getdate(year_floor) if year_floor else date(2024, 7, 1)

    path = _resolve_path(path)
    print(f"  File: {path}")
    print(f"  Suy nam nhat ky: tran {as_of_d} / san {floor_d}")

    rows, unknown_cols = _read_rows(path, limit)
    rs = _Resolver()

    res = {
        "total": len(rows), "created": 0, "duplicates": 0, "failed": 0,
        "notes": 0, "guardians_created": 0, "guardians_reused": 0,
        "errors": [], "warnings": [],
        "created_leads": [], "created_guardians": [], "note_review": [],
    }
    if unknown_cols:
        res["warnings"].append(f"Cot khong nhan dang (bo qua): {', '.join(unknown_cols)}")

    for i, row in enumerate(rows, start=1):
        rnum = row["_row"]
        errs, resolved = _validate(row, rs)
        if errs:
            res["failed"] += 1
            res["errors"].append({"row": rnum, "error": "; ".join(errs)})
            continue

        dup = _existing_lead(resolved["phone"], _txt(row.get("student_name")))
        if dup:
            res["duplicates"] += 1
            res["errors"].append({"row": rnum, "error": f"da co ho so {dup} (trung SDT + ten HS)"})
            continue

        if resolved.get("warn_no_reject_reason"):
            res["warnings"].append(f"dong {rnum}: Tu choi nhung trong Ly do tu choi")

        if dry_run:
            res["created"] += 1
            entries = _parse_care_entries(_txt(row.get("care_log")))
            entries = [e for e in entries if e["text"].strip()]
            if entries:
                _, susp = _assign_years(entries, as_of_d, floor_d)
                res["notes"] += len(entries)
                if susp:
                    res["note_review"].append(
                        {"row": rnum, "student": _txt(row.get("student_name")),
                         "entries": len(entries)})
            g, how = _get_or_create_guardian(row, dry_run=1)
            if g:
                res["guardians_created" if how == "created" else "guardians_reused"] += 1
            continue

        savepoint = f"imp_qlead_{i}"
        try:
            frappe.db.savepoint(savepoint)

            doc = _build_lead(row, resolved, rs)
            doc.insert(ignore_permissions=True)

            gname, how = _get_or_create_guardian(row, dry_run=0)
            if gname:
                doc.append("lead_guardians", {
                    "guardian": gname,
                    "relationship_type": normalize_relationship(row.get("g2_relationship")),
                    "is_primary_contact": 0,
                    "display_order": 2,
                })
                doc.save(ignore_permissions=True)
                res["guardians_created" if how == "created" else "guardians_reused"] += 1
                if how == "created":
                    res["created_guardians"].append(gname)

            from erp.api.crm.student_code import ensure_student_code_for_qlead_status
            if not doc.student_code:
                ensure_student_code_for_qlead_status(doc)
                if doc.student_code:
                    doc.save(ignore_permissions=True)

            n_notes, susp = _add_care_notes(
                doc, row.get("care_log"), resolved["pic"], as_of_d, floor_d)
            res["notes"] += n_notes
            if susp:
                res["note_review"].append(
                    {"row": rnum, "lead": doc.name,
                     "student": doc.student_name, "entries": n_notes})

            from erp.api.crm.pipeline import _log_step_change
            _log_step_change(
                doc.name, "", STEP, "", doc.status,
                reject_reason=doc.reject_reason if doc.status == "Tu choi" else None,
                reject_detail=doc.reject_detail if doc.status == "Tu choi" else None,
            )

            res["created"] += 1
            res["created_leads"].append(doc.name)
        except Exception as e:
            try:
                frappe.db.rollback(save_point=savepoint)
            except Exception:
                pass
            res["failed"] += 1
            res["errors"].append({"row": rnum, "error": str(e)[:300]})
            frappe.log_error(
                message=frappe.get_traceback() or str(e),
                title=f"import_qlead_excel dong {rnum}",
            )

        if res["created"] and res["created"] % commit_every == 0:
            frappe.db.commit()

    if not dry_run:
        frappe.db.commit()
        res["log_path"] = _write_log(path, res)

    _print(res, dry_run)
    return res


_BACKFILL_FIELDS = {
    "student_gender": "g",              # 'g' = chuoi thuong, controller tu chuan hoa Nu/Nam
    "student_dob": "d",                 # 'd' = ngay dd/mm/yyyy -> YYYY-MM-DD
    "current_grade": "s",
    "target_grade": "s",
    "current_school": "s",
    "student_note": "s",
}


def backfill(path, dry_run=1, fields=None, overwrite=0, clear_missing=None,
             commit_every=200):
    """
    Cap nhat them thong tin cho ho so DA import (khong tao moi).

        bench --site <site> execute erp.api.crm.import_qlead_excel.backfill \
            --kwargs "{'path': '/private/files/....xlsx', 'dry_run': 0}"

    path       file .xlsx (sheet «QLead») — chap nhan ca duong dan Frappe
    dry_run    1 = chi bao cao (mac dinh)
    fields     danh sach field muon cap nhat; bo trong = ca 5:
               student_gender, student_dob, current_grade, target_grade, current_school
    overwrite  0 = chi dien vao o dang TRONG tren ho so (mac dinh, an toan)
               1 = ghi de ca khi ho so da co gia tri khac
    clear_missing  danh sach field: khi FILE khong co du lieu ma ho so lai dang co
               gia tri thi XOA gia tri do. Dung cho truong hop Frappe tu gan mac dinh:
               Select nao khong co dong trong dau `options` se duoc
               frappe.model.create_new gan option dau tien cho MOI doc moi
               (vd student_gender = 'Nam'), tao ra du lieu khong ai nhap.

    Doi chieu ho so bang cap (SDT da chuan hoa + ten hoc sinh) — dung rule voi
    `run()`, nen anh chi em dung chung SDT khong bi lan.
    """
    dry_run, overwrite = int(dry_run), int(overwrite)
    commit_every = int(commit_every) or 200
    use = list(fields) if fields else list(_BACKFILL_FIELDS)
    bad = [f for f in use if f not in _BACKFILL_FIELDS]
    if bad:
        frappe.throw(f"Field khong ho tro: {bad}. Cho phep: {list(_BACKFILL_FIELDS)}")
    clear = set(clear_missing or [])
    if not clear <= set(use):
        frappe.throw(f"clear_missing phai nam trong fields: {sorted(clear - set(use))}")

    path = _resolve_path(path)
    print(f"  File: {path}")
    print(f"  Field: {', '.join(use)} | overwrite={overwrite}"
          + (f" | clear_missing={sorted(clear)}" if clear else ""))

    rows, _ = _read_rows(path, 0)
    res = {"total": len(rows), "matched": 0, "updated": 0, "unchanged": 0,
           "not_found": 0, "errors": [], "per_field": {f: 0 for f in use},
           "cleared": {}}

    # cot Excel -> field CRM (chi 5 field duoc phep)
    src_key = {f: f for f in _BACKFILL_FIELDS}

    for i, row in enumerate(rows, start=1):
        rnum = row["_row"]
        phone = _phone(row.get("g1_phone_1"))
        sname = _txt(row.get("student_name"))
        if not phone or not sname:
            continue
        lead = _existing_lead(phone, sname)
        if not lead:
            res["not_found"] += 1
            continue
        res["matched"] += 1

        changes = {}
        for f in use:
            raw = row.get(src_key[f])
            kind = _BACKFILL_FIELDS[f]
            val = _date(raw) if kind == "d" else _txt(raw)
            cur = frappe.db.get_value("CRM Lead", lead, f)
            if not val:
                # file khong co du lieu: mac dinh bo qua, tru khi duoc yeu cau don
                # gia tri rac do Frappe tu gan (xem clear_missing).
                if f in clear and cur:
                    changes[f] = ""
                    res["cleared"][f] = res["cleared"].get(f, 0) + 1
                continue
            cur_s = str(cur)[:10] if (kind == "d" and cur) else (str(cur).strip() if cur else "")
            if cur_s and not overwrite:
                continue
            if cur_s == val:
                continue
            changes[f] = val

        if not changes:
            res["unchanged"] += 1
            continue

        for f in changes:
            res["per_field"][f] += 1
        res["updated"] += 1
        if dry_run:
            continue

        savepoint = f"bf_{i}"
        try:
            frappe.db.savepoint(savepoint)
            doc = frappe.get_doc("CRM Lead", lead)
            for f, v in changes.items():
                doc.set(f, v)
            # giu validate de controller chuan hoa gioi tinh (Nữ -> Nu) va check Select
            doc.flags.ignore_mandatory = True
            doc.save(ignore_permissions=True)
        except Exception as e:
            try:
                frappe.db.rollback(save_point=savepoint)
            except Exception:
                pass
            res["updated"] -= 1
            for f in changes:
                res["per_field"][f] -= 1
            res["errors"].append({"row": rnum, "lead": lead, "error": str(e)[:250]})
            frappe.log_error(message=frappe.get_traceback() or str(e),
                             title=f"backfill CRM Lead dong {rnum}")

        if res["updated"] and res["updated"] % commit_every == 0:
            frappe.db.commit()

    if not dry_run:
        frappe.db.commit()

    print("")
    print("=" * 62)
    print("  DRY-RUN — khong ghi gi vao DB" if dry_run else "  DA GHI VAO DB")
    print("=" * 62)
    print(f"  Dong doc duoc      : {res['total']}")
    print(f"  Khop ho so         : {res['matched']}")
    print(f"  Khong tim thay     : {res['not_found']}")
    print(f"  {'Se cap nhat' if dry_run else 'Da cap nhat':<18} : {res['updated']}")
    print(f"  Khong doi          : {res['unchanged']}")
    print("  Theo field:")
    for f, n in res["per_field"].items():
        cl = res["cleared"].get(f, 0)
        extra = f"   (trong do XOA gia tri rac: {cl})" if cl else ""
        print(f"    {f:<18} {n}{extra}")
    if res["errors"]:
        print(f"\n  Loi ({len(res['errors'])}):")
        for e in res["errors"][:20]:
            print(f"    dong {e['row']:>5} {e['lead']}: {e['error']}")
    print("")
    return res


def _write_log(path, res):
    """Ghi danh sach ban ghi da tao ra <path>.import-log.json — dau vao cua undo()."""
    import json

    log_path = f"{path}.import-log.json"
    payload = {
        "source": path,
        "created_leads": res["created_leads"],
        "created_guardians": res["created_guardians"],
        "counts": {k: res[k] for k in
                   ("total", "created", "duplicates", "failed", "notes")},
        "errors": res["errors"],
        "note_review": res.get("note_review") or [],
    }
    try:
        with open(log_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        return log_path
    except OSError as e:
        frappe.log_error(message=str(e), title="import_qlead_excel: khong ghi duoc log")
        return None


def undo(log_path, dry_run=1):
    """
    Go bo dung nhung ban ghi lan chay truoc da tao (doc tu <file>.import-log.json).

        bench --site <site> execute erp.api.crm.import_qlead_excel.undo \
            --kwargs "{'log_path': '/duong/dan/file.xlsx.import-log.json', 'dry_run': 0}"

    Xoa theo thu tu: CRM Lead Note -> CRM Lead Step History -> CRM Lead -> CRM Guardian.
    Guardian nao da duoc ho so KHAC tro toi thi giu lai.
    """
    import json

    dry_run = int(dry_run)
    with open(log_path, encoding="utf-8") as f:
        payload = json.load(f)

    leads = payload.get("created_leads") or []
    guardians = payload.get("created_guardians") or []
    out = {"leads": 0, "notes": 0, "history": 0, "guardians": 0, "guardians_kept": 0}

    for name in leads:
        if not frappe.db.exists("CRM Lead", name):
            continue
        out["notes"] += frappe.db.count("CRM Lead Note", {"lead": name})
        out["history"] += frappe.db.count("CRM Lead Step History", {"lead": name})
        out["leads"] += 1
        if dry_run:
            continue
        frappe.db.delete("CRM Lead Note", {"lead": name})
        frappe.db.delete("CRM Lead Step History", {"lead": name})
        frappe.delete_doc("CRM Lead", name, force=1, ignore_permissions=True,
                          delete_permanently=True)

    for g in guardians:
        if not frappe.db.exists("CRM Guardian", g):
            continue
        still_used = frappe.db.sql(
            """SELECT 1 FROM `tabCRM Lead Guardian` WHERE guardian = %s LIMIT 1""", (g,)
        )
        if still_used:
            out["guardians_kept"] += 1
            continue
        out["guardians"] += 1
        if not dry_run:
            frappe.delete_doc("CRM Guardian", g, force=1, ignore_permissions=True,
                              delete_permanently=True)

    if not dry_run:
        frappe.db.commit()

    print("")
    print("  DRY-RUN undo — khong xoa gi" if dry_run else "  DA XOA")
    print(f"    CRM Lead            : {out['leads']}")
    print(f"    CRM Lead Note       : {out['notes']}")
    print(f"    CRM Lead Step History: {out['history']}")
    print(f"    CRM Guardian        : {out['guardians']} (giu lai vi con ho so khac dung: {out['guardians_kept']})")
    print("")
    return out


def _print(res, dry_run):
    print("")
    print("=" * 62)
    print("  DRY-RUN — khong ghi gi vao DB" if dry_run else "  DA GHI VAO DB")
    print("=" * 62)
    print(f"  Tong dong doc duoc      : {res['total']}")
    print(f"  {'Se tao' if dry_run else 'Da tao':<24}: {res['created']}")
    print(f"  Trung (bo qua)          : {res['duplicates']}")
    print(f"  Loi                     : {res['failed']}")
    print(f"  Ghi chu cham soc (muc)  : {res['notes']}")
    print(f"  CRM Guardian moi / dung lai: {res['guardians_created']} / {res['guardians_reused']}")
    nr = res.get("note_review") or []
    if nr:
        print(f"\n  Ho so co ngay nhat ky DANG NGO ({len(nr)}) — nen review tay:")
        for x in nr[:25]:
            print(f"    dong {x['row']:>5}: {x.get('student','')[:34]:<36} {x['entries']} muc")
        if len(nr) > 25:
            print(f"    … con {len(nr) - 25}")
    if res["warnings"]:
        print(f"\n  Canh bao ({len(res['warnings'])}):")
        for w in res["warnings"][:20]:
            print(f"    - {w}")
        if len(res["warnings"]) > 20:
            print(f"    … con {len(res['warnings']) - 20}")
    if res["errors"]:
        print(f"\n  Dong khong vao duoc ({len(res['errors'])}):")
        for e in res["errors"][:40]:
            print(f"    dong {e['row']:>5}: {e['error']}")
        if len(res["errors"]) > 40:
            print(f"    … con {len(res['errors']) - 40}")
    print("")


# ==================================================================================
# Bo sung ho so gia dinh tu sheet «HSM_ NEW ENROLLED STDS»
# (danh sach hoc sinh moi nhap hoc — nguon chuan cho thong tin bo/me/NGH/anh chi em)
# ==================================================================================

HSM_SHEET = "HSM_ NEW ENROLLED STDS"
HSM_HEADER_ROW = 3

# Cot theo chi so (1-based) — sheet co header gop nhieu tang nen bam vi tri cho chac.
_HSM_STUDENT = 6
_HSM_HEALTH_NOTE = 94                       # «Ghi chú đặc biệt (Sức khoẻ/Tâm lý)»
_HSM_SIBLING_BASES = (77, 80, 83)           # moi khoi: ten / ngay sinh / truong hoc
_HSM_GUARDIANS = (
    # (nhan, cot ten, sdt, cccd, email, nghe nghiep, chuc vu, noi lam viec, quan he)
    ("Bố", 53, 56, 55, 57, 58, 59, 60, None),
    ("Mẹ", 62, 65, 64, 66, 67, 68, 69, None),
    ("Người giám hộ", 72, 75, 74, 76, None, None, None, 73),
)


def _hsm_phone(v):
    """SDT trong HSM: Excel hay cat mat so 0 dau, va co o la '#VALUE!'."""
    if v in (None, ""):
        return ""
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        d = str(int(v))
    else:
        s = str(v)
        if s.strip().startswith("#"):        # #VALUE! / #N/A
            return ""
        d = re.sub(r"\D", "", re.split(r"[/,;\n]", s)[0])
    if d.startswith("84") and len(d) == 11:
        d = "0" + d[2:]
    elif len(d) == 9:
        d = "0" + d
    return d if re.fullmatch(r"0\d{9}", d) else ""


def _name_key(s):
    """Khoa doi chieu ten: bo dau, bo phan trong ngoac (biet danh), gop khoang trang."""
    s = re.sub(r"\(.*?\)", " ", str(s or ""))
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.replace("đ", "d").replace("Đ", "D")
    return " ".join(re.sub(r"[^A-Za-z0-9 ]", " ", s).split()).lower()


def _add_child_contact(doc, table, key_field, value):
    """
    Them mot dong vao bang con lien lac cua CRM Guardian neu chua co.

    Phai tu lam thay vi trong cho CRMGuardian.before_save(): ta save voi
    flags.ignore_validate = True, ma Frappe `return` ngay tai co do trong
    run_before_save_methods() nen ca validate() LAN before_save() bi bo qua —
    _migrate_flat_contacts_if_needed() khong chay, bang con `emails`/`phone_numbers`
    rong, giao dien hien "Chua co du lieu" du field vo huong da co gia tri.
    """
    v = _txt(value)
    if not v:
        return False
    cur = [_txt(r.get(key_field)).lower() for r in (doc.get(table) or [])]
    if v.lower() in cur:
        return False
    doc.append(table, {key_field: v, "is_primary": 0 if cur else 1})
    return True


def _guardian_doc_for(name, phone, cccd, email, job, pos, work, overwrite=0):
    """
    get-or-create CRM Guardian theo SDT (khoa tu nhien, dong bo add_lead_guardian).

    Guardian dung chung giua anh chi em nen khi tim thay ban ghi cu:
      overwrite=0 -> chi dien them field dang trong,
      overwrite=1 -> HSM ghi de.
    """
    # CRM Guardian.phone_number luu dang +84xxxxxxxxx (erp_sis.guardian
    # .validate_vietnamese_phone_number doi 0... -> +84...). File Excel dung 0...,
    # neu tra cuu/ghi bang 0... thi KHONG BAO GIO tim thay ban ghi cu -> lan nao cung
    # tao guardian moi va email/CCCD tren ban ghi that khong duoc cap nhat.
    e164 = normalize_phone_number(phone) or phone
    # Tra theo thu tu: field phang dang +84... -> field phang dang cu '0...' (ban ghi
    # do chinh script nay tao sai dinh dang o lan chay truoc) -> bang con
    # `CRM Guardian Phone` (so thu hai cua chinh nguoi do). Thieu buoc bang con thi
    # phu huynh khai them so se bi tao thanh Guardian thu hai.
    matches = guardian_phone_matches(phone)
    existing = matches[0]["guardian"] if matches else None
    # Chi chuan hoa lai field phang khi khop o CHINH field do va no dang dinh dang cu.
    # Khop qua bang con nghia la e164 chi la so PHU cua nguoi ta -> ghi vao
    # `phone_number` se xoa mat so chinh.
    legacy = bool(existing) and matches[0]["source"] == "flat" \
        and matches[0]["phone"] != e164
    if existing:
        g = frappe.get_doc("CRM Guardian", existing)
        changed = False
        if legacy and _txt(g.phone_number) != e164:
            g.phone_number = e164          # chuan hoa lai ban ghi cu
            changed = True
        for val, fld in ((name, "guardian_name"), (email, "email"), (cccd, "id_number"),
                         (job, "occupation"), (pos, "position"), (work, "workplace")):
            if val and (overwrite or not _txt(g.get(fld))) and _txt(g.get(fld)) != val:
                g.set(fld, val)
                changed = True
        if _add_child_contact(g, "emails", "email_address", email):
            changed = True
        if _add_child_contact(g, "phone_numbers", "phone_number", e164):
            changed = True
        if changed:
            g.flags.ignore_validate = True
            g.flags.ignore_mandatory = True
            g.save(ignore_permissions=True)
        return existing, "reused"
    gid = f"{_slug(name)}-{frappe.generate_hash(length=6)}"
    while frappe.db.exists("CRM Guardian", {"guardian_id": gid}):
        gid = f"{_slug(name)}-{frappe.generate_hash(length=6)}"
    doc = frappe.get_doc({
        "doctype": "CRM Guardian", "guardian_id": gid, "guardian_name": name,
        # GHI e164, khong ghi chuoi '0...' cua Excel: ta save voi ignore_validate nen
        # erp_sis.guardian.validate_vietnamese_phone_number KHONG chay de tu chuan hoa.
        # Ghi dang cu la tu sinh ra dung lop ban ghi ma fix_guardian_phone_duplicates
        # phai di gop lai, va lech voi dong `phone_numbers` ngay ben duoi.
        "phone_number": e164, "email": email or "", "id_number": cccd or "",
        "occupation": job or "", "position": pos or "", "workplace": work or "",
    })
    _add_child_contact(doc, "emails", "email_address", email)
    _add_child_contact(doc, "phone_numbers", "phone_number", e164)
    doc.flags.ignore_validate = True
    doc.flags.ignore_mandatory = True
    doc.insert(ignore_permissions=True)
    return doc.name, "created"


def backfill_hsm(path, sheet=None, dry_run=1, status="Dong phi", overwrite=0,
                 commit_every=50):
    """
    Dong bo thong tin gia dinh tu sheet HSM vao ho so CRM da co.

        bench --site <site> execute erp.api.crm.import_qlead_excel.backfill_hsm \
            --kwargs "{'path': '/private/files/....xlsx'}"

    Cong them:
      - Bo / Me / Nguoi giam ho CO SDT  -> CRM Guardian + dong trong `lead_guardians`
      - SDT cua ho                      -> them vao `phone_numbers` neu chua co
      - Me (khong co thi Bo)            -> khoi guardian_* phang
      - Anh chi em                      -> `lead_siblings` (gop theo ten)
      - Ghi chu suc khoe/tam ly         -> `student_health_notes`
      - NGUOI THIEU SDT                 -> KHONG tao Guardian; ten + thong tin duoc
                                           ghi noi tiep vao `student_note`

    overwrite  0 (mac dinh) = CRM thang: field nao CRM da co gia tri thi giu nguyen,
                 HSM chi dien vao cho dang trong.
               1 = HSM thang: ghi de khoi guardian_* phang, ghi chu suc khoe, va
                 cap nhat lai document CRM Guardian dung chung khi HSM co du lieu moi.

    Du overwrite=1 van KHONG BAO GIO xoa dong: `lead_guardians`, `phone_numbers`,
    `lead_siblings` chi duoc them; anh chi em / phu huynh chi co o CRM ma HSM khong
    co thi giu nguyen.

    Doi chieu hoc sinh theo ten da chuan hoa, trong pham vi buoc QLead + status
    truyen vao (mac dinh «Dong phi») nen khong lan sang ho so khac.
    """
    import openpyxl

    dry_run = int(dry_run)
    ow = int(overwrite)
    commit_every = int(commit_every) or 50
    path = _resolve_path(path)
    print(f"  File: {path}")
    print(f"  overwrite={ow} ({'HSM thang' if ow else 'CRM thang — chi dien cho trong'})")

    wb = openpyxl.load_workbook(path, data_only=True)
    sname = sheet or HSM_SHEET
    if sname not in wb.sheetnames:
        frappe.throw(f"Khong co sheet {sname!r}. Co: {wb.sheetnames}")
    ws = wb[sname]

    # index ho so dich
    leads = frappe.get_all("CRM Lead", filters={"step": STEP, "status": status},
                           fields=["name", "student_name", "crm_code"],
                           limit_page_length=0)
    idx = {}
    for l in leads:
        idx.setdefault(_name_key(l["student_name"]), []).append(l)
    print(f"  Ho so {STEP}/{status}: {len(leads)}")

    res = {"hsm_rows": 0, "matched": 0, "not_found": [], "ambiguous": [],
           "guardians_created": 0, "guardians_reused": 0, "guardian_rows": 0,
           "phones_added": 0, "siblings": 0, "health_notes": 0,
           "no_phone_noted": 0, "flat_filled": 0, "updated": 0, "errors": []}

    for r in range(HSM_HEADER_ROW + 1, ws.max_row + 1):
        sn = _txt(ws.cell(r, _HSM_STUDENT).value)
        if not sn:
            continue
        res["hsm_rows"] += 1
        hit = idx.get(_name_key(sn)) or []
        if not hit:
            res["not_found"].append({"row": r, "student": sn})
            continue
        if len(hit) > 1:
            res["ambiguous"].append({"row": r, "student": sn,
                                     "leads": [x["crm_code"] for x in hit]})
            continue
        lead_name = hit[0]["name"]
        res["matched"] += 1

        # ---- gom nguoi
        with_phone, without_phone = [], []
        for lbl, c_name, c_tel, c_cccd, c_mail, c_job, c_pos, c_work, c_rel in _HSM_GUARDIANS:
            nm = _txt(ws.cell(r, c_name).value)
            if not nm:
                continue
            rel = _txt(ws.cell(r, c_rel).value) if c_rel else lbl
            info = {
                "label": lbl, "rel": normalize_relationship(rel or lbl), "name": nm,
                "phone": _hsm_phone(ws.cell(r, c_tel).value),
                "cccd": _txt(ws.cell(r, c_cccd).value) if c_cccd else "",
                "email": (_emails(ws.cell(r, c_mail).value) or [""])[0] if c_mail else "",
                "job": _txt(ws.cell(r, c_job).value) if c_job else "",
                "pos": _txt(ws.cell(r, c_pos).value) if c_pos else "",
                "work": _txt(ws.cell(r, c_work).value) if c_work else "",
            }
            (with_phone if info["phone"] else without_phone).append(info)

        siblings = []
        for base in _HSM_SIBLING_BASES:
            snm = _txt(ws.cell(r, base).value)
            if not snm:
                continue
            siblings.append({
                "sibling_name": snm,
                "dob": _date(ws.cell(r, base + 1).value),
                "school": _txt(ws.cell(r, base + 2).value),
            })
        health = _txt(ws.cell(r, _HSM_HEALTH_NOTE).value)

        if dry_run:
            res["guardian_rows"] += len(with_phone)
            res["no_phone_noted"] += len(without_phone)
            res["siblings"] += len(siblings)
            res["health_notes"] += 1 if health else 0
            res["updated"] += 1
            continue

        sp = f"hsm_{r}"
        try:
            frappe.db.savepoint(sp)
            doc = frappe.get_doc("CRM Lead", lead_name)
            touched = False

            linked = {g.guardian for g in (doc.lead_guardians or [])}
            have_phones = {(p.phone_number or "").strip()
                           for p in (doc.phone_numbers or [])}
            for i, g in enumerate(with_phone):
                gname, how = _guardian_doc_for(g["name"], g["phone"], g["cccd"],
                                               g["email"], g["job"], g["pos"], g["work"],
                                               overwrite=ow)
                res["guardians_created" if how == "created" else "guardians_reused"] += 1
                if gname not in linked:
                    doc.append("lead_guardians", {
                        "guardian": gname, "relationship_type": g["rel"],
                        "is_primary_contact": 0, "display_order": i + 1,
                    })
                    linked.add(gname)
                    res["guardian_rows"] += 1
                    touched = True
                e164 = normalize_phone_number(g["phone"])
                if e164 and e164 not in have_phones:
                    doc.append("phone_numbers", {"phone_number": e164, "is_primary": 0})
                    have_phones.add(e164)
                    res["phones_added"] += 1
                    touched = True

            # khoi guardian_* phang: uu tien Me, khong co thi Bo
            # Chot chan o CAP FIELD (xem ghi chu trong backfill_family).
            if with_phone:
                pick = next((g for g in with_phone if g["label"] == "Mẹ"), with_phone[0])
                wrote = False
                for val, fld in (
                    (pick["name"], "guardian_name"), (pick["rel"], "relationship"),
                    (pick["email"], "guardian_email"), (pick["cccd"], "guardian_id_number"),
                    (pick["job"], "guardian_occupation"), (pick["pos"], "guardian_position"),
                    (pick["work"], "guardian_workplace"),
                ):
                    if val and (ow or not _txt(doc.get(fld))) and _txt(doc.get(fld)) != val:
                        doc.set(fld, val)
                        wrote = True
                if wrote:
                    res["flat_filled"] += 1
                    touched = True

            # Gop theo TEN thay vi bo qua ca bang khi da co dong: giu nguyen anh chi em
            # CRM dang co, dong thoi khong danh roi nguoi ma HSM co them.
            have_sib = {_name_key(s.sibling_name): s for s in (doc.lead_siblings or [])}
            for s in siblings:
                k = _name_key(s["sibling_name"])
                row = have_sib.get(k)
                if row is not None:
                    # da co dong nay: chi bo sung ngay sinh / truong con thieu
                    # (hoac ghi de khi overwrite) — khong bao gio xoa dong.
                    for fld in ("dob", "school"):
                        v = s.get(fld)
                        if v and (ow or not _txt(row.get(fld))) and _txt(row.get(fld)) != v:
                            row.set(fld, v)
                            touched = True
                    continue
                doc.append("lead_siblings", s)
                have_sib[k] = doc.lead_siblings[-1]
                res["siblings"] += 1
                touched = True

            if health and (ow or not _txt(doc.student_health_notes)) \
                    and _txt(doc.student_health_notes) != health:
                doc.student_health_notes = health
                res["health_notes"] += 1
                touched = True

            if without_phone:
                lines = ["[HSM] Người thân KHÔNG có số điện thoại — chưa tạo hồ sơ phụ huynh:"]
                for g in without_phone:
                    bits = [f"{g['rel']}: {g['name']}"]
                    for lbl2, v in (("CCCD", g["cccd"]), ("email", g["email"]),
                                    ("nghề", g["job"]), ("nơi làm việc", g["work"])):
                        if v:
                            bits.append(f"{lbl2} {v}")
                    lines.append("  - " + " | ".join(bits))
                block = "\n".join(lines)
                cur = _txt(doc.student_note)
                if "[HSM] Người thân KHÔNG có số điện thoại" not in cur:
                    doc.student_note = (cur + "\n\n" + block).strip() if cur else block
                    res["no_phone_noted"] += len(without_phone)
                    touched = True

            if touched:
                doc.flags.ignore_mandatory = True
                doc.save(ignore_permissions=True)
                res["updated"] += 1
        except Exception as e:
            try:
                frappe.db.rollback(save_point=sp)
            except Exception:
                pass
            res["errors"].append({"row": r, "student": sn, "error": str(e)[:250]})
            frappe.log_error(message=frappe.get_traceback() or str(e),
                             title=f"backfill_hsm dong {r}")

        if res["updated"] and res["updated"] % commit_every == 0:
            frappe.db.commit()

    if not dry_run:
        frappe.db.commit()

    print("")
    print("=" * 62)
    print("  DRY-RUN — khong ghi gi vao DB" if dry_run else "  DA GHI VAO DB")
    print("=" * 62)
    print(f"  Dong HSM doc duoc       : {res['hsm_rows']}")
    print(f"  Khop ho so              : {res['matched']}")
    print(f"  Khong tim thay ho so    : {len(res['not_found'])}")
    print(f"  Ten trung (bo qua)      : {len(res['ambiguous'])}")
    print(f"  {'Se cap nhat' if dry_run else 'Da cap nhat':<23} : {res['updated']}")
    print(f"  Dong lead_guardians     : {res['guardian_rows']}")
    if not dry_run:
        print(f"  CRM Guardian moi/dung lai: {res['guardians_created']} / {res['guardians_reused']}")
        print(f"  SDT them vao ho so      : {res['phones_added']}")
        print(f"  Dien khoi guardian phang: {res['flat_filled']}")
    print(f"  Dong anh chi em         : {res['siblings']}")
    print(f"  Ghi chu suc khoe        : {res['health_notes']}")
    print(f"  Nguoi THIEU SDT -> ghi chu: {res['no_phone_noted']}")
    for lbl, key_ in (("Khong tim thay", "not_found"), ("Ten trung", "ambiguous")):
        if res[key_]:
            print(f"\n  {lbl}:")
            for x in res[key_][:20]:
                print(f"    HSM r{x['row']:>4} {x['student'][:40]}")
    if res["errors"]:
        print(f"\n  Loi ({len(res['errors'])}):")
        for e in res["errors"][:20]:
            print(f"    r{e['row']:>4} {e['student'][:28]:<30} {e['error']}")
    print("")
    return res


# ==================================================================================
# Bo sung ho so gia dinh tu CHINH file mau (sheet QLead) — dung khi du lieu HSM da
# duoc ghep san vao file bang cac cot PH1/PH2/PH3, ACE1-3, «Ghi chú sức khoẻ/tâm lý».
# ==================================================================================

_FAM_GUARDIAN_SLOTS = (
    (1, "PH1: SĐT chính *"),
    (2, "PH2: SĐT chính"),
    (3, "PH3: SĐT chính"),
)
# Bac cua buoc — trung ten thi uu tien ho so da di xa nhat trong pipeline.
_STEP_RANK = {s: i for i, s in enumerate(
    ["Draft", "Verify", "Lead", "QLead", "Enrolled", "Nghi hoc"])}


def _fam_cell(ws, r, J, header):
    j = J.get(header)
    return _txt(ws.cell(r, j).value) if j else ""


def _find_lead_flexible(phones, student_name, cache):
    """
    Tim ho so bang BAT KY so nao tren dong + ten hoc sinh; khong duoc thi doi chieu
    rieng theo ten.

    KHONG loc theo `step`: ho so «Nop phi» sau khi len Hoc sinh chinh thuc se doi
    sang step Enrolled, loc cung se tra ve rong. Trung ten thi lay ho so o buoc xa
    nhat (Enrolled > QLead > ...), thay vi bo qua.
    """
    key = " ".join((student_name or "").split()).lower()
    if not key:
        return None, "thieu ten hoc sinh"

    for p in phones:
        rows = frappe.db.sql(
            """
            SELECT DISTINCT cl.name, cl.step FROM `tabCRM Lead` cl
            INNER JOIN `tabCRM Lead Phone` clp ON clp.parent = cl.name
            WHERE clp.phone_number = %s
              AND LOWER(TRIM(IFNULL(cl.student_name,''))) = %s
            """,
            (normalize_phone_number(p), key), as_dict=True)
        if len(rows) == 1:
            return rows[0]["name"], None
        if len(rows) > 1:
            rows.sort(key=lambda x: _STEP_RANK.get(x["step"], -1), reverse=True)
            return rows[0]["name"], None

    if key not in cache:
        cache[key] = frappe.db.sql(
            """
            SELECT name, step FROM `tabCRM Lead`
            WHERE LOWER(TRIM(IFNULL(student_name,''))) = %s
            """, (key,), as_dict=True)
    rows = cache[key]
    if not rows:
        return None, "khong tim thay ho so"
    rows.sort(key=lambda x: _STEP_RANK.get(x["step"], -1), reverse=True)
    return rows[0]["name"], None


def backfill_family(path, dry_run=1, overwrite=0, commit_every=50):
    """
    Doc sheet «QLead» cua file mau v3 va bo sung ho so gia dinh cho lead da co.

        bench --site <site> execute erp.api.crm.import_qlead_excel.backfill_family \
            --kwargs "{'path': '/private/files/....xlsx'}"

    Doc: PH1/PH2/PH3 (ho ten + SDT chinh/phu + email/CCCD/nghe/chuc vu/noi cong tac),
    ACE1-3 (anh chi em), «Ghi chú sức khoẻ/tâm lý».

    Ghi (khong bao gio xoa dong):
      - moi khoi PH co ten + SDT -> CRM Guardian + dong `lead_guardians`
      - moi SDT tren dong        -> `phone_numbers` neu chua co
      - PH1                      -> khoi guardian_* phang
      - ACE                      -> `lead_siblings` (gop theo ten)
      - ghi chu SK               -> `student_health_notes`

    overwrite  0 = CRM thang (chi dien cho trong) | 1 = file thang.

    Doi chieu qua `_find_lead_flexible`: thu moi SDT tren dong, khong loc theo buoc.
    """
    import openpyxl

    dry_run, ow = int(dry_run), int(overwrite)
    commit_every = int(commit_every) or 50
    path = _resolve_path(path)
    print(f"  File: {path}")
    print(f"  overwrite={ow} ({'file thang' if ow else 'CRM thang'})")

    ws = openpyxl.load_workbook(path, data_only=True)["QLead"]
    J = {_txt(ws.cell(HEADER_ROW, j).value): j for j in range(1, ws.max_column + 1)}
    missing_cols = [h for _, h in _FAM_GUARDIAN_SLOTS if h not in J]
    if missing_cols:
        frappe.throw(f"File thieu cot: {missing_cols}")

    res = {"rows": 0, "matched": 0, "not_found": [], "updated": 0,
           "guardian_rows": 0, "guardians_created": 0, "guardians_reused": 0,
           "phones_added": 0, "siblings": 0, "health_notes": 0, "errors": []}
    name_cache = {}

    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        sn = _fam_cell(ws, r, J, "Họ tên học sinh *")
        if not sn or sn.startswith(EXAMPLE_PREFIX):
            continue
        res["rows"] += 1

        phones = []
        for slot, main_h in _FAM_GUARDIAN_SLOTS:
            for h in (main_h, f"PH{slot}: SĐT phụ"):
                v = _fam_cell(ws, r, J, h)
                if v and validate_phone_number(v) and v not in phones:
                    phones.append(v)

        lead, err = _find_lead_flexible(phones, sn, name_cache)
        if not lead:
            res["not_found"].append({"row": r, "student": sn, "error": err})
            continue
        res["matched"] += 1

        people = []
        for slot, main_h in _FAM_GUARDIAN_SLOTS:
            nm = _fam_cell(ws, r, J, f"PH{slot}: Họ tên phụ huynh")
            ph = _fam_cell(ws, r, J, main_h)
            if not nm or not ph or not validate_phone_number(ph):
                continue
            people.append({
                "slot": slot, "name": nm, "phone": ph,
                "rel": normalize_relationship(_fam_cell(ws, r, J, f"PH{slot}: Quan hệ với HS")),
                "email": (_emails(_fam_cell(ws, r, J, f"PH{slot}: Email")) or [""])[0],
                "cccd": _fam_cell(ws, r, J, f"PH{slot}: Số CCCD/Hộ chiếu"),
                "job": _fam_cell(ws, r, J, f"PH{slot}: Nghề nghiệp"),
                "pos": _fam_cell(ws, r, J, f"PH{slot}: Chức vụ"),
                "work": _fam_cell(ws, r, J, f"PH{slot}: Nơi công tác"),
            })
        sibs = []
        for i in (1, 2, 3):
            nm = _fam_cell(ws, r, J, f"ACE{i}: Họ và tên")
            if nm:
                sibs.append({
                    "sibling_name": nm,
                    "dob": _date(_fam_cell(ws, r, J, f"ACE{i}: Ngày sinh (dd/mm/yyyy)")),
                    "school": _fam_cell(ws, r, J, f"ACE{i}: Trường học"),
                })
        health = _fam_cell(ws, r, J, "Ghi chú sức khoẻ/tâm lý")

        if dry_run:
            res["guardian_rows"] += len(people)
            res["siblings"] += len(sibs)
            res["health_notes"] += 1 if health else 0
            res["updated"] += 1
            continue

        sp = f"fam_{r}"
        try:
            frappe.db.savepoint(sp)
            doc = frappe.get_doc("CRM Lead", lead)
            touched = False

            linked = {g.guardian for g in (doc.lead_guardians or [])}
            have = {(p.phone_number or "").strip() for p in (doc.phone_numbers or [])}
            for g in people:
                gname, how = _guardian_doc_for(g["name"], g["phone"], g["cccd"], g["email"],
                                               g["job"], g["pos"], g["work"], overwrite=ow)
                res["guardians_created" if how == "created" else "guardians_reused"] += 1
                if gname not in linked:
                    doc.append("lead_guardians", {
                        "guardian": gname, "relationship_type": g["rel"],
                        "is_primary_contact": 1 if g["slot"] == 1 else 0,
                        "display_order": g["slot"]})
                    linked.add(gname)
                    res["guardian_rows"] += 1
                    touched = True
            for p in phones:
                e164 = normalize_phone_number(p)
                if e164 and e164 not in have:
                    doc.append("phone_numbers", {"phone_number": e164, "is_primary": 0})
                    have.add(e164)
                    res["phones_added"] += 1
                    touched = True

            p1 = next((g for g in people if g["slot"] == 1), None)
            # Chot chan «CRM thang» phai o CAP FIELD, khong phai cap khoi: neu gac ca
            # khoi bang guardian_name thi ho so nao da co ten se khong bao gio duoc
            # dien guardian_email/CCCD/nghe nghiep dang trong.
            if p1:
                for val, fld in ((p1["name"], "guardian_name"), (p1["rel"], "relationship"),
                                 (p1["email"], "guardian_email"),
                                 (p1["cccd"], "guardian_id_number"),
                                 (p1["job"], "guardian_occupation"),
                                 (p1["pos"], "guardian_position"),
                                 (p1["work"], "guardian_workplace")):
                    if val and (ow or not _txt(doc.get(fld))) and _txt(doc.get(fld)) != val:
                        doc.set(fld, val)
                        touched = True

            have_sib = {_name_key(s.sibling_name): s for s in (doc.lead_siblings or [])}
            for s in sibs:
                k = _name_key(s["sibling_name"])
                if k in have_sib:
                    row_ = have_sib[k]
                    for fld in ("dob", "school"):
                        v = s.get(fld)
                        if v and (ow or not _txt(row_.get(fld))) and _txt(row_.get(fld)) != v:
                            row_.set(fld, v)
                            touched = True
                    continue
                doc.append("lead_siblings", s)
                have_sib[k] = doc.lead_siblings[-1]
                res["siblings"] += 1
                touched = True

            if health and (ow or not _txt(doc.student_health_notes)) \
                    and _txt(doc.student_health_notes) != health:
                doc.student_health_notes = health
                res["health_notes"] += 1
                touched = True

            if touched:
                doc.flags.ignore_mandatory = True
                doc.save(ignore_permissions=True)
                res["updated"] += 1
        except Exception as e:
            try:
                frappe.db.rollback(save_point=sp)
            except Exception:
                pass
            res["errors"].append({"row": r, "student": sn, "error": str(e)[:250]})
            frappe.log_error(message=frappe.get_traceback() or str(e),
                             title=f"backfill_family dong {r}")

        if res["updated"] and res["updated"] % commit_every == 0:
            frappe.db.commit()

    if not dry_run:
        frappe.db.commit()

    print("")
    print("=" * 62)
    print("  DRY-RUN — khong ghi gi vao DB" if dry_run else "  DA GHI VAO DB")
    print("=" * 62)
    print(f"  Dong doc duoc        : {res['rows']}")
    print(f"  Khop ho so           : {res['matched']}")
    print(f"  Khong tim thay       : {len(res['not_found'])}")
    print(f"  {'Se cap nhat' if dry_run else 'Da cap nhat':<20} : {res['updated']}")
    print(f"  Dong lead_guardians  : {res['guardian_rows']}")
    if not dry_run:
        print(f"  Guardian moi/dung lai: {res['guardians_created']} / {res['guardians_reused']}")
        print(f"  SDT them vao         : {res['phones_added']}")
    print(f"  Dong anh chi em      : {res['siblings']}")
    print(f"  Ghi chu suc khoe     : {res['health_notes']}")
    for x in res["not_found"][:20]:
        print(f"    r{x['row']:>4} {x['student'][:34]:<36} {x['error']}")
    if res["errors"]:
        print(f"\n  Loi ({len(res['errors'])}):")
        for e in res["errors"][:20]:
            print(f"    r{e['row']:>4} {e['student'][:28]:<30} {e['error']}")
    print("")
    return res


# ==================================================================================
# Don du lieu: guardian bi nhan doi do bug dinh dang SDT
# ==================================================================================

# Moi cho trong he thong tro tro toi CRM Guardian — phai chuyen het truoc khi xoa
# ban ghi trung, neu khong se de lai Link mo coi.
_GUARDIAN_REFS = (
    ("CRM Family Relationship", "guardian"),
    ("CRM Info Confirmation Log", "guardian"),
    ("CRM Issue", "guardian"),
    ("CRM Issue Guardian", "guardian"),
    ("CRM Lead", "info_confirmed_by"),
    ("CRM Lead Guardian", "guardian"),
    ("FaceID Person", "crm_guardian"),
    ("Feedback", "guardian"),
    ("Portal API Error", "guardian"),
    ("Portal Guardian Activity", "guardian"),
    ("SIS Re-enrollment", "guardian_id"),
    ("SIS Scholarship Application", "guardian_id"),
    ("SIS Student Leave Request", "parent_id"),
)


def fix_guardian_phone_duplicates(dry_run=1, commit_every=50):
    """
    Gop cac CRM Guardian bi nhan doi vi SDT luu sai dinh dang.

        bench --site <site> execute \
            erp.api.crm.import_qlead_excel.fix_guardian_phone_duplicates \
            --kwargs "{'dry_run': 0}"

    Lan chay dau cua backfill tra cuu guardian bang chuoi '0xxxxxxxxx' trong khi he
    thong luu '+84xxxxxxxxx', nen da tao ban ghi moi cho nguoi da co san. Ket qua:
    hai CRM Guardian cung mot nguoi, va ho so co hai dong `lead_guardians`.

    Voi moi ban ghi con SDT dang '0...':
      - co ban ghi '+84...' tuong ung -> chuyen moi tham chieu sang ban ghi do,
        bo sung field con trong cho no, roi XOA ban ghi '0...',
      - khong co -> chi chuan hoa phone_number ve '+84...'.
    Cuoi cung don cac dong `lead_guardians` trung (cung ho so + cung guardian).
    """
    dry_run = int(dry_run)
    commit_every = int(commit_every) or 50

    olds = frappe.db.sql(
        """SELECT name, guardian_name, phone_number, email, id_number, occupation,
                  position, workplace
           FROM `tabCRM Guardian` WHERE phone_number LIKE '0%'""", as_dict=True)

    res = {"scanned": len(olds), "merged": 0, "normalized": 0, "refs_moved": 0,
           "dup_rows_removed": 0, "details": [], "errors": []}
    print(f"  Guardian con SDT dang '0...': {len(olds)}")

    for i, o in enumerate(olds, start=1):
        e164 = normalize_phone_number(o["phone_number"])
        # CO Y khong dung find_guardian_by_phone o day, du no la ham dedup dung chung:
        #   - phai so sanh HAI DINH DANG voi nhau, ham kia coi '0...' va '+84...' la
        #     mot nen se tu khop chinh no,
        #   - ham kia con tra bang con `CRM Guardian Phone`; o day keeper bi GOP VA
        #     XOA VINH VIEN, ma khop bang con chi chung to so nay la so PHU cua nguoi
        #     khac -> gop hai nguoi khac nhau. Chi field phang moi la danh tinh.
        keeper = frappe.db.get_value(
            "CRM Guardian", {"phone_number": e164, "name": ["!=", o["name"]]}, "name")

        if not keeper:
            res["normalized"] += 1
            res["details"].append(
                {"action": "normalize", "guardian": o["name"],
                 "name": o["guardian_name"], "phone": f'{o["phone_number"]} -> {e164}'})
            if not dry_run:
                frappe.db.set_value("CRM Guardian", o["name"], "phone_number", e164,
                                    update_modified=False)
            continue

        moved = 0
        for dt, fld in _GUARDIAN_REFS:
            try:
                moved += frappe.db.count(dt, {fld: o["name"]})
            except Exception:
                continue
        res["merged"] += 1
        res["refs_moved"] += moved
        res["details"].append(
            {"action": "merge", "from": o["name"], "into": keeper,
             "name": o["guardian_name"], "phone": e164, "refs": moved})
        if dry_run:
            continue

        sp = f"gmerge_{i}"
        try:
            frappe.db.savepoint(sp)
            # bo sung field con trong cho ban ghi giu lai
            k = frappe.get_doc("CRM Guardian", keeper)
            ch = False
            for src, fld in (("guardian_name", "guardian_name"), ("email", "email"),
                             ("id_number", "id_number"), ("occupation", "occupation"),
                             ("position", "position"), ("workplace", "workplace")):
                v = _txt(o.get(src))
                if v and not _txt(k.get(fld)):
                    k.set(fld, v)
                    ch = True
            if ch:
                k.flags.ignore_validate = True
                k.flags.ignore_mandatory = True
                k.save(ignore_permissions=True)

            for dt, fld in _GUARDIAN_REFS:
                try:
                    frappe.db.sql(
                        f"UPDATE `tab{dt}` SET `{fld}` = %s WHERE `{fld}` = %s",
                        (keeper, o["name"]))
                except Exception:
                    continue
            frappe.delete_doc("CRM Guardian", o["name"], force=1,
                              ignore_permissions=True, delete_permanently=True)
        except Exception as e:
            try:
                frappe.db.rollback(save_point=sp)
            except Exception:
                pass
            res["merged"] -= 1
            res["errors"].append({"guardian": o["name"], "error": str(e)[:250]})
            frappe.log_error(message=frappe.get_traceback() or str(e),
                             title=f"fix_guardian_phone_duplicates {o['name']}")

        if res["merged"] and res["merged"] % commit_every == 0:
            frappe.db.commit()

    # --- don dong lead_guardians trung (cung parent + cung guardian)
    dups = frappe.db.sql(
        """SELECT parent, guardian, COUNT(*) n, MIN(name) keep
           FROM `tabCRM Lead Guardian`
           GROUP BY parent, guardian HAVING n > 1""", as_dict=True)
    res["dup_rows_removed"] = sum(d["n"] - 1 for d in dups)
    if not dry_run:
        for d in dups:
            frappe.db.sql(
                """DELETE FROM `tabCRM Lead Guardian`
                   WHERE parent = %s AND guardian = %s AND name != %s""",
                (d["parent"], d["guardian"], d["keep"]))
        frappe.db.commit()

    print("")
    print("=" * 62)
    print("  DRY-RUN — khong ghi gi vao DB" if dry_run else "  DA GHI VAO DB")
    print("=" * 62)
    print(f"  Guardian quet             : {res['scanned']}")
    print(f"  {'Se gop' if dry_run else 'Da gop':<25} : {res['merged']}")
    print(f"  Chi chuan hoa SDT         : {res['normalized']}")
    print(f"  Tham chieu chuyen sang    : {res['refs_moved']}")
    print(f"  Dong lead_guardians trung : {res['dup_rows_removed']}")
    for d in res["details"][:30]:
        if d["action"] == "merge":
            print(f"    GOP  {d['name'][:26]:<28} {d['phone']:<16} {d['from']} -> {d['into']} ({d['refs']} ref)")
        else:
            print(f"    CHUAN {d['name'][:26]:<28} {d['phone']}")
    if len(res["details"]) > 30:
        print(f"    … con {len(res['details']) - 30}")
    if res["errors"]:
        print(f"\n  Loi ({len(res['errors'])}):")
        for e in res["errors"][:20]:
            print(f"    {e['guardian']}: {e['error']}")
    print("")
    return res


# ==================================================================================
# Doc file «DANH SÁCH HỌC SINH XẾP LỚP» (TSxTiH) — cap nhat lien lac phu huynh.
#
# Doi chieu bang MA HOC SINH (cot ID / WS...). Tim PH uu tien quan he CRM Family
# (Me/Bo/NGH), fallback SDT. Ghi ten, SDT, email (+ nghe nghiep/dia chi neu co).
# KHONG tao CRM Guardian moi, KHONG them quan he gia dinh moi.
# ==================================================================================

TIH_SHEET = "1. TỔNG HS ĐÓNG PHÍ_ENROLLED  S"
TIH_HEADER_ROW = 3

# tieu de (da gop khoang trang) -> khoa noi bo. Dat truoc cac nhan de sau vi
# «Email Bố & Mẹ» phai duoc phan biet voi «Email Bố».
_TIH_HEADERS = {
    "id": ("ID",),
    "student": ("HỌ VÀ TÊN HỌC SINH/ Sts Name",),
    "gender": ("Giới tính| Gender",),
    "dob": ("Ngày sinh| D.O.B",),
    "cur_grade": ("Khối Grade 25-26",),
    "cur_school": ("Trường School 25-26",),
    "m_name": ("Họ tên Mẹ| Mother",),
    "m_phone": ("SĐT Mẹ| Mobile",),
    "m_email": ("Email mẹ",),
    "m_job": ("Nghề nghiệp Mẹ",),
    "f_name": ("Họ & Tên Bố| Father",),
    "f_phone": ("SĐT Bố| Mobile",),
    "f_email": ("Email Bố",),
    "f_job": ("Nghề nghiệp Bố",),
    "g_name": ("Thông tin Người giám hộ| Guardian",),
    "g_rel": ("Quan hệ với HS| Relationship",),
    "g_phone": ("SĐT Người Giám Hộ| Mobile",),
    "addr": ("Địa chỉ nhà/Khu vực",),
    "district": ("Quận/huyện",),
    "province": ("Tỉnh/Thành phố",),
    "both_email": ("Email Bố & Mẹ",),
}


def _tih_resolve_cols(ws):
    """Tra cot theo tieu de dong 3 — ba sheet trong file co bo cot lech nhau."""
    seen = {}
    for j in range(1, ws.max_column + 1):
        h = " ".join(str(ws.cell(TIH_HEADER_ROW, j).value or "").split())
        if h and h not in seen:
            seen[h] = j
    out = {}
    for key, names in _TIH_HEADERS.items():
        for n in names:
            if n in seen:
                out[key] = seen[n]
                break
    return out, seen


# Me / Bo / NGH — quan he mac dinh tren CRM Family Relationship
_TIH_ROLES = (
    ("m", "m_name", "m_phone", "m_email", "m_job", ("mother",)),
    ("f", "f_name", "f_phone", "f_email", "f_job", ("father",)),
    ("g", "g_name", "g_phone", None, None, None),
)


def _is_skip(v):
    """Bo qua o trong hoac ghi '0' (placeholder Excel)."""
    if v is None:
        return True
    if isinstance(v, (int, float)) and v == 0:
        return True
    s = _txt(v)
    return not s or s == "0"


def _phones(v):
    """
    Tu o SDT -> (list +84... hop le, list chuoi loi).
    Mot o co the chua nhieu SDT (phay, xuong dong, ...).
    """
    if _is_skip(v):
        return [], []
    if isinstance(v, float) and v.is_integer():
        raw = str(int(v))
    else:
        raw = _txt(v)
    valid, invalid, seen = [], [], set()
    chunks = re.split(r"[\s,;/|]+", raw)
    candidates = []
    for chunk in chunks:
        chunk = chunk.strip()
        if not chunk or chunk == "0":
            continue
        if validate_phone_number(chunk):
            candidates.append(chunk)
            continue
        for m in re.findall(r"\d{8,11}", chunk):
            candidates.append(m)
    for cand in candidates:
        if not validate_phone_number(cand):
            if cand not in invalid:
                invalid.append(cand)
            continue
        norm = normalize_phone_number(cand)
        if norm and norm not in seen:
            seen.add(norm)
            valid.append(norm)
    return valid, invalid


def _tih_find_guardian_for_student(student_id, rel_codes, phones):
    """
    Tim CRM Guardian theo quan he tren CRM Family Relationship, fallback SDT.
    Tra (docname|None, cach_khop, rel_row|None).
    """
    rel_codes = [c for c in (rel_codes or []) if c]
    rows = []
    if student_id and rel_codes:
        rows = frappe.get_all(
            "CRM Family Relationship",
            filters={"student": student_id, "relationship_type": ["in", rel_codes]},
            fields=["name", "guardian", "relationship_type"],
            limit_page_length=0,
        ) or []
    if len(rows) == 1:
        return rows[0]["guardian"], "relationship", rows[0]
    if len(rows) > 1:
        if phones:
            for row in rows:
                for ph in phones:
                    matches = guardian_phone_matches(ph)
                    if any(m["guardian"] == row["guardian"] for m in matches):
                        return row["guardian"], "relationship_phone", row
        return rows[0]["guardian"], "relationship_first", rows[0]
    for ph in phones or []:
        gname = find_guardian_by_phone(ph)
        if gname:
            return gname, "phone_fallback", None
    return None, "missing", None


def _tih_would_update_field(cur, new, overwrite):
    """Co ghi field vo huong khong (overwrite hoac dang trong)."""
    new = _txt(new)
    if not new:
        return False
    return overwrite or not _txt(cur)


def _tih_preview_guardian_update(gname, person, overwrite, counters):
    """Dem thay doi khi dry_run."""
    cur = frappe.db.get_value(
        "CRM Guardian", gname,
        ["guardian_name", "phone_number", "email", "occupation", "address"],
        as_dict=True,
    ) or {}
    if _tih_would_update_field(cur.get("guardian_name"), person["name"], overwrite):
        counters["set_name"] += 1
    phones = person.get("phones") or []
    if phones and _tih_would_update_field(cur.get("phone_number"), phones[0], overwrite):
        counters["set_phone"] += 1
    if person.get("emails") and _tih_would_update_field(cur.get("email"), person["emails"][0], overwrite):
        counters["set_email"] += 1
    if person.get("job") and _tih_would_update_field(cur.get("occupation"), person["job"], overwrite):
        counters["set_job"] += 1
    if person.get("addr") and _tih_would_update_field(cur.get("address"), person["addr"], overwrite):
        counters["set_addr"] += 1


def _apply_guardian_contact_update(doc, person, overwrite, counters):
    """Cap nhat ten / SDT / email (+ child tables) len CRM Guardian."""
    changed = False
    name = person.get("name") or ""
    phones = person.get("phones") or []
    emails = person.get("emails") or []

    if name and _tih_would_update_field(doc.guardian_name, name, overwrite) \
            and _txt(doc.guardian_name) != name:
        doc.guardian_name = name
        counters["set_name"] += 1
        changed = True

    if phones:
        primary = phones[0]
        if _tih_would_update_field(doc.phone_number, primary, overwrite) \
                and _txt(doc.phone_number) != primary:
            doc.phone_number = primary
            counters["set_phone"] += 1
            changed = True
        for ph in phones:
            if _add_child_contact(doc, "phone_numbers", "phone_number", ph):
                counters["child_phone"] += 1
                changed = True

    if emails:
        primary = emails[0]
        if _tih_would_update_field(doc.email, primary, overwrite) \
                and _txt(doc.email) != primary:
            doc.email = primary
            counters["set_email"] += 1
            changed = True
        for em in emails:
            if _add_child_contact(doc, "emails", "email_address", em):
                counters["child_email"] += 1
                changed = True

    for val, fld, counter in (
        (person.get("job"), "occupation", "set_job"),
        (person.get("addr"), "address", "set_addr"),
    ):
        if val and _tih_would_update_field(doc.get(fld), val, overwrite) \
                and _txt(doc.get(fld)) != val:
            doc.set(fld, val)
            counters[counter] += 1
            changed = True
    return changed


def backfill_tih(path, sheet=None, dry_run=1, overwrite=0, commit_every=100):
    """
    Cap nhat lien lac phu huynh (ten / SDT / email) tu file
    «DANH SÁCH HỌC SINH XẾP LỚP», doi chieu ma HS (cot ID / WS...).

        bench --site <site> execute erp.api.crm.import_qlead_excel.backfill_tih \
            --kwargs "{'path': '/private/files/....xlsx', 'dry_run': 1}"

    sheet      ten sheet (mac dinh «1. TỔNG HS ĐÓNG PHÍ_ENROLLED  S»).
    overwrite  0 = chi dien o dang trong | 1 = file thang.

    Tim PH: uu tien quan he tren CRM Family (Me/Bo/NGH), fallback SDT.
    Ghi CRM Guardian.guardian_name, .phone_number, .email + bang con phone_numbers/emails.
    Bo qua gia tri '0'. Mot o co the chua nhieu SDT/email.

    KHONG tao CRM Guardian moi, KHONG them lead_guardians / quan he gia dinh moi.
    """
    import openpyxl

    dry_run, ow = int(dry_run), int(overwrite)
    commit_every = int(commit_every) or 100
    path = _resolve_path(path)
    print(f"  File: {path}")
    print(f"  overwrite={ow} ({'file thang' if ow else 'CRM thang'})")
    print("  Ghi: ten / SDT / email (+ nghe nghiep / dia chi neu co)")

    wb = openpyxl.load_workbook(path, data_only=True)
    sname = sheet or TIH_SHEET
    if sname not in wb.sheetnames:
        frappe.throw(f"Khong co sheet {sname!r}. Co: {wb.sheetnames}")
    ws = wb[sname]
    C, seen = _tih_resolve_cols(ws)
    print(f"  Sheet: {sname!r}")
    missing = [k for k in ("id", "student") if k not in C]
    if missing:
        frappe.throw(f"Sheet thieu cot: {missing}. Tieu de tim thay: {list(seen)[:25]}")

    def cell(r, key):
        j = C.get(key)
        return _txt(ws.cell(r, j).value) if j else ""

    def cell_raw(r, key):
        j = C.get(key)
        return ws.cell(r, j).value if j else None

    res = {
        "rows": 0, "people": 0, "guardian_found": 0, "guardian_missing": [],
        "set_name": 0, "set_phone": 0, "set_email": 0, "set_job": 0, "set_addr": 0,
        "child_phone": 0, "child_email": 0, "set_relationship": 0,
        "lead_filled": 0, "no_code": [], "student_missing": [],
        "invalid_phones": [], "errors": [],
    }
    n_done = 0

    for r in range(TIH_HEADER_ROW + 1, ws.max_row + 1):
        sn = cell(r, "student")
        if not sn:
            continue
        res["rows"] += 1
        code = cell(r, "id")
        if not re.fullmatch(r"WS\d+", code):
            res["no_code"].append({"row": r, "student": sn, "code": code})
            continue

        student_id = frappe.db.get_value("CRM Student", {"student_code": code}, "name")
        if not student_id:
            res["student_missing"].append({"row": r, "student": sn, "code": code})

        addr = " - ".join([x for x in (cell(r, "addr"), cell(r, "district"),
                                       cell(r, "province")) if x and not _is_skip(x)])
        both = _emails(cell(r, "both_email"))
        people = []
        for tag, kn, kp, ke, kj, default_rels in _TIH_ROLES:
            nm_raw = cell_raw(r, kn)
            nm = "" if _is_skip(nm_raw) else _txt(nm_raw)
            phones, bad_phones = _phones(cell_raw(r, kp)) if kp else ([], [])
            for bad in bad_phones:
                res["invalid_phones"].append(
                    {"row": r, "student": sn, "tag": tag, "phone": bad})
            mails = []
            if ke and not _is_skip(cell_raw(r, ke)):
                mails = _emails(cell(r, ke))
            if not mails and both and tag in ("m", "f"):
                mails = list(both)
            job_raw = cell_raw(r, kj) if kj else None
            job = "" if _is_skip(job_raw) else _txt(job_raw)

            if tag == "g":
                rel_raw = cell(r, "g_rel")
                rel = normalize_relationship(rel_raw) if not _is_skip(rel_raw) else "guardian"
                rel_codes = [rel] if rel else ["guardian"]
            else:
                rel_codes = list(default_rels)
                rel = rel_codes[0] if rel_codes else ""

            if not nm and not phones and not mails and not job:
                continue
            people.append({
                "tag": tag, "name": nm, "phones": phones, "emails": mails,
                "job": job, "addr": addr, "rel_codes": rel_codes,
                "rel": rel,
            })
        if not people:
            continue
        res["people"] += len(people)

        lead = frappe.db.get_value("CRM Lead", {"student_code": code}, "name")

        for g in people:
            gname, how, rel_row = _tih_find_guardian_for_student(
                student_id, g["rel_codes"], g["phones"])
            if not gname:
                res["guardian_missing"].append({
                    "row": r, "student": sn, "tag": g["tag"], "name": g["name"],
                    "phones": ", ".join(g["phones"][:2]),
                })
                continue
            res["guardian_found"] += 1
            if dry_run:
                _tih_preview_guardian_update(gname, g, ow, res)
                if g["tag"] == "g" and rel_row and g.get("rel"):
                    cur_rel = rel_row.get("relationship_type") or ""
                    if ow or not cur_rel:
                        if cur_rel != g["rel"]:
                            res["set_relationship"] += 1
                continue

            sp = f"tih_{r}_{g['tag']}"
            try:
                frappe.db.savepoint(sp)
                doc = frappe.get_doc("CRM Guardian", gname)
                changed = _apply_guardian_contact_update(doc, g, ow, res)
                if changed:
                    doc.flags.ignore_validate = True
                    doc.flags.ignore_mandatory = True
                    doc.save(ignore_permissions=True)
                    n_done += 1

                # Cap nhat quan he NGH tren CRM Family neu cot AE khac DB
                if g["tag"] == "g" and rel_row and g.get("rel"):
                    cur_rel = rel_row.get("relationship_type") or ""
                    if (ow or not cur_rel) and cur_rel != g["rel"]:
                        frappe.db.set_value(
                            "CRM Family Relationship", rel_row["name"],
                            "relationship_type", g["rel"], update_modified=True)
                        res["set_relationship"] += 1
            except Exception as e:
                try:
                    frappe.db.rollback(save_point=sp)
                except Exception:
                    pass
                res["errors"].append({"row": r, "student": sn, "error": str(e)[:220]})
                frappe.log_error(message=frappe.get_traceback() or str(e),
                                 title=f"backfill_tih guardian dong {r}")

        # CRM Lead: bu guardian_* phang, uu tien Me
        if lead:
            pick = next((g for g in people if g["tag"] == "m"), people[0])
            pairs = (
                ((pick["emails"][0] if pick.get("emails") else ""), "guardian_email"),
                (pick.get("job") or "", "guardian_occupation"),
                (pick.get("addr") or "", "guardian_address"),
            )
            if dry_run:
                cur = frappe.db.get_value(
                    "CRM Lead", lead,
                    ["guardian_email", "guardian_occupation", "guardian_address"],
                    as_dict=True) or {}
                if any(v and (ow or not _txt(cur.get(f))) for v, f in pairs):
                    res["lead_filled"] += 1
            else:
                try:
                    doc = frappe.get_doc("CRM Lead", lead)
                    ch = False
                    for val, fld in pairs:
                        if val and (ow or not _txt(doc.get(fld))) and _txt(doc.get(fld)) != val:
                            doc.set(fld, val)
                            ch = True
                    if ch:
                        doc.flags.ignore_mandatory = True
                        doc.save(ignore_permissions=True)
                        res["lead_filled"] += 1
                except Exception as e:
                    res["errors"].append({"row": r, "student": sn, "error": str(e)[:220]})

        if n_done and n_done % commit_every == 0:
            frappe.db.commit()

    if not dry_run:
        frappe.db.commit()

    print("")
    print("=" * 62)
    print("  DRY-RUN — khong ghi gi vao DB" if dry_run else "  DA GHI VAO DB")
    print("=" * 62)
    print(f"  Dong doc duoc            : {res['rows']}")
    print(f"  Phu huynh co du lieu     : {res['people']}")
    print(f"  Tim thay CRM Guardian    : {res['guardian_found']}")
    print(f"  KHONG tim thay PH        : {len(res['guardian_missing'])}")
    print(f"  HS khong co CRM Student  : {len(res['student_missing'])}")
    print(f"  {'Se ghi ten' if dry_run else 'Da ghi ten':<24} : {res['set_name']}")
    print(f"  {'Se ghi SDT chinh' if dry_run else 'Da ghi SDT chinh':<24} : {res['set_phone']}")
    print(f"  {'Se ghi email' if dry_run else 'Da ghi email':<24} : {res['set_email']}")
    print(f"  {'Se ghi nghe nghiep' if dry_run else 'Da ghi nghe nghiep':<24} : {res['set_job']}")
    print(f"  {'Se ghi dia chi' if dry_run else 'Da ghi dia chi':<24} : {res['set_addr']}")
    if not dry_run:
        print(f"  Dong phone_numbers       : {res['child_phone']}")
        print(f"  Dong CRM Guardian Email  : {res['child_email']}")
    print(f"  Cap nhat quan he NGH     : {res['set_relationship']}")
    print(f"  Ho so CRM Lead duoc bu   : {res['lead_filled']}")
    if res["no_code"]:
        print(f"\n  Ma HS khong hop le ({len(res['no_code'])}):")
        for x in res["no_code"][:10]:
            print(f"    r{x['row']:>4} {x['code'][:14]:<16} {x['student'][:34]}")
    if res["student_missing"]:
        print(f"\n  Khong co CRM Student ({len(res['student_missing'])}):")
        for x in res["student_missing"][:10]:
            print(f"    r{x['row']:>4} {x['code']:<12} {x['student'][:34]}")
    if res["guardian_missing"]:
        print(f"\n  Khong tim thay guardian ({len(res['guardian_missing'])}):")
        for x in res["guardian_missing"][:20]:
            print(f"    r{x['row']:>4} [{x['tag']}] {x['name'][:22]:<24} {x['phones']:<14} (HS {x['student'][:18]})")
        if len(res["guardian_missing"]) > 20:
            print(f"    … con {len(res['guardian_missing']) - 20}")
    if res["invalid_phones"]:
        print(f"\n  SDT sai dinh dang ({len(res['invalid_phones'])}):")
        for x in res["invalid_phones"][:15]:
            print(f"    r{x['row']:>4} [{x['tag']}] {x['phone']}")
    if res["errors"]:
        print(f"\n  Loi ({len(res['errors'])}):")
        for e in res["errors"][:20]:
            print(f"    r{e['row']:>4} {e['student'][:28]:<30} {e['error']}")
    print("")
    return res
