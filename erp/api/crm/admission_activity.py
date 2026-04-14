"""
CRM Admission Activity API - CRUD Sự kiện và Khoá học tuyển sinh
"""

from collections import defaultdict

import frappe
from erp.utils.api_response import (
    success_response,
    error_response,
    single_item_response,
    list_response,
    validation_error_response,
    not_found_response,
)
from erp.api.crm.utils import check_crm_permission, get_request_data


def _resolve_crm_lead_name(ref):
    """
    Chuẩn hóa tham chiếu CRM Lead: docname (name) hoặc mã crm_code.
    Tab CRM có thể gửi URL param là crm_code; bản ghi Event/Course Student luôn lưu link theo name.
    """
    if not ref:
        return None
    ref = (ref or "").strip()
    if not ref:
        return None
    if frappe.db.exists("CRM Lead", ref):
        return ref
    return frappe.db.get_value("CRM Lead", {"crm_code": ref}, "name")


# ========== KHOÁ HỌC — helper bảng lớp ==========

COURSE_CLASS_TYPE_LABELS = {"regular": "Lớp chính quy", "running": "Lớp chạy"}


def _append_course_classes_from_payload(doc, payload):
    """Gán bảng con course_classes — mỗi dòng: class_name, class_type (regular|running)"""
    if payload is None:
        return
    doc.set("course_classes", [])
    for row in payload or []:
        if not isinstance(row, dict):
            continue
        cn = (row.get("class_name") or "").strip()
        ct = (row.get("class_type") or "").strip()
        if not cn or ct not in ("regular", "running"):
            continue
        doc.append("course_classes", {"class_name": cn, "class_type": ct})


def _course_classes_catalog(course_id):
    """Danh sách lớp trên khoá — id, tên, loại (cho FE)"""
    if not course_id or not frappe.db.exists("CRM Admission Course", course_id):
        return []
    doc = frappe.get_doc("CRM Admission Course", course_id)
    out = []
    for r in doc.course_classes:
        ct = r.class_type or ""
        out.append(
            {
                "name": r.name,
                "class_name": r.class_name or "",
                "class_type": ct,
                "class_type_label": COURSE_CLASS_TYPE_LABELS.get(ct, ct),
            }
        )
    return out


def _enrich_course_students_with_classes(items):
    """Bổ sung regular_class_name, running_class_ids, running_class_names, class_summary"""
    if not items:
        return items
    ids = [i["name"] for i in items]
    rc_rows = frappe.get_all(
        "CRM Admission Course Student Running",
        filters={"parent": ["in", ids]},
        fields=["parent", "course_class"],
    )
    by_parent = defaultdict(list)
    for r in rc_rows:
        by_parent[r.parent].append(r.course_class)

    all_cc_ids = set()
    for i in items:
        if i.get("regular_class"):
            all_cc_ids.add(i["regular_class"])
    for ccs in by_parent.values():
        for cc in ccs:
            all_cc_ids.add(cc)

    cc_names = {}
    if all_cc_ids:
        for row in frappe.get_all(
            "CRM Admission Course Class",
            filters={"name": ["in", list(all_cc_ids)]},
            fields=["name", "class_name"],
        ):
            cc_names[row.name] = row.class_name or row.name

    for i in items:
        rc = i.get("regular_class")
        i["regular_class_name"] = cc_names.get(rc) if rc else None
        runs = by_parent.get(i["name"], [])
        i["running_class_ids"] = runs
        i["running_class_names"] = [cc_names.get(x, x) for x in runs]
        parts = []
        if i["regular_class_name"]:
            parts.append(f"Lớp CQ: {i['regular_class_name']}")
        if i["running_class_names"]:
            parts.append("Lớp chạy: " + ", ".join(i["running_class_names"]))
        i["class_summary"] = " | ".join(parts) if parts else ""
    return items


def _set_course_student_class_fields(doc, regular_class=None, running_class_ids=None):
    """Gán regular_class và các dòng running_classes trước insert/save"""
    doc.regular_class = regular_class or None
    doc.set("running_classes", [])
    seen = set()
    for rid in running_class_ids or []:
        if not rid or rid in seen:
            continue
        seen.add(rid)
        doc.append("running_classes", {"course_class": rid})


def _course_has_regular_classes(course_id):
    return (
        frappe.db.count(
            "CRM Admission Course Class",
            filters={
                "parent": course_id,
                "parenttype": "CRM Admission Course",
                "class_type": "regular",
            },
        )
        > 0
    )


# ========== SỰ KIỆN (CRM Admission Event) ==========


def _append_event_promotions_from_payload(doc, promotions_payload):
    """Gán bảng con promotions — mỗi dòng Link tới CRM Promotion (danh mục có sẵn)"""
    if promotions_payload is None:
        return
    doc.set("promotions", [])
    seen = set()
    for row in promotions_payload or []:
        if not isinstance(row, dict):
            continue
        prom = (row.get("promotion") or "").strip()
        if not prom or prom in seen:
            continue
        if not frappe.db.exists("CRM Promotion", prom):
            continue
        seen.add(prom)
        doc.append("promotions", {"promotion": prom})


def _event_promotions_meta(event_id):
    """Danh sách ưu đãi trên sự kiện — đọc tên/phân loại/% từ CRM Promotion"""
    doc = frappe.get_doc("CRM Admission Event", event_id)
    out = []
    for r in doc.promotions:
        pid = r.promotion
        if not pid:
            continue
        p = frappe.db.get_value(
            "CRM Promotion",
            pid,
            ["promotion_name", "category", "value"],
            as_dict=True,
        )
        if not p:
            continue
        out.append(
            {
                "promotion_id": pid,
                "promotion_name": p.get("promotion_name") or pid,
                "category": p.get("category") or "",
                "value": p.get("value"),
            }
        )
    return out


def _enrich_modified_by_name(items, modified_by_field="modified_by"):
    """Bổ sung modified_by_name (full_name từ User) cho mỗi item"""
    for item in items:
        user_id = item.get(modified_by_field)
        if user_id:
            full_name = frappe.db.get_value("User", user_id, "full_name")
            item["modified_by_name"] = full_name or user_id
        else:
            item["modified_by_name"] = None
    return items


@frappe.whitelist()
def get_events():
    """Lấy danh sách sự kiện, filter theo school_year_id nếu có"""
    check_crm_permission()
    school_year_id = frappe.request.args.get("school_year_id")
    filters = {}
    if school_year_id and school_year_id != "all":
        filters["school_year_id"] = school_year_id
    items = frappe.get_all(
        "CRM Admission Event",
        filters=filters,
        fields=["name", "event_name", "event_date", "student_count", "is_active", "school_year_id", "modified", "modified_by"],
        order_by="modified desc",
    )
    _enrich_modified_by_name(items)
    return list_response(items)


@frappe.whitelist()
def get_event(event_id=None):
    """Lấy chi tiết 1 sự kiện"""
    check_crm_permission()
    event_id = event_id or frappe.request.args.get("event_id")
    if not event_id:
        return validation_error_response("Thiếu event_id", {"event_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Event", event_id):
        return not_found_response("Không tìm thấy sự kiện")
    doc = frappe.get_doc("CRM Admission Event", event_id)
    data = doc.as_dict()
    if doc.modified_by:
        data["modified_by_name"] = frappe.db.get_value("User", doc.modified_by, "full_name") or doc.modified_by
    # FE chi tiết bảng: tên / phân loại / % từ danh mục CRM Promotion
    data["promotions_catalog"] = _event_promotions_meta(event_id)
    return single_item_response(data, "Thành công")


@frappe.whitelist(methods=["POST"])
def create_event():
    """Tạo sự kiện mới"""
    check_crm_permission()
    data = get_request_data()
    if not data.get("event_name"):
        return validation_error_response("Thiếu event_name", {"event_name": ["Bắt buộc"]})
    try:
        doc = frappe.new_doc("CRM Admission Event")
        doc.event_name = data.get("event_name", "").strip()
        doc.event_date = data.get("event_date") or None
        doc.student_count = data.get("student_count", 0) or 0
        doc.is_active = 1 if data.get("is_active", True) else 0
        doc.school_year_id = data.get("school_year_id") or None
        if "promotions" in data:
            _append_event_promotions_from_payload(doc, data.get("promotions"))
        doc.insert(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Tạo sự kiện thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi tạo sự kiện: {str(e)}")


@frappe.whitelist(methods=["POST"])
def update_event():
    """Cập nhật sự kiện"""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Event", name):
        return not_found_response("Không tìm thấy sự kiện")
    try:
        doc = frappe.get_doc("CRM Admission Event", name)
        if "event_name" in data:
            doc.event_name = data["event_name"].strip()
        if "event_date" in data:
            doc.event_date = data["event_date"] or None
        if "student_count" in data:
            doc.student_count = data["student_count"] or 0
        if "is_active" in data:
            doc.is_active = 1 if data["is_active"] else 0
        if "school_year_id" in data:
            doc.school_year_id = data["school_year_id"] or None
        if "promotions" in data:
            _append_event_promotions_from_payload(doc, data.get("promotions"))
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Cập nhật sự kiện thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi cập nhật sự kiện: {str(e)}")


@frappe.whitelist(methods=["POST"])
def toggle_event_active():
    """Bật/tắt trạng thái sự kiện"""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    is_active = data.get("is_active", True)
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Event", name):
        return not_found_response("Không tìm thấy sự kiện")
    try:
        doc = frappe.get_doc("CRM Admission Event", name)
        doc.is_active = 1 if is_active else 0
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Cập nhật trạng thái thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi cập nhật trạng thái: {str(e)}")


@frappe.whitelist(methods=["POST"])
def delete_event():
    """Xóa sự kiện"""
    check_crm_permission()
    name = get_request_data().get("name")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Event", name):
        return not_found_response("Không tìm thấy sự kiện")
    try:
        frappe.delete_doc("CRM Admission Event", name, ignore_permissions=True)
        frappe.db.commit()
        return success_response(message="Xóa sự kiện thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi xóa sự kiện: {str(e)}")


# ========== HỌC SINH SỰ KIỆN (CRM Admission Event Student) ==========
# Trạng thái: registered, attended, not_attended (không có paid)
EVENT_STATUS_MAP = {
    "registered": "Đã đăng ký",
    "attended": "Đã tham gia",
    "not_attended": "Không tham gia",
}


@frappe.whitelist()
def get_lead_events():
    """Lấy mọi sự kiện đã gắn lead (mọi trạng thái: registered / attended / not_attended) — DealSection CRM."""
    check_crm_permission()
    crm_lead_raw = frappe.request.args.get("crm_lead_id")
    if not crm_lead_raw:
        return validation_error_response("Thiếu crm_lead_id", {"crm_lead_id": ["Bắt buộc"]})
    crm_lead_id = _resolve_crm_lead_name(crm_lead_raw)
    if not crm_lead_id:
        return not_found_response("Không tìm thấy CRM Lead")

    event_students = frappe.get_all(
        "CRM Admission Event Student",
        filters={"crm_lead_id": crm_lead_id},
        fields=["name", "event_id", "status", "modified"],
        order_by="modified desc",
    )
    if not event_students:
        return list_response([], "Thành công")

    event_ids = list({r["event_id"] for r in event_students})
    events = frappe.get_all(
        "CRM Admission Event",
        filters={"name": ["in", event_ids]},
        fields=["name", "event_name", "event_date", "modified"],
    )
    event_map = {e["name"]: e for e in events}
    result = []
    for es in event_students:
        ev = event_map.get(es["event_id"])
        if ev:
            result.append({
                "name": es["name"],
                "event_id": es["event_id"],
                "event_name": ev.get("event_name") or "-",
                "event_date": ev.get("event_date"),
                "status": es["status"],
                "status_label": EVENT_STATUS_MAP.get(es["status"], es["status"]),
                "modified": es["modified"],
            })
    return list_response(result, "Thành công")


@frappe.whitelist()
def get_lead_courses():
    """Lấy mọi khoá học/CLB đã gắn lead (mọi trạng thái) — DealSection CRM."""
    check_crm_permission()
    crm_lead_raw = frappe.request.args.get("crm_lead_id")
    if not crm_lead_raw:
        return validation_error_response("Thiếu crm_lead_id", {"crm_lead_id": ["Bắt buộc"]})
    crm_lead_id = _resolve_crm_lead_name(crm_lead_raw)
    if not crm_lead_id:
        return not_found_response("Không tìm thấy CRM Lead")

    course_students = frappe.get_all(
        "CRM Admission Course Student",
        filters={"crm_lead_id": crm_lead_id},
        fields=["name", "course_id", "status", "modified", "regular_class"],
        order_by="modified desc",
    )
    if not course_students:
        return list_response([], "Thành công")

    _enrich_course_students_with_classes(course_students)

    course_ids = list({r["course_id"] for r in course_students})
    courses = frappe.get_all(
        "CRM Admission Course",
        filters={"name": ["in", course_ids]},
        fields=["name", "course_name", "event_date", "modified"],
    )
    course_map = {c["name"]: c for c in courses}
    result = []
    for cs in course_students:
        co = course_map.get(cs["course_id"])
        if co:
            st = cs.get("status") or ""
            result.append({
                "name": cs["name"],
                "course_id": cs["course_id"],
                "course_name": co.get("course_name") or "-",
                "event_date": co.get("event_date"),
                "status": st,
                "status_label": STATUS_MAP.get(st, st),
                "modified": cs["modified"],
                "class_summary": cs.get("class_summary") or "",
            })
    return list_response(result, "Thành công")


def _get_event_student_summary(event_id):
    """Tính tổng, đã đăng ký, đã tham gia, không tham gia"""
    filters = {"event_id": event_id}
    total = frappe.db.count("CRM Admission Event Student", filters=filters)
    registered = frappe.db.count("CRM Admission Event Student", filters={**filters, "status": "registered"})
    attended = frappe.db.count("CRM Admission Event Student", filters={**filters, "status": "attended"})
    not_attended = frappe.db.count("CRM Admission Event Student", filters={**filters, "status": "not_attended"})
    return {"total": total, "registered": registered, "attended": attended, "not_attended": not_attended}


@frappe.whitelist()
def get_event_students():
    """Lấy danh sách học sinh trong sự kiện"""
    check_crm_permission()
    event_id = frappe.request.args.get("event_id")
    if not event_id:
        return validation_error_response("Thiếu event_id", {"event_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Event", event_id):
        return not_found_response("Không tìm thấy sự kiện")

    search = frappe.request.args.get("search")
    status_filter = frappe.request.args.get("status")

    filters = {"event_id": event_id}
    if status_filter and status_filter in ("registered", "attended", "not_attended"):
        filters["status"] = status_filter

    or_filters = None
    if search and search.strip():
        lead_ids = frappe.db.sql("""
            SELECT name FROM `tabCRM Lead`
            WHERE name LIKE %(s)s OR crm_code LIKE %(s)s OR student_name LIKE %(s)s
        """, {"s": f"%{search.strip()}%"}, as_dict=True)
        lead_names = [r["name"] for r in lead_ids]
        if not lead_names:
            return list_response(
                [],
                "Thành công",
                meta={
                    "summary": _get_event_student_summary(event_id),
                    "event_promotions": _event_promotions_meta(event_id),
                },
            )
        or_filters = {"crm_lead_id": ["in", lead_names]}

    items = frappe.get_all(
        "CRM Admission Event Student",
        filters=filters,
        or_filters=or_filters,
        fields=["name", "event_id", "crm_lead_id", "status", "modified", "modified_by"],
        order_by="modified desc",
    )

    flags_by_student = defaultdict(dict)
    if items:
        names = [i["name"] for i in items]
        for pr in frappe.get_all(
            "CRM Admission Event Student Promotion",
            filters={"parent": ["in", names]},
            fields=["parent", "promotion", "selected"],
        ):
            pid = pr.get("promotion")
            if pid:
                flags_by_student[pr.parent][pid] = bool(pr.selected)

    for item in items:
        lead = frappe.db.get_value(
            "CRM Lead",
            item["crm_lead_id"],
            ["crm_code", "student_name", "student_dob"],
            as_dict=True,
        )
        if lead:
            item["crm_code"] = lead.get("crm_code") or item["crm_lead_id"]
            item["student_name"] = lead.get("student_name") or "-"
            item["student_dob"] = lead.get("student_dob")
        else:
            item["crm_code"] = item["crm_lead_id"]
            item["student_name"] = "-"
            item["student_dob"] = None
        if item.get("modified_by"):
            item["modified_by_name"] = frappe.db.get_value("User", item["modified_by"], "full_name") or item["modified_by"]
        else:
            item["modified_by_name"] = None

        item["promotion_flags"] = dict(flags_by_student.get(item["name"], {}))

    return list_response(
        items,
        "Thành công",
        meta={
            "summary": _get_event_student_summary(event_id),
            "event_promotions": _event_promotions_meta(event_id),
        },
    )


@frappe.whitelist(methods=["POST"])
def add_event_student():
    """Thêm 1 học sinh (CRM Lead) vào sự kiện"""
    check_crm_permission()
    data = get_request_data()
    event_id = data.get("event_id")
    crm_lead_raw = data.get("crm_lead_id")
    if not event_id or not crm_lead_raw:
        return validation_error_response("Thiếu event_id hoặc crm_lead_id", {"event_id": ["Bắt buộc"], "crm_lead_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Event", event_id):
        return not_found_response("Không tìm thấy sự kiện")
    crm_lead_id = _resolve_crm_lead_name(crm_lead_raw)
    if not crm_lead_id:
        return not_found_response("Không tìm thấy CRM Lead")

    existing = frappe.db.exists(
        "CRM Admission Event Student",
        {"event_id": event_id, "crm_lead_id": crm_lead_id},
    )
    if existing:
        return validation_error_response("Học sinh đã có trong sự kiện", {"crm_lead_id": ["Đã tồn tại"]})

    try:
        doc = frappe.new_doc("CRM Admission Event Student")
        doc.event_id = event_id
        doc.crm_lead_id = crm_lead_id
        doc.status = data.get("status", "registered")
        doc.insert(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Thêm học sinh thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi thêm học sinh: {str(e)}")


@frappe.whitelist(methods=["POST"])
def add_event_students_excel():
    """Thêm nhiều học sinh từ Excel - 1 cột CRM Lead (hoặc crm_code), trạng thái mặc định Đã đăng ký"""
    check_crm_permission()
    import io
    import openpyxl

    event_id = frappe.form_dict.get("event_id")
    if not event_id:
        return validation_error_response("Thiếu event_id", {"event_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Event", event_id):
        return not_found_response("Không tìm thấy sự kiện")

    file = frappe.request.files.get("file")
    if not file:
        return validation_error_response("Thiếu file", {"file": ["Bắt buộc"]})

    try:
        content = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()
    except Exception as e:
        return error_response(f"Lỗi đọc file Excel: {str(e)}")

    if not rows:
        return error_response("File Excel trống")

    headers = [str(c).strip().lower() if c else "" for c in rows[0]]
    crm_col_idx = None
    for i, h in enumerate(headers):
        if h in ("crm_lead", "crm_lead_id", "crm_code", "crm id"):
            crm_col_idx = i
            break
    if crm_col_idx is None:
        return error_response("Không tìm thấy cột CRM Lead trong file. Cần có cột: CRM Lead, CRM Lead ID hoặc CRM Code")

    success_count = 0
    error_count = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        val = row[crm_col_idx] if crm_col_idx < len(row) else None
        if not val or not str(val).strip():
            continue
        lead_id = str(val).strip()

        lead = frappe.db.get_value("CRM Lead", lead_id, "name")
        if not lead:
            lead = frappe.db.get_value("CRM Lead", {"crm_code": lead_id}, "name")
        if not lead:
            error_count += 1
            errors.append(f"Dòng {row_idx}: Không tìm thấy CRM Lead '{lead_id}'")
            continue

        existing = frappe.db.exists(
            "CRM Admission Event Student",
            {"event_id": event_id, "crm_lead_id": lead},
        )
        if existing:
            error_count += 1
            errors.append(f"Dòng {row_idx}: Học sinh đã có trong sự kiện")
            continue

        try:
            doc = frappe.new_doc("CRM Admission Event Student")
            doc.event_id = event_id
            doc.crm_lead_id = lead
            doc.status = "registered"
            doc.insert(ignore_permissions=True)
            success_count += 1
        except Exception:
            error_count += 1
            errors.append(f"Dòng {row_idx}: Lỗi khi thêm")

    frappe.db.commit()
    return success_response(
        message=f"Import: {success_count} thành công, {error_count} lỗi",
        data={"success_count": success_count, "error_count": error_count, "errors": errors[:50]},
    )


@frappe.whitelist()
def export_event_students_template():
    """Xuất template Excel cho nhập liệu trạng thái - danh sách học sinh kèm trạng thái hiện tại"""
    check_crm_permission()
    event_id = frappe.request.args.get("event_id")
    if not event_id:
        return validation_error_response("Thiếu event_id", {"event_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Event", event_id):
        return not_found_response("Không tìm thấy sự kiện")

    items = frappe.get_all(
        "CRM Admission Event Student",
        filters={"event_id": event_id},
        fields=["name", "crm_lead_id", "status"],
        order_by="modified desc",
    )
    for item in items:
        lead = frappe.db.get_value(
            "CRM Lead",
            item["crm_lead_id"],
            ["crm_code", "student_name", "student_dob"],
            as_dict=True,
        )
        if lead:
            item["crm_code"] = lead.get("crm_code") or item["crm_lead_id"]
            item["student_name"] = lead.get("student_name") or ""
            item["student_dob"] = lead.get("student_dob")
        else:
            item["crm_code"] = item["crm_lead_id"]
            item["student_name"] = ""
            item["student_dob"] = None

    return success_response(
        message="OK",
        data={
            "headers": ["crm_lead_id", "crm_code", "student_name", "student_dob", "status"],
            "header_labels": ["CRM Lead ID", "Mã CRM", "Tên học sinh", "Ngày sinh", "Trạng thái"],
            "rows": [
                {
                    "crm_lead_id": r["crm_lead_id"],
                    "crm_code": r.get("crm_code", ""),
                    "student_name": r.get("student_name", ""),
                    "student_dob": str(r["student_dob"]) if r.get("student_dob") else "",
                    "status": r.get("status", "registered"),
                }
                for r in items
            ],
        },
    )


@frappe.whitelist(methods=["POST"])
def import_event_students_status():
    """Nhập liệu trạng thái - upload Excel với crm_lead_id + status mới"""
    check_crm_permission()
    import io
    import openpyxl

    event_id = frappe.form_dict.get("event_id")
    if not event_id:
        return validation_error_response("Thiếu event_id", {"event_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Event", event_id):
        return not_found_response("Không tìm thấy sự kiện")

    file = frappe.request.files.get("file")
    if not file:
        return validation_error_response("Thiếu file", {"file": ["Bắt buộc"]})

    try:
        content = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()
    except Exception as e:
        return error_response(f"Lỗi đọc file Excel: {str(e)}")

    if not rows or len(rows) < 2:
        return error_response("File Excel trống hoặc không có dữ liệu")

    headers = [str(c).strip().lower() if c else "" for c in rows[0]]
    crm_col = next((i for i, h in enumerate(headers) if h in ("crm_lead_id", "crm_lead", "crm_code", "crm id")), None)
    status_col = next((i for i, h in enumerate(headers) if h == "status" or "trạng thái" in (h or "")), None)

    if crm_col is None or status_col is None:
        return error_response("Không tìm thấy cột CRM Lead ID và Status. Cần tải template từ nút Nhập liệu.")

    valid_statuses = {"registered", "attended", "not_attended"}
    success_count = 0
    error_count = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        crm_val = row[crm_col] if crm_col < len(row) else None
        status_val = row[status_col] if status_col < len(row) else None
        if not crm_val or not str(crm_val).strip():
            continue
        lead_id = str(crm_val).strip()
        status = str(status_val).strip().lower() if status_val else ""
        if status not in valid_statuses:
            status_map_vn = {"registered": "đã đăng ký", "attended": "đã tham gia", "not_attended": "không tham gia"}
            if status in status_map_vn.values():
                rev = {v: k for k, v in status_map_vn.items()}
                status = rev.get(status, "registered")
            else:
                status = "registered"

        rec = frappe.db.get_value(
            "CRM Admission Event Student",
            {"event_id": event_id, "crm_lead_id": lead_id},
            "name",
        )
        if not rec:
            lead = frappe.db.get_value("CRM Lead", {"crm_code": lead_id}, "name")
            if lead:
                rec = frappe.db.get_value(
                    "CRM Admission Event Student",
                    {"event_id": event_id, "crm_lead_id": lead},
                    "name",
                )
        if not rec:
            error_count += 1
            errors.append(f"Dòng {row_idx}: Không tìm thấy bản ghi cho CRM Lead '{lead_id}'")
            continue

        try:
            doc = frappe.get_doc("CRM Admission Event Student", rec)
            doc.status = status
            doc.save(ignore_permissions=True)
            success_count += 1
        except Exception:
            error_count += 1
            errors.append(f"Dòng {row_idx}: Lỗi cập nhật")

    frappe.db.commit()
    return success_response(
        message=f"Nhập liệu: {success_count} thành công, {error_count} lỗi",
        data={"success_count": success_count, "error_count": error_count, "errors": errors[:50]},
    )


def _promo_export_column_labels(promo_meta):
    """
    (header Excel, promotion_id) — tiêu đề cột theo tên ưu đãi + phân loại + %;
    trùng tên thì thêm [promotion_id].
    """
    used_labels = set()
    out = []
    for p in promo_meta:
        pid = p.get("promotion_id") or ""
        name = (p.get("promotion_name") or pid or "").strip()
        parts = [name]
        if p.get("category"):
            parts.append(str(p["category"]))
        if p.get("value") is not None:
            parts.append(f"{p['value']}%")
        label = " — ".join(parts)
        if label in used_labels:
            label = f"{label} [{pid}]"
        used_labels.add(label)
        out.append((label, pid))
    return out


@frappe.whitelist()
def export_event_report():
    """Xuất báo cáo sự kiện - danh sách học sinh kèm trạng thái và cột ưu đãi (CRM Promotion)"""
    check_crm_permission()
    event_id = frappe.request.args.get("event_id")
    if not event_id:
        return validation_error_response("Thiếu event_id", {"event_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Event", event_id):
        return not_found_response("Không tìm thấy sự kiện")

    promo_meta = _event_promotions_meta(event_id)
    promo_cols = _promo_export_column_labels(promo_meta)

    items = frappe.get_all(
        "CRM Admission Event Student",
        filters={"event_id": event_id},
        fields=["name", "crm_lead_id", "status", "modified", "modified_by"],
        order_by="modified desc",
    )

    flags_by_student = defaultdict(dict)
    if items:
        names = [i["name"] for i in items]
        for pr in frappe.get_all(
            "CRM Admission Event Student Promotion",
            filters={"parent": ["in", names]},
            fields=["parent", "promotion", "selected"],
        ):
            pid = pr.get("promotion")
            if pid:
                flags_by_student[pr.parent][pid] = bool(pr.selected)

    for item in items:
        lead = frappe.db.get_value(
            "CRM Lead",
            item["crm_lead_id"],
            ["crm_code", "student_name", "student_dob"],
            as_dict=True,
        )
        if lead:
            item["crm_code"] = lead.get("crm_code") or item["crm_lead_id"]
            item["student_name"] = lead.get("student_name") or ""
            item["student_dob"] = lead.get("student_dob")
        else:
            item["crm_code"] = item["crm_lead_id"]
            item["student_name"] = ""
            item["student_dob"] = None
        item["status_label"] = EVENT_STATUS_MAP.get(item.get("status"), item.get("status", ""))
        if item.get("modified_by"):
            item["modified_by_name"] = frappe.db.get_value("User", item["modified_by"], "full_name") or item["modified_by"]

        fl = flags_by_student.get(item["name"], {})
        for label, pid in promo_cols:
            item[label] = "Có" if fl.get(pid) else "Không"

    base_headers = [
        "crm_lead_id",
        "crm_code",
        "student_name",
        "student_dob",
        "status",
        "status_label",
        "modified",
        "modified_by_name",
    ]
    headers = base_headers + [c[0] for c in promo_cols]

    return success_response(
        message="OK",
        data={
            "headers": headers,
            "rows": items,
        },
    )


@frappe.whitelist(methods=["POST"])
def update_event_student_status():
    """Cập nhật trạng thái 1 học sinh"""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    status = data.get("status")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if status not in ("registered", "attended", "not_attended"):
        return validation_error_response("Trạng thái không hợp lệ", {"status": ["Phải là registered, attended hoặc not_attended"]})
    if not frappe.db.exists("CRM Admission Event Student", name):
        return not_found_response("Không tìm thấy bản ghi")

    try:
        doc = frappe.get_doc("CRM Admission Event Student", name)
        doc.status = status
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Cập nhật trạng thái thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi cập nhật: {str(e)}")


@frappe.whitelist(methods=["POST"])
def set_event_student_promotion():
    """Bật/tắt ưu đãi cho 1 học sinh trong sự kiện (checkbox cột động) — promotion = name CRM Promotion"""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    promotion = (data.get("promotion") or data.get("promotion_uid") or "").strip()
    selected = bool(data.get("selected"))
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not promotion:
        return validation_error_response("Thiếu promotion", {"promotion": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Event Student", name):
        return not_found_response("Không tìm thấy bản ghi")

    try:
        doc = frappe.get_doc("CRM Admission Event Student", name)
        ev = frappe.get_doc("CRM Admission Event", doc.event_id)
        valid_ids = {r.promotion for r in ev.promotions if r.promotion}
        if promotion not in valid_ids:
            return validation_error_response(
                "Chương trình ưu đãi không thuộc sự kiện này",
                {"promotion": ["Không hợp lệ"]},
            )

        found = False
        for row in doc.promotion_selections:
            if row.promotion == promotion:
                row.selected = 1 if selected else 0
                found = True
                break
        if not found:
            doc.append(
                "promotion_selections",
                {"promotion": promotion, "selected": 1 if selected else 0},
            )
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Cập nhật ưu đãi thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi cập nhật: {str(e)}")


@frappe.whitelist(methods=["POST"])
def delete_event_student():
    """Xóa học sinh khỏi sự kiện"""
    check_crm_permission()
    name = get_request_data().get("name")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Event Student", name):
        return not_found_response("Không tìm thấy bản ghi")
    try:
        frappe.delete_doc("CRM Admission Event Student", name, ignore_permissions=True)
        frappe.db.commit()
        return success_response(message="Xóa học sinh khỏi sự kiện thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi xóa: {str(e)}")


# ========== KHOÁ HỌC (CRM Admission Course) ==========


@frappe.whitelist()
def get_courses():
    """Lấy danh sách khoá học, filter theo school_year_id nếu có"""
    check_crm_permission()
    school_year_id = frappe.request.args.get("school_year_id")
    filters = {}
    if school_year_id and school_year_id != "all":
        filters["school_year_id"] = school_year_id
    items = frappe.get_all(
        "CRM Admission Course",
        filters=filters,
        fields=["name", "course_name", "event_date", "student_count", "is_active", "school_year_id", "modified", "modified_by"],
        order_by="modified desc",
    )
    _enrich_modified_by_name(items)
    return list_response(items)


@frappe.whitelist()
def get_course(course_id=None):
    """Lấy chi tiết 1 khoá học"""
    check_crm_permission()
    course_id = course_id or frappe.request.args.get("course_id")
    if not course_id:
        return validation_error_response("Thiếu course_id", {"course_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Course", course_id):
        return not_found_response("Không tìm thấy khoá học")
    doc = frappe.get_doc("CRM Admission Course", course_id)
    data = doc.as_dict()
    if doc.modified_by:
        data["modified_by_name"] = frappe.db.get_value("User", doc.modified_by, "full_name") or doc.modified_by
    data["classes_catalog"] = _course_classes_catalog(course_id)
    return single_item_response(data, "Thành công")


@frappe.whitelist(methods=["POST"])
def create_course():
    """Tạo khoá học mới"""
    check_crm_permission()
    data = get_request_data()
    if not data.get("course_name"):
        return validation_error_response("Thiếu course_name", {"course_name": ["Bắt buộc"]})
    try:
        doc = frappe.new_doc("CRM Admission Course")
        doc.course_name = data.get("course_name", "").strip()
        doc.event_date = data.get("event_date") or None
        doc.student_count = data.get("student_count", 0) or 0
        doc.is_active = 1 if data.get("is_active", True) else 0
        doc.school_year_id = data.get("school_year_id") or None
        if "course_classes" in data:
            _append_course_classes_from_payload(doc, data.get("course_classes"))
        doc.insert(ignore_permissions=True)
        frappe.db.commit()
        out = doc.as_dict()
        out["classes_catalog"] = _course_classes_catalog(doc.name)
        return single_item_response(out, "Tạo khoá học thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi tạo khoá học: {str(e)}")


@frappe.whitelist(methods=["POST"])
def update_course():
    """Cập nhật khoá học"""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Course", name):
        return not_found_response("Không tìm thấy khoá học")
    try:
        doc = frappe.get_doc("CRM Admission Course", name)
        if "course_name" in data:
            doc.course_name = data["course_name"].strip()
        if "event_date" in data:
            doc.event_date = data["event_date"] or None
        if "student_count" in data:
            doc.student_count = data["student_count"] or 0
        if "is_active" in data:
            doc.is_active = 1 if data["is_active"] else 0
        if "school_year_id" in data:
            doc.school_year_id = data["school_year_id"] or None
        if "course_classes" in data:
            _append_course_classes_from_payload(doc, data.get("course_classes"))
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        out = doc.as_dict()
        out["classes_catalog"] = _course_classes_catalog(name)
        return single_item_response(out, "Cập nhật khoá học thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi cập nhật khoá học: {str(e)}")


@frappe.whitelist(methods=["POST"])
def toggle_course_active():
    """Bật/tắt trạng thái khoá học"""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    is_active = data.get("is_active", True)
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Course", name):
        return not_found_response("Không tìm thấy khoá học")
    try:
        doc = frappe.get_doc("CRM Admission Course", name)
        doc.is_active = 1 if is_active else 0
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Cập nhật trạng thái thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi cập nhật trạng thái: {str(e)}")


@frappe.whitelist(methods=["POST"])
def delete_course():
    """Xóa khoá học"""
    check_crm_permission()
    name = get_request_data().get("name")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Course", name):
        return not_found_response("Không tìm thấy khoá học")
    try:
        frappe.delete_doc("CRM Admission Course", name, ignore_permissions=True)
        frappe.db.commit()
        return success_response(message="Xóa khoá học thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi xóa khoá học: {str(e)}")


# ========== HỌC SINH KHOÁ HỌC (CRM Admission Course Student) ==========

# Trạng thái học sinh khoá học (mã Select trên CRM Admission Course Student)
COURSE_STUDENT_STATUSES = (
    "registered_interest",
    "trial",
    "paid",
    "attended",
    "transferred",
    "refunded",
)

STATUS_MAP = {
    "registered_interest": "Đã đăng ký/Quan tâm",
    "trial": "Học thử",
    "paid": "Đã đóng tiền",
    "attended": "Đã tham gia",
    "transferred": "Chuyển nhượng",
    "refunded": "Hoàn phí",
}


def _coerce_course_student_status(val):
    """Chuẩn hoá mã trạng thái từ API/Excel (mã snake_case, nhãn tiếng Việt, giá trị cũ)."""
    if val is None:
        return "registered_interest"
    s = str(val).strip()
    if not s:
        return "registered_interest"
    sl = s.lower()
    for st in COURSE_STUDENT_STATUSES:
        if sl == st:
            return st
    legacy = {"registered": "registered_interest", "not_attended": "registered_interest"}
    if sl in legacy:
        return legacy[sl]
    for st, label in STATUS_MAP.items():
        if s.lower() == (label or "").lower():
            return st
    extras = {
        "đã đăng ký": "registered_interest",
        "không tham gia": "refunded",
    }
    if s.lower() in extras:
        return extras[s.lower()]
    return "registered_interest"


def _get_course_student_summary(course_id):
    """Đếm theo từng trạng thái khoá học + tổng."""
    filters = {"course_id": course_id}
    total = frappe.db.count("CRM Admission Course Student", filters=filters)
    out = {"total": total}
    for st in COURSE_STUDENT_STATUSES:
        out[st] = frappe.db.count("CRM Admission Course Student", filters={**filters, "status": st})
    return out


@frappe.whitelist()
def get_course_students():
    """Lấy danh sách học sinh trong khoá học"""
    check_crm_permission()
    course_id = frappe.request.args.get("course_id")
    if not course_id:
        return validation_error_response("Thiếu course_id", {"course_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Course", course_id):
        return not_found_response("Không tìm thấy khoá học")

    search = frappe.request.args.get("search")
    status_filter = frappe.request.args.get("status")

    filters = {"course_id": course_id}
    if status_filter and status_filter in COURSE_STUDENT_STATUSES:
        filters["status"] = status_filter

    or_filters = None
    if search and search.strip():
        # Tìm theo tên lead, crm_code, student_name
        lead_ids = frappe.db.sql("""
            SELECT name FROM `tabCRM Lead`
            WHERE name LIKE %(s)s OR crm_code LIKE %(s)s OR student_name LIKE %(s)s
        """, {"s": f"%{search.strip()}%"}, as_dict=True)
        lead_names = [r["name"] for r in lead_ids]
        if not lead_names:
            return list_response([], "Thành công", meta={"summary": _get_course_student_summary(course_id)})
        or_filters = {"crm_lead_id": ["in", lead_names]}

    items = frappe.get_all(
        "CRM Admission Course Student",
        filters=filters,
        or_filters=or_filters,
        fields=["name", "course_id", "crm_lead_id", "status", "regular_class", "modified", "modified_by"],
        order_by="modified desc",
    )

    _enrich_course_students_with_classes(items)

    # Bổ sung thông tin từ CRM Lead: student_name, student_dob (student_dob trong CRM Lead)
    for item in items:
        lead = frappe.db.get_value(
            "CRM Lead",
            item["crm_lead_id"],
            ["crm_code", "student_name", "student_dob"],
            as_dict=True,
        )
        if lead:
            item["crm_code"] = lead.get("crm_code") or item["crm_lead_id"]
            item["student_name"] = lead.get("student_name") or "-"
            item["student_dob"] = lead.get("student_dob")
        else:
            item["crm_code"] = item["crm_lead_id"]
            item["student_name"] = "-"
            item["student_dob"] = None
        # modified_by_name
        if item.get("modified_by"):
            item["modified_by_name"] = frappe.db.get_value("User", item["modified_by"], "full_name") or item["modified_by"]
        else:
            item["modified_by_name"] = None

    return list_response(
        items,
        "Thành công",
        meta={"summary": _get_course_student_summary(course_id)},
    )


@frappe.whitelist(methods=["POST"])
def add_course_student():
    """Thêm 1 học sinh (CRM Lead) vào khoá học"""
    check_crm_permission()
    data = get_request_data()
    course_id = data.get("course_id")
    crm_lead_raw = data.get("crm_lead_id")
    if not course_id or not crm_lead_raw:
        return validation_error_response("Thiếu course_id hoặc crm_lead_id", {"course_id": ["Bắt buộc"], "crm_lead_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Course", course_id):
        return not_found_response("Không tìm thấy khoá học")
    crm_lead_id = _resolve_crm_lead_name(crm_lead_raw)
    if not crm_lead_id:
        return not_found_response("Không tìm thấy CRM Lead")

    # Kiểm tra trùng
    existing = frappe.db.exists(
        "CRM Admission Course Student",
        {"course_id": course_id, "crm_lead_id": crm_lead_id},
    )
    if existing:
        return validation_error_response("Học sinh đã có trong khoá học", {"crm_lead_id": ["Đã tồn tại"]})

    try:
        doc = frappe.new_doc("CRM Admission Course Student")
        doc.course_id = course_id
        doc.crm_lead_id = crm_lead_id
        doc.status = _coerce_course_student_status(data.get("status", "registered_interest"))
        regular_class = (data.get("regular_class") or "").strip() or None
        running_class_ids = data.get("running_class_ids") or []
        if not isinstance(running_class_ids, list):
            running_class_ids = []
        _set_course_student_class_fields(doc, regular_class, running_class_ids)
        doc.insert(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Thêm học sinh thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi thêm học sinh: {str(e)}")


@frappe.whitelist(methods=["POST"])
def update_course_student_classes():
    """Cập nhật lớp chính quy / các lớp chạy cho học sinh trong khoá"""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Course Student", name):
        return not_found_response("Không tìm thấy bản ghi")
    try:
        doc = frappe.get_doc("CRM Admission Course Student", name)
        if "regular_class" in data:
            doc.regular_class = data.get("regular_class") or None
        if "running_class_ids" in data:
            doc.set("running_classes", [])
            seen = set()
            for rid in data.get("running_class_ids") or []:
                if not rid or rid in seen:
                    continue
                seen.add(rid)
                doc.append("running_classes", {"course_class": rid})
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Cập nhật lớp thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi cập nhật lớp: {str(e)}")


@frappe.whitelist(methods=["POST"])
def add_course_students_excel():
    """Thêm nhiều học sinh từ Excel - 1 cột CRM Lead (hoặc crm_code), trạng thái mặc định Đã đăng ký"""
    check_crm_permission()
    import io
    import openpyxl

    course_id = frappe.form_dict.get("course_id")
    if not course_id:
        return validation_error_response("Thiếu course_id", {"course_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Course", course_id):
        return not_found_response("Không tìm thấy khoá học")

    if _course_has_regular_classes(course_id):
        return error_response(
            "Khoá học có lớp chính quy — không thể import hàng loạt. Vui lòng thêm học sinh từ giao diện và gán lớp."
        )

    file = frappe.request.files.get("file")
    if not file:
        return validation_error_response("Thiếu file", {"file": ["Bắt buộc"]})

    try:
        content = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()
    except Exception as e:
        return error_response(f"Lỗi đọc file Excel: {str(e)}")

    if not rows:
        return error_response("File Excel trống")

    # Dòng 1: header - tìm cột CRM Lead / crm_lead_id / crm_code
    headers = [str(c).strip().lower() if c else "" for c in rows[0]]
    crm_col_idx = None
    for i, h in enumerate(headers):
        if h in ("crm_lead", "crm_lead_id", "crm_code", "crm id"):
            crm_col_idx = i
            break
    if crm_col_idx is None:
        return error_response("Không tìm thấy cột CRM Lead trong file. Cần có cột: CRM Lead, CRM Lead ID hoặc CRM Code")

    success_count = 0
    error_count = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        val = row[crm_col_idx] if crm_col_idx < len(row) else None
        if not val or not str(val).strip():
            continue
        lead_id = str(val).strip()

        # Tìm CRM Lead theo name hoặc crm_code
        lead = frappe.db.get_value("CRM Lead", lead_id, "name")
        if not lead:
            lead = frappe.db.get_value("CRM Lead", {"crm_code": lead_id}, "name")
        if not lead:
            error_count += 1
            errors.append(f"Dòng {row_idx}: Không tìm thấy CRM Lead '{lead_id}'")
            continue

        existing = frappe.db.exists(
            "CRM Admission Course Student",
            {"course_id": course_id, "crm_lead_id": lead},
        )
        if existing:
            error_count += 1
            errors.append(f"Dòng {row_idx}: Học sinh đã có trong khoá học")
            continue

        try:
            doc = frappe.new_doc("CRM Admission Course Student")
            doc.course_id = course_id
            doc.crm_lead_id = lead
            doc.status = "registered_interest"
            doc.insert(ignore_permissions=True)
            success_count += 1
        except Exception:
            error_count += 1
            errors.append(f"Dòng {row_idx}: Lỗi khi thêm")

    frappe.db.commit()
    return success_response(
        message=f"Import: {success_count} thành công, {error_count} lỗi",
        data={"success_count": success_count, "error_count": error_count, "errors": errors[:50]},
    )


@frappe.whitelist()
def export_course_students_template():
    """Xuất template Excel cho nhập liệu trạng thái - danh sách học sinh kèm trạng thái hiện tại"""
    check_crm_permission()
    course_id = frappe.request.args.get("course_id")
    if not course_id:
        return validation_error_response("Thiếu course_id", {"course_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Course", course_id):
        return not_found_response("Không tìm thấy khoá học")

    items = frappe.get_all(
        "CRM Admission Course Student",
        filters={"course_id": course_id},
        fields=["name", "crm_lead_id", "status", "regular_class"],
        order_by="modified desc",
    )
    _enrich_course_students_with_classes(items)
    for item in items:
        lead = frappe.db.get_value(
            "CRM Lead",
            item["crm_lead_id"],
            ["crm_code", "student_name", "student_dob"],
            as_dict=True,
        )
        if lead:
            item["crm_code"] = lead.get("crm_code") or item["crm_lead_id"]
            item["student_name"] = lead.get("student_name") or ""
            item["student_dob"] = lead.get("student_dob")
        else:
            item["crm_code"] = item["crm_lead_id"]
            item["student_name"] = ""
            item["student_dob"] = None

    return success_response(
        message="OK",
        data={
            "headers": ["crm_lead_id", "crm_code", "student_name", "student_dob", "status", "class_summary"],
            "header_labels": ["CRM Lead ID", "Mã CRM", "Tên học sinh", "Ngày sinh", "Trạng thái", "Lớp (tham khảo)"],
            "rows": [
                {
                    "crm_lead_id": r["crm_lead_id"],
                    "crm_code": r.get("crm_code", ""),
                    "student_name": r.get("student_name", ""),
                    "student_dob": str(r["student_dob"]) if r.get("student_dob") else "",
                    "status": r.get("status", "registered_interest"),
                    "class_summary": r.get("class_summary") or "",
                }
                for r in items
            ],
        },
    )


@frappe.whitelist(methods=["POST"])
def import_course_students_status():
    """Nhập liệu trạng thái - upload Excel với crm_lead_id + status mới"""
    check_crm_permission()
    import io
    import openpyxl

    course_id = frappe.form_dict.get("course_id")
    if not course_id:
        return validation_error_response("Thiếu course_id", {"course_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Course", course_id):
        return not_found_response("Không tìm thấy khoá học")

    file = frappe.request.files.get("file")
    if not file:
        return validation_error_response("Thiếu file", {"file": ["Bắt buộc"]})

    try:
        content = file.read()
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(min_row=1, values_only=True))
        wb.close()
    except Exception as e:
        return error_response(f"Lỗi đọc file Excel: {str(e)}")

    if not rows or len(rows) < 2:
        return error_response("File Excel trống hoặc không có dữ liệu")

    headers = [str(c).strip().lower() if c else "" for c in rows[0]]
    crm_col = next((i for i, h in enumerate(headers) if h in ("crm_lead_id", "crm_lead", "crm_code", "crm id")), None)
    status_col = next((i for i, h in enumerate(headers) if h == "status" or "trạng thái" in (h or "")), None)

    if crm_col is None or status_col is None:
        return error_response("Không tìm thấy cột CRM Lead ID và Status. Cần tải template từ nút Nhập liệu.")

    success_count = 0
    error_count = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        crm_val = row[crm_col] if crm_col < len(row) else None
        status_val = row[status_col] if status_col < len(row) else None
        if not crm_val or not str(crm_val).strip():
            continue
        lead_id = str(crm_val).strip()
        status = _coerce_course_student_status(status_val)

        rec = frappe.db.get_value(
            "CRM Admission Course Student",
            {"course_id": course_id, "crm_lead_id": lead_id},
            "name",
        )
        if not rec:
            lead = frappe.db.get_value("CRM Lead", {"crm_code": lead_id}, "name")
            if lead:
                rec = frappe.db.get_value(
                    "CRM Admission Course Student",
                    {"course_id": course_id, "crm_lead_id": lead},
                    "name",
                )
        if not rec:
            error_count += 1
            errors.append(f"Dòng {row_idx}: Không tìm thấy bản ghi cho CRM Lead '{lead_id}'")
            continue

        try:
            doc = frappe.get_doc("CRM Admission Course Student", rec)
            doc.status = status
            doc.save(ignore_permissions=True)
            success_count += 1
        except Exception:
            error_count += 1
            errors.append(f"Dòng {row_idx}: Lỗi cập nhật")

    frappe.db.commit()
    return success_response(
        message=f"Nhập liệu: {success_count} thành công, {error_count} lỗi",
        data={"success_count": success_count, "error_count": error_count, "errors": errors[:50]},
    )


@frappe.whitelist()
def export_course_report():
    """Xuất báo cáo khoá học - danh sách học sinh kèm trạng thái"""
    check_crm_permission()
    course_id = frappe.request.args.get("course_id")
    if not course_id:
        return validation_error_response("Thiếu course_id", {"course_id": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Course", course_id):
        return not_found_response("Không tìm thấy khoá học")

    items = frappe.get_all(
        "CRM Admission Course Student",
        filters={"course_id": course_id},
        fields=["name", "crm_lead_id", "status", "regular_class", "modified", "modified_by"],
        order_by="modified desc",
    )
    _enrich_course_students_with_classes(items)
    for item in items:
        lead = frappe.db.get_value(
            "CRM Lead",
            item["crm_lead_id"],
            ["crm_code", "student_name", "student_dob"],
            as_dict=True,
        )
        if lead:
            item["crm_code"] = lead.get("crm_code") or item["crm_lead_id"]
            item["student_name"] = lead.get("student_name") or ""
            item["student_dob"] = lead.get("student_dob")
        else:
            item["crm_code"] = item["crm_lead_id"]
            item["student_name"] = ""
            item["student_dob"] = None
        item["status_label"] = STATUS_MAP.get(item.get("status"), item.get("status", ""))
        if item.get("modified_by"):
            item["modified_by_name"] = frappe.db.get_value("User", item["modified_by"], "full_name") or item["modified_by"]

    return success_response(
        message="OK",
        data={
            "headers": [
                "crm_lead_id",
                "crm_code",
                "student_name",
                "student_dob",
                "status",
                "status_label",
                "class_summary",
                "modified",
                "modified_by_name",
            ],
            "rows": items,
        },
    )


@frappe.whitelist(methods=["POST"])
def update_course_student_status():
    """Cập nhật trạng thái 1 học sinh"""
    check_crm_permission()
    data = get_request_data()
    name = data.get("name")
    status = data.get("status")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if status not in COURSE_STUDENT_STATUSES:
        return validation_error_response(
            "Trạng thái không hợp lệ",
            {"status": [f"Phải là một trong: {', '.join(COURSE_STUDENT_STATUSES)}"]},
        )
    if not frappe.db.exists("CRM Admission Course Student", name):
        return not_found_response("Không tìm thấy bản ghi")

    try:
        doc = frappe.get_doc("CRM Admission Course Student", name)
        doc.status = status
        doc.save(ignore_permissions=True)
        frappe.db.commit()
        return single_item_response(doc.as_dict(), "Cập nhật trạng thái thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi cập nhật: {str(e)}")


@frappe.whitelist(methods=["POST"])
def delete_course_student():
    """Xóa học sinh khỏi khoá học"""
    check_crm_permission()
    name = get_request_data().get("name")
    if not name:
        return validation_error_response("Thiếu name", {"name": ["Bắt buộc"]})
    if not frappe.db.exists("CRM Admission Course Student", name):
        return not_found_response("Không tìm thấy bản ghi")
    try:
        frappe.delete_doc("CRM Admission Course Student", name, ignore_permissions=True)
        frappe.db.commit()
        return success_response(message="Xóa học sinh khỏi khoá học thành công")
    except Exception as e:
        frappe.db.rollback()
        return error_response(f"Lỗi xóa: {str(e)}")
