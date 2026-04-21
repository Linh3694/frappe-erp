# Copyright (c) 2024, Wellspring International School and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.utils import nowdate, get_datetime, get_fullname
from frappe.exceptions import LinkExistsError
import json
from typing import Dict, Any
from erp.utils.campus_utils import get_current_campus_from_context, get_campus_id_from_user_roles
from erp.utils.api_response import success_response, error_response, validation_error_response, not_found_response
from erp.api.erp_administrative.room_activity_log import log_room_activity
try:
    import pandas as pd
except ImportError:
    pd = None


def _enrich_room_responsible_users_employee_codes(rooms):
    """Gắn employee_code từ User (ưu tiên mã NV, không thì username) cho cột PIC trên FE."""
    if not rooms:
        return
    all_uids = set()
    for r in rooms:
        for u in r.get("responsible_users") or []:
            uid = (u.get("user") or "").strip()
            if uid:
                all_uids.add(uid)
    if not all_uids:
        return
    user_meta = frappe.get_meta("User")
    fields = ["name"]
    for fn in ("employee_code", "username"):
        if user_meta.has_field(fn):
            fields.append(fn)
    user_rows = frappe.get_all(
        "User",
        filters={"name": ["in", list(all_uids)]},
        fields=fields,
    )
    uid_map = {}
    for ur in user_rows:
        name = ur.get("name")
        ec = (ur.get("employee_code") or "").strip()
        un = (ur.get("username") or "").strip()
        uid_map[name] = ec or un or ""
    for r in rooms:
        for u in r.get("responsible_users") or []:
            uid = (u.get("user") or "").strip()
            if uid and uid in uid_map and uid_map[uid]:
                u["employee_code"] = uid_map[uid]


@frappe.whitelist(allow_guest=False)
def get_all_rooms():
    """Get all rooms with basic information - SIMPLE VERSION"""
    try:
        # Get current user's campus information from roles
        campus_id = get_current_campus_from_context()

        if not campus_id:
            # Fallback: try to get first available campus from database
            try:
                first_campus = frappe.db.get_value("SIS Campus", {}, "name", order_by="creation asc")
                campus_id = first_campus or "CAMPUS-00001"
            except Exception:
                # Final fallback to known campus
                campus_id = "CAMPUS-00001"
            frappe.logger().warning(f"No campus found for user {frappe.session.user}, using fallback: {campus_id}")
        
        # Get buildings for this campus to filter rooms
        building_filters = {"campus_id": campus_id}
        buildings = frappe.get_all(
            "ERP Administrative Building",
            fields=["name"],
            filters=building_filters
        )
        
        building_ids = [b.name for b in buildings]
        
        if not building_ids:
            return success_response(
                data=[],
                message="No buildings found for this campus",
                meta={"total_count": 0}
            )
        
        # Get rooms that belong to buildings in this campus
        rooms = frappe.get_all(
            "ERP Administrative Room",
            fields=[
                "name",
                "title_vn",
                "title_en",
                "short_title",
                "physical_code",
                "room_number",
                "is_active",
                "capacity",
                "room_type",
                "building_id",
                "creation",
                "modified",
            ],
            filters={"building_id": ["in", building_ids]},
            order_by="title_vn asc",
        )

        # Get unique building IDs from rooms
        room_building_ids = list(set(room["building_id"] for room in rooms if room.get("building_id")))

        # Get building information for populate
        buildings = []
        if room_building_ids:
            buildings = frappe.get_all(
                "ERP Administrative Building",
                fields=[
                    "name",
                    "title_vn",
                    "title_en",
                    "short_title",
                    "campus_id"
                ],
                filters={"name": ["in", room_building_ids]}
            )

        # Create building mapping for quick lookup
        building_map = {building["name"]: building for building in buildings}

        # Add building information to each room
        for room in rooms:
            building_id = room.get("building_id")
            if building_id and building_id in building_map:
                room["building"] = building_map[building_id]
            else:
                room["building"] = None

        # Người phụ trách phòng (child table) — dùng cho FacilityLiteRoom
        if rooms:
            room_names = [r["name"] for r in rooms]
            child_rows = frappe.get_all(
                "ERP Administrative Room Responsible User",
                filters={"parent": ["in", room_names]},
                fields=["parent", "user", "full_name", "user_image", "designation"],
                order_by="idx asc",
            )
            by_room = {}
            for cr in child_rows:
                by_room.setdefault(cr.parent, []).append(
                    {
                        "user": cr.user,
                        "full_name": cr.full_name or "",
                        "user_image": cr.user_image or "",
                        "designation": cr.designation or "",
                    }
                )
            for room in rooms:
                room["responsible_users"] = by_room.get(room["name"], [])

        # Gán snapshot theo năm học (lọc FE qua ?school_year_id=...) — comment: hiển thị tên lớp / GVCN theo YA
        # GET axios: tham số nằm trong query string (request.args), không phải lúc nào cũng có trong form_dict
        sy_id = (
            (getattr(frappe.request, "args", None) or {}).get("school_year_id")
            or frappe.form_dict.get("school_year_id")
            or ""
        )
        sy_id = (sy_id or "").strip()
        if sy_id and rooms:
            rnames = [r["name"] for r in rooms]
            ya_rows = frappe.get_all(
                "ERP Administrative Room Yearly Assignment",
                filters={"room": ["in", rnames], "school_year_id": sy_id},
                fields=[
                    "name",
                    "room",
                    "display_title_vn",
                    "display_title_en",
                    "homeroom_teacher_name",
                    "status",
                ],
            )
            ya_map = {y["room"]: y for y in ya_rows}
            parent_to_room = {y["name"]: y["room"] for y in ya_rows}
            ya_room_ids = {y["room"] for y in ya_rows}
            yearly_pic_by_room = {}
            ya_parent_ids = [y["name"] for y in ya_rows]
            if ya_parent_ids:
                pic_rows = frappe.get_all(
                    "ERP Administrative Room Yearly PIC",
                    filters={"parent": ["in", ya_parent_ids]},
                    fields=["parent", "user", "full_name", "role_label"],
                    order_by="idx asc",
                )
                uids = list({r.user for r in pic_rows if r.get("user")})
                user_images = {}
                if uids:
                    for urow in frappe.get_all("User", filters={"name": ["in", uids]}, fields=["name", "user_image"]):
                        user_images[urow.name] = urow.user_image or ""
                for pr in pic_rows:
                    rid = parent_to_room.get(pr.parent)
                    if not rid:
                        continue
                    yearly_pic_by_room.setdefault(rid, []).append(
                        {
                            "user": pr.user,
                            "full_name": pr.full_name or "",
                            "user_image": user_images.get(pr.user, ""),
                            "designation": (pr.role_label or "").strip(),
                        }
                    )
            for room in rooms:
                y = ya_map.get(room["name"])
                room["yearly_assignment_display"] = y.get("display_title_vn") if y else None
                room["yearly_assignment_display_en"] = y.get("display_title_en") if y else None
                room["yearly_homeroom_name"] = y.get("homeroom_teacher_name") if y else None
                room["yearly_assignment_status"] = y.get("status") if y else None
                # PIC theo gán năm (Room Yearly Assignment) — ưu tiên Yearly PIC khi đã có dòng
                # Nếu có YA nhưng chưa nhập PIC trong Yearly PIC: giữ PIC bảng con trên Room (tránh cột trống trên FE)
                if room["name"] in ya_room_ids:
                    yearly = yearly_pic_by_room.get(room["name"], [])
                    room["responsible_users"] = yearly if yearly else by_room.get(room["name"], [])
        else:
            for room in rooms:
                room["yearly_assignment_display"] = None
                room["yearly_assignment_display_en"] = None
                room["yearly_homeroom_name"] = None
                room["yearly_assignment_status"] = None

        _enrich_room_responsible_users_employee_codes(rooms)

        return success_response(
            data=rooms,
            message="Rooms fetched successfully",
            meta={"total_count": len(rooms)}
        )
        
    except Exception as e:
        frappe.log_error(f"Error fetching rooms: {str(e)}")
        return {
            "success": False,
            "data": [],
            "total_count": 0,
            "message": f"Error fetching rooms: {str(e)}"
        }


@frappe.whitelist(allow_guest=False)
def get_room_by_id():
    """Get a specific room by ID"""
    try:
        # Get room_id from JSON payload or form_dict  
        room_id = None
        if frappe.request.data:
            try:
                json_data = json.loads(frappe.request.data)
                if json_data and 'room_id' in json_data:
                    room_id = json_data['room_id']
            except (json.JSONDecodeError, TypeError):
                pass

        # Fallback to form_dict
        if not room_id:
            room_id = frappe.local.form_dict.get('room_id')
            
        if not room_id:
            return error_response("Room ID is required")
        
        # Get current user's campus
        campus_id = get_current_campus_from_context()

        if not campus_id:
            # Fallback: try to get first available campus from database
            try:
                first_campus = frappe.db.get_value("SIS Campus", {}, "name", order_by="creation asc")
                campus_id = first_campus or "CAMPUS-00001"
            except Exception:
                # Final fallback to known campus
                campus_id = "CAMPUS-00001"
        
        # Get room and check if it belongs to a building in this campus
        room = frappe.get_doc("ERP Administrative Room", room_id)
        
        if not room:
            return error_response("Room not found")
        
        # Check if the room's building belongs to this campus
        building_exists = frappe.db.exists(
            "ERP Administrative Building",
            {
                "name": room.building_id,
                "campus_id": campus_id
            }
        )
        
        if not building_exists:
            return error_response("Room not found or access denied")
        
        return success_response(
            data={
                "name": room.name,
                "title_vn": room.title_vn,
                "title_en": room.title_en,
                "short_title": room.short_title,
                "physical_code": getattr(room, "physical_code", None),
                "room_number": getattr(room, "room_number", None),
                "needs_review": getattr(room, "needs_review", 0),
                "is_active": getattr(room, "is_active", 1),
                "capacity": room.capacity,
                "room_type": room.room_type,
                "building_id": room.building_id,
            },
            message="Room fetched successfully",
        )
        
    except Exception as e:
        frappe.log_error(f"Error fetching room: {str(e)}")
        return error_response(f"Error fetching room: {str(e)}")


@frappe.whitelist(allow_guest=False)
def create_room():
    """Create a new room - SIMPLE VERSION"""
    try:
        # Get data from request - follow Education Stage pattern
        data = {}
        
        # First try to get JSON data from request body
        if frappe.request.data:
            try:
                json_data = json.loads(frappe.request.data)
                if json_data:
                    data = json_data
                    frappe.logger().info(f"Received JSON data for create_room: {data}")
                else:
                    data = frappe.local.form_dict
                    frappe.logger().info(f"Received form data for create_room: {data}")
            except (json.JSONDecodeError, TypeError):
                # If JSON parsing fails, use form_dict
                data = frappe.local.form_dict
                frappe.logger().info(f"JSON parsing failed, using form data for create_room: {data}")
        else:
            # Fallback to form_dict
            data = frappe.local.form_dict
            frappe.logger().info(f"No request data, using form_dict for create_room: {data}")
        
        # Extract values from data
        title_vn = data.get("title_vn")
        title_en = data.get("title_en")
        short_title = data.get("short_title")
        capacity = data.get("capacity")
        room_type = data.get("room_type")
        building_id = data.get("building_id")
        room_number = (data.get("room_number") or "").strip()
        is_active = 1 if data.get("is_active", True) not in (False, "0", 0, None, "") else 0

        # Luồng mới: building + room_number -> physical_code tự sinh
        if room_number and building_id and room_type:
            pass
        elif not title_vn or not short_title or not room_type or not building_id:
            return {
                "success": False,
                "data": {},
                "message": "Title VN, short title, room type, and building are required (or use room_number + building + room_type)",
            }
        
        # Get campus from user context
        campus_id = get_current_campus_from_context()

        if not campus_id:
            # Fallback: try to get first available campus from database
            try:
                first_campus = frappe.db.get_value("SIS Campus", {}, "name", order_by="creation asc")
                campus_id = first_campus or "CAMPUS-00001"
            except Exception:
                # Final fallback to known campus
                campus_id = "CAMPUS-00001"
            frappe.logger().warning(f"No campus found for user {frappe.session.user}, using fallback: {campus_id}")
        
        # Get building details to extract campus_id
        building_doc = frappe.get_doc("ERP Administrative Building", building_id)
        building_campus_id = building_doc.campus_id

        # Verify building exists and belongs to same campus (if we have campus context)
        if campus_id and building_campus_id != campus_id:
            return {
                "success": False,
                "data": {},
                "message": "Selected building does not belong to your campus"
            }

        if room_number:
            bst = frappe.db.get_value("ERP Administrative Building", building_id, "short_title") or ""
            pc = f"{bst}.{room_number.upper()}".strip() if bst else ""
            if pc and frappe.db.exists(
                "ERP Administrative Room", {"campus_id": building_campus_id, "physical_code": pc}
            ):
                return {
                    "success": False,
                    "data": {},
                    "message": f"Room physical code '{pc}' already exists on this campus",
                }
            room_doc = frappe.get_doc(
                {
                    "doctype": "ERP Administrative Room",
                    "building_id": building_id,
                    "room_number": room_number.upper(),
                    "title_vn": pc,
                    "title_en": title_en or pc,
                    "short_title": short_title or pc,
                    "capacity": capacity or 0,
                    "room_type": room_type,
                    "campus_id": building_campus_id,
                    "is_active": is_active,
                }
            )
        else:
            existing = frappe.db.exists(
                "ERP Administrative Room",
                {"title_vn": title_vn, "building_id": building_id},
            )

            if existing:
                return {
                    "success": False,
                    "data": {},
                    "message": f"Room with title '{title_vn}' already exists in this building",
                }

            room_doc = frappe.get_doc(
                {
                    "doctype": "ERP Administrative Room",
                    "title_vn": title_vn,
                    "title_en": title_en,
                    "short_title": short_title,
                    "capacity": capacity or 0,
                    "room_type": room_type,
                    "building_id": building_id,
                    "campus_id": building_campus_id,
                }
            )
        
        room_doc.insert()
        frappe.db.commit()
        
        # Return the created data - follow StandardApiResponse pattern
        return success_response(
            data={
                "name": room_doc.name,
                "title_vn": room_doc.title_vn,
                "title_en": room_doc.title_en,
                "short_title": room_doc.short_title,
                "physical_code": getattr(room_doc, "physical_code", None),
                "room_number": getattr(room_doc, "room_number", None),
                "capacity": room_doc.capacity,
                "room_type": room_doc.room_type,
                "building_id": room_doc.building_id,
                "is_active": getattr(room_doc, "is_active", 1),
            },
            message="Room created successfully",
        )
        
    except Exception as e:
        frappe.log_error(f"Error creating room: {str(e)}")
        return error_response(f"Error creating room: {str(e)}")


@frappe.whitelist(allow_guest=False)
def update_room():
    """Update an existing room - SIMPLE VERSION with JSON payload support"""
    try:
        # Get data from request - follow Building pattern
        data = {}
        
        # First try to get JSON data from request body
        if frappe.request.data:
            try:
                json_data = json.loads(frappe.request.data)
                if json_data:
                    data = json_data
            except (json.JSONDecodeError, TypeError):
                # If JSON parsing fails, use form_dict
                data = frappe.local.form_dict
        else:
            # Fallback to form_dict
            data = frappe.local.form_dict
        
        room_id = data.get('room_id')
        if not room_id:
            return error_response("Room ID is required")
        
        # Get campus from user context
        campus_id = get_current_campus_from_context()

        if not campus_id:
            # Fallback: try to get first available campus from database
            try:
                first_campus = frappe.db.get_value("SIS Campus", {}, "name", order_by="creation asc")
                campus_id = first_campus or "CAMPUS-00001"
            except Exception:
                # Final fallback to known campus
                campus_id = "CAMPUS-00001"
        
        # Get existing document and verify access
        try:
            room_doc = frappe.get_doc("ERP Administrative Room", room_id)
            
            # Check if the room's building belongs to this campus
            building_exists = frappe.db.exists(
                "ERP Administrative Building",
                {
                    "name": room_doc.building_id,
                    "campus_id": campus_id
                }
            )
            
            if not building_exists:
                return error_response("Access denied: You don't have permission to modify this room")
                
        except frappe.DoesNotExistError:
            return error_response("Room not found")
        
        # Update fields if provided
        title_vn = data.get('title_vn')
        title_en = data.get('title_en')
        short_title = data.get('short_title')
        capacity = data.get('capacity')
        room_type = data.get('room_type')
        building_id = data.get('building_id')
        
        if title_vn and title_vn != room_doc.title_vn:
            # Check for duplicate room title in the same building
            existing = frappe.db.exists(
                "ERP Administrative Room",
                {
                    "title_vn": title_vn,
                    "building_id": room_doc.building_id,
                    "name": ["!=", room_id]
                }
            )
            if existing:
                return error_response(f"Room with title '{title_vn}' already exists in this building")
            room_doc.title_vn = title_vn
        
        if title_en and title_en != room_doc.title_en:
            room_doc.title_en = title_en
            
        if short_title and short_title != room_doc.short_title:
            room_doc.short_title = short_title
            
        if capacity is not None:
            room_doc.capacity = capacity
            
        if room_type and room_type != room_doc.room_type:
            room_doc.room_type = room_type
            
        if building_id and building_id != room_doc.building_id:
            # Get new building details to extract campus_id
            new_building_doc = frappe.get_doc("ERP Administrative Building", building_id)
            new_building_campus_id = new_building_doc.campus_id

            # Verify new building belongs to same campus (if we have campus context)
            if campus_id and new_building_campus_id != campus_id:
                return error_response("Selected building does not belong to your campus")

            room_doc.building_id = building_id
            room_doc.campus_id = new_building_campus_id  # Update campus_id when building changes

        room_number = data.get("room_number")
        if room_number is not None and str(room_number).strip():
            room_doc.room_number = str(room_number).strip().upper()

        if "is_active" in data:
            room_doc.is_active = 0 if data.get("is_active") in (False, "0", 0, None, "") else 1

        room_doc.save()
        frappe.db.commit()
        
        return success_response(
            data={
                "name": room_doc.name,
                "title_vn": room_doc.title_vn,
                "title_en": room_doc.title_en,
                "short_title": room_doc.short_title,
                "capacity": room_doc.capacity,
                "room_type": room_doc.room_type,
                "building_id": room_doc.building_id
            },
            message="Room updated successfully"
        )
        
    except Exception as e:
        frappe.log_error(f"Error updating room: {str(e)}")
        return error_response(f"Error updating room: {str(e)}")


@frappe.whitelist(allow_guest=False) 
def delete_room():
    """Delete a room"""
    try:
        # Get data from request - follow Building pattern
        data = {}
        
        # First try to get JSON data from request body
        if frappe.request.data:
            try:
                json_data = json.loads(frappe.request.data)
                if json_data:
                    data = json_data
                    room_id = data.get('room_id')
            except (json.JSONDecodeError, TypeError):
                # If JSON parsing fails, use form_dict
                data = frappe.local.form_dict
                room_id = data.get('room_id')
        else:
            # Fallback to form_dict
            data = frappe.local.form_dict
            room_id = data.get('room_id')
        
        if not room_id:
            return error_response("Room ID is required")
        
        # Get campus from user context
        campus_id = get_current_campus_from_context()

        if not campus_id:
            # Fallback: try to get first available campus from database
            try:
                first_campus = frappe.db.get_value("SIS Campus", {}, "name", order_by="creation asc")
                campus_id = first_campus or "CAMPUS-00001"
            except Exception:
                # Final fallback to known campus
                campus_id = "CAMPUS-00001"
        
        # Get existing document and verify access
        try:
            room_doc = frappe.get_doc("ERP Administrative Room", room_id)
            
            # Check if the room's building belongs to this campus
            building_exists = frappe.db.exists(
                "ERP Administrative Building",
                {
                    "name": room_doc.building_id,
                    "campus_id": campus_id
                }
            )
            
            if not building_exists:
                return error_response("Access denied: You don't have permission to delete this room")
                
        except frappe.DoesNotExistError:
            return error_response("Room not found")
        
        # Delete the document
        frappe.delete_doc("ERP Administrative Room", room_id)
        frappe.db.commit()
        
        return success_response(message="Room deleted successfully")

    except LinkExistsError as e:
        # Handle case where room is linked to classes
        error_msg = "Không thể xóa phòng vì nó đang được sử dụng bởi các lớp học. Vui lòng gỡ bỏ liên kết trước khi xóa."
        return error_response(error_msg)

    except Exception as e:
        # Truncate error message to avoid CharacterLengthExceededError when logging
        error_str = str(e)
        if len(error_str) > 200:
            error_str = error_str[:200] + "..."

        frappe.log_error(f"Error deleting room: {error_str}")
        return error_response("Có lỗi xảy ra khi xóa phòng. Vui lòng thử lại sau.")


class RoomExcelImporter:
    """Handle Excel import for Rooms with validation and mapping"""

    def __init__(self, campus_id: str):
        self.campus_id = campus_id
        self.errors: list = []
        self.warnings: list = []
        self.building_mapping = {}

    def validate_excel_structure(self, df):
        """Validate Excel file structure for room import

        Required columns: title_vn, title_en, short_title, room_type, building_title
        Optional columns: capacity
        """
        # Check if dataframe is empty after reading
        if df is None or df.empty:
            self.errors.append("File Excel không thể đọc được hoặc file trống. Vui lòng kiểm tra: 1) File có đúng định dạng .xlsx hoặc .xls không? 2) File có bị hỏng không? 3) Bạn đã điền dữ liệu vào file mẫu chưa?")
            return False

        # Normalize column names
        df = self.normalize_columns(df)

        required_cols = ['title_vn', 'title_en', 'short_title', 'room_type', 'building_title']
        cols_lower = [str(c).strip().lower() for c in df.columns]

        missing_cols = []
        for col in required_cols:
            if col not in cols_lower:
                missing_cols.append(col)

        if missing_cols:
            self.errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            return False

        # Check if there's at least one data row (skip empty rows)
        data_rows = 0
        for idx, row in df.iterrows():
            # Check if at least one required field has data
            has_data = any(str(row.get(col, '')).strip() for col in required_cols)
            if has_data:
                data_rows += 1

        if data_rows == 0:
            self.errors.append("No valid data rows found. Please ensure data starts from row 2 and required columns are filled")
            return False

        return True

    def normalize_columns(self, df):
        """Rename Vietnamese/variant headers to canonical English headers"""
        try:
            canonical_map = {
                'tên tiếng việt': 'title_vn',
                'tên vn': 'title_vn',
                'vietnamese name': 'title_vn',
                'tên tiếng anh': 'title_en',
                'tên en': 'title_en',
                'english name': 'title_en',
                'ký hiệu': 'short_title',
                'short title': 'short_title',
                'mã phòng': 'short_title',
                'loại phòng': 'room_type',
                'room type': 'room_type',
                'tòa nhà': 'building_title',
                'toà nhà': 'building_title',
                'building': 'building_title',
                'tên tòa nhà': 'building_title',
                'tên toà nhà': 'building_title',
                'sức chứa': 'capacity',
                'capacity': 'capacity',
                'số lượng': 'capacity'
            }
            rename_map = {}
            for col in df.columns:
                key = str(col).strip().lower()
                if key in canonical_map:
                    rename_map[col] = canonical_map[key]
            if rename_map:
                df = df.rename(columns=rename_map)
        except Exception:
            pass
        return df

    def normalize_room_type(self, room_type_str):
        """Convert Vietnamese room types to English codes"""
        if not room_type_str:
            return None

        room_type_lower = str(room_type_str).strip().lower()

        # Map Vietnamese names to English codes
        room_type_mapping = {
            'phòng lớp học': 'classroom_room',
            'phòng học': 'classroom_room',
            'lớp học': 'classroom_room',
            'phòng họp': 'meeting_room',
            'họp': 'meeting_room',
            'hội trường': 'auditorium',
            'auditorium': 'auditorium',
            'sân ngoài trời': 'outdoor',
            'sân': 'outdoor',
            'phòng làm việc': 'office',
            'văn phòng': 'office',
            'phòng chức năng': 'function_room',
            'chức năng': 'function_room'
        }

        # Direct English mapping
        direct_mapping = {
            'classroom_room': 'classroom_room',
            'meeting_room': 'meeting_room',
            'auditorium': 'auditorium',
            'outdoor': 'outdoor',
            'office': 'office',
            'function_room': 'function_room'
        }

        # Try Vietnamese mapping first, then direct English
        if room_type_lower in room_type_mapping:
            return room_type_mapping[room_type_lower]
        elif room_type_lower in direct_mapping:
            return direct_mapping[room_type_lower]

        return None

    def load_building_mapping(self):
        """Load building mapping for the campus"""
        try:
            buildings = frappe.get_all(
                "ERP Administrative Building",
                filters={"campus_id": self.campus_id},
                fields=["name", "title_vn", "title_en", "short_title"]
            )

            for building in buildings:
                # Map by title_vn, title_en, and short_title
                self.building_mapping[building.title_vn.lower()] = building.name
                self.building_mapping[building.title_en.lower()] = building.name
                self.building_mapping[building.short_title.lower()] = building.name

        except Exception as e:
            self.errors.append(f"Error loading building mapping: {str(e)}")

    def find_building_id(self, building_title):
        """Find building ID by title"""
        if not building_title:
            return None

        title_lower = str(building_title).strip().lower()
        return self.building_mapping.get(title_lower)

    def validate_room_data(self, row_data):
        """Validate individual room data"""
        errors = []

        # Required fields
        if not row_data.get('title_vn'):
            errors.append("Tên tiếng Việt là bắt buộc")
        if not row_data.get('short_title'):
            errors.append("Ký hiệu phòng là bắt buộc")

        # Room type validation
        room_type = self.normalize_room_type(row_data.get('room_type'))
        if not room_type:
            errors.append(f"Invalid room_type: {row_data.get('room_type')}")
        else:
            row_data['room_type'] = room_type  # Update with normalized value

        # Building validation (optional)
        building_id = self.find_building_id(row_data.get('building_title'))
        if building_id:
            row_data['building_id'] = building_id
        else:
            # Building not found, log warning but don't fail validation
            if row_data.get('building_title'):
                self.warnings.append(f"Toà nhà '{row_data.get('building_title')}' không tìm thấy trong hệ thống. Phòng sẽ được tạo mà không gán tòa nhà.")
            row_data['building_id'] = None

        # Capacity validation (optional)
        capacity = row_data.get('capacity')
        if capacity is not None and str(capacity).strip():
            try:
                # Handle NaN values from pandas
                if str(capacity).lower() == 'nan' or (hasattr(capacity, 'isnan') and capacity.isnan()):
                    # Keep as None for empty capacity
                    row_data['capacity'] = None
                else:
                    capacity = int(float(capacity))
                    if capacity < 0:
                        errors.append("Sức chứa phải là số không âm")
                    row_data['capacity'] = capacity
            except (ValueError, TypeError):
                errors.append("Sức chứa phải là số hợp lệ")
        else:
            # Keep as None for empty capacity
            row_data['capacity'] = None

        return errors

    def process_excel_import(self, file_path, dry_run=False):
        """Process Excel import for rooms"""
        if not pd:
            return {"success": False, "message": "pandas library not available"}

        try:
            # Debug: Check if file exists and readable
            import os
            debug_info = []
            debug_info.append(f"Excel file path: {file_path}")
            debug_info.append(f"File exists: {os.path.exists(file_path)}")

            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                debug_info.append(f"File size: {file_size} bytes")

                if file_size == 0:
                    return {
                        "success": False,
                        "message": "Excel file is empty (0 bytes)",
                        "errors": ["Uploaded file is empty"],
                        "debug_info": debug_info,
                        "warnings": self.warnings
                    }
            else:
                return {
                    "success": False,
                    "message": "Excel file not found",
                    "errors": ["File was not uploaded successfully"],
                    "debug_info": debug_info,
                    "warnings": self.warnings
                }

            # Load building mapping
            self.load_building_mapping()

            # Basic file validation before pandas
            debug_info.append("Basic file validation...")
            try:
                with open(file_path, 'rb') as f:
                    file_header = f.read(512)  # Read first 512 bytes

                file_size = len(file_header)
                debug_info.append(f"File header size: {file_size} bytes")

                if file_size < 4:
                    return {
                        "success": False,
                        "message": "File quá nhỏ, có thể không phải file Excel hợp lệ.",
                        "errors": ["File size too small"],
                        "debug_info": debug_info,
                        "warnings": self.warnings
                    }

                # Check for Excel signatures
                header_hex = file_header[:8].hex()
                debug_info.append(f"File header (hex): {header_hex}")

                # Check for common Excel signatures
                if header_hex.startswith('504b0304'):  # PK\x03\x04 - ZIP/XLSX
                    debug_info.append("Detected ZIP/XLSX file signature")
                elif header_hex.startswith('d0cf11e0a1b11ae1'):  # XLS signature
                    debug_info.append("Detected XLS file signature")
                else:
                    debug_info.append("Warning: File does not have standard Excel signature")

            except Exception as file_error:
                debug_info.append(f"Error reading file header: {str(file_error)}")
                return {
                    "success": False,
                    "message": f"Không thể đọc header của file: {str(file_error)}",
                    "errors": [f"File header read error: {str(file_error)}"],
                    "debug_info": debug_info,
                    "warnings": self.warnings
                }

            # Read Excel file (skip first row if it's sample data)
            debug_info.append("Attempting to read Excel file with pandas...")
            try:
                # First check if pandas is available
                if not pd:
                    return {
                        "success": False,
                        "message": "Pandas library not available on server",
                        "errors": ["Server configuration error: pandas library missing"],
                        "debug_info": debug_info,
                        "warnings": self.warnings
                    }

                # Try reading Excel file with multiple approaches
                df = None
                excel_error = None

                # Approach 1: Standard read_excel with openpyxl
                try:
                    df = pd.read_excel(file_path, header=0, engine='openpyxl')
                    debug_info.append("Successfully read with pandas + openpyxl engine")
                except Exception as e1:
                    debug_info.append(f"Pandas + openpyxl failed: {str(e1)}")
                    excel_error = e1

                    # Approach 2: Try direct openpyxl if available
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(file_path, data_only=True)
                        sheet = wb.active

                        # Convert to pandas DataFrame manually
                        data = []
                        headers = []
                        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                            if row_idx == 0:  # First row is headers
                                headers = [str(cell) if cell is not None else '' for cell in row]
                            else:
                                data.append([cell for cell in row])

                        if headers and data:
                            df = pd.DataFrame(data, columns=headers)
                            debug_info.append("Successfully read with direct openpyxl")
                        else:
                            raise Exception("No headers or data found")

                    except ImportError:
                        debug_info.append("openpyxl not available for direct reading")
                    except Exception as e2:
                        debug_info.append(f"Direct openpyxl failed: {str(e2)}")

                        # Approach 3: Try xlrd engine for .xls files
                        try:
                            df = pd.read_excel(file_path, header=0, engine='xlrd')
                            debug_info.append("Successfully read with pandas + xlrd engine")
                        except Exception as e3:
                            debug_info.append(f"Pandas + xlrd also failed: {str(e3)}")

                            # Approach 4: Try without specifying engine
                            try:
                                df = pd.read_excel(file_path, header=0)
                                debug_info.append("Successfully read without specifying engine")
                            except Exception as e4:
                                debug_info.append(f"All pandas approaches failed. Last error: {str(e4)}")
                                excel_error = e4

                if df is None:
                    error_msg = f"Không thể đọc file Excel bằng bất kỳ phương pháp nào. Lỗi cuối cùng: {str(excel_error) if excel_error else 'Unknown error'}. Vui lòng kiểm tra: 1) File có đúng định dạng Excel (.xlsx hoặc .xls) không? 2) File có bị hỏng hoặc có mật khẩu không? 3) File có được lưu từ Excel thật sự không?"
                    return {
                        "success": False,
                        "message": error_msg,
                        "errors": [error_msg],
                        "debug_info": debug_info,
                        "warnings": self.warnings
                    }

                debug_info.append(f"Excel file read successfully. Shape: {df.shape}")
                debug_info.append(f"Columns: {list(df.columns)}")

                # Check if dataframe has any data
                if df.empty:
                    debug_info.append(f"DataFrame is empty. Shape: {df.shape}")
                    return {
                        "success": False,
                        "message": "File Excel được đọc thành công nhưng không có dữ liệu. Vui lòng kiểm tra: 1) Bạn đã điền dữ liệu vào file mẫu chưa? 2) Dữ liệu có bắt đầu từ dòng 2 trở đi không? 3) File có bị xóa nhầm dữ liệu không?",
                        "errors": ["File Excel trống hoặc không có dữ liệu hợp lệ"],
                        "debug_info": debug_info,
                        "warnings": self.warnings
                    }

                # Check if there are any non-empty rows (skip header row)
                data_rows = 0
                for idx, row in df.iterrows():
                    # Check if row has any meaningful data
                    has_data = any(str(val).strip() for val in row.values if pd.notna(val))
                    if has_data:
                        data_rows += 1

                debug_info.append(f"Found {data_rows} rows with data (excluding header)")

                if data_rows == 0:
                    # Show sample of what was read
                    sample_rows = []
                    for idx, row in df.head(5).iterrows():
                        sample_rows.append(f"Row {idx + 1}: {dict(row)}")

                    debug_info.extend([
                        "Sample rows from file:",
                        *sample_rows
                    ])

                    return {
                        "success": False,
                        "message": f"File Excel có {len(df)} dòng nhưng không có dữ liệu hợp lệ. Vui lòng kiểm tra: 1) Dữ liệu có bắt đầu từ dòng 2 không? 2) Các cột có đúng tên không? 3) Ô dữ liệu có bị trống không?",
                        "errors": [f"Không tìm thấy dữ liệu hợp lệ trong {len(df)} dòng của file"],
                        "debug_info": debug_info,
                        "warnings": self.warnings
                    }

                debug_info.append(f"First 3 rows: {df.head(3).to_dict('records') if len(df) >= 3 else 'Less than 3 rows'}")

            except Exception as excel_error:
                debug_info.append(f"Unexpected error reading Excel file: {str(excel_error)}")
                error_msg = f"Lỗi không mong muốn khi đọc file Excel: {str(excel_error)}. Vui lòng thử lại hoặc liên hệ hỗ trợ kỹ thuật."
                return {
                    "success": False,
                    "message": error_msg,
                    "errors": [error_msg],
                    "debug_info": debug_info,
                    "warnings": self.warnings
                }

            if not self.validate_excel_structure(df):
                return {
                    "success": False,
                    "message": "Excel structure validation failed",
                    "errors": self.errors,
                    "warnings": self.warnings,
                    "debug_info": debug_info
                }

            # Normalize columns
            df = self.normalize_columns(df)

            # Process each row
            processed_rooms = []
            validation_errors = []

            for index, row in df.iterrows():
                row_data = {
                    'title_vn': row.get('title_vn'),
                    'title_en': row.get('title_en'),
                    'short_title': row.get('short_title'),
                    'room_type': row.get('room_type'),
                    'building_title': row.get('building_title'),
                    'capacity': row.get('capacity')
                }

                # Validate row data
                row_errors = self.validate_room_data(row_data)
                if row_errors:
                    validation_errors.append({
                        "row": index + 2,  # +2 because Excel is 1-indexed and has header
                        "data": row_data,
                        "errors": row_errors
                    })
                else:
                    processed_rooms.append(row_data)

            if validation_errors and dry_run:
                return {
                    "success": False,
                    "message": "Validation failed",
                    "validation_errors": validation_errors,
                    "total_rows": len(df),
                    "valid_rows": len(processed_rooms),
                    "debug_info": debug_info
                }

            if not dry_run and processed_rooms:
                # Create rooms in database
                created_count = 0
                for room_data in processed_rooms:
                    try:
                        room_data_dict = {
                            "doctype": "ERP Administrative Room",
                            "title_vn": room_data['title_vn'],
                            "title_en": room_data.get('title_en') or room_data['title_vn'],  # Fallback to title_vn if empty
                            "short_title": room_data['short_title'],
                            "room_type": room_data['room_type'],
                            "campus_id": self.campus_id
                        }

                        # Only add building_id if it exists
                        if room_data.get('building_id'):
                            room_data_dict["building_id"] = room_data['building_id']

                        # Only add capacity if it has a value
                        if room_data.get('capacity') is not None:
                            room_data_dict["capacity"] = room_data['capacity']

                        room_doc = frappe.get_doc(room_data_dict)
                        room_doc.insert()
                        created_count += 1
                    except Exception as e:
                        self.errors.append(f"Error creating room {room_data.get('title_vn')}: {str(e)}")

                frappe.db.commit()

            result = {
                "success": True,
                "message": f"Import completed successfully. Created {created_count} rooms." if not dry_run else "Dry run completed successfully",
                "total_rows": len(df),
                "created_count": created_count if not dry_run else 0,
                "valid_rows": len(processed_rooms),
                "errors": self.errors,
                "warnings": self.warnings,
                "debug_info": debug_info
            }

            # Add validation errors if any
            if validation_errors:
                result["validation_errors"] = validation_errors
                result["invalid_rows"] = len(validation_errors)

            return result

        except Exception as e:
            debug_info.append(f"Exception occurred: {str(e)}")
            return {
                "success": False,
                "message": f"Error processing Excel import: {str(e)}",
                "errors": self.errors,
                "debug_info": debug_info
            }


def save_uploaded_file(file_data, filename):
    """Save uploaded file temporarily"""
    import os
    import uuid

    # Create temp directory if it doesn't exist
    temp_dir = "/tmp/frappe_uploads"
    if not os.path.exists(temp_dir):
        os.makedirs(temp_dir)

    # Generate unique filename
    unique_filename = f"{uuid.uuid4()}_{filename}"
    file_path = os.path.join(temp_dir, unique_filename)

    # Debug: Check file_data type and size
    file_size = len(file_data) if not hasattr(file_data, 'read') else "unknown (stream)"
    frappe.logger().info(f"Saving file: {filename}, size: {file_size}, type: {type(file_data)}")

    # Save file
    try:
        with open(file_path, 'wb') as f:
            if hasattr(file_data, 'read'):
                content = file_data.read()
                f.write(content)
                actual_size = len(content)
            else:
                f.write(file_data)
                actual_size = len(file_data)

        # Verify file was saved
        if os.path.exists(file_path):
            saved_size = os.path.getsize(file_path)
            frappe.logger().info(f"File saved successfully: {file_path}, saved size: {saved_size}")
            if saved_size == 0:
                frappe.logger().error("File was saved but is empty!")
        else:
            frappe.logger().error(f"File was not saved: {file_path}")

    except Exception as save_error:
        frappe.logger().error(f"Error saving file: {str(save_error)}")
        raise save_error

    return file_path


@frappe.whitelist(allow_guest=False)
def import_rooms():
    """Import rooms from Excel file"""
    try:
        # Get current user's campus
        campus_id = get_current_campus_from_context()

        if not campus_id:
            # Fallback: try to get first available campus from database
            try:
                first_campus = frappe.db.get_value("SIS Campus", {}, "name", order_by="creation asc")
                campus_id = first_campus or "CAMPUS-00001"
            except Exception:
                campus_id = "CAMPUS-00001"

        # Check for dry_run parameter
        dry_run = frappe.local.form_dict.get("dry_run", "false").lower() == "true"

        # Process Excel import if file is provided
        files = frappe.request.files

        if files and 'file' in files:
            file_data = files['file']
            if not file_data:
                return validation_error_response("No file uploaded", {"file": ["No file uploaded"]})

            # Debug file information before saving
            frappe.logger().info(f"File data type: {type(file_data)}")
            frappe.logger().info(f"File data attributes: {dir(file_data) if hasattr(file_data, '__dict__') else 'No __dict__'}")

            if hasattr(file_data, 'filename'):
                frappe.logger().info(f"Original filename: {file_data.filename}")
            if hasattr(file_data, 'content_type'):
                frappe.logger().info(f"Content type: {file_data.content_type}")

            # Save file temporarily
            file_path = save_uploaded_file(file_data, "rooms_import.xlsx")

            # Process Excel import
            importer = RoomExcelImporter(campus_id)
            result = importer.process_excel_import(file_path, dry_run)

            # Add file save info to debug_info if available
            if 'debug_info' not in result:
                result['debug_info'] = []
            result['debug_info'].insert(0, f"File saved to: {file_path}")

            # Clean up temp file
            try:
                import os
                os.remove(file_path)
            except Exception:
                pass

            if result["success"]:
                return success_response(result, result["message"])
            else:
                return validation_error_response(result["message"], {
                    "errors": result.get("errors", []),
                    "validation_errors": result.get("validation_errors", []),
                    "warnings": result.get("warnings", [])
                })
        else:
            return validation_error_response({"file": ["No file uploaded"]})

    except Exception as e:
        frappe.log_error(f"Error importing rooms: {str(e)}")
        return error_response(f"Error importing rooms: {str(e)}")


@frappe.whitelist(methods=["GET"])
def get_import_job_status():
    """Get the status/result of room import background job"""
    try:
        cache_key = f"room_import_result_{frappe.session.user}"
        cached_result = frappe.cache().get_value(cache_key)

        if cached_result:
            frappe.cache().delete_value(cache_key)
            return success_response(cached_result, "Import job completed")
        else:
            return success_response({"status": "processing"}, "Import job is still processing")

    except Exception as e:
        frappe.log_error(f"Error getting import job status: {str(e)}")
        return error_response(f"Error getting import job status: {str(e)}")


@frappe.whitelist(allow_guest=False)
def get_buildings_for_selection():
    """Get buildings for dropdown selection"""
    try:
        # Get current user's campus information from roles
        campus_id = get_current_campus_from_context()
        
        if not campus_id:
            campus_id = "campus-1"
        
        filters = {"campus_id": campus_id}
            
        buildings = frappe.get_all(
            "ERP Administrative Building",
            fields=[
                "name",
                "title_vn",
                "title_en",
                "short_title"
            ],
            filters=filters,
            order_by="title_vn asc"
        )
        
        return success_response(
            data=buildings,
            message="Buildings fetched successfully"
        )
        
    except Exception as e:
        frappe.log_error(f"Error fetching buildings for selection: {str(e)}")
        return {
            "success": False,
            "data": [],
            "message": f"Error fetching buildings: {str(e)}"
        }


class BuildingExcelImporter:
    """Handle Excel import for Buildings with validation and mapping"""

    def __init__(self, campus_id: str):
        self.campus_id = campus_id
        self.errors: list = []
        self.warnings: list = []

    def validate_excel_structure(self, df):
        """Validate Excel file structure for building import

        Required columns: title_vn, title_en, short_title
        Optional columns: description
        """
        # Check if dataframe is empty after reading
        if df is None or df.empty:
            self.errors.append("File Excel không thể đọc được hoặc file trống. Vui lòng kiểm tra: 1) File có đúng định dạng .xlsx hoặc .xls không? 2) File có bị hỏng không? 3) Bạn đã điền dữ liệu vào file mẫu chưa?")
            return False

        # Normalize column names
        df = self.normalize_columns(df)

        required_cols = ['title_vn', 'title_en', 'short_title']
        cols_lower = [str(c).strip().lower() for c in df.columns]

        missing_cols = []
        for col in required_cols:
            if col not in cols_lower:
                missing_cols.append(col)

        if missing_cols:
            self.errors.append(f"Missing required columns: {', '.join(missing_cols)}")
            return False

        # Check if there's at least one data row (skip empty rows)
        data_rows = 0
        for idx, row in df.iterrows():
            # Check if at least one required field has data
            has_data = any(str(row.get(col, '')).strip() for col in required_cols)
            if has_data:
                data_rows += 1

        if data_rows == 0:
            self.errors.append("No valid data rows found. Please ensure data starts from row 2 and required columns are filled")
            return False

        return True

    def normalize_columns(self, df):
        """Rename Vietnamese/variant headers to canonical English headers"""
        try:
            canonical_map = {
                'tên tiếng việt': 'title_vn',
                'tên vn': 'title_vn',
                'vietnamese name': 'title_vn',
                'tên tiếng anh': 'title_en',
                'tên en': 'title_en',
                'english name': 'title_en',
                'ký hiệu': 'short_title',
                'short title': 'short_title',
                'mã tòa nhà': 'short_title',
                'mô tả': 'description',
                'description': 'description'
            }
            rename_map = {}
            for col in df.columns:
                key = str(col).strip().lower()
                if key in canonical_map:
                    rename_map[col] = canonical_map[key]
            if rename_map:
                df = df.rename(columns=rename_map)
        except Exception:
            pass
        return df

    def process_excel_import(self, file_path, dry_run=False):
        """Process Excel import for buildings"""
        if not pd:
            return {"success": False, "message": "pandas library not available"}

        try:
            # Debug: Check if file exists and readable
            import os
            debug_info = []
            debug_info.append(f"Excel file path: {file_path}")
            debug_info.append(f"File exists: {os.path.exists(file_path)}")

            if os.path.exists(file_path):
                file_size = os.path.getsize(file_path)
                debug_info.append(f"File size: {file_size} bytes")

                if file_size == 0:
                    return {
                        "success": False,
                        "message": "Excel file is empty (0 bytes)",
                        "errors": ["Uploaded file is empty"],
                        "debug_info": debug_info,
                        "warnings": self.warnings
                    }
            else:
                return {
                    "success": False,
                    "message": "Excel file not found",
                    "errors": ["File was not uploaded successfully"],
                    "debug_info": debug_info,
                    "warnings": self.warnings
                }

            # Basic file validation before pandas
            debug_info.append("Basic file validation...")
            try:
                with open(file_path, 'rb') as f:
                    file_header = f.read(512)  # Read first 512 bytes

                file_size = len(file_header)
                debug_info.append(f"File header size: {file_size} bytes")

                if file_size < 4:
                    return {
                        "success": False,
                        "message": "File quá nhỏ, có thể không phải file Excel hợp lệ.",
                        "errors": ["File size too small"],
                        "debug_info": debug_info,
                        "warnings": self.warnings
                    }

                # Check for Excel signatures
                header_hex = file_header[:8].hex()
                debug_info.append(f"File header (hex): {header_hex}")

                # Check for common Excel signatures
                if header_hex.startswith('504b0304'):  # PK\x03\x04 - ZIP/XLSX
                    debug_info.append("Detected ZIP/XLSX file signature")
                elif header_hex.startswith('d0cf11e0a1b11ae1'):  # XLS signature
                    debug_info.append("Detected XLS file signature")
                else:
                    debug_info.append("Warning: File does not have standard Excel signature")

            except Exception as file_error:
                debug_info.append(f"Error reading file header: {str(file_error)}")
                return {
                    "success": False,
                    "message": f"Không thể đọc header của file: {str(file_error)}",
                    "errors": [f"File header read error: {str(file_error)}"],
                    "debug_info": debug_info,
                    "warnings": self.warnings
                }

            # Read Excel file (skip first row if it's sample data)
            debug_info.append("Attempting to read Excel file with pandas...")
            try:
                # First check if pandas is available
                if not pd:
                    return {
                        "success": False,
                        "message": "Pandas library not available on server",
                        "errors": ["Server configuration error: pandas library missing"],
                        "debug_info": debug_info,
                        "warnings": self.warnings
                    }

                # Try reading Excel file with multiple approaches
                df = None
                excel_error = None

                # Approach 1: Standard read_excel with openpyxl
                try:
                    df = pd.read_excel(file_path, header=0, engine='openpyxl')
                    debug_info.append("Successfully read with pandas + openpyxl engine")
                except Exception as e1:
                    debug_info.append(f"Pandas + openpyxl failed: {str(e1)}")
                    excel_error = e1

                    # Approach 2: Try direct openpyxl if available
                    try:
                        import openpyxl
                        wb = openpyxl.load_workbook(file_path, data_only=True)
                        sheet = wb.active

                        # Convert to pandas DataFrame manually
                        data = []
                        headers = []
                        for row_idx, row in enumerate(sheet.iter_rows(values_only=True)):
                            if row_idx == 0:  # First row is headers
                                headers = [str(cell) if cell is not None else '' for cell in row]
                            else:
                                data.append([cell for cell in row])

                        if headers and data:
                            df = pd.DataFrame(data, columns=headers)
                            debug_info.append("Successfully read with direct openpyxl")
                        else:
                            raise Exception("No headers or data found")

                    except ImportError:
                        debug_info.append("openpyxl not available for direct reading")
                    except Exception as e2:
                        debug_info.append(f"Direct openpyxl failed: {str(e2)}")

                        # Approach 3: Try xlrd engine for .xls files
                        try:
                            df = pd.read_excel(file_path, header=0, engine='xlrd')
                            debug_info.append("Successfully read with pandas + xlrd engine")
                        except Exception as e3:
                            debug_info.append(f"Pandas + xlrd also failed: {str(e3)}")

                            # Approach 4: Try without specifying engine
                            try:
                                df = pd.read_excel(file_path, header=0)
                                debug_info.append("Successfully read without specifying engine")
                            except Exception as e4:
                                debug_info.append(f"All pandas approaches failed. Last error: {str(e4)}")
                                excel_error = e4

                if df is None:
                    error_msg = f"Không thể đọc file Excel bằng bất kỳ phương pháp nào. Lỗi cuối cùng: {str(excel_error) if excel_error else 'Unknown error'}. Vui lòng kiểm tra: 1) File có đúng định dạng Excel (.xlsx hoặc .xls) không? 2) File có bị hỏng hoặc có mật khẩu không? 3) File có được lưu từ Excel thật sự không?"
                    return {
                        "success": False,
                        "message": error_msg,
                        "errors": [error_msg],
                        "debug_info": debug_info,
                        "warnings": self.warnings
                    }

                debug_info.append(f"Excel file read successfully. Shape: {df.shape}")
                debug_info.append(f"Columns: {list(df.columns)}")

                # Check if dataframe has any data
                if df.empty:
                    debug_info.append(f"DataFrame is empty. Shape: {df.shape}")
                    return {
                        "success": False,
                        "message": "File Excel được đọc thành công nhưng không có dữ liệu. Vui lòng kiểm tra: 1) Bạn đã điền dữ liệu vào file mẫu chưa? 2) Dữ liệu có bắt đầu từ dòng 2 trở đi không? 3) File có bị xóa nhầm dữ liệu không?",
                        "errors": ["File Excel trống hoặc không có dữ liệu hợp lệ"],
                        "debug_info": debug_info,
                        "warnings": self.warnings
                    }

                # Check if there are any non-empty rows (skip header row)
                data_rows = 0
                for idx, row in df.iterrows():
                    # Check if row has any meaningful data
                    has_data = any(str(val).strip() for val in row.values if pd.notna(val))
                    if has_data:
                        data_rows += 1

                debug_info.append(f"Found {data_rows} rows with data (excluding header)")

                if data_rows == 0:
                    # Show sample of what was read
                    sample_rows = []
                    for idx, row in df.head(5).iterrows():
                        sample_rows.append(f"Row {idx + 1}: {dict(row)}")

                    debug_info.extend([
                        "Sample rows from file:",
                        *sample_rows
                    ])

                    return {
                        "success": False,
                        "message": f"File Excel có {len(df)} dòng nhưng không có dữ liệu hợp lệ. Vui lòng kiểm tra: 1) Dữ liệu có bắt đầu từ dòng 2 không? 2) Các cột có đúng tên không? 3) Ô dữ liệu có bị trống không?",
                        "errors": [f"Không tìm thấy dữ liệu hợp lệ trong {len(df)} dòng của file"],
                        "debug_info": debug_info,
                        "warnings": self.warnings
                    }

                debug_info.append(f"First 3 rows: {df.head(3).to_dict('records') if len(df) >= 3 else 'Less than 3 rows'}")

            except Exception as excel_error:
                debug_info.append(f"Unexpected error reading Excel file: {str(excel_error)}")
                error_msg = f"Lỗi không mong muốn khi đọc file Excel: {str(excel_error)}. Vui lòng thử lại hoặc liên hệ hỗ trợ kỹ thuật."
                return {
                    "success": False,
                    "message": error_msg,
                    "errors": [error_msg],
                    "debug_info": debug_info,
                    "warnings": self.warnings
                }

            if not self.validate_excel_structure(df):
                return {
                    "success": False,
                    "message": "Excel structure validation failed",
                    "errors": self.errors,
                    "warnings": self.warnings,
                    "debug_info": debug_info
                }

            # Normalize columns
            df = self.normalize_columns(df)

            # Process each row
            processed_buildings = []
            validation_errors = []

            for index, row in df.iterrows():
                row_data = {
                    'title_vn': row.get('title_vn'),
                    'title_en': row.get('title_en'),
                    'short_title': row.get('short_title'),
                    'description': row.get('description')
                }

                # Validate row data
                row_errors = self.validate_building_data(row_data)
                if row_errors:
                    validation_errors.append({
                        "row": index + 2,  # +2 because Excel is 1-indexed and has header
                        "data": row_data,
                        "errors": row_errors
                    })
                else:
                    processed_buildings.append(row_data)

            if validation_errors and dry_run:
                return {
                    "success": False,
                    "message": "Validation failed",
                    "validation_errors": validation_errors,
                    "total_rows": len(df),
                    "valid_rows": len(processed_buildings),
                    "debug_info": debug_info
                }

            if not dry_run and processed_buildings:
                # Create buildings in database
                created_count = 0
                for building_data in processed_buildings:
                    try:
                        building_doc = frappe.get_doc({
                            "doctype": "ERP Administrative Building",
                            "title_vn": building_data['title_vn'],
                            "title_en": building_data['title_en'],
                            "short_title": building_data['short_title'],
                            "description": building_data.get('description', ''),
                            "campus_id": self.campus_id
                        })
                        building_doc.insert()
                        created_count += 1
                    except Exception as e:
                        self.errors.append(f"Error creating building {building_data.get('title_vn')}: {str(e)}")

                frappe.db.commit()

            return {
                "success": True,
                "message": f"Import completed successfully. Created {created_count} buildings." if not dry_run else "Dry run completed successfully",
                "total_rows": len(df),
                "created_count": created_count if not dry_run else 0,
                "errors": self.errors,
                "warnings": self.warnings,
                "debug_info": debug_info
            }

        except Exception as e:
            debug_info.append(f"Exception occurred: {str(e)}")
            return {
                "success": False,
                "message": f"Error processing Excel import: {str(e)}",
                "errors": self.errors,
                "debug_info": debug_info
            }

    def validate_building_data(self, row_data):
        """Validate building row data"""
        errors = []

        # Required fields validation
        if not row_data.get('title_vn') or str(row_data['title_vn']).strip() == '':
            errors.append("Tên tiếng Việt không được để trống")

        if not row_data.get('title_en') or str(row_data['title_en']).strip() == '':
            errors.append("Tên tiếng Anh không được để trống")

        if not row_data.get('short_title') or str(row_data['short_title']).strip() == '':
            errors.append("Ký hiệu tòa nhà không được để trống")

        # Check for duplicates in the same import
        if hasattr(self, '_temp_buildings'):
            for existing in self._temp_buildings:
                if (existing['title_vn'] == row_data['title_vn'] or
                    existing['title_en'] == row_data['title_en'] or
                    existing['short_title'] == row_data['short_title']):
                    errors.append("Tòa nhà đã tồn tại trong file import này")
                    break

        # Check against existing buildings in database
        if row_data.get('title_vn'):
            existing = frappe.db.exists("ERP Administrative Building", {
                "title_vn": row_data['title_vn'],
                "campus_id": self.campus_id
            })
            if existing:
                errors.append(f"Tòa nhà với tên '{row_data['title_vn']}' đã tồn tại")

        if row_data.get('short_title'):
            existing = frappe.db.exists("ERP Administrative Building", {
                "short_title": row_data['short_title'],
                "campus_id": self.campus_id
            })
            if existing:
                errors.append(f"Tòa nhà với ký hiệu '{row_data['short_title']}' đã tồn tại")

        return errors


@frappe.whitelist(allow_guest=False)
def import_buildings():
    """Import buildings from Excel file"""
    try:
        # Get current user's campus
        campus_id = get_current_campus_from_context()

        if not campus_id:
            # Fallback: try to get first available campus from database
            try:
                first_campus = frappe.db.get_value("SIS Campus", {}, "name", order_by="creation asc")
                campus_id = first_campus or "CAMPUS-00001"
            except Exception:
                campus_id = "CAMPUS-00001"

        # Check for dry_run parameter
        dry_run = frappe.local.form_dict.get("dry_run", "false").lower() == "true"

        # Process Excel import if file is provided
        files = frappe.request.files

        if files and 'file' in files:
            file_data = files['file']
            if not file_data:
                return validation_error_response("No file uploaded", {"file": ["No file uploaded"]})

            # Debug file information before saving
            frappe.logger().info(f"File data type: {type(file_data)}")
            frappe.logger().info(f"File data attributes: {dir(file_data) if hasattr(file_data, '__dict__') else 'No __dict__'}")

            if hasattr(file_data, 'filename'):
                frappe.logger().info(f"Original filename: {file_data.filename}")
            if hasattr(file_data, 'content_type'):
                frappe.logger().info(f"Content type: {file_data.content_type}")

            # Save file temporarily
            file_path = save_uploaded_file(file_data, "buildings_import.xlsx")

            # Process Excel import
            importer = BuildingExcelImporter(campus_id)
            result = importer.process_excel_import(file_path, dry_run)

            # Add file save info to debug_info if available
            if 'debug_info' not in result:
                result['debug_info'] = []
            result['debug_info'].insert(0, f"File saved to: {file_path}")

            # Clean up temp file
            try:
                import os
                os.remove(file_path)
            except Exception:
                pass

            if result["success"]:
                return success_response(result, result["message"])
            else:
                return validation_error_response(result["message"], {
                    "errors": result.get("errors", []),
                    "validation_errors": result.get("validation_errors", []),
                    "warnings": result.get("warnings", [])
                })
        else:
            return validation_error_response({"file": ["No file uploaded"]})

    except Exception as e:
        frappe.log_error(f"Error importing buildings: {str(e)}")
        return error_response(f"Error importing buildings: {str(e)}")


@frappe.whitelist(allow_guest=False)
def get_timetable_subjects_for_room_class(education_grade: str = None):
    """Get timetable subjects filtered by education grade for room class assignment"""
    try:
        # Handle both GET and POST requests
        if not education_grade:
            form = frappe.local.form_dict or {}
            education_grade = form.get("education_grade")
            if not education_grade and frappe.request and frappe.request.args:
                education_grade = frappe.request.args.get('education_grade')
            # Also check in request body for POST requests
            if not education_grade and frappe.request and frappe.request.data:
                try:
                    body = frappe.request.data.decode('utf-8') if isinstance(frappe.request.data, bytes) else frappe.request.data
                    data = json.loads(body or '{}')
                    education_grade = data.get('education_grade')
                except Exception:
                    pass

        if not education_grade:
            return validation_error_response("education_grade is required", {"education_grade": ["education_grade is required"]})

        # Get education stage from education grade
        education_grade_doc = frappe.get_all(
            "SIS Education Grade",
            fields=["education_stage_id"],
            filters={"name": education_grade},
            limit=1
        )

        if not education_grade_doc or not education_grade_doc[0].get("education_stage_id"):
            return validation_error_response("Education grade not found or has no education stage", {"education_grade": ["Invalid education grade"]})

        education_stage_id = education_grade_doc[0]["education_stage_id"]

        # Get timetable subjects filtered by education stage
        subjects = frappe.get_all(
            "SIS Timetable Subject",
            fields=[
                "name",
                "title_vn",
                "title_en",
                "education_stage_id"
            ],
            filters={"education_stage_id": education_stage_id},
            order_by="title_vn asc"
        )

        # Enhance with additional info
        enhanced_subjects = []
        for subject in subjects:
            enhanced_subject = {
                "name": subject.name,
                "subject_name": subject.title_vn or subject.title_en,
                "subject_code": subject.name,  # Use name as code for now
                "education_grade": education_grade,
                "education_stage_id": subject.education_stage_id
            }

            # Get education stage title if education_stage_id exists
            if subject.education_stage_id:
                stage_info = frappe.get_all(
                    "SIS Education Stage",
                    fields=["title_vn"],
                    filters={"name": subject.education_stage_id},
                    limit=1
                )
                if stage_info:
                    enhanced_subject["education_stage_title"] = stage_info[0].get("title_vn")

            enhanced_subjects.append(enhanced_subject)

        return success_response(
            data=enhanced_subjects,
            message="Timetable subjects retrieved successfully"
        )

    except Exception as e:
        frappe.log_error(f"Error getting timetable subjects: {str(e)}")
        return error_response(f"Error getting timetable subjects: {str(e)}")


@frappe.whitelist(allow_guest=False)
def available_rooms_for_class_selection(class_type: str):
    """Get available rooms for class selection based on class type"""
    try:
        if not class_type:
            return validation_error_response("class_type is required", {"class_type": ["class_type is required"]})

        # Get all rooms
        rooms = frappe.get_all(
            "ERP Administrative Room",
            fields=[
                "name",
                "title_vn",
                "title_en",
                "short_title",
                "room_type",
                "building_id",
                "capacity"
            ],
            filters={"campus_id": frappe.local.session.get("campus_id")},
            order_by="title_vn asc"
        )

        enhanced_rooms = []

        for room in rooms:
            # Get building info
            building_title = None
            if room.building_id:
                building_info = frappe.get_all(
                    "ERP Administrative Building",
                    fields=["title_vn"],
                    filters={"name": room.building_id},
                    limit=1
                )
                if building_info:
                    building_title = building_info[0].get("title_vn")

            enhanced_room = {
                "name": room.name,
                "title": room.title_vn,
                "short_title": room.short_title,
                "room_type": room.room_type,
                "building_title": building_title,
                "capacity": room.capacity,
                "suggested_usage": None,
                "suggested_usage_display": None,
                "available": True,
                "reason": None
            }

            # Apply filtering logic based on class_type
            if class_type == "regular":
                # Only classroom rooms available and check if already has homeroom
                if room.room_type != "classroom_room":
                    enhanced_room["available"] = False
                    enhanced_room["reason"] = "Chỉ phòng học mới có thể làm lớp chủ nhiệm"
                else:
                    # Check if room already has a homeroom class
                    room_doc = frappe.get_doc("ERP Administrative Room", room.name)
                    has_homeroom = False
                    if hasattr(room_doc, 'room_classes'):
                        for room_class in room_doc.room_classes:
                            if room_class.usage_type == "homeroom":
                                has_homeroom = True
                                break

                    if has_homeroom:
                        enhanced_room["available"] = False
                        enhanced_room["reason"] = "Phòng này đã có lớp chủ nhiệm"
                    else:
                        enhanced_room["suggested_usage"] = "homeroom"
                        enhanced_room["suggested_usage_display"] = "Lớp chủ nhiệm"

            else:  # mixed or club classes
                # All rooms available, will be functional
                enhanced_room["suggested_usage"] = "functional"
                enhanced_room["suggested_usage_display"] = "Lớp chức năng"

            enhanced_rooms.append(enhanced_room)

        # Filter to only available rooms for regular classes, show all for others
        if class_type == "regular":
            result_rooms = [room for room in enhanced_rooms if room["available"]]
        else:
            result_rooms = enhanced_rooms

        return success_response(
            data=result_rooms,
            message="Available rooms retrieved successfully"
        )

    except Exception as e:
        frappe.log_error(f"Error getting available rooms: {str(e)}")
        return error_response(f"Error getting available rooms: {str(e)}")


@frappe.whitelist(allow_guest=False)
def get_room_classes(room_id: str = None):
    """Get all classes assigned to a room"""
    frappe.logger().info("get_room_classes function called")
    try:
        frappe.logger().info(f"Initial room_id parameter: {room_id}")
        frappe.logger().info(f"frappe.local.form_dict: {frappe.local.form_dict}")
        frappe.logger().info(f"frappe.request: {frappe.request}")
        if frappe.request and hasattr(frappe.request, 'data'):
            frappe.logger().info(f"frappe.request.data: {frappe.request.data}")

        if not room_id:
            form = frappe.local.form_dict or {}
            room_id = form.get("room_id") or form.get("name")
            frappe.logger().info(f"room_id from form: {room_id}")
            if not room_id and frappe.request and frappe.request.args:
                room_id = frappe.request.args.get('room_id') or frappe.request.args.get('name')
                frappe.logger().info(f"room_id from args: {room_id}")
            # Also check in request body for POST requests
            if not room_id and frappe.request and frappe.request.data:
                try:
                    body = frappe.request.data.decode('utf-8') if isinstance(frappe.request.data, bytes) else frappe.request.data
                    data = json.loads(body or '{}')
                    room_id = data.get('room_id')
                except Exception:
                    pass
        if not room_id:
            return validation_error_response("Room ID is required", {"room_id": ["Room ID is required"]})

        frappe.logger().info(f"get_room_classes processing room_id: {room_id}")

        # Lọc theo năm học (optional) — body JSON POST
        sy_filter = None
        if frappe.request and frappe.request.data:
            try:
                _body = frappe.request.data.decode("utf-8") if isinstance(frappe.request.data, bytes) else frappe.request.data
                _data = json.loads(_body or "{}")
                sy_filter = _data.get("school_year_id") or None
            except Exception:
                sy_filter = None

        # Check if room exists
        if not frappe.db.exists("ERP Administrative Room", room_id):
            return not_found_response("Room not found")

        # Get room classes from child table (new method)
        enhanced_classes = []
        child_table_has_data = False

        try:
            # Use frappe.get_all to safely fetch child table records
            room_classes_data = frappe.get_all(
                "ERP Administrative Room Class",
                fields=["name", "class_id", "usage_type", "subject_id", "class_title", "school_year_id", "education_grade", "academic_program", "homeroom_teacher"],
                filters={"parent": room_id, "parenttype": "ERP Administrative Room"},
                order_by="creation asc"
            )
            
            if room_classes_data:
                child_table_has_data = True
                frappe.logger().info(f"Found {len(room_classes_data)} classes in child table")
                
                for room_class in room_classes_data:
                    frappe.logger().info(f"Processing room class: {room_class.class_id}, usage: {room_class.usage_type}")
                    try:
                        class_doc = frappe.get_doc("SIS Class", room_class.class_id)
                        frappe.logger().info(f"Got class doc: {class_doc.name}")
                    except Exception as e:
                        frappe.logger().error(f"Failed to get class doc {room_class.class_id}: {str(e)}")
                        continue

                    # Get education grade title
                    education_grade_title = None
                    if class_doc.education_grade:
                        grade_info = frappe.get_all(
                            "SIS Education Grade",
                            fields=["title_vn"],
                            filters={"name": class_doc.education_grade},
                            limit=1
                        )
                        if grade_info:
                            education_grade_title = grade_info[0].get("title_vn")

                    # Get academic program title
                    academic_program_title = None
                    if class_doc.academic_program:
                        program_info = frappe.get_all(
                            "SIS Academic Program",
                            fields=["title_vn"],
                            filters={"name": class_doc.academic_program},
                            limit=1
                        )
                        if program_info:
                            academic_program_title = program_info[0].get("title_vn")

                    enhanced_class = {
                        "name": class_doc.name,
                        "title": class_doc.title,
                        "short_title": class_doc.short_title,
                        "class_type": class_doc.class_type,
                        "school_year_id": class_doc.school_year_id,
                        "campus_id": class_doc.campus_id,
                        "education_grade": education_grade_title or class_doc.education_grade,
                        "academic_program": academic_program_title or class_doc.academic_program,
                        "homeroom_teacher": class_doc.homeroom_teacher,
                        "creation": class_doc.creation,
                        "modified": class_doc.modified,
                        "usage_type": room_class.usage_type,
                        "usage_type_display": "Lớp chủ nhiệm" if room_class.usage_type == "homeroom" else "Lớp chức năng",
                        "subject_id": room_class.subject_id,
                        "subject_name": None,
                        "subject_code": None
                    }

                    # Get subject info if subject_id exists
                    if room_class.subject_id:
                        try:
                            subject_doc = frappe.get_doc("SIS Timetable Subject", room_class.subject_id)
                            enhanced_class["subject_name"] = subject_doc.title_vn or subject_doc.title_en
                            enhanced_class["subject_code"] = subject_doc.name
                        except Exception as e:
                            frappe.logger().warning(f"Failed to get subject {room_class.subject_id}: {str(e)}")

                    # Add teacher info
                    if class_doc.homeroom_teacher:
                        teacher_info = frappe.get_all(
                            "SIS Teacher",
                            fields=["user_id"],
                            filters={"name": class_doc.homeroom_teacher},
                            limit=1
                        )
                        if teacher_info and teacher_info[0].get("user_id"):
                            user_info = frappe.get_all(
                                "User",
                                fields=["full_name"],
                                filters={"name": teacher_info[0]["user_id"]},
                                limit=1
                            )
                            if user_info:
                                enhanced_class["homeroom_teacher_name"] = user_info[0].get("full_name")

                    enhanced_class["vice_homeroom_teacher"] = getattr(class_doc, "vice_homeroom_teacher", None)
                    if getattr(class_doc, "vice_homeroom_teacher", None):
                        vinfo = frappe.get_all(
                            "SIS Teacher",
                            fields=["user_id"],
                            filters={"name": class_doc.vice_homeroom_teacher},
                            limit=1
                        )
                        if vinfo and vinfo[0].get("user_id"):
                            uinfo = frappe.get_all(
                                "User",
                                fields=["full_name"],
                                filters={"name": vinfo[0]["user_id"]},
                                limit=1
                            )
                            if uinfo:
                                enhanced_class["vice_homeroom_teacher_name"] = uinfo[0].get("full_name")

                    enhanced_classes.append(enhanced_class)

        except Exception as e:
            frappe.logger().warning(f"Error fetching from child table: {str(e)}")
            child_table_has_data = False

        # If child table has no data or failed, fallback to legacy method
        if not child_table_has_data:
            frappe.logger().info(f"Falling back to legacy method for room {room_id}")
            # Fallback to legacy method (classes with room field)
            classes = frappe.get_all(
                "SIS Class",
                fields=[
                    "name",
                    "title",
                    "short_title",
                    "class_type",
                    "school_year_id",
                    "campus_id",
                    "education_grade",
                    "academic_program",
                    "homeroom_teacher",
                    "vice_homeroom_teacher",
                    "creation",
                    "modified"
                ],
                filters={"room": room_id},
                order_by="title asc"
            )

            frappe.logger().info(f"Found {len(classes)} classes via legacy method")
            for class_data in classes:
                # Skip if already added from child table (to avoid duplicates)
                if any(ec["name"] == class_data["name"] for ec in enhanced_classes):
                    continue

                # Get education grade title
                education_grade_title = None
                if class_data.get("education_grade"):
                    grade_info = frappe.get_all(
                        "SIS Education Grade",
                        fields=["title_vn"],
                        filters={"name": class_data["education_grade"]},
                        limit=1
                    )
                    if grade_info:
                        education_grade_title = grade_info[0].get("title_vn")

                # Get academic program title
                academic_program_title = None
                if class_data.get("academic_program"):
                    program_info = frappe.get_all(
                        "SIS Academic Program",
                        fields=["title_vn"],
                        filters={"name": class_data["academic_program"]},
                        limit=1
                    )
                    if program_info:
                        academic_program_title = program_info[0].get("title_vn")

                enhanced_class = class_data.copy()
                enhanced_class["education_grade"] = education_grade_title or class_data.get("education_grade")
                enhanced_class["academic_program"] = academic_program_title or class_data.get("academic_program")

                # Determine usage type based on class_type
                if class_data.get("class_type") == "regular":
                    enhanced_class["usage_type"] = "homeroom"
                    enhanced_class["usage_type_display"] = "Lớp chủ nhiệm"
                else:
                    enhanced_class["usage_type"] = "functional"
                    enhanced_class["usage_type_display"] = "Lớp chức năng"

                # Add teacher info
                if class_data.get("homeroom_teacher"):
                    teacher_info = frappe.get_all(
                        "SIS Teacher",
                        fields=["user_id"],
                        filters={"name": class_data["homeroom_teacher"]},
                        limit=1
                    )
                    if teacher_info and teacher_info[0].get("user_id"):
                        user_info = frappe.get_all(
                            "User",
                            fields=["full_name"],
                            filters={"name": teacher_info[0]["user_id"]},
                            limit=1
                        )
                        if user_info:
                            enhanced_class["homeroom_teacher_name"] = user_info[0].get("full_name")

                if class_data.get("vice_homeroom_teacher"):
                    vinfo = frappe.get_all(
                        "SIS Teacher",
                        fields=["user_id"],
                        filters={"name": class_data["vice_homeroom_teacher"]},
                        limit=1
                    )
                    if vinfo and vinfo[0].get("user_id"):
                        uinfo = frappe.get_all(
                            "User",
                            fields=["full_name"],
                            filters={"name": vinfo[0]["user_id"]},
                            limit=1
                        )
                        if uinfo:
                            enhanced_class["vice_homeroom_teacher_name"] = uinfo[0].get("full_name")

                enhanced_classes.append(enhanced_class)

        # Lọc theo năm học (body JSON) — không “nới” khi rỗng: nếu năm đã chọn chưa gán lớp nào
        # thì trả về danh sách rỗng (tránh UI hiển thị lớp của năm khác).
        if sy_filter:
            enhanced_classes = [c for c in enhanced_classes if (c.get("school_year_id") or "") == sy_filter]

        # Add debug info to response
        debug_info = {
            "room_id": room_id,
            "child_table_has_data": child_table_has_data,
            "school_year_filter": sy_filter,
            "school_year_filter_relaxed": False,
            "total_classes_found": len(enhanced_classes),
            "classes": [{"name": c["name"], "usage_type": c["usage_type"]} for c in enhanced_classes]
        }

        frappe.logger().info(f"Returning {len(enhanced_classes)} classes for room {room_id}")

        return success_response(
            data=enhanced_classes,
            message="Room classes fetched successfully",
            debug_info=debug_info
        )

    except Exception as e:
        frappe.log_error(f"Error fetching room classes: {str(e)}")
        return error_response(f"Error fetching room classes: {str(e)}")


# Nhãn vai trò do hệ thống gán từ lớp CN — mỗi lần đồng bộ sẽ thay thế, không trộn với PIC tay cùng nhãn cũ
ROOM_PIC_AUTO_ROLE_LABELS = frozenset({"Giáo viên chủ nhiệm", "Giáo viên phó chủ nhiệm"})


def _sync_homeroom_teachers_to_room_yearly_pic(room_id: str, class_id: str):
    """
    Phòng lớp học (classroom_room): đồng bộ PIC năm học từ User GVCN + phó CN.
    Giữ các PIC gán tay (nhãn khác hai nhãn trên); hai nhãn trên luôn khớp lớp hiện tại.
    """
    room_rt = (frappe.db.get_value("ERP Administrative Room", room_id, "room_type") or "").lower()
    if room_rt != "classroom_room":
        return

    class_doc = frappe.get_doc("SIS Class", class_id)
    sy = class_doc.school_year_id
    if not sy:
        return

    def _teacher_link_to_user(sis_teacher_id):
        if not sis_teacher_id:
            return None
        uid = frappe.db.get_value("SIS Teacher", sis_teacher_id, "user_id")
        if uid and frappe.db.exists("User", uid):
            return uid
        return None

    pairs = []
    if class_doc.homeroom_teacher:
        u = _teacher_link_to_user(class_doc.homeroom_teacher)
        if u:
            pairs.append((u, "Giáo viên chủ nhiệm"))
    vice = getattr(class_doc, "vice_homeroom_teacher", None)
    if vice:
        u = _teacher_link_to_user(vice)
        if u:
            pairs.append((u, "Giáo viên phó chủ nhiệm"))

    seen = set()
    unique_pairs = []
    for uid, label in pairs:
        if uid in seen:
            continue
        seen.add(uid)
        unique_pairs.append((uid, label))

    ya_name = frappe.db.get_value(
        "ERP Administrative Room Yearly Assignment",
        {"room": room_id, "school_year_id": sy},
        "name",
    )

    if not ya_name and not unique_pairs:
        return

    if ya_name:
        doc = frappe.get_doc("ERP Administrative Room Yearly Assignment", ya_name)
    else:
        doc = frappe.get_doc(
            {
                "doctype": "ERP Administrative Room Yearly Assignment",
                "room": room_id,
                "school_year_id": sy,
                "usage_type": "homeroom_class",
                "status": "active",
            }
        )

    doc.class_id = class_id
    if class_doc.homeroom_teacher:
        doc.homeroom_teacher_id = class_doc.homeroom_teacher
    if vice:
        doc.vice_homeroom_teacher_id = vice

    manual_rows = []
    for r in doc.responsible_users or []:
        uid = getattr(r, "user", None)
        rl = (r.role_label or "").strip()
        if rl in ROOM_PIC_AUTO_ROLE_LABELS:
            continue
        if not uid:
            continue
        fn = r.full_name or frappe.db.get_value("User", uid, "full_name") or ""
        manual_rows.append({"user": uid, "role_label": rl, "full_name": fn})

    manual_users = {m["user"] for m in manual_rows}

    doc.responsible_users = []
    for m in manual_rows:
        doc.append("responsible_users", m)

    for uid, role_label in unique_pairs:
        if uid in manual_users:
            continue
        fn = frappe.db.get_value("User", uid, "full_name") or ""
        doc.append("responsible_users", {"user": uid, "role_label": role_label, "full_name": fn})

    if doc.is_new():
        doc.insert(ignore_permissions=True)
    else:
        doc.save(ignore_permissions=True)
    frappe.db.commit()


def sync_class_homeroom_teachers_to_room_pic(doc, method):
    """
    Hook SIS Class on_update: khi đổi GVCN / phó CN, cập nhật PIC phòng lớp học đã gán lớp CN.
    """
    try:
        before = doc.get_doc_before_save()
        if before:
            h_old = before.get("homeroom_teacher") or ""
            h_new = doc.homeroom_teacher or ""
            v_old = before.get("vice_homeroom_teacher") or ""
            v_new = getattr(doc, "vice_homeroom_teacher", None) or ""
            if h_old == h_new and v_old == v_new:
                return
    except Exception:
        pass

    room_ids = set()
    for r in frappe.get_all(
        "ERP Administrative Room Class",
        fields=["parent"],
        filters={"class_id": doc.name, "usage_type": "homeroom"},
    ):
        room_ids.add(r.parent)
    if getattr(doc, "room", None):
        room_ids.add(doc.room)

    for rid in room_ids:
        if not rid or not frappe.db.exists("ERP Administrative Room", rid):
            continue
        rt = (frappe.db.get_value("ERP Administrative Room", rid, "room_type") or "").lower()
        if rt != "classroom_room":
            continue
        try:
            _sync_homeroom_teachers_to_room_yearly_pic(rid, doc.name)
        except Exception:
            frappe.log_error(frappe.get_traceback(), "sync_class_homeroom_teachers_to_room_pic")

        # Đồng bộ snapshot homeroom_teacher trên dòng Room Class (child)
        try:
            rdoc = frappe.get_doc("ERP Administrative Room", rid)
            if hasattr(rdoc, "room_classes"):
                dirty = False
                for rc in rdoc.room_classes:
                    if rc.class_id == doc.name and rc.usage_type == "homeroom":
                        if rc.homeroom_teacher != (doc.homeroom_teacher or ""):
                            rc.homeroom_teacher = doc.homeroom_teacher
                            dirty = True
                if dirty:
                    rdoc.save(ignore_permissions=True)
                    frappe.db.commit()
        except Exception:
            frappe.logger().warning("sync room_class homeroom_teacher snapshot failed", exc_info=True)


@frappe.whitelist(allow_guest=False, methods=['POST'])
def add_room_class():
    """Add a class to a room"""
    frappe.logger().info("add_room_class called")
    try:
        data = {}
        if frappe.request and frappe.request.data:
            try:
                body = frappe.request.data.decode('utf-8') if isinstance(frappe.request.data, bytes) else frappe.request.data
                data = json.loads(body or '{}')
            except Exception:
                data = frappe.local.form_dict or {}
        else:
            data = frappe.local.form_dict or {}

        required = ["room_id", "class_id"]
        for field in required:
            if not data.get(field):
                return validation_error_response({field: [f"{field} is required"]})

        room_id = data.get("room_id")
        class_id = data.get("class_id")
        usage_type = data.get("usage_type")
        subject_id = data.get("subject_id")  # Optional for functional classes

        # Validate room exists
        if not frappe.db.exists("ERP Administrative Room", room_id):
            return not_found_response("Room not found")

        # Validate class exists
        if not frappe.db.exists("SIS Class", class_id):
            return not_found_response("Class not found")

        # Get class info
        class_info = frappe.get_doc("SIS Class", class_id)

        # Auto-determine usage_type if not provided
        if not usage_type:
            usage_type = "homeroom" if class_info.class_type == "regular" else "functional"
            frappe.logger().info(f"Auto-determined usage_type for class {class_id}: {usage_type} (class_type: {class_info.class_type})")

        # Validate usage_type
        if usage_type not in ["homeroom", "functional"]:
            return validation_error_response("Invalid usage type", {"usage_type": ["Usage type must be 'homeroom' or 'functional'"]})

        # Get room document
        room_doc = frappe.get_doc("ERP Administrative Room", room_id)

        # Check if class is already assigned to this room
        existing_assignment = None
        if hasattr(room_doc, 'room_classes'):
            for room_class in room_doc.room_classes:
                if room_class.class_id == class_id:
                    existing_assignment = room_class
                    break

        if existing_assignment:
            return validation_error_response(
                "Lớp đã được gán cho phòng này",
                {"class_id": ["Lớp này đã được gán cho phòng này rồi."]}
            )

        # Business logic: sync with SIS Class room field for homeroom
        if usage_type == "homeroom":
            # For homeroom usage, only allow if class_type is "regular"
            if class_info.class_type != "regular":
                return validation_error_response(
                    "Không thể đặt lớp này làm lớp chủ nhiệm",
                    {"usage_type": ["Chỉ lớp có loại 'regular' mới có thể làm lớp chủ nhiệm"]}
                )

            # Check if room already has a homeroom class (both in child table and legacy room field)
            # Một phòng chỉ 1 lớp CN / mỗi năm học (theo school_year_id của lớp)
            sy = class_info.school_year_id
            if hasattr(room_doc, 'room_classes'):
                for room_class in room_doc.room_classes:
                    if room_class.usage_type != "homeroom":
                        continue
                    rc_sy = room_class.school_year_id or frappe.db.get_value(
                        "SIS Class", room_class.class_id, "school_year_id"
                    )
                    if rc_sy == sy:
                        return validation_error_response(
                            "Phòng đã có lớp chủ nhiệm",
                            {"room_id": ["Phòng này đã có lớp chủ nhiệm cho năm học này."]}
                        )

            existing_homeroom = frappe.get_all(
                "SIS Class",
                fields=["name", "title"],
                filters={"room": room_id, "class_type": "regular", "school_year_id": sy},
                limit=1
            )

            if existing_homeroom:
                return validation_error_response(
                    "Phòng đã có lớp chủ nhiệm",
                    {"room_id": [f"Phòng này đã có lớp chủ nhiệm ({existing_homeroom[0].title}) cho năm học này."]}
                )

            # Check if class is already homeroom for another room (legacy check)
            if class_info.room and class_info.room != room_id:
                return validation_error_response(
                    "Lớp đã là chủ nhiệm cho phòng khác",
                    {"class_id": ["Lớp này đã là lớp chủ nhiệm cho phòng khác. Không thể gán làm chủ nhiệm cho phòng này."]}
                )

            # Update class.room to this room (legacy compatibility)
            frappe.db.set_value("SIS Class", class_id, "room", room_id)
            frappe.logger().info(f"Updated SIS Class {class_id} room to {room_id} for homeroom usage")

        else:  # functional usage
            # For functional usage, require subject_id
            if not subject_id:
                return validation_error_response(
                    "Môn học là bắt buộc",
                    {"subject_id": ["Phải chọn môn học cho lớp chức năng"]}
                )

            # Validate subject exists
            if not frappe.db.exists("SIS Timetable Subject", subject_id):
                return validation_error_response("Môn học không tồn tại", {"subject_id": ["Môn học không tồn tại"]})

            # TODO: Add validation for subject compatibility with class education grade
            # Currently skipping validation as relationship needs to be clarified

            # For functional usage, check if class is already a homeroom somewhere (legacy check)
            if class_info.room and class_info.room != room_id and class_info.class_type == "regular":
                return validation_error_response(
                    "Lớp đã là chủ nhiệm",
                    {"class_id": ["Lớp này đang là lớp chủ nhiệm cho phòng khác. Không thể thêm làm phòng chức năng."]}
                )

            # For functional usage, we don't update class.room field to avoid conflicts with homeroom assignments

        # Add to Room Classes child table
        try:
            frappe.logger().info(f"Attempting to add class {class_id} to room {room_id}")

            # Get the room document
            room_doc = frappe.get_doc("ERP Administrative Room", room_id)
            frappe.logger().info(f"Got room doc: {room_doc.name}")

            # Append new room class to the child table
            room_class_data = {
                "class_id": class_id,
                "usage_type": usage_type,
                "class_title": class_info.title,
                "school_year_id": class_info.school_year_id,
                "education_grade": class_info.education_grade,
                "academic_program": class_info.academic_program,
                "homeroom_teacher": class_info.homeroom_teacher,
                "subject_id": subject_id if usage_type == "functional" else None
            }

            frappe.logger().info(f"Appending room class data: {room_class_data}")
            room_doc.append("room_classes", room_class_data)

            frappe.logger().info(f"Saving room doc...")
            room_doc.save(ignore_permissions=True)
            frappe.logger().info(f"Room doc saved successfully")

            frappe.db.commit()  # Ensure the transaction is committed
            frappe.logger().info(f"Added class {class_id} to room {room_id} with usage {usage_type}")

        except Exception as e:
            frappe.logger().error(f"Failed to add to child table: {str(e)}")
            frappe.logger().error(f"Exception type: {type(e).__name__}")
            import traceback
            frappe.logger().error(f"Traceback: {traceback.format_exc()}")
            return error_response(f"Không thể thêm lớp vào phòng: {str(e)}")

        # Phòng lớp học: đồng bộ PIC năm từ GVCN + phó CN (User) vào bản ghi gán năm
        if usage_type == "homeroom":
            try:
                _sync_homeroom_teachers_to_room_yearly_pic(room_id, class_id)
            except Exception:
                frappe.log_error(frappe.get_traceback(), "sync_homeroom_teachers_to_room_yearly_pic")

        frappe.db.commit()

        return success_response(
            data={
                "room_id": room_id,
                "class_id": class_id,
                "usage_type": usage_type,
                "subject_id": subject_id,
                "class_title": class_info.title,
                "class_type": class_info.class_type,
                "class_school_year_id": class_info.school_year_id,
            },
            message="Class added to room successfully"
        )
        frappe.logger().info(f"add_room_class completed successfully for class {class_id}")

    except Exception as e:
        frappe.log_error(f"Error adding room class: {str(e)}")
        return error_response(f"Error adding room class: {str(e)}")


def sync_class_room_assignment(doc, method):
    """Sync room assignment when SIS Class.room field is updated"""
    try:
        # Only process if room field was changed
        if not doc.has_value_changed('room'):
            return

        old_room = doc.get_doc_before_save().get('room') if doc.get_doc_before_save() else None
        new_room = doc.room

        # Remove from old room if exists
        if old_room:
            try:
                old_room_doc = frappe.get_doc("ERP Administrative Room", old_room)
                if hasattr(old_room_doc, 'room_classes'):
                    # Find and remove this class from old room
                    for i, room_class in enumerate(old_room_doc.room_classes):
                        if room_class.class_id == doc.name and room_class.usage_type == "homeroom":
                            old_room_doc.room_classes.pop(i)
                            old_room_doc.save()
                            frappe.logger().info(f"Removed class {doc.name} from old room {old_room}")
                            break
            except Exception as e:
                frappe.logger().warning(f"Could not remove class {doc.name} from old room {old_room}: {str(e)}")

        # Add to new room if exists
        if new_room:
            try:
                new_room_doc = frappe.get_doc("ERP Administrative Room", new_room)

                # Check if already exists
                existing = False
                if hasattr(new_room_doc, 'room_classes'):
                    for room_class in new_room_doc.room_classes:
                        if room_class.class_id == doc.name:
                            existing = True
                            break

                if not existing:
                    # Determine usage type
                    usage_type = "homeroom" if doc.class_type == "regular" else "functional"

                    # Add to child table
                    new_room_doc.append("room_classes", {
                        "class_id": doc.name,
                        "usage_type": usage_type,
                        "class_title": doc.title,
                        "school_year_id": doc.school_year_id,
                        "education_grade": doc.education_grade,
                        "academic_program": doc.academic_program,
                        "homeroom_teacher": doc.homeroom_teacher
                    })
                    new_room_doc.save()
                    frappe.logger().info(f"Added class {doc.name} to new room {new_room} with usage {usage_type}")

            except Exception as e:
                frappe.logger().warning(f"Could not add class {doc.name} to new room {new_room}: {str(e)}")

    except Exception as e:
        frappe.logger().error(f"Error syncing class room assignment: {str(e)}")


@frappe.whitelist(allow_guest=False, methods=['POST'])
def remove_room_class():
    """Remove a class from a room"""
    try:
        data = {}
        if frappe.request and frappe.request.data:
            try:
                body = frappe.request.data.decode('utf-8') if isinstance(frappe.request.data, bytes) else frappe.request.data
                data = json.loads(body or '{}')
            except Exception:
                data = frappe.local.form_dict or {}
        else:
            data = frappe.local.form_dict or {}

        required = ["room_id", "class_id"]
        for field in required:
            if not data.get(field):
                return validation_error_response({field: [f"{field} is required"]})

        room_id = data.get("room_id")
        class_id = data.get("class_id")

        # Validate room exists
        if not frappe.db.exists("ERP Administrative Room", room_id):
            return not_found_response("Room not found")

        # Validate class exists
        if not frappe.db.exists("SIS Class", class_id):
            return not_found_response("Class not found")

        # Get class info
        class_info = frappe.get_doc("SIS Class", class_id)

        # Get room document
        room_doc = frappe.get_doc("ERP Administrative Room", room_id)

        # Remove from Room Classes child table
        removed = False
        if hasattr(room_doc, 'room_classes'):
            for i, room_class in enumerate(room_doc.room_classes):
                if room_class.class_id == class_id:
                    # Remove the child record
                    frappe.delete_doc("ERP Administrative Room Class", room_class.name)
                    removed = True
                    frappe.logger().info(f"Removed class {class_id} from room {room_id} child table")
                    break

        # Fallback: If this room is set as the class's room (legacy), remove it
        if class_info.room == room_id:
            frappe.db.set_value("SIS Class", class_id, "room", None)
            frappe.logger().info(f"Removed room {room_id} from SIS Class {class_id} (legacy)")

        if not removed:
            return validation_error_response(
                "Lớp không được gán cho phòng này",
                {"class_id": ["Lớp này không được gán cho phòng này."]}
            )

        frappe.db.commit()

        return success_response(
            data={
                "room_id": room_id,
                "class_id": class_id,
                "class_title": class_info.title
            },
            message="Class removed from room successfully"
        )

    except Exception as e:
        frappe.log_error(f"Error removing room class: {str(e)}")
        return error_response(f"Error removing room class: {str(e)}")


@frappe.whitelist(allow_guest=False)
def get_available_classes_for_room(room_id: str = None, school_year_id: str = None):
    """Get classes that can be assigned to a room (not already assigned to other rooms)"""
    try:
        if not room_id:
            form = frappe.local.form_dict or {}
            room_id = form.get("room_id")
            if not room_id and frappe.request and frappe.request.args:
                room_id = frappe.request.args.get('room_id')
            # Also check in request body for POST requests
            if not room_id and frappe.request and frappe.request.data:
                try:
                    body = frappe.request.data.decode('utf-8') if isinstance(frappe.request.data, bytes) else frappe.request.data
                    data = json.loads(body or '{}')
                    room_id = data.get('room_id')
                except Exception:
                    pass
        if not room_id:
            return validation_error_response("Room ID is required", {"room_id": ["Room ID is required"]})

        # school_year_id — bắt buộc lọc theo năm (tránh chọn nhầm lớp năm cũ)
        if not school_year_id:
            form = frappe.local.form_dict or {}
            school_year_id = form.get("school_year_id")
        if not school_year_id and frappe.request and getattr(frappe.request, "args", None):
            school_year_id = frappe.request.args.get("school_year_id")
        if not school_year_id and frappe.request and frappe.request.data:
            try:
                body = frappe.request.data.decode("utf-8") if isinstance(frappe.request.data, bytes) else frappe.request.data
                data = json.loads(body or "{}")
                school_year_id = data.get("school_year_id")
            except Exception:
                pass
        if not school_year_id:
            try:
                from erp.api.erp_administrative.administrative_ticket import _active_school_year_id_api

                school_year_id = _active_school_year_id_api(None)
            except Exception:
                school_year_id = None
        if not school_year_id:
            return validation_error_response(
                _("Không xác định được năm học để lọc danh sách lớp."),
                {"school_year_id": ["required"]},
            )

        # Get current campus
        campus_id = get_current_campus_from_context()

        filters = {"campus_id": campus_id or "campus-1", "school_year_id": school_year_id}

        # Get all classes for this campus/year
        classes = frappe.get_all(
            "SIS Class",
            fields=[
                "name",
                "title",
                "short_title",
                "class_type",
                "room",
                "education_grade"
            ],
            filters=filters,
            order_by="title asc"
        )


        # Check if room already has a homeroom class (both child table and legacy)
        room_has_homeroom = False

        try:
            # Check child table first
            room_doc = frappe.get_doc("ERP Administrative Room", room_id)
            if hasattr(room_doc, 'room_classes'):
                for room_class in room_doc.room_classes:
                    if room_class.usage_type == "homeroom":
                        room_has_homeroom = True
                        break
        except Exception:
            pass

        # Fallback to legacy check
        if not room_has_homeroom:
            existing_homeroom = frappe.get_all(
                "SIS Class",
                fields=["name", "title"],
                filters={"room": room_id, "class_type": "regular"},
                limit=1
            )
            room_has_homeroom = len(existing_homeroom) > 0

        # Filter classes that are available
        available_classes = []
        for class_data in classes:
            is_current_room_class = class_data.get("room") == room_id
            has_room = bool(class_data.get("room"))
            is_regular_class = class_data.get("class_type") == "regular"

            # Logic for available classes:
            # 1. Classes without any room assigned
            # 2. Classes assigned to current room (for editing)
            # 3. For homeroom usage: only regular classes that are not homeroom for other rooms
            # 4. For functional usage: all classes except those that are homeroom for other rooms

            can_be_homeroom = False
            can_be_functional = False

            if not has_room:
                # Class has no room - can be both homeroom and functional
                can_be_homeroom = is_regular_class
                can_be_functional = True
            elif is_current_room_class:
                # Class is assigned to current room - can edit usage type
                can_be_homeroom = is_regular_class
                can_be_functional = True
            else:
                # Class is assigned to another room
                # Can only be functional, and only if it's not a regular class (not homeroom elsewhere)
                can_be_homeroom = False
                can_be_functional = not is_regular_class

            # Apply the room-has-homeroom rule: if room already has homeroom, don't suggest homeroom for other classes
            if room_has_homeroom and not is_current_room_class:
                can_be_homeroom = False

            if can_be_homeroom or can_be_functional:
                available_class = class_data.copy()


                # Determine suggested usage based on availability and room state
                if can_be_homeroom and not room_has_homeroom:
                    available_class["suggested_usage"] = "homeroom"
                    available_class["suggested_usage_display"] = "Lớp chủ nhiệm"
                elif can_be_functional:
                    available_class["suggested_usage"] = "functional"
                    available_class["suggested_usage_display"] = "Lớp chức năng"
                else:
                    # Fallback - should not happen with current logic
                    available_class["suggested_usage"] = "functional"
                    available_class["suggested_usage_display"] = "Lớp chức năng"

                available_classes.append(available_class)

        return success_response(
            data=available_classes,
            message="Available classes fetched successfully"
        )

    except Exception as e:
        frappe.log_error(f"Error fetching available classes: {str(e)}")
        return error_response(f"Error fetching available classes: {str(e)}")


def get_room_for_class_subject(class_id: str, subject_title: str = None) -> Dict[str, Any]:
    """Get room information for a class and subject combination.

    Args:
        class_id: SIS Class ID
        subject_title: Subject title to match (optional)

    Returns:
        Dict with room_id, room_name, room_type ('functional' or 'homeroom')
    """
    try:
        # Get all room assignments for this class
        # Use SQL raw query to avoid Frappe auto-resolving Link fields to non-existent columns
        room_assignments_query = """
            SELECT 
                name,
                parent,
                usage_type,
                subject_id
            FROM `tabERP Administrative Room Class`
            WHERE class_id = %s
        """
        room_assignments = frappe.db.sql(room_assignments_query, (class_id,), as_dict=True)

        frappe.logger().info(f"🏫 ROOM DEBUG: Found {len(room_assignments)} room assignments for class {class_id}")
        
        if room_assignments:
            frappe.logger().info(f"🏫 ROOM DEBUG: Room assignments: {[(a.get('parent'), a.get('usage_type'), a.get('subject_id')) for a in room_assignments]}")

        # Priority 1: Try to find functional room matching subject if subject is provided
        if subject_title:
            for assignment in room_assignments:
                if assignment.get("usage_type") == "functional" and assignment.get("subject_id"):
                    # Get subject title from subject_id
                    try:
                        subject_doc = frappe.get_doc("SIS Timetable Subject", assignment["subject_id"])
                        subject_name = subject_doc.title_vn or subject_doc.title_en or ""
                        frappe.logger().info(f"🏫 ROOM DEBUG: Checking subject '{subject_name}' against '{subject_title}'")

                        if subject_name and subject_title.lower() in subject_name.lower():
                            # Get room details
                            room_doc = frappe.get_doc("ERP Administrative Room", assignment["parent"])
                            frappe.logger().info(f"🏫 ROOM DEBUG: Found matching functional room {assignment['parent']} for subject '{subject_title}'")
                            return {
                                "room_id": assignment["parent"],
                                "room_name": room_doc.title_vn or room_doc.title_en or room_doc.name,
                                "room_type": "functional"
                            }
                    except Exception as subj_error:
                        frappe.logger().warning(f"Error getting subject {assignment.get('subject_id')}: {str(subj_error)}")

        # Priority 2: Try to find homeroom room (preferred for any subject)
        homeroom_room = None
        for assignment in room_assignments:
            if assignment.get("usage_type") == "homeroom" and assignment.get("parent"):
                homeroom_room = assignment
                break
        
        if homeroom_room:
            try:
                room_doc = frappe.get_doc("ERP Administrative Room", homeroom_room["parent"])
                frappe.logger().info(f"🏫 ROOM DEBUG: Found homeroom room {homeroom_room['parent']} for class {class_id}")
                return {
                    "room_id": homeroom_room["parent"],
                    "room_name": room_doc.title_vn or room_doc.title_en or room_doc.name,
                    "room_type": "homeroom"
                }
            except Exception as room_error:
                frappe.logger().warning(f"Error getting homeroom room {homeroom_room.get('parent')}: {str(room_error)}")

        # Priority 3: Try to find any room assignment (fallback to any functional room)
        for assignment in room_assignments:
            if assignment.get("parent"):
                try:
                    room_doc = frappe.get_doc("ERP Administrative Room", assignment["parent"])
                    room_type = assignment.get("usage_type") or "homeroom"
                    frappe.logger().info(f"🏫 ROOM DEBUG: Found room {assignment['parent']} ({room_type}) for class {class_id}")
                    return {
                        "room_id": assignment["parent"],
                        "room_name": room_doc.title_vn or room_doc.title_en or room_doc.name,
                        "room_type": room_type
                    }
                except Exception as room_error:
                    frappe.logger().warning(f"Error getting room {assignment.get('parent')}: {str(room_error)}")
                    continue

        # Final fallback: get homeroom from class directly
        class_doc = frappe.get_doc("SIS Class", class_id)
        if class_doc.room:
            room_doc = frappe.get_doc("ERP Administrative Room", class_doc.room)
            frappe.logger().info(f"🏫 ROOM DEBUG: Using homeroom {class_doc.room} from class {class_id}")
            return {
                "room_id": class_doc.room,
                "room_name": room_doc.title_vn or room_doc.title_en or room_doc.name,
                "room_type": "homeroom"
            }

        # No room found
        frappe.logger().info(f"🏫 ROOM DEBUG: No room found for class {class_id}")
        return {
            "room_id": None,
            "room_name": "Chưa có phòng",
            "room_type": None
        }

    except Exception as e:
        frappe.logger().error(f"Error getting room for class {class_id}, subject {subject_title}: {str(e)}")
        return {
            "room_id": None,
            "room_name": "Lỗi tải phòng",
            "room_type": None
        }


@frappe.whitelist(allow_guest=False)
def get_all_rooms_for_sync():
    """
    Get ALL rooms for inventory service sync (không bị giới hạn bởi campus)
    Dành riêng cho microservice sync - không filter bởi user's campus
    """
    try:
        # Fetch ALL rooms từ database - không filter campus
        rooms = frappe.get_all(
            "ERP Administrative Room",
            fields=[
                "name",
                "title_vn as room_name",
                "title_en as room_name_en",
                "short_title",
                "building_id as building",
                "capacity",
                "room_type",
                "creation",
                "modified"
            ],
            order_by="name asc"
        )
        
        frappe.logger().info(f"[Room Sync] Fetched {len(rooms)} rooms for sync")
        
        return success_response(
            data=rooms,
            message="All rooms fetched successfully for sync",
            meta={"total_count": len(rooms)}
        )
        
    except Exception as e:
        frappe.logger().error(f"Error fetching rooms for sync: {str(e)}")
        return error_response(f"Error fetching rooms: {str(e)}")


def _room_api_json_body():
    """Đọc JSON body cho API phòng (room responsible users)."""
    data = {}
    if frappe.request and frappe.request.data:
        try:
            body = (
                frappe.request.data.decode("utf-8")
                if isinstance(frappe.request.data, bytes)
                else frappe.request.data
            )
            data = json.loads(body or "{}")
        except Exception:
            data = frappe.local.form_dict or {}
    else:
        data = frappe.local.form_dict or {}
    return data


def _responsible_user_row_payload(user_id):
    """Chuẩn dict một dòng người phụ trách từ User."""
    if not user_id or not frappe.db.exists("User", user_id):
        return None
    # User chuẩn Frappe không có designation; job_title thường từ custom/HR — chỉ SELECT field thật sự có trên DB
    user_meta = frappe.get_meta("User")
    fieldnames = ["full_name", "user_image"]
    for fn in ("job_title", "designation"):
        if user_meta.has_field(fn):
            fieldnames.append(fn)
    u = frappe.db.get_value("User", user_id, fieldnames, as_dict=True)
    if not u:
        return None
    designation = (u.get("job_title") or "").strip() or (u.get("designation") or "").strip()
    return {
        "user": user_id,
        "full_name": (u.get("full_name") or "").strip() or get_fullname(user_id) or user_id,
        "user_image": (u.get("user_image") or "").strip(),
        "designation": designation,
    }


@frappe.whitelist(allow_guest=False)
def get_room_responsible_users():
    """Danh sách người phụ trách theo room_id (Frappe room name)."""
    try:
        data = _room_api_json_body()
        room_id = data.get("room_id")
        if not room_id or not frappe.db.exists("ERP Administrative Room", room_id):
            return validation_error_response(_("Phòng không hợp lệ"), {"room_id": ["invalid"]})
        rows = frappe.get_all(
            "ERP Administrative Room Responsible User",
            filters={"parent": room_id},
            fields=["user", "full_name", "user_image", "designation"],
            order_by="idx asc",
        )
        out = []
        for r in rows:
            out.append(
                {
                    "user": r.user,
                    "full_name": r.full_name or "",
                    "user_image": r.user_image or "",
                    "designation": r.designation or "",
                }
            )
        return success_response(data=out, message=_("OK"))
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "room.get_room_responsible_users")
        return error_response(str(e))


@frappe.whitelist(allow_guest=False)
def add_room_responsible_user():
    """Thêm User vào bảng người phụ trách phòng."""
    try:
        data = _room_api_json_body()
        room_id = data.get("room_id")
        user_id = data.get("user") or data.get("user_email")
        if not room_id or not frappe.db.exists("ERP Administrative Room", room_id):
            return validation_error_response(_("Phòng không hợp lệ"), {"room_id": ["invalid"]})
        if not user_id:
            return validation_error_response(_("Thiếu user"), {"user": ["required"]})
        if not frappe.db.exists("User", user_id):
            return validation_error_response(_("User không tồn tại"), {"user": ["invalid"]})

        room = frappe.get_doc("ERP Administrative Room", room_id)
        for row in room.responsible_users or []:
            if row.user == user_id:
                return validation_error_response(_("User đã là người phụ trách phòng này"), {"user": ["duplicate"]})

        payload = _responsible_user_row_payload(user_id)
        if not payload:
            return validation_error_response(_("Không đọc được thông tin User"), {"user": ["invalid"]})

        room.append(
            "responsible_users",
            {
                "user": user_id,
                "full_name": payload["full_name"],
                "user_image": payload["user_image"],
                "designation": payload["designation"],
            },
        )
        room.save(ignore_permissions=False)
        frappe.db.commit()
        try:
            log_room_activity(
                room_id,
                "user_assigned",
                user=frappe.session.user,
                target_user=user_id,
                note="",
            )
            frappe.db.commit()
        except Exception:
            frappe.log_error(frappe.get_traceback(), "room.add_room_responsible_user.activity_log")
        return success_response(data=payload, message=_("Đã thêm người phụ trách"))
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "room.add_room_responsible_user")
        return error_response(str(e))


@frappe.whitelist(allow_guest=False)
def remove_room_responsible_user():
    """Xóa User khỏi danh sách người phụ trách phòng."""
    try:
        data = _room_api_json_body()
        room_id = data.get("room_id")
        user_id = data.get("user") or data.get("user_email")
        if not room_id or not frappe.db.exists("ERP Administrative Room", room_id):
            return validation_error_response(_("Phòng không hợp lệ"), {"room_id": ["invalid"]})
        if not user_id:
            return validation_error_response(_("Thiếu user"), {"user": ["required"]})

        room = frappe.get_doc("ERP Administrative Room", room_id)
        kept = [r for r in (room.responsible_users or []) if r.user != user_id]
        if len(kept) == len(room.responsible_users or []):
            return not_found_response(_("Không tìm thấy user trong danh sách phụ trách"))
        room.responsible_users = []
        for row in kept:
            room.append(
                "responsible_users",
                {
                    "user": row.user,
                    "full_name": row.full_name,
                    "user_image": row.user_image,
                    "designation": row.designation,
                },
            )
        room.save(ignore_permissions=False)
        frappe.db.commit()
        try:
            log_room_activity(
                room_id,
                "user_removed",
                user=frappe.session.user,
                target_user=user_id,
                note="",
            )
            frappe.db.commit()
        except Exception:
            frappe.log_error(frappe.get_traceback(), "room.remove_room_responsible_user.activity_log")
        return success_response(data={"removed": user_id}, message=_("Đã xóa"))
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "room.remove_room_responsible_user")
        return error_response(str(e))


@frappe.whitelist(allow_guest=False)
def get_users_for_room_responsible():
    """Danh sách User để chọn người phụ trách (tương tự teacher.get_users_for_selection)."""
    try:
        # Chỉ email nội bộ — loại phụ huynh / tài khoản email cá nhân khỏi pool chọn PIC phòng
        internal_email_domain = "@wellspring.edu.vn"
        user_meta = frappe.get_meta("User")
        fieldnames = [
            "name",
            "email",
            "full_name",
            "first_name",
            "last_name",
            "user_image",
            "employee_code",
            "employee_id",
        ]
        for fn in ("job_title", "designation"):
            if user_meta.has_field(fn):
                fieldnames.append(fn)
        users = frappe.get_all(
            "User",
            fields=fieldnames,
            filters={"enabled": 1, "email": ["like", f"%{internal_email_domain}"]},
            order_by="full_name asc",
        )
        processed = []
        for u in users:
            designation = (u.get("job_title") or "").strip() or (u.get("designation") or "").strip()
            row = {**u, "user_id": u.get("name"), "designation": designation}
            processed.append(row)
        return success_response(data=processed, message=_("OK"))
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "room.get_users_for_room_responsible")
        return error_response(str(e))


@frappe.whitelist(allow_guest=False)
def check_physical_code_unique():
    """Kiểm tra trùng mã vật lý trên campus: building_id, room_number, exclude_room_id (optional)."""
    try:
        data = _room_api_json_body()
        campus_id = data.get("campus_id") or get_current_campus_from_context()
        building_id = data.get("building_id")
        room_number = (data.get("room_number") or "").strip().upper()
        exclude_room = data.get("exclude_room_id")
        if not campus_id or not building_id or not room_number:
            return validation_error_response(_("Thiếu campus / building / room_number"), {})
        bst = frappe.db.get_value("ERP Administrative Building", building_id, "short_title") or ""
        pc = f"{bst}.{room_number}" if bst else ""
        if not pc:
            return success_response(data={"unique": True, "physical_code": ""}, message="OK")
        filt = {"campus_id": campus_id, "physical_code": pc}
        if exclude_room:
            filt["name"] = ["!=", exclude_room]
        exists = frappe.db.exists("ERP Administrative Room", filt)
        return success_response(data={"unique": not bool(exists), "physical_code": pc}, message="OK")
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "room.check_physical_code_unique")
        return error_response(str(e))


@frappe.whitelist(allow_guest=False, methods=["POST"])
def sync_room_homeroom_pic_if_classroom():
    """
    Phòng classroom_room: đồng bộ PIC năm từ lớp CN theo school_year_id (gọi khi vào chi tiết / làm mới).
    Bổ sung cho hook SIS Class khi người dùng chưa F5 sau khi sửa lớp.
    """
    try:
        data = _room_api_json_body()
        room_id = data.get("room_id") or data.get("room")
        sy = data.get("school_year_id")
        if not room_id or not sy:
            return validation_error_response(_("Thiếu room_id hoặc school_year_id"), {})
        if not frappe.db.exists("ERP Administrative Room", room_id):
            return not_found_response(_("Không tìm thấy phòng"))
        rt = (frappe.db.get_value("ERP Administrative Room", room_id, "room_type") or "").lower()
        if rt != "classroom_room":
            return success_response(data={"skipped": True}, message="OK")
        class_id = None
        for rc in frappe.get_all(
            "ERP Administrative Room Class",
            fields=["class_id"],
            filters={"parent": room_id, "usage_type": "homeroom"},
        ):
            cid = rc.class_id
            cy = frappe.db.get_value("SIS Class", cid, "school_year_id")
            if cy == sy:
                class_id = cid
                break
        if not class_id:
            return success_response(data={"skipped": True, "reason": "no_homeroom_class"}, message="OK")
        _sync_homeroom_teachers_to_room_yearly_pic(room_id, class_id)
        return success_response(data={"class_id": class_id}, message="OK")
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "room.sync_room_homeroom_pic_if_classroom")
        return error_response(str(e))


@frappe.whitelist(allow_guest=False)
def list_room_yearly_assignments():
    """Danh sách gán năm theo phòng: room_id."""
    try:
        data = _room_api_json_body()
        room_id = data.get("room_id")
        if not room_id:
            return validation_error_response(_("Thiếu room_id"), {})
        rows = frappe.get_all(
            "ERP Administrative Room Yearly Assignment",
            filters={"room": room_id},
            fields=["*"],
            order_by="school_year_id desc",
        )
        return success_response(data=rows, message="OK")
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "room.list_room_yearly_assignments")
        return error_response(str(e))


@frappe.whitelist(allow_guest=False)
def get_room_yearly_assignment():
    """Một bản ghi gán năm đầy đủ (kèm child PIC) theo room + school_year_id."""
    try:
        data = _room_api_json_body()
        room_id = data.get("room_id") or data.get("room")
        sy = data.get("school_year_id")
        if not room_id or not sy:
            return validation_error_response(_("Thiếu room_id hoặc school_year_id"), {})
        name = frappe.db.get_value(
            "ERP Administrative Room Yearly Assignment",
            {"room": room_id, "school_year_id": sy},
            "name",
        )
        if not name:
            return success_response(data=None, message=_("Chưa có gán năm"))
        doc = frappe.get_doc("ERP Administrative Room Yearly Assignment", name)
        return success_response(data=doc.as_dict(), message="OK")
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "room.get_room_yearly_assignment")
        return error_response(str(e))


@frappe.whitelist(allow_guest=False)
def save_room_yearly_assignment():
    """
    Tạo/cập nhật ERP Administrative Room Yearly Assignment.
    Body: name (optional), room, school_year_id, usage_type, display_title_vn, class_id, homeroom_teacher_id, ...
    """
    try:
        data = _room_api_json_body()
        name = data.get("name")
        room_id = data.get("room_id") or data.get("room")
        sy = data.get("school_year_id")
        if not room_id or not sy:
            return validation_error_response(_("Thiếu room hoặc school_year_id"), {})

        existing_ya = frappe.db.get_value(
            "ERP Administrative Room Yearly Assignment",
            {"room": room_id, "school_year_id": sy},
            "name",
        )
        if name and frappe.db.exists("ERP Administrative Room Yearly Assignment", name):
            doc = frappe.get_doc("ERP Administrative Room Yearly Assignment", name)
        elif existing_ya:
            doc = frappe.get_doc("ERP Administrative Room Yearly Assignment", existing_ya)
        else:
            doc = frappe.get_doc({"doctype": "ERP Administrative Room Yearly Assignment"})

        doc.room = room_id
        doc.school_year_id = sy
        if data.get("usage_type"):
            doc.usage_type = data.get("usage_type")
        if "display_title_vn" in data:
            doc.display_title_vn = data.get("display_title_vn") or ""
        if "display_title_en" in data:
            doc.display_title_en = data.get("display_title_en") or ""
        if "display_short_title" in data:
            doc.display_short_title = data.get("display_short_title") or ""
        if data.get("class_id") is not None:
            doc.class_id = data.get("class_id") or None
        if data.get("homeroom_teacher_id") is not None:
            doc.homeroom_teacher_id = data.get("homeroom_teacher_id") or None
        if data.get("vice_homeroom_teacher_id") is not None:
            doc.vice_homeroom_teacher_id = data.get("vice_homeroom_teacher_id") or None
        if data.get("status"):
            doc.status = data.get("status")
        if data.get("notes") is not None:
            doc.notes = data.get("notes") or ""

        pic_list = data.get("responsible_users") or data.get("pic_list")
        if isinstance(pic_list, list):
            doc.responsible_users = []
            for p in pic_list:
                uid = p.get("user") if isinstance(p, dict) else p
                if not uid:
                    continue
                fn = frappe.db.get_value("User", uid, "full_name") or ""
                doc.append(
                    "responsible_users",
                    {"user": uid, "role_label": (p.get("role_label") if isinstance(p, dict) else "") or "", "full_name": fn},
                )

        if doc.is_new():
            doc.insert()
        else:
            doc.save()
        frappe.db.commit()
        return success_response(data={"name": doc.name}, message=_("Đã lưu"))
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "room.save_room_yearly_assignment")
        return error_response(str(e))


@frappe.whitelist(allow_guest=False)
def get_room_history():
    """Timeline: yearly assignments + handovers + inventory (handover incoming + IC) + tickets."""
    try:
        data = _room_api_json_body()
        room_id = data.get("room_id")
        sy_filter = data.get("school_year_id")
        if not room_id:
            return validation_error_response(_("Thiếu room_id"), {})

        ya_filters = {"room": room_id}
        if sy_filter:
            ya_filters["school_year_id"] = sy_filter
        yearly = frappe.get_all(
            "ERP Administrative Room Yearly Assignment",
            filters=ya_filters,
            fields=["name", "school_year_id", "display_title_vn", "status", "creation"],
            order_by="creation desc",
            limit=50,
        )

        ho_filters = {"room": room_id, "direction": "outgoing"}
        if sy_filter:
            ho_filters["school_year_id"] = sy_filter
        handovers = frappe.get_all(
            "ERP Administrative Facility Handover",
            filters=ho_filters,
            fields=["name", "school_year_id", "status", "class_id", "sent_on", "confirmed_on"],
            order_by="creation desc",
            limit=50,
        )

        ic_filters = {"room": room_id}
        if sy_filter:
            ic_filters["school_year_id"] = sy_filter
        inv_incoming = frappe.get_all(
            "ERP Administrative Facility Handover",
            filters={**ic_filters, "direction": "incoming"},
            fields=["name", "school_year_id", "status", "sent_on", "reviewed_on"],
            order_by="creation desc",
            limit=50,
        )
        inv_ic = frappe.get_all(
            "ERP Administrative Inventory Check",
            filters=ic_filters,
            fields=["name", "school_year_id", "status", "submitted_on", "reviewed_on"],
            order_by="creation desc",
            limit=50,
        )

        tickets = frappe.get_all(
            "ERP Administrative Ticket",
            filters={"room_id": room_id},
            fields=["name", "ticket_code", "title", "status", "creation"],
            order_by="creation desc",
            limit=30,
        )

        return success_response(
            data={
                "yearly_assignments": yearly,
                "handovers_out": handovers,
                "inventory_handover_in": inv_incoming,
                "inventory_checks": inv_ic,
                "tickets": tickets,
            },
            message="OK",
        )
    except Exception as e:
        frappe.log_error(frappe.get_traceback(), "room.get_room_history")
        return error_response(str(e))
