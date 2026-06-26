#!/usr/bin/env python3
"""
Import "Vấn đề chung" (CRM Issue) thẳng vào CSDL từ file Excel tổng hợp vấn đề/feedback PHHS.

ĐÂY LÀ SCRIPT ONE-OFF (không phải tính năng import trên UI, không whitelist API).
Dùng Frappe ORM (frappe.get_doc/insert) — KHÔNG raw SQL — để giữ autoname, child table, validation.

Nguồn: chỉ đọc sheet "Data báo cáo sự vụfeedback".
  - Header ở row 5, data từ row 6.
  - Cột map theo NHÃN header (không hardcode chỉ số) để bền với thay đổi layout.

Cách chạy (sau khi đã sửa bench env):

    # 0) Pre-flight: thống kê khớp/không khớp module, phòng ban, PIC, học sinh (không ghi gì)
    bench --site admin.sis.localhost execute \
        erp.scripts.import_crm_issues_from_excel.check_master \
        --kwargs "{'path': '/Users/admin/Downloads/TSxIT_Tổng hợp vấn đề và feedback PHHS (1).xlsx'}"

    # 1) Tạo sẵn các Loại vấn đề (CRM Issue Module) còn thiếu — dry-run trước
    bench --site admin.sis.localhost execute \
        erp.scripts.import_crm_issues_from_excel.ensure_modules \
        --kwargs "{'path': '.../(1).xlsx', 'commit': False}"
    # rồi commit=True để tạo thật

    # 2) Import — dry-run (chỉ resolve + in báo cáo, KHÔNG ghi)
    bench --site admin.sis.localhost execute \
        erp.scripts.import_crm_issues_from_excel.run \
        --kwargs "{'path': '.../(1).xlsx', 'commit': False}"
    # rồi commit=True để ghi thật

Hoặc trong `bench --site ... console`:
    from erp.scripts.import_crm_issues_from_excel import run, ensure_modules, check_master
    run(path="...(1).xlsx", commit=False)
"""

import re
from datetime import datetime, date, timedelta

import frappe

from erp.api.crm.issue import (
    _sync_issue_students,
    _set_issue_departments,
    _sync_issue_guardians,
)
from erp.api.crm.utils import normalize_phone_number

# ----------------------------------------------------------------------------- config

DATA_SHEET = "Data báo cáo sự vụfeedback"
HEADER_ROW = 5            # nhãn cột ở row 5; data bắt đầu row 6

REPORT_FILE = "import_crm_issues_report.txt"

# Gộp các Loại vấn đề gần trùng nghĩa (áp dụng cho cả ensure_modules lẫn resolve khi import).
MODULE_ALIASES = {
    "Học bổng": "Học bổng - Khen thưởng",
}

# Ánh xạ tên "Bộ phận liên quan" trong Excel -> unit_name_vn thật trên org chart.
# Tên nào khớp y hệt (Trường Tiểu học, Phòng Kế toán, Phòng Tuyển sinh & Dịch vụ trường học)
# thì không cần liệt kê.
DEPT_ALIASES = {
    "Trường THCS": "Trường Trung học cơ sở",
    "Trường THPT": "Trường Trung học phổ thông",
    "Khối HCTH & DVHS": "Khối hành chính tổng hợp và Dịch vụ học sinh",
    "Ban Đào tạo": "Ban đào tạo",
    "Khoa Phát triển Học sinh": "Khoa Phát triển Sinh viên",
    "Khối vận hành": "Khối Vận hành",
    "Phòng Công nghệ thông tin": "Phòng IT",
}

# Map nhãn cột -> khoá nội bộ. Match bằng "chứa chuỗi con" trên header đã chuẩn hoá khoảng trắng.
# Thứ tự quan trọng: 'pic_email' phải kiểm trước 'submitter_email' (đều chứa "email/mã nv").
COLUMN_MATCHERS = [
    ("status",          ["tình trạng xử lý"]),
    ("received_date",   ["ngày tiếp nhận"]),
    ("submitter_name",  ["người gửi ý kiến"]),
    ("pic_name",        ["người phụ trách"]),
    ("pic_email",       ["email/mã nv pic"]),
    ("module",          ["nhóm ý kiến"]),
    ("code",            ["mã sự vụ"]),
    ("priority",        ["mức độ ưu tiên"]),
    ("related_dept",    ["bộ phận liên quan"]),
    ("received_dept",   ["bộ phận tiếp nhận"]),
    ("content",         ["nội dung"]),
    ("student_code",    ["mã học sinh"]),
    ("student_name",    ["học sinh liên quan"]),
    ("grade",           ["lớp/grade", "lớp/ grade"]),
    ("parent_name",     ["phhs liên quan"]),
    ("parent_phone",    ["sđt phhs"]),
    ("handled_school",  ["thông tin xử lý của bộ phận tiếp nhận"]),
    ("handled_dept",    ["kết quả xử lý của bộ phận liên quan"]),
    ("link",            ["link liên quan"]),
    ("satisfaction",    ["theo dõi mức độ hài lòng"]),
    ("ts_note",         ["note của ts"]),
]

PRIORITY_MAP = {"high": "Cao", "normal": "Trung binh", "medium": "Trung binh", "low": "Thap"}
DEFAULT_PRIORITY = "Trung binh"

STATUS_MAP = {
    "chờ phê duyệt": "Cho duyet",
    "tiếp nhận": "Tiep nhan",
    "đang xử lý": "Dang xu ly",
    "hoàn thành": "Hoan thanh",
    "đóng": "Dong",
}
DEFAULT_STATUS = "Tiep nhan"

# ----------------------------------------------------------------------------- helpers

def _norm(s):
    """Chuẩn hoá: gộp mọi khoảng trắng/newline thành 1 space, strip, lower."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _txt(v):
    """Lấy chuỗi gọn từ ô (giữ nguyên dấu)."""
    if v is None:
        return ""
    return re.sub(r"[ \t]+", " ", str(v)).strip()


def _files_dirs(sub):
    """Thư mục files khả dĩ (anchor cả tương đối lẫn tuyệt đối) để né lệch cwd."""
    import os
    dirs = []
    try:
        dirs.append(frappe.get_site_path(sub, "files"))
    except Exception:
        pass
    try:
        site = getattr(frappe.local, "site", None)
        bench = frappe.utils.get_bench_path()
        if site and bench:
            dirs.append(os.path.join(bench, "sites", site, sub, "files"))
    except Exception:
        pass
    return dirs


def _resolve_path(path):
    """
    Cho phép truyền:
      - URL file Frappe: '/private/files/x.xlsx' hoặc '/files/x.xlsx'
      - path filesystem (tuyệt đối/tương đối).
    Bền với: cwd khác sites, lệch chuẩn hoá Unicode (NFC/NFD), %20, Frappe đổi tên file.
    Khi không match đúng tên, tìm theo TIỀN TỐ ASCII (vd 'TSxIT') trong thư mục files.
    """
    import os
    import glob
    import unicodedata
    from urllib.parse import unquote

    # 1) path filesystem trực tiếp
    for c in (path, unquote(path)):
        if c and os.path.isfile(c):
            return c

    if path.startswith("/private/files/") or path.startswith("/files/"):
        sub = "private" if path.startswith("/private/files/") else "public"
        base = unquote(path.split("/")[-1])
        name_variants = {base,
                        unicodedata.normalize("NFC", base),
                        unicodedata.normalize("NFD", base)}

        dirs = _files_dirs(sub)

        # 2) thử tên chính xác (mọi biến thể chuẩn hoá) trong mọi thư mục
        for d in dirs:
            for nm in name_variants:
                cand = os.path.join(d, nm)
                if os.path.isfile(cand):
                    return cand

        # 3) glob theo tiền tố ASCII đầu (chống lệch Unicode + đổi tên)
        ascii_prefix = ""
        for ch in base:
            if ord(ch) < 128 and ch not in "*?[]":
                ascii_prefix += ch
            else:
                break
        if len(ascii_prefix) >= 3:
            ext = os.path.splitext(base)[1] or ".xlsx"
            for d in dirs:
                hits = (glob.glob(os.path.join(d, ascii_prefix + "*" + ext))
                       or glob.glob(os.path.join(d, ascii_prefix + "*")))
                xlsx = [h for h in hits if h.lower().endswith(".xlsx")]
                if len(xlsx) == 1:
                    return xlsx[0]
                if len(hits) == 1:
                    return hits[0]
                if xlsx:
                    return xlsx[0]

        # 4) tra DocType File theo file_url rồi lấy full path
        for url in {path, unquote(path)}:
            fn = frappe.db.get_value("File", {"file_url": url}, "name")
            if fn:
                try:
                    fp = frappe.get_doc("File", fn).get_full_path()
                    if os.path.isfile(fp):
                        return fp
                except Exception:
                    pass

    return path


def _open_data_sheet(path):
    import openpyxl
    real = _resolve_path(path)
    wb = openpyxl.load_workbook(real, data_only=True, read_only=True)
    if DATA_SHEET not in wb.sheetnames:
        # fallback: sheet bắt đầu bằng "Data"
        cand = [n for n in wb.sheetnames if n.startswith("Data")]
        if not cand:
            raise ValueError(f"Không tìm thấy sheet '{DATA_SHEET}'. Có: {wb.sheetnames}")
        ws = wb[cand[0]]
    else:
        ws = wb[DATA_SHEET]
    return wb, ws


def _build_colmap(ws):
    """Đọc row HEADER_ROW, trả về dict khoá_nội_bộ -> chỉ số cột (0-based)."""
    header_cells = list(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))[0]
    headers = [_norm(c) for c in header_cells]
    colmap = {}
    for key, needles in COLUMN_MATCHERS:
        if key in colmap:
            continue
        for ci, h in enumerate(headers):
            if not h:
                continue
            if any(n in h for n in needles):
                # tránh map nhầm submitter_email vào cột PIC ("email/mã nv pic")
                colmap[key] = ci
                break
    return colmap, header_cells


def _cell(row, colmap, key):
    ci = colmap.get(key)
    if ci is None or ci >= len(row):
        return None
    return row[ci]


def parse_date(v):
    """Trả về 'YYYY-MM-DD' hoặc None. Hỗ trợ datetime/date, chuỗi DD/MM/YYYY, serial Excel."""
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        return (date(1899, 12, 30) + timedelta(days=int(v))).strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    try:
        return (date(1899, 12, 30) + timedelta(days=int(float(s)))).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def _escape_html(s):
    return (
        (s or "")
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
    )


def _to_html(s):
    s = _txt(s)
    if not s:
        return ""
    return "<p>" + _escape_html(s).replace("\n", "<br>") + "</p>"


def _split_multi(s):
    """Tách giá trị nhiều mục ngăn bởi dấu phẩy (và xuống dòng)."""
    if not s:
        return []
    parts = re.split(r"[,\n]+", str(s))
    return [p.strip() for p in parts if p and p.strip()]


def _resolve_module(text):
    """Trả về docname CRM Issue Module theo module_name (áp alias). None nếu không có."""
    name = _txt(text)
    if not name:
        return None
    name = MODULE_ALIASES.get(name, name)
    return frappe.db.get_value("CRM Issue Module", {"module_name": name}, "name")


def _resolve_students(code_cell, name_cell):
    """Trả về (list docname CRM Student, list mã/tên không khớp)."""
    ids, missing = [], []
    codes = _split_multi(code_cell)
    for c in codes:
        sid = frappe.db.get_value("CRM Student", {"student_code": c}, "name")
        if sid and sid not in ids:
            ids.append(sid)
        elif not sid:
            missing.append(c)
    if not ids:
        # fallback theo tên (có thể nhiều người trùng tên -> lấy bản đầu)
        for nm in _split_multi(name_cell):
            sid = frappe.db.get_value("CRM Student", {"student_name": nm}, "name")
            if sid and sid not in ids:
                ids.append(sid)
            elif not sid:
                missing.append(nm)
    return ids, missing


def _phone_strings(cell):
    """
    Từ ô SĐT -> list chuỗi số. Xử lý: float Excel ('903296856.0', rớt số 0 đầu),
    nhiều số (xuống dòng/phẩy), và ghi chú chữ ('934567755 (số Hân)').
    Chỉ lấy các run 8–11 chữ số.
    """
    if cell is None or cell == "":
        return []
    s = str(int(cell)) if isinstance(cell, (int, float)) else str(cell)
    out, seen = [], set()
    for m in re.findall(r"\d{8,11}", s):
        if m not in seen:
            seen.add(m)
            out.append(m)
    return out


def _resolve_guardians(phone_cell, name_cell):
    """
    Trả về (list docname CRM Guardian, list giá trị không khớp).
    Ưu tiên khớp theo SĐT đã chuẩn hoá (+84...); fallback theo guardian_name.
    """
    ids, missing = [], []
    for raw in _phone_strings(phone_cell):
        norm = normalize_phone_number(raw)
        gid = None
        if norm:
            gid = (frappe.db.get_value("CRM Guardian", {"phone_number": norm}, "name")
                   or frappe.db.get_value("CRM Guardian Phone", {"phone_number": norm}, "parent"))
        if gid and gid not in ids:
            ids.append(gid)
        elif not gid:
            missing.append(raw)
    if not ids:
        for nm in _split_multi(name_cell):
            gid = frappe.db.get_value("CRM Guardian", {"guardian_name": nm}, "name")
            if gid and gid not in ids:
                ids.append(gid)
            elif not gid:
                missing.append(nm)
    return ids, missing


def _resolve_dept_name(nm):
    """Docname ERP Organization Unit từ tên Excel (áp DEPT_ALIASES). None nếu không có."""
    name = DEPT_ALIASES.get(nm, nm)
    return frappe.db.get_value("ERP Organization Unit", {"unit_name_vn": name}, "name")


def _unique_issue_code(code, seen):
    """
    Trả về (code_duy_nhất, đã_đổi?). Nếu code đã có (trong file `seen` hoặc trong CSDL),
    thêm hậu tố -2, -3, ... cho tới khi không trùng (vd ISSUE-04 -> ISSUE-04-2).
    """
    def taken(c):
        return c in seen or bool(frappe.db.exists("CRM Issue", {"issue_code": c}))

    if not taken(code):
        return code, False
    i = 2
    while taken(f"{code}-{i}"):
        i += 1
    return f"{code}-{i}", True


def _resolve_departments(text):
    """Trả về (list docname ERP Organization Unit, list tên không khớp)."""
    ids, missing = [], []
    for nm in _split_multi(text):
        did = _resolve_dept_name(nm)
        if did and did not in ids:
            ids.append(did)
        elif not did:
            missing.append(nm)
    return ids, missing


def _map_priority(v):
    return PRIORITY_MAP.get(_norm(v), DEFAULT_PRIORITY) if _txt(v) else DEFAULT_PRIORITY


def _map_status(v):
    return STATUS_MAP.get(_norm(v), DEFAULT_STATUS) if _txt(v) else DEFAULT_STATUS


def _apply_satisfaction(sat, status):
    """
    (result, status) theo quy tắc nghiệp vụ:
      - 'Hài lòng'                  -> result Hai long
      - 'Đồng ý, nhưng chưa hài lòng' -> status Hoan thanh, result Chua hai long
      - 'Tiếp tục theo dõi'         -> status Dang xu ly, result rỗng
      - khác/rỗng                   -> result rỗng (giữ status đã map)
    """
    s = _norm(sat)
    if s == "hài lòng":
        return "Hai long", status
    if s == "đồng ý, nhưng chưa hài lòng":
        return "Chua hai long", "Hoan thanh"
    if s == "tiếp tục theo dõi":
        return "", "Dang xu ly"
    return "", status


def _row_is_empty(row, colmap):
    keys = ("code", "content", "module", "status", "received_date", "student_code", "student_name")
    return not any(_txt(_cell(row, colmap, k)) for k in keys)


def _write_report(lines):
    text = "\n".join(lines)
    try:
        path = frappe.get_site_path(REPORT_FILE)
    except Exception:
        path = REPORT_FILE
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(text)
        print(f"\n📝 Báo cáo đã ghi: {path}")
    except Exception as e:
        print(f"\n⚠️ Không ghi được file báo cáo ({e}). In ra console:")
    return path


# ----------------------------------------------------------------------------- pre-flight

def check_master(path):
    """Pre-flight: liệt kê module / phòng ban / PIC / mã HS trong sheet, đánh dấu khớp/không khớp. Không ghi gì."""
    wb, ws = _open_data_sheet(path)
    colmap, headers = _build_colmap(ws)
    print("Header phát hiện:", {k: headers[v] for k, v in colmap.items()})

    modules, depts, pics, stu_codes, phones = {}, {}, {}, {}, {}
    n = 0
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if _row_is_empty(row, colmap):
            continue
        n += 1
        m = _txt(_cell(row, colmap, "module"))
        if m:
            modules[m] = modules.get(m, 0) + 1
        for d in _split_multi(_cell(row, colmap, "related_dept")):
            depts[d] = depts.get(d, 0) + 1
        p = _txt(_cell(row, colmap, "pic_email"))
        if p:
            pics[p] = pics.get(p, 0) + 1
        for c in _split_multi(_cell(row, colmap, "student_code")):
            stu_codes[c] = stu_codes.get(c, 0) + 1
        for ph in _phone_strings(_cell(row, colmap, "parent_phone")):
            phones[ph] = phones.get(ph, 0) + 1

    def _report_group(title, counter, resolver):
        print(f"\n=== {title} ({len(counter)} giá trị, {sum(counter.values())} lượt) ===")
        ok = miss = 0
        for val, cnt in sorted(counter.items(), key=lambda x: -x[1]):
            hit = resolver(val)
            flag = "✅" if hit else "❌ THIẾU"
            if hit:
                ok += 1
            else:
                miss += 1
            print(f"  {flag}  {val!r}  x{cnt}" + (f"  -> {hit}" if hit else ""))
        print(f"  Tổng: {ok} khớp / {miss} thiếu")

    print(f"\nSố dòng data: {n}")
    _report_group("Loại vấn đề (CRM Issue Module)", modules, _resolve_module)
    _report_group("Phòng ban liên quan (ERP Organization Unit)", depts, _resolve_dept_name)
    _report_group("PIC (User by email)", pics,
                  lambda v: v if frappe.db.exists("User", v) else None)
    _report_group("Mã học sinh (CRM Student)", stu_codes,
                  lambda v: frappe.db.get_value("CRM Student", {"student_code": v}, "name"))
    _report_group("SĐT PHHS (CRM Guardian)", phones,
                  lambda v: (frappe.db.get_value("CRM Guardian", {"phone_number": normalize_phone_number(v)}, "name")
                             or frappe.db.get_value("CRM Guardian Phone", {"phone_number": normalize_phone_number(v)}, "parent")))
    wb.close()


def ensure_modules(path, commit=False):
    """Tạo các CRM Issue Module còn thiếu (module_name = Nhóm Ý kiến, áp alias). Dry-run nếu commit=False."""
    wb, ws = _open_data_sheet(path)
    colmap, _ = _build_colmap(ws)
    names = {}
    for row in ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True):
        if _row_is_empty(row, colmap):
            continue
        m = _txt(_cell(row, colmap, "module"))
        if m:
            m = MODULE_ALIASES.get(m, m)
            names[m] = names.get(m, 0) + 1
    wb.close()

    created, existed = [], []
    for nm, cnt in sorted(names.items(), key=lambda x: -x[1]):
        if frappe.db.get_value("CRM Issue Module", {"module_name": nm}, "name"):
            existed.append(nm)
            continue
        if commit:
            doc = frappe.new_doc("CRM Issue Module")
            doc.module_name = nm
            doc.is_active = 1
            doc.flags.ignore_permissions = True
            doc.insert()
            created.append(f"{nm} -> {doc.name}")
        else:
            created.append(f"{nm} (SẼ TẠO)")

    if commit:
        frappe.db.commit()

    print(f"\n=== ensure_modules (commit={commit}) ===")
    print(f"Đã có: {len(existed)}")
    for x in existed:
        print("  ✅", x)
    print(f"{'Đã tạo' if commit else 'Sẽ tạo'}: {len(created)}")
    for x in created:
        print("  ➕", x)


def wipe_imported(commit=False, prefixes=("ISSUE-", "CARE-")):
    """
    Xoá các CRM Issue đã import (theo prefix issue_code) để nhập lại sạch.
    Dùng delete_doc -> dọn luôn child table (issue_students/departments/guardians/process_logs).
    commit=False: chỉ liệt kê. CHỈ DÙNG TRÊN MÔI TRƯỜNG TEST.
    """
    names = []
    for p in prefixes:
        names += frappe.get_all("CRM Issue", filters={"issue_code": ["like", f"{p}%"]}, pluck="name")
    names = sorted(set(names))
    print(f"\n=== wipe_imported (commit={commit}) — khớp {len(names)} issue (prefix {prefixes}) ===")
    if commit:
        for nm in names:
            frappe.delete_doc("CRM Issue", nm, force=True, ignore_permissions=True)
        frappe.db.commit()
        print(f"Đã xoá {len(names)}.")
    else:
        for nm in names[:20]:
            print("  -", nm)
        if len(names) > 20:
            print(f"  ... và {len(names) - 20} nữa")
        print("(commit=False — chưa xoá gì)")
    return {"matched": len(names), "deleted": len(names) if commit else 0}


# ----------------------------------------------------------------------------- import

def run(path, commit=False, limit=None, default_date=None):
    """
    Import sheet Data -> CRM Issue.
    commit=False (mặc định): chỉ resolve + báo cáo, KHÔNG ghi CSDL (dry-run).
    commit=True: insert thật + commit.
    limit: giới hạn số dòng (để test nhanh).
    default_date: 'YYYY-MM-DD' dùng cho dòng thiếu ngày tiếp nhận (mặc định None = bỏ dòng).
    """
    wb, ws = _open_data_sheet(path)
    colmap, headers = _build_colmap(ws)

    required = ["status", "received_date", "pic_email", "module", "code", "priority",
               "related_dept", "content", "student_code", "satisfaction"]
    missing_cols = [k for k in required if k not in colmap]
    if missing_cols:
        print(f"⚠️ Không tìm thấy cột cho khoá: {missing_cols}. Header={headers}")

    report = [f"IMPORT CRM ISSUE — commit={commit} — file={path}", "=" * 70]
    created, skipped, warnings = [], [], []
    seen_codes = set()
    n = 0

    for ridx, row in enumerate(ws.iter_rows(min_row=HEADER_ROW + 1, values_only=True), start=HEADER_ROW + 1):
        if _row_is_empty(row, colmap):
            continue
        if limit and n >= limit:
            break
        n += 1

        code = _txt(_cell(row, colmap, "code"))
        rowlbl = f"row {ridx} (code={code or '∅'})"

        # --- validate bắt buộc & trùng mã ---
        occurred = parse_date(_cell(row, colmap, "received_date"))
        if not occurred:
            if default_date:
                occurred = default_date
                warnings.append(f"{rowlbl}: ngày tiếp nhận trống -> dùng default_date {default_date}.")
            else:
                skipped.append(f"{rowlbl}: BỎ — ngày tiếp nhận trống/không parse được "
                               f"({_cell(row, colmap, 'received_date')!r})")
                continue

        module_name = _resolve_module(_cell(row, colmap, "module"))
        if not module_name:
            skipped.append(f"{rowlbl}: BỎ — Loại vấn đề không khớp "
                           f"({_txt(_cell(row, colmap, 'module'))!r}). Tạo module trước (ensure_modules).")
            continue

        if code:
            new_code, renamed = _unique_issue_code(code, seen_codes)
            if renamed:
                warnings.append(f"{rowlbl}: trùng mã -> đổi thành {new_code}")
            code = new_code
            seen_codes.add(code)

        # --- giá trị ---
        priority = _map_priority(_cell(row, colmap, "priority"))
        status = _map_status(_cell(row, colmap, "status"))
        result, status = _apply_satisfaction(_cell(row, colmap, "satisfaction"), status)

        content_html = _to_html(_cell(row, colmap, "content"))
        if not content_html:
            content_html = "<p>(Không có nội dung)</p>"
            warnings.append(f"{rowlbl}: nội dung trống -> dùng placeholder.")

        # --- resolve links ---
        student_ids, stu_missing = _resolve_students(
            _cell(row, colmap, "student_code"), _cell(row, colmap, "student_name"))
        if stu_missing:
            warnings.append(f"{rowlbl}: HS không khớp -> để trống: {stu_missing}")

        dept_ids, dept_missing = _resolve_departments(_cell(row, colmap, "related_dept"))
        if dept_missing:
            warnings.append(f"{rowlbl}: Phòng ban không khớp -> để trống: {dept_missing}")

        pic_email = _txt(_cell(row, colmap, "pic_email"))
        pic = ""
        if pic_email:
            if frappe.db.exists("User", pic_email):
                pic = pic_email
            else:
                warnings.append(f"{rowlbl}: PIC user không tồn tại -> để trống: {pic_email!r}")

        guardian_ids, guardian_missing = _resolve_guardians(
            _cell(row, colmap, "parent_phone"), _cell(row, colmap, "parent_name"))
        if guardian_missing:
            warnings.append(f"{rowlbl}: PHHS không khớp -> để trống: {guardian_missing}")

        submitter = _txt(_cell(row, colmap, "submitter_name"))

        if not commit:
            created.append(f"{rowlbl}: OK (module={module_name}, prio={priority}, status={status}, "
                           f"result={result or '∅'}, students={len(student_ids)}, depts={len(dept_ids)}, "
                           f"guardians={len(guardian_ids)}, pic={'✓' if pic else '∅'}) [DRY-RUN, chưa ghi]")
            continue

        # --- build & insert ---
        try:
            doc = frappe.new_doc("CRM Issue")
            doc.title = (submitter or _txt(_cell(row, colmap, "student_name")) or module_name)[:140]
            doc.content = content_html
            doc.issue_module = module_name
            doc.issue_group = "Sự vụ"
            doc.issue_code = code or None  # nếu trống để autoname/quy ước hệ thống xử lý
            doc.occurred_at = occurred
            doc.priority = priority
            doc.status = status
            doc.result = result
            doc.approval_status = "Da duyet"
            doc.created_by_user = pic  # Người tạo = PIC (theo yêu cầu)
            doc.pic = pic

            link = _txt(_cell(row, colmap, "link"))
            if link.lower().startswith("http"):
                doc.attachment = link

            _sync_issue_students(doc, {"students": student_ids})
            _set_issue_departments(doc, dept_ids)
            _sync_issue_guardians(doc, {"guardians": guardian_ids})

            if student_ids:
                campus = frappe.db.get_value("CRM Student", student_ids[0], "campus_id")
                if campus:
                    doc.campus_id = campus

            # process logs từ các cột xử lý
            for title, key in (("Thông tin Xử lý của Bộ phận tiếp nhận", "handled_school"),
                               ("Kết quả Xử lý của Bộ phận liên quan", "handled_dept"),
                               ("Note của TS/Quản lý", "ts_note")):
                val = _txt(_cell(row, colmap, key))
                if val:
                    doc.append("process_logs", {"title": title, "content": _to_html(val)})

            doc.flags.ignore_permissions = True
            doc.insert(ignore_permissions=True)

            # Ngày yêu cầu (frontend = creation) phải = ngày tiếp nhận -> set creation = occurred_at
            frappe.db.set_value("CRM Issue", doc.name, "creation", occurred,
                                update_modified=False)

            created.append(f"{rowlbl}: TẠO {doc.name} (issue_code={doc.issue_code})")
        except Exception as e:
            skipped.append(f"{rowlbl}: LỖI insert — {e}")

    wb.close()
    if commit:
        frappe.db.commit()

    # --- báo cáo ---
    report.append(f"\nĐọc {n} dòng | {'TẠO' if commit else 'sẽ tạo'}: {len(created)} | "
                  f"BỎ: {len(skipped)} | cảnh báo: {len(warnings)}")
    report.append("\n----- THÀNH CÔNG / DỰ KIẾN -----")
    report.extend(created)
    report.append("\n----- BỎ DÒNG -----")
    report.extend(skipped or ["(không có)"])
    report.append("\n----- CẢNH BÁO (vẫn tạo, link để trống) -----")
    report.extend(warnings or ["(không có)"])

    _write_report(report)
    print(f"\n✅ Xong. commit={commit} | đọc {n} | {'tạo' if commit else 'dự kiến'} {len(created)} "
          f"| bỏ {len(skipped)} | cảnh báo {len(warnings)}")
    return {"read": n, "created": len(created), "skipped": len(skipped), "warnings": len(warnings)}
