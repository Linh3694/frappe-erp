#!/usr/bin/env python3
"""
Import ngân sách năm học 2025-2026 từ file Excel tổng hợp của Phòng Tài chính.

ĐÂY LÀ SCRIPT ONE-OFF (không phải tính năng import trên UI, không whitelist API).
Dùng Frappe ORM (frappe.get_doc/insert/save) — KHÔNG raw SQL — để giữ autoname,
child table, NestedSet (lft/rgt) và validation của ERP Budget Code.

Nguồn: sheet "Sheet1", header ở row 8-9, data từ row 10.
  - Cột E "Mã ngân sách mới"  -> ERP Budget Code.budget_code   (KHÔNG dùng mã cũ ở cột B/D)
  - Cột F "Khoản mục/Items"   -> ERP Budget Code.account_item
  - Các cột "Số thực hiện"    -> ERP Budget Settlement.settlement_amount (năm học 2025-2026)
  - Các cột "Số kế hoạch"     -> BỎ QUA (đây là số thực hiện, không còn là plan)

Cây mã: cha = mã có tiền tố dài nhất đang tồn tại trong file (vd S11011 -> S1101 -> S11 -> S1).
Tối đa 4 cấp theo ERP Budget Code.MAX_LEVEL.

CHỈ DÒNG LÁ mới được ghi số tiền — dòng nhóm trong Excel là dòng cộng, ghi cả hai sẽ
nhân đôi. Script tự kiểm chứng: tổng các dòng lá phải khớp dòng tổng (row 10) của từng phòng.

Cách chạy:

    # 0) Pre-flight: đọc file, dựng cây, đối chiếu phòng ban + năm học với CSDL (KHÔNG ghi gì)
    bench --site admin.sis.localhost execute \
        erp.scripts.import_budget_2526_from_excel.check_master \
        --kwargs "{'path': '/Users/admin/Downloads/Import budget 24.07.26.xlsx'}"

    # 1) Import mã ngân sách — dry-run trước
    bench --site admin.sis.localhost execute \
        erp.scripts.import_budget_2526_from_excel.import_codes \
        --kwargs "{'path': '.../Import budget 24.07.26.xlsx', 'commit': False}"

    # 2) Import số thực hiện — dry-run trước
    bench --site admin.sis.localhost execute \
        erp.scripts.import_budget_2526_from_excel.import_settlements \
        --kwargs "{'path': '.../Import budget 24.07.26.xlsx', 'commit': False}"

    # 3) Chạy cả hai (mã trước, số sau) — đổi commit=True để ghi thật
    bench --site admin.sis.localhost execute \
        erp.scripts.import_budget_2526_from_excel.run \
        --kwargs "{'path': '.../Import budget 24.07.26.xlsx', 'commit': True}"

Hoặc trong `bench --site ... console`:
    from erp.scripts.import_budget_2526_from_excel import run, check_master
    check_master(path="...")
    run(path="...", commit=False)

Phòng ban nhận diện theo MÃ ĐƠN VỊ (unit_code) ở dòng 8 — cùng chuẩn với chức năng
import mã ngân sách trên UI. check_master() in ra unit_code có thật để đối chiếu.

Tham số hay dùng:
    school_year     : name/tiêu đề SIS School Year 2025-2026. Bỏ trống = tự dò.
    set_departments : True (mặc định) = gán "Phòng ban áp dụng" cho mã CẤP 4 theo đúng
                      các phòng có phát sinh số liệu trong file. False = không đụng tới.
    strict          : True (mặc định) = DỪNG nếu file còn mã rỗng/#N/A/0, mã lặp dòng,
                      hoặc mã thiếu cấp cha. Đặt False để cố tình bỏ qua.
"""

import os

import frappe

# ----------------------------------------------------------------------------- config

DATA_SHEET = "Sheet1"
HEADER_TOP_ROW = 8        # dòng tên phòng ban (có ô gộp)
HEADER_SUB_ROW = 9        # dòng "Số kế hoạch" / "Số thực hiện"
FIRST_DATA_ROW = 10
TOTAL_ROW = 10            # dòng tổng toàn bộ — dùng để đối chiếu, KHÔNG import

COL_CODE = 5              # cột E "Mã ngân sách mới"
COL_ITEM = 6              # cột F "Khoản mục/Items"
COL_AMOUNT_FROM = 8       # cột H — bắt đầu vùng số tiền
COL_AMOUNT_TO = 20        # cột T — hết vùng số tiền

ACTUAL_LABEL = "số thực hiện"
PLAN_LABEL = "số kế hoạch"

SCHOOL_YEAR_HINTS = ("2025-2026", "2025 - 2026", "2025–2026", "2025/2026")

REPORT_FILE = "import_budget_2526_report.txt"

CODE_DT = "ERP Budget Code"
SETTLEMENT_DT = "ERP Budget Settlement"
ORG_UNIT_DT = "ERP Organization Unit"
SCHOOL_YEAR_DT = "SIS School Year"

MAX_LEVEL = 4             # khớp ERP Budget Code.MAX_LEVEL

# Giá trị ở cột mã KHÔNG phải mã thật (lỗi VLOOKUP / ô rác) -> bỏ dòng.
BAD_CODES = {"#n/a", "#N/A", "0", "-", "na", "n/a"}

# Phòng ban nhận diện theo MÃ ĐƠN VỊ (unit_code) trên Sơ đồ tổ chức — cùng chuẩn với
# chức năng import mã ngân sách trên UI (erp/api/erp_sis/budget/budget_code_import.py).
#
# File chuẩn ghi thẳng unit_code ở dòng 8 (vd "TH", "THCS", "BSK"...). Map dưới đây chỉ để
# tương thích với file đang ghi TÊN phòng; nhãn nào không có trong map thì script thử dùng
# chính nhãn đó làm unit_code. check_master() in ra toàn bộ unit_code có thật để đối chiếu.
DEPT_LABEL_TO_UNIT_CODE = {
    "tiểu học": "TH",
    "thcs": "THCS",
    "thpt": "THPT",
    "ban sự kiện": "BSK",
    "bsk": "BSK",
    "ban đào tạo": "BDT",
    "bđt": "BDT",
    "trung tâm pttn": "PTTN",
    "pttn": "PTTN",
    "pths": "PTHS",
}


# ----------------------------------------------------------------------------- helpers

def _norm(s):
    """Chuẩn hoá nhãn: bỏ khoảng trắng thừa + hạ chữ thường."""
    if s is None:
        return ""
    return " ".join(str(s).split()).strip().lower()


def _txt(v):
    """Ô Excel -> chuỗi; giữ số nguyên không đuôi .0 (mã ngân sách có thể bị đọc thành số)."""
    if v is None:
        return ""
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v).strip()
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    return "" if s.lower() == "nan" else s


def _num(v):
    """Ô Excel -> float; ô chữ/rỗng -> 0."""
    if isinstance(v, bool):
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    return 0.0


def _files_dirs(sub):
    """Thư mục files khả dĩ (anchor cả tương đối lẫn tuyệt đối) để né lệch cwd."""
    dirs = []
    try:
        dirs.append(frappe.get_site_path(sub, "files"))
    except Exception:
        pass
    try:
        site = getattr(frappe.local, "site", None)
        bench = frappe.utils.get_bench_path()
        if site and bench:
            dirs.append(os.path.join(bench, "sites", site, sub, "files"))
    except Exception:
        pass
    return dirs


def _resolve_path(path):
    """
    Cho phép truyền URL file Frappe ('/private/files/x.xlsx', '/files/x.xlsx') hoặc path
    filesystem. Bền với lệch cwd, chuẩn hoá Unicode (NFC/NFD) và %20.
    """
    import glob
    import unicodedata
    from urllib.parse import unquote

    # 1) path filesystem trực tiếp
    for c in (path, unquote(path)):
        if c and os.path.isfile(c):
            return c

    # 2) URL file Frappe -> thư mục files của site
    base = os.path.basename(unquote(path or ""))
    subs = ("private", "public")
    for sub in subs:
        for d in _files_dirs(sub):
            cand = os.path.join(d, base)
            if os.path.isfile(cand):
                return cand
            # lệch NFC/NFD
            for form in ("NFC", "NFD"):
                c2 = os.path.join(d, unicodedata.normalize(form, base))
                if os.path.isfile(c2):
                    return c2

    # 3) tìm theo tiền tố ASCII trong thư mục files (Frappe hay đổi tên file)
    prefix = "".join(ch for ch in base if ch.isascii() and (ch.isalnum() or ch in " _-"))[:12].strip()
    if prefix:
        for sub in subs:
            for d in _files_dirs(sub):
                hits = sorted(glob.glob(os.path.join(d, prefix + "*")))
                if hits:
                    return hits[0]

    raise FileNotFoundError(f"Không tìm thấy file: {path}")


def _open_sheet(path):
    """Mở workbook + trả về sheet dữ liệu (đọc giá trị đã tính, không đọc công thức)."""
    import openpyxl

    real = _resolve_path(path)
    wb = openpyxl.load_workbook(real, data_only=True, read_only=False)
    if DATA_SHEET in wb.sheetnames:
        ws = wb[DATA_SHEET]
    else:
        ws = wb.worksheets[0]
        print(f"⚠️ Không thấy sheet {DATA_SHEET!r}, dùng sheet đầu tiên {ws.title!r}")
    return wb, ws, real


def _write_report(lines):
    text = "\n".join(lines)
    try:
        path = frappe.get_site_path(REPORT_FILE)
    except Exception:
        path = REPORT_FILE
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(text)
        print(f"\n📝 Báo cáo đã ghi: {path}")
    except Exception as e:
        print(f"\n⚠️ Không ghi được file báo cáo ({e}). In ra console:")
        print(text)
    return path


def _fmt(n):
    """Định dạng số tiền kiểu vi-VN cho báo cáo."""
    try:
        return f"{float(n):,.0f}".replace(",", ".")
    except (TypeError, ValueError):
        return str(n)


# ----------------------------------------------------------------------------- đọc file

def _build_amount_columns(ws):
    """
    Dựng map cột số tiền: {col_index: (tên phòng theo Excel, 'plan'|'actual')}.

    Row 8 (tên phòng) không đồng nhất: có phòng dùng ô gộp và bỏ trống cột thứ 2
    (Tiểu học chiếm H8:I8), có phòng lặp lại tên viết tắt ở cột thứ 2 ("Ban sự kiện" | "BSK").
    Row 9 phân biệt Số kế hoạch / Số thực hiện.

    Quy tắc gom nhóm — một phòng = 1 cột kế hoạch + 1 cột thực hiện liền nhau:
      - ô row 8 trống            -> nối tiếp phòng đang xét
      - ô row 8 có chữ, cột trước là 'Số kế hoạch' -> vẫn là phòng đó (chữ viết tắt)
      - ô row 8 có chữ, cột trước là 'Số thực hiện' (hoặc chưa có gì) -> PHÒNG MỚI
    Nhờ vậy cột T "PTHS" (chỉ có Số thực hiện) được nhận là phòng riêng, còn "BSK"/"BĐT"
    thì gộp vào đúng "Ban sự kiện"/"Ban đào tạo".
    """
    cols = {}
    current_dept = ""
    prev_kind = None
    for c in range(COL_AMOUNT_FROM, COL_AMOUNT_TO + 1):
        sub = _norm(ws.cell(HEADER_SUB_ROW, c).value)
        if sub == ACTUAL_LABEL:
            kind = "actual"
        elif sub == PLAN_LABEL:
            kind = "plan"
        else:
            continue

        top = _txt(ws.cell(HEADER_TOP_ROW, c).value)
        if top and prev_kind != "plan":
            current_dept = top          # phòng mới
        elif not current_dept:
            current_dept = top          # phòng đầu tiên

        if current_dept:
            cols[c] = (current_dept, kind)
        prev_kind = kind
    return cols


def _read_rows(ws, amount_cols):
    """Đọc dòng dữ liệu -> list dict. Bỏ dòng tổng (row 10) và dòng không có mã hợp lệ."""
    rows = []
    dropped = []
    last_row = ws.max_row or FIRST_DATA_ROW

    for r in range(FIRST_DATA_ROW, last_row + 1):
        code = _txt(ws.cell(r, COL_CODE).value)
        item = _txt(ws.cell(r, COL_ITEM).value)
        amounts = {c: _num(ws.cell(r, c).value) for c in amount_cols}
        has_money = any(v for v in amounts.values())

        if r == TOTAL_ROW:
            continue  # dòng tổng toàn bộ — chỉ dùng để đối chiếu
        if not code and not item and not has_money:
            continue  # dòng trống
        if not code or code.lower() in BAD_CODES:
            actual_money = sum(
                v for c, v in amounts.items() if amount_cols[c][1] == "actual"
            )
            dropped.append(
                {
                    "row": r,
                    "code": code,
                    "item": item,
                    "actual": actual_money,
                    "reason": "thiếu mã ngân sách mới" if not code else f"mã không hợp lệ {code!r}",
                }
            )
            continue

        rows.append({"row": r, "code": code, "item": item, "amounts": amounts})

    return rows, dropped


def _build_tree(rows):
    """
    Dựng cây từ tiền tố mã: cha = mã có tiền tố dài nhất đang tồn tại.
    Trả về (codes, order) — codes: {code: {...}}, order: danh sách code sắp theo level tăng dần.
    Mã lặp lại nhiều dòng được gộp: cộng dồn số tiền, lấy khoản mục đầu tiên có giá trị.
    """
    codeset = {r["code"] for r in rows}

    def parent_of(c):
        for length in range(len(c) - 1, 0, -1):
            if c[:length] in codeset:
                return c[:length]
        return None

    codes = {}
    for r in rows:
        c = r["code"]
        entry = codes.get(c)
        if entry is None:
            entry = codes[c] = {
                "code": c,
                "item": r["item"],
                "rows": [],
                "amounts": {},
                "parent": parent_of(c),
            }
        if not entry["item"] and r["item"]:
            entry["item"] = r["item"]
        entry["rows"].append(r["row"])
        for col, val in r["amounts"].items():
            entry["amounts"][col] = entry["amounts"].get(col, 0.0) + val

    # level = số cấp trên + 1
    def level_of(c):
        lvl, cur, seen = 1, c, {c}
        while codes[cur]["parent"]:
            cur = codes[cur]["parent"]
            if cur in seen:
                break
            seen.add(cur)
            lvl += 1
            if lvl > MAX_LEVEL + 2:
                break
        return lvl

    for c in codes:
        codes[c]["level"] = level_of(c)

    parents = {e["parent"] for e in codes.values() if e["parent"]}
    for c, e in codes.items():
        e["is_group"] = c in parents
        e["is_leaf"] = c not in parents

    order = sorted(codes, key=lambda c: (codes[c]["level"], c))
    return codes, order


def _parse(path):
    """Đọc + dựng cây + trả về mọi thứ cần cho cả 2 pha import."""
    wb, ws, real = _open_sheet(path)
    amount_cols = _build_amount_columns(ws)
    actual_cols = {c: d for c, (d, k) in amount_cols.items() if k == "actual"}
    rows, dropped = _read_rows(ws, amount_cols)
    codes, order = _build_tree(rows)

    # Tổng đối chiếu lấy từ dòng tổng (row 10) của file
    totals = {c: _num(ws.cell(TOTAL_ROW, c).value) for c in actual_cols}
    wb.close()

    return {
        "file": real,
        "amount_cols": amount_cols,
        "actual_cols": actual_cols,   # {col: tên phòng theo Excel}
        "rows": rows,
        "dropped": dropped,
        "codes": codes,
        "order": order,
        "totals": totals,
    }


def _reconcile(parsed):
    """
    Kiểm chứng: tổng số thực hiện của các mã LÁ phải khớp dòng tổng của file.
    Lệch => Excel không phải cây cộng dồn thuần -> DỪNG, không import.
    """
    codes, actual_cols, totals = parsed["codes"], parsed["actual_cols"], parsed["totals"]
    result = []
    for col, dept in actual_cols.items():
        leaf_sum = sum(e["amounts"].get(col, 0.0) for e in codes.values() if e["is_leaf"])
        expected = totals.get(col, 0.0)
        result.append(
            {
                "dept": dept,
                "leaf_sum": leaf_sum,
                "expected": expected,
                "diff": leaf_sum - expected,
                "ok": abs(leaf_sum - expected) <= 1.0,
            }
        )
    return result


# ----------------------------------------------------------------------------- CSDL

def _resolve_department(excel_label):
    """
    Nhãn cột Excel -> (docname ERP Organization Unit, cách khớp). (None, lý do) nếu trượt.

    Thứ tự: unit_code == nhãn -> unit_code theo map -> unit_name_vn khớp HỆT.
    KHÔNG khớp mờ (LIKE) — dữ liệu tài chính, thà báo trượt hơn gán sai phòng.
    """
    label = (excel_label or "").strip()
    if not label:
        return None, "nhãn cột trống"

    hit = frappe.db.get_value(ORG_UNIT_DT, {"unit_code": label}, "name")
    if hit:
        return hit, f"unit_code={label}"

    mapped = DEPT_LABEL_TO_UNIT_CODE.get(_norm(label))
    if mapped:
        hit = frappe.db.get_value(ORG_UNIT_DT, {"unit_code": mapped}, "name")
        if hit:
            return hit, f"unit_code={mapped} (qua map)"

    hit = frappe.db.get_value(ORG_UNIT_DT, {"unit_name_vn": label}, "name")
    if hit:
        return hit, f"unit_name_vn={label}"

    tried = f"unit_code={label!r}"
    if mapped:
        tried += f" / {mapped!r}"
    return None, f"không có đơn vị nào khớp ({tried})"


def _resolve_school_year(school_year=None):
    """Tìm SIS School Year 2025-2026. Nhận thẳng docname hoặc dò theo tiêu đề."""
    if school_year and frappe.db.exists(SCHOOL_YEAR_DT, school_year):
        return school_year
    hints = [school_year] if school_year else list(SCHOOL_YEAR_HINTS)
    for h in hints:
        if not h:
            continue
        for field in ("title_vn", "title_en"):
            hit = frappe.db.get_value(SCHOOL_YEAR_DT, {field: h}, "name")
            if hit:
                return hit
        for field in ("title_vn", "title_en"):
            hit = frappe.db.get_value(
                SCHOOL_YEAR_DT, {field: ("like", f"%{h}%")}, "name"
            )
            if hit:
                return hit
    return None


def _dept_map(parsed):
    """{col: (nhãn Excel, docname phòng ban|None, cách khớp)} cho các cột Số thực hiện."""
    out = {}
    for col, label in parsed["actual_cols"].items():
        dept, how = _resolve_department(label)
        out[col] = (label, dept, how)
    return out


def _org_unit_code_catalog(limit=60):
    """Danh sách unit_code có thật — in ra khi nhãn cột không khớp để đối chiếu."""
    rows = frappe.get_all(
        ORG_UNIT_DT,
        filters={"is_active": 1},
        fields=["unit_code", "unit_name_vn"],
        order_by="unit_code asc",
        limit=limit,
    )
    return [r for r in rows if getattr(r, "unit_code", None)]


# ----------------------------------------------------------------------------- pre-flight

def check_master(path, school_year=None):
    """Đọc file + đối chiếu với CSDL. KHÔNG ghi gì. Chạy cái này trước tiên."""
    parsed = _parse(path)
    codes, dropped = parsed["codes"], parsed["dropped"]
    rep = [f"PRE-FLIGHT IMPORT BUDGET 2025-2026 — file={parsed['file']}", "=" * 78]

    # --- cột số tiền nhận diện được ---
    rep.append("\n----- CỘT SỐ TIỀN NHẬN DIỆN ĐƯỢC -----")
    for col, (dept, kind) in sorted(parsed["amount_cols"].items()):
        mark = "→ IMPORT" if kind == "actual" else "   (bỏ qua)"
        rep.append(f"  col{col:>3} {dept:24} {kind:7} {mark}")

    # --- cây mã ---
    by_level = {}
    for e in codes.values():
        by_level[e["level"]] = by_level.get(e["level"], 0) + 1
    leaves = [e for e in codes.values() if e["is_leaf"]]
    rep.append("\n----- CÂY MÃ NGÂN SÁCH (cột E 'Mã ngân sách mới') -----")
    rep.append(f"  Tổng số mã duy nhất : {len(codes)}")
    rep.append("  Theo cấp            : " + ", ".join(f"cấp {k}={v}" for k, v in sorted(by_level.items())))
    rep.append(f"  Mã nhóm / mã lá     : {len(codes) - len(leaves)} / {len(leaves)}")
    over = [c for c, e in codes.items() if e["level"] > MAX_LEVEL]
    rep.append(f"  Mã vượt {MAX_LEVEL} cấp      : {len(over)} {sorted(over)[:10] if over else ''}")

    dup = {c: e["rows"] for c, e in codes.items() if len(e["rows"]) > 1}
    rep.append(f"\n  {'⛔' if dup else 'OK'} Mã xuất hiện nhiều dòng (sẽ CỘNG DỒN số tiền): {len(dup)}")
    for c, rws in sorted(dup.items()):
        rep.append(f"    {c:10} dòng {rws} — {codes[c]['item'][:50]}")

    orphan = [c for c, e in codes.items() if not e["parent"] and len(c) > 3]
    rep.append(f"\n  {'⛔' if orphan else 'OK'} Mã thiếu cấp cha (sẽ thành mã gốc cấp 1): {len(orphan)}")
    for c in sorted(orphan):
        rep.append(f"    {c:10} cấp{codes[c]['level']} — {codes[c]['item'][:50]}")

    lvl4_leaves = [e for e in leaves if e["level"] == MAX_LEVEL]
    rep.append(
        f"\n  Mã lá ở cấp {MAX_LEVEL} (gán được 'Phòng ban áp dụng'): {len(lvl4_leaves)}"
        f" / {len(leaves)} mã lá"
    )

    # --- dòng bị bỏ ---
    rep.append(f"\n----- {'⛔' if dropped else 'OK'} DÒNG BỊ BỎ (mã không dùng được) -----")
    if not dropped:
        rep.append("  (không có)")
    for d in dropped:
        money = f" — số thực hiện {_fmt(d['actual'])}" if d["actual"] else ""
        rep.append(f"  row{d['row']:>4} [{d['reason']}] {d['item'][:52]}{money}")
    lost = sum(d["actual"] for d in dropped)
    rep.append(f"  Tổng số thực hiện nằm trên dòng bị bỏ: {_fmt(lost)}")
    rep.append("  (Lưu ý: phần lớn là DÒNG CỘNG NHÓM — xem đối chiếu bên dưới để biết có mất tiền thật không.)")

    # --- đối chiếu tổng ---
    rep.append("\n----- ĐỐI CHIẾU: Σ dòng lá vs dòng tổng của file (row 10) -----")
    rec = _reconcile(parsed)
    all_ok = all(r["ok"] for r in rec)
    for r in rec:
        flag = "OK " if r["ok"] else "LỆCH"
        rep.append(
            f"  {flag} {r['dept']:24} Σlá={_fmt(r['leaf_sum']):>18}"
            f"  file={_fmt(r['expected']):>18}  chênh={_fmt(r['diff']):>14}"
        )
    rep.append(f"  => {'ĐẠT — import theo dòng lá là đủ và đúng.' if all_ok else 'KHÔNG ĐẠT — DỪNG, kiểm tra lại file.'}")

    # --- CSDL ---
    rep.append("\n----- ĐỐI CHIẾU CSDL -----")
    sy = _resolve_school_year(school_year)
    if sy:
        title = frappe.db.get_value(SCHOOL_YEAR_DT, sy, "title_vn") or sy
        rep.append(f"  Năm học 2025-2026 -> {sy} ({title})")
    else:
        rep.append("  ⛔ KHÔNG tìm thấy SIS School Year 2025-2026 — truyền tham số school_year='<name>'")
        existing = frappe.get_all(
            SCHOOL_YEAR_DT, fields=["name", "title_vn"], limit=20, order_by="start_date desc"
        )
        for x in existing:
            rep.append(f"       có sẵn: {x.name} — {x.title_vn}")

    rep.append("\n  Phòng ban (khớp theo mã đơn vị / unit_code):")
    dmap = _dept_map(parsed)
    unresolved = []
    for col, (label, dept, how) in sorted(dmap.items()):
        if dept:
            nm = frappe.db.get_value(ORG_UNIT_DT, dept, "unit_name_vn")
            rep.append(f"    OK   {label:24} -> {dept} ({nm}) [{how}]")
        else:
            unresolved.append(label)
            rep.append(f"    ⛔   {label:24} -> {how}")
    if unresolved:
        rep.append(
            "\n    Cách xử lý: ghi đúng unit_code vào dòng 8 của file, HOẶC bổ sung"
            " DEPT_LABEL_TO_UNIT_CODE trong script."
        )
        rep.append("    unit_code đang có trên Sơ đồ tổ chức:")
        for r in _org_unit_code_catalog():
            rep.append(f"      {r.unit_code:12} — {r.unit_name_vn}")

    rep.append("\n  Mã ngân sách đã có sẵn trong hệ thống:")
    existing_codes = set(frappe.get_all(CODE_DT, pluck="budget_code"))
    hit = sorted(set(codes) & existing_codes)
    rep.append(f"    Đã tồn tại (sẽ CẬP NHẬT): {len(hit)}")
    rep.append(f"    Tạo mới                 : {len(set(codes) - existing_codes)}")
    rep.append(f"    Tổng mã đang có trong hệ thống: {len(existing_codes)}")

    # --- số dòng settlement dự kiến ---
    n_settle = sum(
        1
        for e in codes.values()
        if e["is_leaf"]
        for col in parsed["actual_cols"]
        if e["amounts"].get(col, 0.0)
    )
    rep.append(f"\n  Số dòng ERP Budget Settlement dự kiến ghi: {n_settle}")

    blockers = []
    if not all_ok:
        blockers.append("tổng dòng lá không khớp dòng tổng của file")
    if not sy:
        blockers.append("chưa xác định được năm học 2025-2026")
    if unresolved:
        blockers.append(f"{len(unresolved)} cột phòng ban chưa khớp unit_code")
    if dropped:
        blockers.append(f"{len(dropped)} dòng có mã rỗng/#N/A/0")
    if dup:
        blockers.append(f"{len(dup)} mã bị lặp dòng")
    if orphan:
        blockers.append(f"{len(orphan)} mã thiếu cấp cha")
    if over:
        blockers.append(f"{len(over)} mã vượt {MAX_LEVEL} cấp")

    ready = not blockers
    rep.append("\n" + "=" * 78)
    if ready:
        rep.append("SẴN SÀNG IMPORT")
    else:
        rep.append("CHƯA SẴN SÀNG — còn " + str(len(blockers)) + " vấn đề:")
        for b in blockers:
            rep.append(f"  ⛔ {b}")
        rep.append(
            "\nGhi chú: mặc định strict=True nên import sẽ DỪNG khi còn các vấn đề trên."
            " Truyền strict=False nếu cố tình chấp nhận (dòng lỗi bị bỏ, mã lặp bị cộng dồn,"
            " mã thiếu cấp cha thành mã gốc cấp 1)."
        )

    _write_report(rep)
    print("\n".join(rep[-45:]))
    return {
        "codes": len(codes),
        "leaves": len(leaves),
        "dropped": len(dropped),
        "duplicates": len(dup),
        "orphans": len(orphan),
        "reconcile_ok": all_ok,
        "school_year": sy,
        "unresolved_departments": unresolved,
        "blockers": blockers,
        "ready": bool(ready),
    }


def _strict_issues(parsed):
    """Các lỗi dữ liệu mà file chuẩn không được có (dùng cho strict=True)."""
    codes = parsed["codes"]
    issues = []
    if parsed["dropped"]:
        rows = ", ".join(str(d["row"]) for d in parsed["dropped"][:15])
        issues.append(f"{len(parsed['dropped'])} dòng mã rỗng/#N/A/0 (dòng {rows}...)")
    dup = [c for c, e in codes.items() if len(e["rows"]) > 1]
    if dup:
        issues.append(f"{len(dup)} mã lặp dòng: {sorted(dup)[:10]}")
    orphan = [c for c, e in codes.items() if not e["parent"] and len(c) > 3]
    if orphan:
        issues.append(f"{len(orphan)} mã thiếu cấp cha: {sorted(orphan)[:10]}")
    return issues


# ----------------------------------------------------------------------------- pha 1: mã

def import_codes(path, commit=False, set_departments=True, strict=True):
    """
    Pha 1 — tạo/cập nhật cây ERP Budget Code từ cột E.

    set_departments=True: gán 'Phòng ban áp dụng' cho mã CẤP 4 theo đúng các phòng có
    phát sinh số thực hiện trong file (ERP Budget Code chỉ cho gán ở cấp 4).
    strict=True: dừng nếu file còn mã rỗng/#N/A/0, mã lặp dòng, hoặc mã thiếu cấp cha.
    """
    parsed = _parse(path)
    codes, order = parsed["codes"], parsed["order"]

    rec = _reconcile(parsed)
    if not all(r["ok"] for r in rec):
        print("⛔ Đối chiếu tổng KHÔNG khớp — chạy check_master() và xử lý trước.")
        return {"aborted": "reconcile_failed"}

    if strict:
        issues = _strict_issues(parsed)
        if issues:
            print("⛔ File chưa sạch — dừng (strict=True):")
            for i in issues:
                print(f"   - {i}")
            print("   Sửa file rồi chạy lại, hoặc truyền strict=False để chấp nhận.")
            return {"aborted": "strict_data_issues", "issues": issues}

    dmap = _dept_map(parsed) if set_departments else {}
    missing_dept = [lbl for lbl, d, _how in dmap.values() if not d]
    if set_departments and missing_dept:
        print(f"⛔ Chưa khớp phòng ban: {missing_dept}. Chạy check_master() trước.")
        return {"aborted": "departments_unresolved"}

    rep = [
        f"IMPORT MÃ NGÂN SÁCH — commit={commit} set_departments={set_departments}",
        f"file={parsed['file']}",
        "=" * 78,
    ]
    created, updated, failed = [], [], []
    # Dry-run không ghi CSDL nên mã cha chưa tồn tại khi tới lượt mã con.
    # Ghi nhận mã "sẽ tạo" để vẫn kiểm tra được tính toàn vẹn của cây.
    would_create = set()

    for code in order:                      # đã sắp theo level: cha trước con
        e = codes[code]
        if e["level"] > MAX_LEVEL:
            failed.append(f"{code}: vượt {MAX_LEVEL} cấp (cấp {e['level']}) — bỏ")
            continue

        # phòng ban áp dụng: chỉ mã LÁ cấp 4, theo phòng có số liệu
        depts = []
        if set_departments and e["level"] == MAX_LEVEL:
            for col, (_lbl, dept, _how) in dmap.items():
                if dept and e["amounts"].get(col, 0.0):
                    if dept not in depts:
                        depts.append(dept)

        try:
            existing = frappe.db.get_value(CODE_DT, {"budget_code": code}, "name")
            if existing:
                doc = frappe.get_doc(CODE_DT, existing)
                action = "CẬP NHẬT"
            else:
                doc = frappe.new_doc(CODE_DT)
                action = "TẠO"

            doc.budget_code = code
            doc.account_item = e["item"] or doc.get("account_item") or ""
            doc.is_active = 1
            if e["parent"]:
                parent_name = frappe.db.get_value(
                    CODE_DT, {"budget_code": e["parent"]}, "name"
                )
                if not parent_name and not commit and e["parent"] in would_create:
                    parent_name = f"(sẽ tạo: {e['parent']})"
                if not parent_name:
                    failed.append(f"{code}: chưa có mã cha {e['parent']} — bỏ")
                    continue
                doc.parent_budget_code = parent_name
            else:
                doc.parent_budget_code = None

            # ERP Budget Code tự xoá applicable_departments nếu level != 4.
            # Chỉ ghi đè khi file thực sự suy ra được phòng ban — mã không phát sinh số
            # liệu 25-26 thì GIỮ NGUYÊN cấu hình đang có (tránh xoá phần chỉnh tay trên UI).
            if set_departments and e["level"] == MAX_LEVEL and depts:
                doc.set("applicable_departments", [])
                for d in depts:
                    doc.append("applicable_departments", {"department": d})

            if commit:
                doc.save(ignore_permissions=True)
            else:
                would_create.add(code)
            line = f"{action} {code:10} cấp{e['level']} {'nhóm' if e['is_group'] else 'lá  '} — {e['item'][:44]}"
            if depts:
                line += f" | {len(depts)} phòng"
            (created if action == "TẠO" else updated).append(line)
        except Exception as ex:
            failed.append(f"{code}: LỖI — {ex}")

    if commit:
        frappe.db.commit()

    rep.append(
        f"\n{'ĐÃ' if commit else 'SẼ'} tạo: {len(created)} | "
        f"{'ĐÃ' if commit else 'SẼ'} cập nhật: {len(updated)} | lỗi/bỏ: {len(failed)}"
    )
    rep.append("\n----- TẠO MỚI -----")
    rep.extend(created or ["(không có)"])
    rep.append("\n----- CẬP NHẬT -----")
    rep.extend(updated or ["(không có)"])
    rep.append("\n----- LỖI / BỎ -----")
    rep.extend(failed or ["(không có)"])
    _write_report(rep)

    print(
        f"\n✅ Pha 1 xong. commit={commit} | tạo {len(created)} | "
        f"cập nhật {len(updated)} | lỗi {len(failed)}"
    )
    return {"created": len(created), "updated": len(updated), "failed": len(failed)}


# ----------------------------------------------------------------------------- pha 2: tiền

def import_settlements(path, school_year=None, commit=False, strict=True):
    """
    Pha 2 — ghi cột 'Số thực hiện' vào ERP Budget Settlement cho năm học 2025-2026.

    CHỈ ghi mã LÁ (dòng nhóm trong Excel là dòng cộng).
    Idempotent: đã có bản ghi (năm × phòng × mã) thì CẬP NHẬT, dư thì xoá bớt.
    strict=True: dừng nếu file còn mã rỗng/#N/A/0, mã lặp dòng, hoặc mã thiếu cấp cha.
    """
    parsed = _parse(path)
    codes = parsed["codes"]

    rec = _reconcile(parsed)
    if not all(r["ok"] for r in rec):
        print("⛔ Đối chiếu tổng KHÔNG khớp — chạy check_master() và xử lý trước.")
        return {"aborted": "reconcile_failed"}

    if strict:
        issues = _strict_issues(parsed)
        if issues:
            print("⛔ File chưa sạch — dừng (strict=True):")
            for i in issues:
                print(f"   - {i}")
            print("   Sửa file rồi chạy lại, hoặc truyền strict=False để chấp nhận.")
            return {"aborted": "strict_data_issues", "issues": issues}

    sy = _resolve_school_year(school_year)
    if not sy:
        print("⛔ Không tìm thấy SIS School Year 2025-2026 — truyền school_year='<name>'.")
        return {"aborted": "school_year_not_found"}

    dmap = _dept_map(parsed)
    missing = [lbl for lbl, d, _how in dmap.values() if not d]
    if missing:
        print(f"⛔ Chưa khớp phòng ban: {missing}. Chạy check_master() trước.")
        return {"aborted": "departments_unresolved"}

    rep = [
        f"IMPORT SỐ THỰC HIỆN 2025-2026 — commit={commit} — năm học={sy}",
        f"file={parsed['file']}",
        "=" * 78,
    ]
    inserted, changed, skipped, failed = [], [], [], []
    dept_totals = {}

    for code in sorted(codes):
        e = codes[code]
        if not e["is_leaf"]:
            continue

        code_name = frappe.db.get_value(CODE_DT, {"budget_code": code}, "name")
        if not code_name:
            if commit:
                failed.append(f"{code}: chưa có trong ERP Budget Code — chạy pha 1 trước")
                continue
            # Dry-run: pha 1 chưa ghi nên mã chưa tồn tại — vẫn mô phỏng tiếp để
            # báo cáo dry-run phản ánh đúng số dòng sẽ ghi.
            code_name = f"(sẽ tạo ở pha 1: {code})"

        for col, (label, dept, _how) in dmap.items():
            amount = e["amounts"].get(col, 0.0)
            if not amount:
                continue
            dept_totals[label] = dept_totals.get(label, 0.0) + amount

            try:
                existing = frappe.get_all(
                    SETTLEMENT_DT,
                    filters={
                        "school_year_id": sy,
                        "department": dept,
                        "budget_code": code_name,
                    },
                    pluck="name",
                )
                if existing:
                    keep = existing[0]
                    old = frappe.db.get_value(SETTLEMENT_DT, keep, "settlement_amount") or 0
                    if commit:
                        doc = frappe.get_doc(SETTLEMENT_DT, keep)
                        doc.settlement_amount = amount
                        doc.account_item = e["item"] or doc.account_item
                        doc.save(ignore_permissions=True)
                        for dupname in existing[1:]:
                            frappe.delete_doc(
                                SETTLEMENT_DT, dupname, ignore_permissions=True, force=True
                            )
                    if abs(float(old) - amount) > 0.5 or len(existing) > 1:
                        extra = f" (+xoá {len(existing) - 1} bản trùng)" if len(existing) > 1 else ""
                        changed.append(
                            f"{label:22} {code:10} {_fmt(old):>16} -> {_fmt(amount):>16}{extra}"
                        )
                    else:
                        skipped.append(f"{label:22} {code:10} không đổi ({_fmt(amount)})")
                    continue

                if commit:
                    doc = frappe.new_doc(SETTLEMENT_DT)
                    doc.school_year_id = sy
                    doc.department = dept
                    doc.budget_code = code_name
                    doc.account_item = e["item"] or ""
                    doc.settlement_amount = amount
                    doc.campus_id = frappe.db.get_value(ORG_UNIT_DT, dept, "campus_id")
                    doc.insert(ignore_permissions=True)
                inserted.append(f"{label:22} {code:10} {_fmt(amount):>18} — {e['item'][:38]}")
            except Exception as ex:
                failed.append(f"{label} / {code}: LỖI — {ex}")

    if commit:
        frappe.db.commit()

    rep.append("\n----- TỔNG ĐÃ GHI THEO PHÒNG (đối chiếu với dòng tổng của file) -----")
    exp = {parsed["actual_cols"][c]: parsed["totals"].get(c, 0.0) for c in parsed["actual_cols"]}
    for label in sorted(set(list(dept_totals) + list(exp))):
        got, want = dept_totals.get(label, 0.0), exp.get(label, 0.0)
        flag = "OK " if abs(got - want) <= 1.0 else "LỆCH"
        rep.append(f"  {flag} {label:24} ghi={_fmt(got):>18}  file={_fmt(want):>18}")

    rep.append(
        f"\n{'ĐÃ' if commit else 'SẼ'} thêm: {len(inserted)} | "
        f"{'ĐÃ' if commit else 'SẼ'} sửa: {len(changed)} | không đổi: {len(skipped)} | lỗi: {len(failed)}"
    )
    rep.append("\n----- THÊM MỚI -----")
    rep.extend(inserted or ["(không có)"])
    rep.append("\n----- CẬP NHẬT -----")
    rep.extend(changed or ["(không có)"])
    rep.append("\n----- LỖI -----")
    rep.extend(failed or ["(không có)"])
    _write_report(rep)

    print(
        f"\n✅ Pha 2 xong. commit={commit} | thêm {len(inserted)} | sửa {len(changed)} "
        f"| không đổi {len(skipped)} | lỗi {len(failed)}"
    )
    return {
        "inserted": len(inserted),
        "changed": len(changed),
        "unchanged": len(skipped),
        "failed": len(failed),
        "school_year": sy,
    }


# ----------------------------------------------------------------------------- chạy cả 2

def run(path, school_year=None, commit=False, set_departments=True, strict=True):
    """Chạy pha 1 (mã ngân sách) rồi pha 2 (số thực hiện). Dry-run mặc định."""
    print("=" * 78)
    print("PHA 1 — MÃ NGÂN SÁCH")
    print("=" * 78)
    r1 = import_codes(path, commit=commit, set_departments=set_departments, strict=strict)
    if r1.get("aborted"):
        return {"phase1": r1, "phase2": {"aborted": "phase1_failed"}}

    if not commit:
        print("\nℹ️ Dry-run — chưa ghi gì vào CSDL. Chạy lại với commit=True để import thật.")

    print("\n" + "=" * 78)
    print("PHA 2 — SỐ THỰC HIỆN 2025-2026")
    print("=" * 78)
    r2 = import_settlements(path, school_year=school_year, commit=commit, strict=strict)
    return {"phase1": r1, "phase2": r2}
